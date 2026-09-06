import os
import sys
import logging
from datetime import datetime
from pathlib import Path

import requests
import pandas as pd
import yfinance as yf
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════════════════════
# CONFIGURARE
# ══════════════════════════════════════════════════════════════════════════════

EXCEL_PATH = Path(r"C:\\Users\\Marius\\Desktop\\Nu sterge\\Analiza_Piata_Profesionala.xlsx")
FRED_API_KEY = os.environ.get("FRED_API_KEY", "")

INDICI = {
    "S&P 500": "^GSPC", "NASDAQ 100": "^NDX", "NASDAQ Comp.": "^IXIC",
    "Dow Jones": "^DJI", "Russell 2000": "^RUT", "DAX Germany": "^GDAXI",
    "FTSE 100": "^FTSE", "CAC 40": "^FCHI", "Nikkei 225": "^N225",
    "Hang Seng": "^HSI", "Shanghai Comp.": "000001.SS", "MSCI World ETF": "URTH",
    "MSCI EM ETF": "EEM", "BET Romania": "BET.RO",
}

ACTIUNI = {
    "Apple": "AAPL", "Microsoft": "MSFT", "NVIDIA": "NVDA", "Alphabet": "GOOGL",
    "Amazon": "AMZN", "Meta": "META", "Tesla": "TSLA", "AMD": "AMD",
    "Intel": "INTC", "Broadcom": "AVGO", "ASML": "ASML", "Taiwan Semi": "TSM",
    "Palantir": "PLTR", "Salesforce": "CRM", "Oracle": "ORCL", "JPMorgan": "JPM",
    "Goldman Sachs": "GS", "Berkshire B": "BRK-B", "Visa": "V", "Mastercard": "MA",
    "ExxonMobil": "XOM", "Chevron": "CVX", "Shell": "SHEL", "Caterpillar": "CAT",
    "Boeing": "BA", "SPY": "SPY", "QQQ": "QQQ", "GLD ETF": "GLD",
    "TLT Bond ETF": "TLT", "ARKK": "ARKK",
}

CRYPTO = {
    "Bitcoin": "BTC-USD", "Ethereum": "ETH-USD", "BNB": "BNB-USD", "Solana": "SOL-USD",
    "XRP": "XRP-USD", "Cardano": "ADA-USD", "Avalanche": "AVAX-USD", "Dogecoin": "DOGE-USD",
    "Chainlink": "LINK-USD", "Polkadot": "DOT-USD", "Litecoin": "LTC-USD", "Shiba Inu": "SHIB-USD",
    "Polygon": "MATIC-USD", "Uniswap": "UNI-USD", "Cosmos": "ATOM-USD", "Stellar": "XLM-USD",
    "Monero": "XMR-USD", "Tron": "TRX-USD", "Filecoin": "FIL-USD", "Aave": "AAVE-USD",
    "Arbitrum": "ARB-USD", "Optimism": "OP-USD", "Render": "RNDR-USD", "Sui": "SUI-USD",
    "Near Protocol": "NEAR-USD",
}

VALUTE = {
    "EUR/USD": "EURUSD=X", "GBP/USD": "GBPUSD=X", "USD/JPY": "USDJPY=X", "USD/CHF": "USDCHF=X",
    "AUD/USD": "AUDUSD=X", "USD/CAD": "USDCAD=X", "EUR/RON": "EURRON=X", "USD/RON": "USDRON=X",
    "GBP/RON": "GBPRON=X", "EUR/GBP": "EURGBP=X", "USD/CNY": "USDCNY=X", "USD/TRY": "USDTRY=X",
}

MATERII_PRIME = {
    "Gold": "GC=F", "Silver": "SI=F", "Platinum": "PL=F", "Palladium": "PA=F",
    "Oil WTI": "CL=F", "Oil Brent": "BZ=F", "Natural Gas": "NG=F", "Copper": "HG=F",
    "Corn": "ZC=F", "Wheat": "ZW=F", "Soybean": "ZS=F", "Coffee": "KC=F",
    "Sugar": "SB=F", "Cotton": "CT=F",
}

ACTIVE = {**INDICI, **ACTIUNI, **CRYPTO, **VALUTE, **MATERII_PRIME}

MACRO_TICKERS = {
    "VIX": "^VIX", "Yield 10Y US": "^TNX", "Yield 2Y US": "^IRX",
    "Yield 30Y US": "^TYX", "USD Index": "DX-Y.NYB",
}

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
    handlers=[logging.StreamHandler(sys.stdout),
              logging.FileHandler(Path(__file__).parent / "actualizare.log", encoding="utf-8")],
)
log = logging.getLogger(__name__)

# ══════════════════════════════════════════════════════════════════════════════
# CULORI
# ══════════════════════════════════════════════════════════════════════════════

def fill(hex_c):
    return PatternFill("solid", fgColor=hex_c)

def fnt(bold=False, color="000000", size=10):
    return Font(name="Arial", bold=bold, color=color, size=size)

F_GRN = fill("C6EFCE"); FN_GRN = fnt(color="375623")
F_YLW = fill("FFEB9C"); FN_YLW = fnt(color="9C6500")
F_RED = fill("FFC7CE"); FN_RED = fnt(color="9C0006")
F_HDR = fill("1F4E79"); FN_HDR = fnt(bold=True, color="FFFFFF")
F_GRY = fill("F2F2F2"); FN_GRY = fnt(color="595959")
NO_FILL = PatternFill(fill_type=None)

# ══════════════════════════════════════════════════════════════════════════════
# INDICATORI TEHNICI
# ══════════════════════════════════════════════════════════════════════════════

def calc_rsi(prices, period=14):
    delta = prices.diff()
    gain = delta.clip(lower=0).rolling(period).mean()
    loss = (-delta.clip(upper=0)).rolling(period).mean()
    rs = gain / loss.replace(0, 1e-10)
    rsi = 100 - (100 / (1 + rs))
    val = rsi.iloc[-1]
    return round(float(val), 2) if pd.notna(val) else 50.0

def calc_macd(prices):
    ema12 = prices.ewm(span=12, adjust=False).mean()
    ema26 = prices.ewm(span=26, adjust=False).mean()
    ml = ema12 - ema26
    sl = ml.ewm(span=9, adjust=False).mean()
    hist = ml - sl
    m, s, h = float(ml.iloc[-1]), float(sl.iloc[-1]), float(hist.iloc[-1])
    prev_h = float(hist.iloc[-2]) if len(hist) > 1 else 0
    if m > s:
        cross = "Bullish Cross" if prev_h < 0 else "Pozitiv"
    else:
        cross = "Bearish Cross" if prev_h > 0 else "Negativ"
    return {"macd": round(m, 6), "signal": round(s, 6), "histogram": round(h, 6), "cross": cross}

def calc_ma(prices):
    def ma(n):
        return round(float(prices.tail(n).mean()), 6) if len(prices) >= n else None
    ma20, ma50, ma200 = ma(20), ma(50), ma(200)
    cross = "Neutru"
    if ma50 and ma200:
        cross = "Golden Cross" if ma50 > ma200 else "Death Cross"
    return {"ma20": ma20, "ma50": ma50, "ma200": ma200, "macross": cross}

def calc_bollinger(prices, period=20):
    m = prices.rolling(period).mean()
    std = prices.rolling(period).std()
    sup = float((m + 2 * std).iloc[-1])
    inf = float((m - 2 * std).iloc[-1])
    return {"bb_sup": round(sup, 6), "bb_inf": round(inf, 6), "bb_width": round(sup - inf, 6)}

def calc_atr(hist, period=14):
    hi, lo, cl = hist["High"], hist["Low"], hist["Close"]
    tr = pd.concat([(hi - lo), (hi - cl.shift()).abs(), (lo - cl.shift()).abs()], axis=1).max(axis=1)
    atr = tr.rolling(period).mean().iloc[-1]
    return round(float(atr), 6) if pd.notna(atr) else 0.0

def calc_stochastic(hist, period=14):
    lo14 = hist["Low"].rolling(period).min()
    hi14 = hist["High"].rolling(period).max()
    diff = (hi14 - lo14).replace(0, 1e-10)
    k = (hist["Close"] - lo14) / diff * 100
    d = k.rolling(3).mean()
    return {
        "stoch_k": round(float(k.iloc[-1]), 2) if pd.notna(k.iloc[-1]) else 50.0,
        "stoch_d": round(float(d.iloc[-1]), 2) if pd.notna(d.iloc[-1]) else 50.0,
    }

def calc_signal(rsi, macd_cross, ma_cross, rvol):
    score = 0
    if rsi < 35: score += 2
    elif rsi < 45: score += 1
    elif rsi > 75: score -= 2
    elif rsi > 65: score -= 1
    if "Bullish Cross" in macd_cross: score += 2
    elif "Pozitiv" in macd_cross: score += 1
    elif "Bearish Cross" in macd_cross: score -= 2
    elif "Negativ" in macd_cross: score -= 1
    if ma_cross == "Golden Cross": score += 2
    elif ma_cross == "Death Cross": score -= 2
    if rvol > 1.5: score += 1
    elif rvol < 0.6: score -= 1
    confluente = min(abs(score), 5)
    if score >= 3: return "BUY", confluente
    elif score <= -3: return "SELL", confluente
    else: return "WAIT", confluente

def map_rsi_status(rsi: float) -> str:
    if rsi > 70: return "Extrem peste prag"
    if rsi > 55: return "Puternic peste medie"
    if rsi > 45: return "Neutru"
    if rsi > 30: return "Sub medie"
    return "Extrem sub prag"

# ══════════════════════════════════════════════════════════════════════════════
# PRELUARE DATE
# ══════════════════════════════════════════════════════════════════════════════

def get_full_data(name, ticker):
    try:
        t = yf.Ticker(ticker)
        hist = t.history(period="1y", auto_adjust=True)
        if hist.empty or len(hist) < 5:
            log.warning(f"  ⚠ Date insuficiente: {name} ({ticker})")
            return {}

        closes = hist["Close"]
        latest = hist.iloc[-1]
        prev = hist.iloc[-2]

        close_now = float(latest["Close"])
        close_prev = float(prev["Close"])
        close_5d = float(hist.iloc[-5]["Close"]) if len(hist) >= 5 else close_prev
        close_20d = float(hist.iloc[-20]["Close"]) if len(hist) >= 20 else close_prev

        var_zi = (close_now - close_prev) / close_prev * 100 if close_prev else 0
        var_sapt = (close_now - close_5d) / close_5d * 100 if close_5d else 0
        var_luna = (close_now - close_20d) / close_20d * 100 if close_20d else 0

        volume = int(latest.get("Volume", 0))
        avg_vol = int(hist["Volume"].tail(20).mean()) if "Volume" in hist else 0
        rvol = round(volume / avg_vol, 2) if avg_vol > 0 else 1.0

        rsi = calc_rsi(closes)
        macd = calc_macd(closes)
        ma = calc_ma(closes)
        boll = calc_bollinger(closes)
        atr = calc_atr(hist)
        stoch = calc_stochastic(hist)
        mom10 = round(float(closes.pct_change(10).iloc[-1] * 100), 2) if len(closes) > 10 else 0

        price = close_now
        if ma["ma50"] and price > ma["ma50"] * 1.01: trend = "Bullish"
        elif ma["ma50"] and price < ma["ma50"] * 0.99: trend = "Bearish"
        else: trend = "Sideways"

        rsi_status = map_rsi_status(rsi)
        semnal, confluente = calc_signal(rsi, macd["cross"], ma["macross"], rvol)

        sl = round(price - 1.5 * atr, 6) if semnal == "BUY" else \
             round(price + 1.5 * atr, 6) if semnal == "SELL" else \
             round(price - 2.0 * atr, 6)
        tp = round(price + 3.0 * atr, 6) if semnal == "BUY" else \
             round(price - 3.0 * atr, 6) if semnal == "SELL" else \
             round(price + 2.0 * atr, 6)

        prob = min(90, 35 + confluente * 10 + (5 if rvol > 1.2 else 0))

        return {
            "name": name, "ticker": ticker,
            "data": datetime.now().strftime("%d.%m.%Y"),
            "timestamp": datetime.now().strftime("%d.%m.%Y %H:%M"),
            "deschidere": round(float(latest.get("Open", close_now)), 6),
            "maxim": round(float(latest.get("High", close_now)), 6),
            "minim": round(float(latest.get("Low", close_now)), 6),
            "inchidere": round(close_now, 6),
            "var_zi_pct": round(var_zi, 4),
            "var_sapt_pct": round(var_sapt, 4),
            "var_luna_pct": round(var_luna, 4),
            "volum": volume,
            "avg_vol_20": avg_vol,
            "rvol": rvol,
            "rsi": rsi,
            "rsi_status": rsi_status,
            "macd": macd["macd"],
            "macd_signal": macd["signal"],
            "macd_hist": macd["histogram"],
            "macd_cross": macd["cross"],
            "ma20": ma["ma20"],
            "ma50": ma["ma50"],
            "ma200": ma["ma200"],
            "macross": ma["macross"],
            "bb_sup": boll["bb_sup"],
            "bb_inf": boll["bb_inf"],
            "bb_width": boll["bb_width"],
            "atr": atr,
            "stoch_k": stoch["stoch_k"],
            "stoch_d": stoch["stoch_d"],
            "momentum_10z": mom10,
            "trend": trend,
            "semnal": semnal,
            "confluente": confluente,
            "sl": sl,
            "tp": tp,
            "probabilitate": prob,
        }
    except Exception as e:
        log.error(f"  X Eroare {name} ({ticker}): {e}")
        return {}

def get_fear_greed():
    try:
        r = requests.get("https://api.alternative.me/fng/?limit=1", timeout=10)
        d = r.json()["data"][0]
        val = int(d["value"])
        label = d["value_classification"]
        return {
            "value": val,
            "display": f"{val} - {label}",
            "status": "Pozitiv" if val > 60 else ("Negativ" if val < 40 else "Neutru"),
        }
    except Exception as e:
        log.warning(f"  Fear & Greed indisponibil: {e}")
        return {"value": None, "display": "N/A", "status": "Neutru"}

def get_fred(series_id):
    if not FRED_API_KEY:
        return None
    try:
        url = (f"https://api.stlouisfed.org/fred/series/observations"
               f"?series_id={series_id}&api_key={FRED_API_KEY}"
               f"&file_type=json&sort_order=desc&limit=1")
        obs = requests.get(url, timeout=10).json().get("observations", [])
        if obs and obs[0]["value"] != ".":
            return float(obs[0]["value"])
    except Exception as e:
        log.warning(f"  FRED {series_id}: {e}")
    return None

# ══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def color_pnl(cell, value):
    if value > 0: cell.fill = F_GRN; cell.font = FN_GRN
    elif value < 0: cell.fill = F_RED; cell.font = FN_RED
    else: cell.fill = F_YLW; cell.font = FN_YLW

def color_rsi(cell, rsi):
    if rsi < 30: cell.fill = F_GRN; cell.font = FN_GRN
    elif rsi < 50: cell.fill = F_YLW; cell.font = FN_YLW
    elif rsi < 70: cell.fill = fill("E2EFDA"); cell.font = fnt(color="375623")
    else: cell.fill = F_RED; cell.font = FN_RED

def color_signal(cell, semnal):
    if semnal == "BUY": cell.fill = F_GRN; cell.font = FN_GRN
    elif semnal == "SELL": cell.fill = F_RED; cell.font = FN_RED
    else: cell.fill = F_YLW; cell.font = FN_YLW

def color_trend(cell, trend):
    if trend == "Bullish": cell.fill = F_GRN; cell.font = FN_GRN
    elif trend == "Bearish": cell.fill = F_RED; cell.font = FN_RED
    else: cell.fill = F_YLW; cell.font = FN_YLW

def safe_write(ws, row, col, value):
    """Scrie în celulă doar dacă nu e merged"""
    cell = ws.cell(row, col)
    if not isinstance(cell, type(cell).__bases__[0]):  # Check if MergedCell
        try:
            cell.value = value
            return cell
        except:
            pass
    return None

def clear_rows(ws, start_row, end_row=5000):
    for row in ws.iter_rows(min_row=start_row, max_row=end_row):
        for cell in row:
            try:
                if not isinstance(cell, type(cell).__bases__[0]):
                    cell.value = None
                    cell.fill = NO_FILL
            except:
                pass

def write_category_header(ws, row, label, num_cols):
    c = safe_write(ws, row, 1, label)
    if c:
        c.fill = F_HDR
        c.font = FN_HDR
    for col in range(2, num_cols + 1):
        c = ws.cell(row, col)
        try:
            if not isinstance(c, type(c).__bases__[0]):
                c.fill = F_HDR
        except:
            pass

# ══════════════════════════════════════════════════════════════════════════════
# ACTUALIZARE SHEET-URI
# ══════════════════════════════════════════════════════════════════════════════

CATEGORII = [
    ("INDICI BURSIERI", INDICI),
    ("ACTIUNI & ETF", ACTIUNI),
    ("CRYPTOCURRENCY", CRYPTO),
    ("VALUTE FOREX", VALUTE),
    ("MATERII PRIME", MATERII_PRIME),
]

def update_preturi_volume(ws, all_data):
    clear_rows(ws, 3)
    r = 3
    for cat_name, cat_dict in CATEGORII:
        write_category_header(ws, r, f"--- {cat_name} ---", 15)
        r += 1
        for name, ticker in cat_dict.items():
            d = all_data.get(ticker, {})
            safe_write(ws, r, 1, d.get("data", datetime.now().strftime("%d.%m.%Y")))
            c = safe_write(ws, r, 2, name)
            if c: c.font = fnt(bold=True)
            if d:
                c = safe_write(ws, r, 3, d["deschidere"])
                if c: c.number_format = "#,##0.0000"
                c = safe_write(ws, r, 4, d["maxim"])
                if c: c.number_format = "#,##0.0000"
                c = safe_write(ws, r, 5, d["minim"])
                if c: c.number_format = "#,##0.0000"
                c = safe_write(ws, r, 6, d["inchidere"])
                if c: c.number_format = "#,##0.0000"
                c = safe_write(ws, r, 7, d["var_zi_pct"] / 100)
                if c: c.number_format = "0.00%"; color_pnl(c, d["var_zi_pct"])
                c = safe_write(ws, r, 8, d["var_sapt_pct"] / 100)
                if c: c.number_format = "0.00%"; color_pnl(c, d["var_sapt_pct"])
                c = safe_write(ws, r, 9, d["var_luna_pct"] / 100)
                if c: c.number_format = "0.00%"; color_pnl(c, d["var_luna_pct"])
                c = safe_write(ws, r, 10, d["volum"])
                if c: c.number_format = "#,##0"
                c = safe_write(ws, r, 11, d["avg_vol_20"])
                if c: c.number_format = "#,##0"
                c = safe_write(ws, r, 12, d["rvol"])
                if c: c.number_format = "0.00x"
                c = safe_write(ws, r, 15, d["trend"])
                if c: color_trend(c, d["trend"])
            else:
                c = safe_write(ws, r, 6, "N/A")
                if c: c.fill = F_GRY
            r += 1
    log.info(f"  OK PRETURI VOLUME — {r - 3} randuri actualizate")

def update_indicatori_tehnici(ws, all_data):
    clear_rows(ws, 3)
    r = 3
    for cat_name, cat_dict in CATEGORII:
        write_category_header(ws, r, f"--- {cat_name} ---", 23)
        r += 1
        for name, ticker in cat_dict.items():
            d = all_data.get(ticker, {})
            safe_write(ws, r, 1, d.get("data", datetime.now().strftime("%d.%m.%Y")))
            c = safe_write(ws, r, 2, name)
            if c: c.font = fnt(bold=True)
            safe_write(ws, r, 3, ticker)
            if d:
                nums = [
                    (4, d["inchidere"], "#,##0.0000"),
                    (5, d["ma20"], "#,##0.0000"),
                    (6, d["ma50"], "#,##0.0000"),
                    (7, d["ma200"], "#,##0.0000"),
                    (8, d["rsi"], "0.00"),
                    (9, d["rsi_status"], None),
                    (10, d["macd"], "0.000000"),
                    (11, d["macd_signal"], "0.000000"),
                    (12, d["macd_hist"], "0.000000"),
                    (13, d["macd_cross"], None),
                    (14, d["bb_sup"], "#,##0.0000"),
                    (15, d["bb_inf"], "#,##0.0000"),
                    (16, d["bb_width"], "#,##0.0000"),
                    (17, d["atr"], "0.0000"),
                    (18, d["stoch_k"], "0.00"),
                    (19, d["stoch_d"], "0.00"),
                    (20, d["volum"], "#,##0"),
                    (21, d["rvol"], "0.00x"),
                    (22, d["trend"], None),
                    (23, d["macross"], None),
                ]
                for col, val, fmt in nums:
                    c = safe_write(ws, r, col, val)
                    if c and fmt and val is not None:
                        c.number_format = fmt
                c = ws.cell(r, 8)
                if c: color_rsi(c, d["rsi"])
                c = ws.cell(r, 22)
                if c: color_trend(c, d["trend"])
                mc = ws.cell(r, 23)
                if mc:
                    if d["macross"] == "Golden Cross": mc.fill = F_GRN; mc.font = FN_GRN
                    elif d["macross"] == "Death Cross": mc.fill = F_RED; mc.font = FN_RED
                    else: mc.fill = F_YLW; mc.font = FN_YLW
            r += 1
    log.info(f"  OK INDICATORI TEHNICI — {r - 3} randuri actualizate")

def update_semnale(ws, all_data):
    clear_rows(ws, 3)
    r = 3
    for cat_name, cat_dict in CATEGORII:
        write_category_header(ws, r, f"--- {cat_name} ---", 17)
        r += 1
        for name, ticker in cat_dict.items():
            d = all_data.get(ticker, {})
            if not d:
                r += 1
                continue
            conditie = (f"RSI={d['rsi']:.0f} | {d['macd_cross']} | "
                        f"{d['macross']} | RVOL={d['rvol']:.1f}x")
            vol_text = ("Crescut" if d["rvol"] > 1.3
                        else ("Scazut" if d["rvol"] < 0.8 else "Normal"))

            safe_write(ws, r, 1, d["data"])
            c = safe_write(ws, r, 2, name)
            if c: c.font = fnt(bold=True)
            safe_write(ws, r, 3, d["semnal"])
            safe_write(ws, r, 4, conditie)
            c = safe_write(ws, r, 5, d["rsi"])
            if c: c.number_format = "0.00"
            safe_write(ws, r, 6, d["macd_cross"])
            safe_write(ws, r, 7, d["macross"])
            safe_write(ws, r, 8, vol_text)
            c = safe_write(ws, r, 9, d["momentum_10z"] / 100)
            if c: c.number_format = "0.00%"
            safe_write(ws, r, 10, d["confluente"])
            c = safe_write(ws, r, 11, d["inchidere"])
            if c: c.number_format = "#,##0.0000"
            c = safe_write(ws, r, 12, d["sl"])
            if c: c.number_format = "#,##0.0000"
            c = safe_write(ws, r, 13, d["tp"])
            if c: c.number_format = "#,##0.0000"
            c = safe_write(ws, r, 14)
            if c:
                c.value = f"=IFERROR((M{r}-K{r})/(K{r}-L{r}),\"N/A\")"
                c.number_format = "0.00x"
            c = safe_write(ws, r, 15, d["probabilitate"] / 100)
            if c: c.number_format = "0%"
            safe_write(ws, r, 16, "Activ")
            safe_write(ws, r, 17, f"Auto {datetime.now().strftime('%H:%M')}")

            c = ws.cell(r, 3)
            if c: color_signal(c, d["semnal"])
            c = ws.cell(r, 5)
            if c: color_rsi(c, d["rsi"])
            prob_c = ws.cell(r, 15)
            if prob_c:
                if d["probabilitate"] >= 65: prob_c.fill = F_GRN; prob_c.font = FN_GRN
                elif d["probabilitate"] >= 50: prob_c.fill = F_YLW; prob_c.font = FN_YLW
                else: prob_c.fill = F_RED; prob_c.font = FN_RED
            r += 1
    log.info(f"  OK SEMNALE INTRARE — {r - 3} randuri actualizate")

def update_macro(ws, macro_live, fear_greed):
    today = datetime.now().strftime("%d.%m.%Y %H:%M")
    macro_map = {
        "VIX": ("VIX", macro_live.get("VIX")),
        "Yield 10Y US (%)": ("Yield 10Y US", macro_live.get("Yield 10Y US")),
        "Yield 2Y US": ("Yield 2Y US", macro_live.get("Yield 2Y US")),
        "Yield 30Y US": ("Yield 30Y US", macro_live.get("Yield 30Y US")),
        "USD Index (DXY)": ("USD Index", macro_live.get("USD Index")),
        "Fear & Greed (0-100)": ("Fear & Greed", fear_greed.get("value")),
    }
    cnt = 0
    for row in ws.iter_rows(min_row=3, max_row=50):
        ind_cell = row[0]
        if not ind_cell.value:
            continue
        ind_name = str(ind_cell.value).strip()
        if ind_name in macro_map:
            key, val = macro_map[ind_name]
            if val is not None:
                c = safe_write(ws, ind_cell.row, 2, val)
                if c: c.number_format = "0.00"
                safe_write(ws, ind_cell.row, 9, today)
                cnt += 1
    log.info(f"  OK INDICATORI MACRO — {cnt} valori actualizate")

def update_dashboard(ws, all_data, fear_greed):
    """Actualizează DASHBOARD cu activul selectat și agregate piață"""
    # Găsește celula cu activul selectat (rândul 2, coloana H sau I)
    activ_sel = None
    for row in range(1, 10):
        for col in range(1, 15):
            c = ws.cell(row, col)
            try:
                if c.value and "Activ selectat" in str(c.value):
                    # Activul e în celula de lângă
                    activ_sel = ws.cell(row, col + 2).value
                    break
            except:
                pass
        if activ_sel:
            break
    
    if not activ_sel:
        activ_sel = "S&P 500"  # default
    
    # Caută ticker-ul activului selectat
    ticker_sel = None
    for name, ticker in ACTIVE.items():
        if name == activ_sel:
            ticker_sel = ticker
            break
    
    d = all_data.get(ticker_sel, {}) if ticker_sel else {}
    
    # Data actualizare (rândul 2, coloana 1)
    safe_write(ws, 2, 4, datetime.now().strftime("%d.%m.%Y %H:%M"))
    
    # TREND GENERAL, VOLATILITATE, VOLUM RELATIV, MOMENT INTRARE (rândurile 4-5)
    # Căutăm aceste titluri și scriem sub ele
    for row in range(3, 10):
        c = ws.cell(row, 1)
        if c.value and "TREND GENERAL" in str(c.value):
            safe_write(ws, row + 1, 1, d.get("trend", "N/A") if d else "N/A")
        elif c.value and "VOLATILITATE" in str(c.value):
            safe_write(ws, row + 1, 5, f"VIX: {all_data.get('^VIX', {}).get('inchidere', 'N/A')}")
        elif c.value and "VOLUM RELATIV" in str(c.value):
            safe_write(ws, row + 1, 9, f"{d.get('rvol', 'N/A')}x" if d else "N/A")
        elif c.value and "MOMENT INTRARE" in str(c.value):
            safe_write(ws, row + 1, 13, d.get("semnal", "N/A") if d else "N/A")
    
    # SEMNAL PRINCIPAL DE TRADING (completăm datele pentru activul selectat)
    for row in range(8, 25):
        c = ws.cell(row, 1)
        if not c.value:
            continue
        label = str(c.value).strip()
        if "SEMNAL ACTIV" in label:
            safe_write(ws, row, 2, d.get("semnal", "N/A") if d else "N/A")
        elif "Activ analizat" in label:
            safe_write(ws, row, 2, activ_sel)
        elif "Entry Price" in label:
            safe_write(ws, row, 2, d.get("inchidere", "N/A") if d else "N/A")
        elif "Stop Loss" in label:
            safe_write(ws, row, 2, d.get("sl", "N/A") if d else "N/A")
        elif "Take Profit" in label:
            safe_write(ws, row, 2, d.get("tp", "N/A") if d else "N/A")
        elif "Risk/Reward Ratio" in label:
            if d and d.get("sl") and d.get("tp"):
                rr = abs((d["tp"] - d["inchidere"]) / (d["inchidere"] - d["sl"]))
                safe_write(ws, row, 2, round(rr, 2))
        elif "Confluențe aliniate" in label:
            safe_write(ws, row, 2, d.get("confluente", "N/A") if d else "N/A")
        elif "Probabilitate" in label:
            safe_write(ws, row, 2, d.get("probabilitate", "N/A") if d else "N/A")
        elif "Condiție declanșare" in label:
            if d:
                cond = f"RSI={d['rsi']:.0f} | {d['macd_cross']} | {d['macross']} | RVOL={d['rvol']:.1f}x"
                safe_write(ws, row, 2, cond)
    
    # REZUMAT INDICATORI
    for row in range(20, 35):
        c = ws.cell(row, 1)
        if not c.value:
            continue
        ind = str(c.value).strip()
        if ind == "RSI(14)":
            safe_write(ws, row, 2, d.get("rsi", "N/A") if d else "N/A")
        elif ind == "MACD Cross":
            safe_write(ws, row, 2, d.get("macd_cross", "N/A") if d else "N/A")
        elif ind == "MA50 vs MA200":
            safe_write(ws, row, 2, d.get("macross", "N/A") if d else "N/A")
        elif ind == "VIX (global)":
            vix_d = all_data.get("^VIX", {})
            safe_write(ws, row, 2, vix_d.get("inchidere", "N/A") if vix_d else "N/A")
        elif ind == "Fear & Greed":
            safe_write(ws, row, 2, fear_greed.get("display", "N/A"))
        elif ind == "RVOL activ":
            safe_write(ws, row, 2, f"{d.get('rvol', 'N/A')}x" if d else "N/A")
    
    log.info("  OK DASHBOARD actualizat")

def update_fisa_activ(ws, all_data, activ_selectat):
    """Actualizează FISA ACTIV pentru activul selectat din Dashboard"""
    # Găsește ticker-ul
    ticker_sel = None
    for name, ticker in ACTIVE.items():
        if name == activ_selectat:
            ticker_sel = ticker
            break
    
    d = all_data.get(ticker_sel, {}) if ticker_sel else {}
    
    # Activ analizat (rândul 2)
    safe_write(ws, 2, 2, activ_selectat)
    
    if not d:
        log.warning(f"  Nu există date pentru {activ_selectat}")
        return
    
    # 1 — SEMNAL DE INTRARE (găsim secțiunea și completăm)
    for row in range(5, 100):
        c = ws.cell(row, 1)
        if not c.value:
            continue
        label = str(c.value).strip()
        
        if "Semnal (BUY/SELL/WAIT)" in label:
            safe_write(ws, row, 2, d["semnal"])
        elif "Condiție declanșare" in label:
            cond = f"RSI={d['rsi']:.0f} | {d['macd_cross']} | {d['macross']} | RVOL={d['rvol']:.1f}x"
            safe_write(ws, row, 2, cond)
        elif label == "Entry Price":
            safe_write(ws, row, 2, d["inchidere"])
            safe_write(ws, row, 5, d["confluente"])
        elif label == "Stop Loss (SL)":
            safe_write(ws, row, 2, d["sl"])
            safe_write(ws, row, 5, d["tp"])
        elif label == "Risk/Reward Ratio":
            rr = abs((d["tp"] - d["inchidere"]) / (d["inchidere"] - d["sl"])) if d["sl"] != d["inchidere"] else 0
            safe_write(ws, row, 2, round(rr, 2))
            safe_write(ws, row, 5, d["probabilitate"])
        elif label == "Status semnal":
            safe_write(ws, row, 2, "Activ")
            safe_write(ws, row, 5, d["timestamp"])
        
        # 2 — INDICATORI TEHNICI
        elif label == "Preț curent":
            safe_write(ws, row, 2, d["inchidere"])
            safe_write(ws, row, 5, d["trend"])
        elif label == "MA20":
            safe_write(ws, row, 2, d["ma20"])
            safe_write(ws, row, 5, d["macross"])
        elif label == "MA50":
            safe_write(ws, row, 2, d["ma50"])
            safe_write(ws, row, 5, d["macd_cross"])
        elif label == "MA200":
            safe_write(ws, row, 2, d["ma200"])
            safe_write(ws, row, 5, d["macd"])
        elif label == "RSI(14)":
            safe_write(ws, row, 2, d["rsi"])
            safe_write(ws, row, 5, d["macd_signal"])
        elif label == "RSI Status":
            safe_write(ws, row, 2, d["rsi_status"])
            safe_write(ws, row, 5, d["macd_hist"])
        elif label == "BB Superior":
            safe_write(ws, row, 2, d["bb_sup"])
            safe_write(ws, row, 5, d["bb_inf"])
        elif label == "BB Lățime":
            safe_write(ws, row, 2, d["bb_width"])
            safe_write(ws, row, 5, d["atr"])
        elif label == "Stoch %K":
            safe_write(ws, row, 2, d["stoch_k"])
            safe_write(ws, row, 5, d["stoch_d"])
        elif label == "Volum" and row < 50:
            safe_write(ws, row, 2, d["volum"])
            safe_write(ws, row, 5, d["rvol"])
        
        # 3 — PREȚURI & VOLUME
        elif label == "Deschidere":
            safe_write(ws, row, 2, d["deschidere"])
            safe_write(ws, row, 5, d["inchidere"])
        elif label == "Maxim":
            safe_write(ws, row, 2, d["maxim"])
            safe_write(ws, row, 5, d["minim"])
        elif label == "Var. Zi (%)":
            safe_write(ws, row, 2, d["var_zi_pct"])
            safe_write(ws, row, 5, d["var_sapt_pct"])
        elif label == "Var. Lună (%)":
            safe_write(ws, row, 2, d["var_luna_pct"])
            safe_write(ws, row, 5, d["semnal"])
        elif label == "Medie Vol.20z":
            safe_write(ws, row, 2, d["avg_vol_20"])
            safe_write(ws, row, 5, d["trend"])
    
    log.info(f"  OK FISA ACTIV pentru {activ_selectat}")

def update_rezumat_executiv(ws, all_data, fear_greed):
    """Actualizează REZUMAT EXECUTIV cu date agregate"""
    # Data ultimei actualizări (rândul 3)
    safe_write(ws, 3, 4, datetime.now().strftime("%d.%m.%Y %H:%M"))
    
    # Calculează agregate
    total_buy = sum(1 for d in all_data.values() if d.get("semnal") == "BUY")
    total_sell = sum(1 for d in all_data.values() if d.get("semnal") == "SELL")
    total_wait = sum(1 for d in all_data.values() if d.get("semnal") == "WAIT")
    
    avg_rsi = sum(d.get("rsi", 50) for d in all_data.values()) / len(all_data) if all_data else 50
    
    sp500 = all_data.get("^GSPC", {})
    vix_d = all_data.get("^VIX", {})
    
    trend_general = "Bullish ↑" if total_buy > total_sell else "Bearish ↓" if total_sell > total_buy else "Sideways →"
    volatilitate = "Ridicată" if vix_d.get("inchidere", 20) > 25 else "Moderată" if vix_d.get("inchidere", 20) > 15 else "Scăzută"
    volum_tranz = "Crescut" if sp500.get("rvol", 1) > 1.2 else "Scăzut" if sp500.get("rvol", 1) < 0.8 else "Normal"
    sentiment = fear_greed.get("status", "Neutru")
    
    # Scriem în tabel
    for row in range(6, 20):
        c = ws.cell(row, 1)
        if not c.value:
            continue
        ind = str(c.value).strip()
        
        if ind == "Tendință generală":
            safe_write(ws, row, 2, trend_general)
            safe_write(ws, row, 5, f"BUY: {total_buy} | SELL: {total_sell} | WAIT: {total_wait}")
        elif ind == "Volatilitate":
            safe_write(ws, row, 2, volatilitate)
            safe_write(ws, row, 5, f"VIX: {vix_d.get('inchidere', 'N/A')}")
        elif ind == "Volum tranzacții":
            safe_write(ws, row, 2, volum_tranz)
            safe_write(ws, row, 5, f"RVOL S&P500: {sp500.get('rvol', 'N/A')}x")
        elif ind == "Sentiment piață":
            safe_write(ws, row, 2, fear_greed.get("display", "N/A"))
            safe_write(ws, row, 5, f"RSI mediu: {avg_rsi:.1f}")
        elif ind == "Risc sistemic":
            risc = "Ridicat" if vix_d.get("inchidere", 20) > 30 else "Moderat"
            safe_write(ws, row, 2, risc)
    
    log.info("  OK REZUMAT EXECUTIV actualizat")

def update_historic(ws, all_data, fear_greed):
    """Adaugă sau actualizează rândul pentru luna curentă"""
    luna_curenta = datetime.now().strftime("%b %Y")
    
    # Caută dacă există deja
    exista = False
    for row in range(2, 100):
        c = ws.cell(row, 1)
        if c.value == luna_curenta:
            exista = True
            log.info(f"  OK ISTORIC TRENDING — {luna_curenta} exista deja")
            return
    
    # Găsește primul rând gol după header
    r = 2
    while ws.cell(r, 1).value:
        r += 1
    
    # Calculează agregate
    avg_rsi = sum(d.get("rsi", 50) for d in all_data.values()) / len(all_data) if all_data else 50
    sp500 = all_data.get("^GSPC", {})
    vix_d = all_data.get("^VIX", {})
    
    total_buy = sum(1 for d in all_data.values() if d.get("semnal") == "BUY")
    total_sell = sum(1 for d in all_data.values() if d.get("semnal") == "SELL")
    semnal_luna = "BUY" if total_buy > total_sell else "SELL" if total_sell > total_buy else "WAIT"
    
    safe_write(ws, r, 1, luna_curenta)
    safe_write(ws, r, 2, round(avg_rsi, 2))
    safe_write(ws, r, 3, sp500.get("inchidere", "N/A"))
    safe_write(ws, r, 6, vix_d.get("inchidere", "N/A"))
    safe_write(ws, r, 7, semnal_luna)
    safe_write(ws, r, 8, fear_greed.get("value", "N/A"))
    
    log.info(f"  OK ISTORIC TRENDING — adăugat {luna_curenta}")

# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    log.info("=================================================================")
    log.info("  PORNIRE ACTUALIZARE AUTOMATA")
    log.info(f"  {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}")
    log.info(f"  Active: {len(ACTIVE)} total  "
             f"({len(INDICI)} indici | {len(ACTIUNI)} actiuni | "
             f"{len(CRYPTO)} crypto | {len(VALUTE)} valute | {len(MATERII_PRIME)} materii prime)")
    log.info("=================================================================")

    all_data = {}
    log.info("\n  Preia date pentru toate activele (poate dura cateva minute)...\n")
    
    for idx, (name, ticker) in enumerate(ACTIVE.items(), 1):
        d = get_full_data(name, ticker)
        if d:
            all_data[ticker] = d
            log.info(f"  [{idx:3}/{len(ACTIVE)}] {name:30} {d['inchidere']:12.4f}  "
                     f"({d['var_zi_pct']:+.2f}%)  RSI={d['rsi']:5.1f}  [{d['semnal']}]")
        else:
            log.warning(f"  [{idx:3}/{len(ACTIVE)}] {name:30} -- EROARE / N/A --")

    log.info("\n  Preia Fear & Greed Index...")
    fear_greed = get_fear_greed()
    log.info(f"  Fear & Greed: {fear_greed['display']}")

    log.info("\n  Preia date macro...")
    macro_live = {}
    for name, ticker in MACRO_TICKERS.items():
        d = get_full_data(name, ticker)
        if d:
            macro_live[name] = d["inchidere"]
            log.info(f"  {name:25} = {d['inchidere']:.4f}")

    log.info("\n  Scriere Excel...\n")
    
    try:
        wb = load_workbook(EXCEL_PATH)
        
        if "PRETURI VOLUME" in wb.sheetnames:
            update_preturi_volume(wb["PRETURI VOLUME"], all_data)
        
        if "INDICATORI TEHNICI" in wb.sheetnames:
            update_indicatori_tehnici(wb["INDICATORI TEHNICI"], all_data)
        
        if "SEMNALE INTRARE" in wb.sheetnames:
            try:
                update_semnale(wb["SEMNALE INTRARE"], all_data)
            except Exception as e:
                log.error(f"  EROARE SEMNALE INTRARE: {e}")
        
        if "INDICATORI MACRO" in wb.sheetnames:
            update_macro(wb["INDICATORI MACRO"], macro_live, fear_greed)
        
        if "DASHBOARD" in wb.sheetnames:
            try:
                update_dashboard(wb["DASHBOARD"], all_data, fear_greed)
            except Exception as e:
                log.error(f"  EROARE DASHBOARD: {e}")
        
        if "ISTORIC TRENDING" in wb.sheetnames:
            update_historic(wb["ISTORIC TRENDING"], all_data, fear_greed)
        
        if "FISA ACTIV" in wb.sheetnames:
            try:
                # Găsește activul selectat din DASHBOARD
                activ_sel = "S&P 500"  # default
                if "DASHBOARD" in wb.sheetnames:
                    dash = wb["DASHBOARD"]
                    for row in range(1, 10):
                        for col in range(1, 15):
                            c = dash.cell(row, col)
                            try:
                                if c.value and "Activ selectat" in str(c.value):
                                    activ_sel = dash.cell(row, col + 2).value
                                    break
                            except:
                                pass
                
                update_fisa_activ(wb["FISA ACTIV"], all_data, activ_sel)
            except Exception as e:
                log.error(f"  EROARE FISA ACTIV: {e}")
        
        if "REZUMAT EXECUTIV" in wb.sheetnames:
            update_rezumat_executiv(wb["REZUMAT EXECUTIV"], all_data, fear_greed)
        
        wb.save(EXCEL_PATH)
        log.info(f"\n{'=' * 65}")
        log.info(f"  SALVAT: {EXCEL_PATH.name}")
        log.info(f"  Procesate: {len(all_data)}/{len(ACTIVE)} active OK  |  {len(ACTIVE) - len(all_data)} esuate")
        log.info(f"  Finalizat: {datetime.now().strftime('%H:%M:%S')}")
        log.info(f"{'=' * 65}")
        
    except Exception as e:
        log.error(f"  EROARE FATALA: {e}", exc_info=True)
        sys.exit(1)

if __name__ == "__main__":
    main()