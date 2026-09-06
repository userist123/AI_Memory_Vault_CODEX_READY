import os
"""
╔══════════════════════════════════════════════════════════════════════════════╗
║           ACTUALIZARE AUTOMATĂ — Analiză Piață Profesională                 ║
║           80+ active: Indici | Acțiuni | Crypto | Valute | Materii prime   ║
╚══════════════════════════════════════════════════════════════════════════════╝

INSTALARE (o singură dată):
    pip install yfinance openpyxl requests pandas

RULARE:
    python actualizare_piata.py

AUTOMATIZARE:
    Windows  → Task Scheduler → python actualizare_piata.py la 08:30 L-V
    Mac/Linux → crontab: 30 8 * * 1-5 python3 /cale/actualizare_piata.py
"""

import sys
import logging
from datetime import datetime
from pathlib import Path

import requests
import pandas as pd
import yfinance as yf
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# ══════════════════════════════════════════════════════════════════════════════
# CONFIGURARE — modifică doar această secțiune
# ══════════════════════════════════════════════════════════════════════════════

EXCEL_PATH = Path(r"C:\Users\Marius\Desktop\Nu sterge\Analiza_Piata_Profesionala.xlsx")

# Cheie FRED API (gratuit la fred.stlouisfed.org) — lasă "" dacă nu ai
FRED_API_KEY = os.environ.get("FRED_API_KEY", "")

# ── INDICI BURSIERI ───────────────────────────────────────────────────────────
INDICI = {
    "S&P 500":          "^GSPC",
    "NASDAQ 100":       "^NDX",
    "NASDAQ Comp.":     "^IXIC",
    "Dow Jones":        "^DJI",
    "Russell 2000":     "^RUT",
    "DAX Germany":      "^GDAXI",
    "FTSE 100":         "^FTSE",
    "CAC 40":           "^FCHI",
    "Nikkei 225":       "^N225",
    "Hang Seng":        "^HSI",
    "Shanghai Comp.":   "000001.SS",
    "MSCI World ETF":   "URTH",
    "MSCI EM ETF":      "EEM",
    "BET Romania":      "BET.RO",
}

# ── ACȚIUNI & ETF ─────────────────────────────────────────────────────────────
ACTIUNI = {
    # Tech / AI
    "Apple":            "AAPL",
    "Microsoft":        "MSFT",
    "NVIDIA":           "NVDA",
    "Alphabet":         "GOOGL",
    "Amazon":           "AMZN",
    "Meta":             "META",
    "Tesla":            "TSLA",
    "AMD":              "AMD",
    "Intel":            "INTC",
    "Broadcom":         "AVGO",
    "ASML":             "ASML",
    "Taiwan Semi":      "TSM",
    "Palantir":         "PLTR",
    "Salesforce":       "CRM",
    "Oracle":           "ORCL",
    # Finance
    "JPMorgan":         "JPM",
    "Goldman Sachs":    "GS",
    "Berkshire B":      "BRK-B",
    "Visa":             "V",
    "Mastercard":       "MA",
    # Energy
    "ExxonMobil":       "XOM",
    "Chevron":          "CVX",
    "Shell":            "SHEL",
    # Industrial
    "Caterpillar":      "CAT",
    "Boeing":           "BA",
    # ETF-uri
    "SPY":              "SPY",
    "QQQ":              "QQQ",
    "GLD ETF":          "GLD",
    "TLT Bond ETF":     "TLT",
    "ARKK":             "ARKK",
}

# ── CRYPTOCURRENCY (25 monede) ────────────────────────────────────────────────
CRYPTO = {
    "Bitcoin":          "BTC-USD",
    "Ethereum":         "ETH-USD",
    "BNB":              "BNB-USD",
    "Solana":           "SOL-USD",
    "XRP":              "XRP-USD",
    "Cardano":          "ADA-USD",
    "Avalanche":        "AVAX-USD",
    "Dogecoin":         "DOGE-USD",
    "Chainlink":        "LINK-USD",
    "Polkadot":         "DOT-USD",
    "Litecoin":         "LTC-USD",
    "Shiba Inu":        "SHIB-USD",
    "Polygon":          "MATIC-USD",
    "Uniswap":          "UNI-USD",
    "Cosmos":           "ATOM-USD",
    "Stellar":          "XLM-USD",
    "Monero":           "XMR-USD",
    "Tron":             "TRX-USD",
    "Filecoin":         "FIL-USD",
    "Aave":             "AAVE-USD",
    "Arbitrum":         "ARB-USD",
    "Optimism":         "OP-USD",
    "Render":           "RNDR-USD",
    "Sui":              "SUI-USD",
    "Near Protocol":    "NEAR-USD",
}

# ── VALUTE FOREX ──────────────────────────────────────────────────────────────
VALUTE = {
    "EUR/USD":          "EURUSD=X",
    "GBP/USD":          "GBPUSD=X",
    "USD/JPY":          "USDJPY=X",
    "USD/CHF":          "USDCHF=X",
    "AUD/USD":          "AUDUSD=X",
    "USD/CAD":          "USDCAD=X",
    "EUR/RON":          "EURRON=X",
    "USD/RON":          "USDRON=X",
    "GBP/RON":          "GBPRON=X",
    "EUR/GBP":          "EURGBP=X",
    "USD/CNY":          "USDCNY=X",
    "USD/TRY":          "USDTRY=X",
}

# ── MATERII PRIME ─────────────────────────────────────────────────────────────
MATERII_PRIME = {
    "Gold":             "GC=F",
    "Silver":           "SI=F",
    "Platinum":         "PL=F",
    "Palladium":        "PA=F",
    "Oil WTI":          "CL=F",
    "Oil Brent":        "BZ=F",
    "Natural Gas":      "NG=F",
    "Copper":           "HG=F",
    "Corn":             "ZC=F",
    "Wheat":            "ZW=F",
    "Soybean":          "ZS=F",
    "Coffee":           "KC=F",
    "Sugar":            "SB=F",
    "Cotton":           "CT=F",
}

# ── MACRO (indicatori de piata) ───────────────────────────────────────────────
MACRO_TICKERS = {
    "VIX":              "^VIX",
    "Yield 10Y US":     "^TNX",
    "Yield 2Y US":      "^IRX",
    "Yield 30Y US":     "^TYX",
    "USD Index":        "DX-Y.NYB",
}

# Combina toate pentru sheet-urile de preturi si tehnic
ACTIVE = {**INDICI, **ACTIUNI, **CRYPTO, **VALUTE, **MATERII_PRIME}

# ══════════════════════════════════════════════════════════════════════════════
# LOGGING
# ══════════════════════════════════════════════════════════════════════════════

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(
            Path(__file__).parent / "actualizare.log",
            encoding="utf-8"
        ),
    ]
)
log = logging.getLogger(__name__)

# ══════════════════════════════════════════════════════════════════════════════
# CULORI EXCEL
# ══════════════════════════════════════════════════════════════════════════════

def fill(hex_c):
    return PatternFill("solid", start_color=hex_c, fgColor=hex_c)

def fnt(bold=False, color="000000", size=10):
    return Font(name="Arial", bold=bold, color=color, size=size)

F_GRN  = fill("C6EFCE"); FN_GRN  = fnt(color="375623")
F_YLW  = fill("FFEB9C"); FN_YLW  = fnt(color="9C6500")
F_RED  = fill("FFC7CE"); FN_RED  = fnt(color="9C0006")
F_HDR  = fill("1F4E79"); FN_HDR  = fnt(bold=True, color="FFFFFF")
F_GRY  = fill("F2F2F2"); FN_GRY  = fnt(color="595959")
NO_FILL = PatternFill(fill_type=None)

# ══════════════════════════════════════════════════════════════════════════════
# CALCUL INDICATORI TEHNICI
# ══════════════════════════════════════════════════════════════════════════════

def calc_rsi(prices, period=14):
    delta = prices.diff()
    gain  = delta.clip(lower=0).rolling(period).mean()
    loss  = (-delta.clip(upper=0)).rolling(period).mean()
    rs    = gain / loss.replace(0, 1e-10)
    rsi   = 100 - (100 / (1 + rs))
    val   = rsi.iloc[-1]
    return round(float(val), 2) if pd.notna(val) else 50.0


def calc_macd(prices):
    ema12 = prices.ewm(span=12, adjust=False).mean()
    ema26 = prices.ewm(span=26, adjust=False).mean()
    ml    = ema12 - ema26
    sl    = ml.ewm(span=9, adjust=False).mean()
    hist  = ml - sl
    m, s, h = float(ml.iloc[-1]), float(sl.iloc[-1]), float(hist.iloc[-1])
    prev_h  = float(hist.iloc[-2]) if len(hist) > 1 else 0
    if m > s:
        cross = "Bullish Cross" if prev_h < 0 else "Pozitiv"
    else:
        cross = "Bearish Cross" if prev_h > 0 else "Negativ"
    return {"macd": round(m, 6), "signal": round(s, 6),
            "histogram": round(h, 6), "cross": cross}


def calc_ma(prices):
    def ma(n):
        return round(float(prices.tail(n).mean()), 6) if len(prices) >= n else None
    ma20, ma50, ma200 = ma(20), ma(50), ma(200)
    cross = "Neutru"
    if ma50 and ma200:
        cross = "Golden Cross" if ma50 > ma200 else "Death Cross"
    return {"ma20": ma20, "ma50": ma50, "ma200": ma200, "macross": cross}


def calc_bollinger(prices, period=20):
    m   = prices.rolling(period).mean()
    std = prices.rolling(period).std()
    sup = float((m + 2 * std).iloc[-1])
    inf = float((m - 2 * std).iloc[-1])
    return {"bb_sup": round(sup, 6), "bb_inf": round(inf, 6),
            "bb_width": round(sup - inf, 6)}


def calc_atr(hist, period=14):
    hi, lo, cl = hist["High"], hist["Low"], hist["Close"]
    tr  = pd.concat([(hi - lo), (hi - cl.shift()).abs(),
                     (lo - cl.shift()).abs()], axis=1).max(axis=1)
    atr = tr.rolling(period).mean().iloc[-1]
    return round(float(atr), 6) if pd.notna(atr) else 0.0


def calc_stochastic(hist, period=14):
    lo14  = hist["Low"].rolling(period).min()
    hi14  = hist["High"].rolling(period).max()
    diff  = (hi14 - lo14).replace(0, 1e-10)
    k     = (hist["Close"] - lo14) / diff * 100
    d     = k.rolling(3).mean()
    return {
        "stoch_k": round(float(k.iloc[-1]), 2) if pd.notna(k.iloc[-1]) else 50.0,
        "stoch_d": round(float(d.iloc[-1]), 2) if pd.notna(d.iloc[-1]) else 50.0,
    }


def calc_signal(rsi, macd_cross, ma_cross, rvol):
    score = 0
    if rsi < 35:                     score += 2
    elif rsi < 45:                   score += 1
    elif rsi > 75:                   score -= 2
    elif rsi > 65:                   score -= 1
    if "Bullish Cross" in macd_cross: score += 2
    elif "Pozitiv" in macd_cross:    score += 1
    elif "Bearish Cross" in macd_cross: score -= 2
    elif "Negativ" in macd_cross:    score -= 1
    if ma_cross == "Golden Cross":   score += 2
    elif ma_cross == "Death Cross":  score -= 2
    if rvol > 1.5:                   score += 1
    elif rvol < 0.6:                 score -= 1
    confluente = min(abs(score), 5)
    if score >= 3:    return "BUY",  confluente
    elif score <= -3: return "SELL", confluente
    else:             return "WAIT", confluente

# ══════════════════════════════════════════════════════════════════════════════
# PRELUARE DATE
# ══════════════════════════════════════════════════════════════════════════════

def get_full_data(name, ticker):
    try:
        t    = yf.Ticker(ticker)
        hist = t.history(period="1y", auto_adjust=True)
        if hist.empty or len(hist) < 5:
            log.warning(f"  ⚠ Date insuficiente: {name} ({ticker})")
            return {}

        closes = hist["Close"]
        latest = hist.iloc[-1]
        prev   = hist.iloc[-2]

        close_now  = float(latest["Close"])
        close_prev = float(prev["Close"])
        close_5d   = float(hist.iloc[-5]["Close"])  if len(hist) >= 5  else close_prev
        close_20d  = float(hist.iloc[-20]["Close"]) if len(hist) >= 20 else close_prev

        var_zi   = (close_now - close_prev) / close_prev * 100 if close_prev else 0
        var_sapt = (close_now - close_5d)   / close_5d   * 100 if close_5d   else 0
        var_luna = (close_now - close_20d)  / close_20d  * 100 if close_20d  else 0

        volume   = int(latest.get("Volume", 0))
        avg_vol  = int(hist["Volume"].tail(20).mean()) if "Volume" in hist else 0
        rvol     = round(volume / avg_vol, 2) if avg_vol > 0 else 1.0

        rsi   = calc_rsi(closes)
        macd  = calc_macd(closes)
        ma    = calc_ma(closes)
        boll  = calc_bollinger(closes)
        atr   = calc_atr(hist)
        stoch = calc_stochastic(hist)
        mom10 = round(float(closes.pct_change(10).iloc[-1] * 100), 2) if len(closes) > 10 else 0

        price = close_now
        if ma["ma50"] and price > ma["ma50"] * 1.01:   trend = "Bullish"
        elif ma["ma50"] and price < ma["ma50"] * 0.99: trend = "Bearish"
        else:                                            trend = "Sideways"

        rsi_status = (
            "Supravandut ▼" if rsi > 70 else
            "Puternic"      if rsi > 55 else
            "Neutru"        if rsi > 45 else
            "Slab"          if rsi > 30 else
            "Supravendut ▲"
        )

        semnal, confluente = calc_signal(rsi, macd["cross"], ma["macross"], rvol)

        sl = round(price - 1.5 * atr, 6) if semnal == "BUY"  else \
             round(price + 1.5 * atr, 6) if semnal == "SELL" else \
             round(price - 2.0 * atr, 6)
        tp = round(price + 3.0 * atr, 6) if semnal == "BUY"  else \
             round(price - 3.0 * atr, 6) if semnal == "SELL" else \
             round(price + 2.0 * atr, 6)

        prob = min(90, 35 + confluente * 10 + (5 if rvol > 1.2 else 0))

        return {
            "name": name, "ticker": ticker,
            "data": datetime.now().strftime("%d.%m.%Y"),
            "timestamp": datetime.now().strftime("%d.%m.%Y %H:%M"),
            "deschidere":   round(float(latest.get("Open", close_now)), 6),
            "maxim":        round(float(latest.get("High", close_now)), 6),
            "minim":        round(float(latest.get("Low",  close_now)), 6),
            "inchidere":    round(close_now, 6),
            "var_zi_pct":   round(var_zi,   4),
            "var_sapt_pct": round(var_sapt, 4),
            "var_luna_pct": round(var_luna, 4),
            "volum":        volume,
            "avg_vol_20":   avg_vol,
            "rvol":         rvol,
            "rsi":          rsi,
            "rsi_status":   rsi_status,
            "macd":         macd["macd"],
            "macd_signal":  macd["signal"],
            "macd_hist":    macd["histogram"],
            "macd_cross":   macd["cross"],
            "ma20":         ma["ma20"],
            "ma50":         ma["ma50"],
            "ma200":        ma["ma200"],
            "macross":      ma["macross"],
            "bb_sup":       boll["bb_sup"],
            "bb_inf":       boll["bb_inf"],
            "bb_width":     boll["bb_width"],
            "atr":          atr,
            "stoch_k":      stoch["stoch_k"],
            "stoch_d":      stoch["stoch_d"],
            "momentum_10z": mom10,
            "trend":        trend,
            "semnal":       semnal,
            "confluente":   confluente,
            "sl":           sl,
            "tp":           tp,
            "probabilitate": prob,
        }
    except Exception as e:
        log.error(f"  X Eroare {name} ({ticker}): {e}")
        return {}


def get_fear_greed():
    try:
        r = requests.get("https://api.alternative.me/fng/?limit=1", timeout=10)
        d = r.json()["data"][0]
        val   = int(d["value"])
        label = d["value_classification"]
        return {
            "value":   val,
            "display": f"{val} - {label}",
            "status":  "Pozitiv" if val > 60 else ("Negativ" if val < 40 else "Neutru"),
        }
    except Exception as e:
        log.warning(f"  Fear & Greed indisponibil: {e}")
        return {"value": None, "display": "N/A", "status": "Neutru"}


def get_fred(series_id):
    if not FRED_API_KEY:
        return None
    try:
        url = (f"https://api.stlouisfed.org/fred/series/observations"
               f"?series_id={series_id}&api_key={FRED_API_KEY}"
               f"&file_type=json&sort_order=desc&limit=1")
        obs = requests.get(url, timeout=10).json().get("observations", [])
        if obs and obs[0]["value"] != ".":
            return float(obs[0]["value"])
    except Exception as e:
        log.warning(f"  FRED {series_id}: {e}")
    return None

# ══════════════════════════════════════════════════════════════════════════════
# HELPERS COLORARE
# ══════════════════════════════════════════════════════════════════════════════

def color_pnl(cell, value):
    if value > 0:   cell.fill = F_GRN; cell.font = FN_GRN
    elif value < 0: cell.fill = F_RED; cell.font = FN_RED
    else:           cell.fill = F_YLW; cell.font = FN_YLW

def color_rsi(cell, rsi):
    if rsi < 30:        cell.fill = F_GRN; cell.font = FN_GRN
    elif rsi < 50:      cell.fill = F_YLW; cell.font = FN_YLW
    elif rsi < 70:
        cell.fill = fill("E2EFDA"); cell.font = fnt(color="375623")
    else:               cell.fill = F_RED; cell.font = FN_RED

def color_signal(cell, semnal):
    if semnal == "BUY":   cell.fill = F_GRN; cell.font = FN_GRN
    elif semnal == "SELL": cell.fill = F_RED; cell.font = FN_RED
    else:                  cell.fill = F_YLW; cell.font = FN_YLW

def color_trend(cell, trend):
    if trend == "Bullish":   cell.fill = F_GRN; cell.font = FN_GRN
    elif trend == "Bearish": cell.fill = F_RED; cell.font = FN_RED
    else:                    cell.fill = F_YLW; cell.font = FN_YLW

def clear_rows(ws, start_row, end_row=500):
    for row in ws.iter_rows(min_row=start_row, max_row=end_row):
        for cell in row:
            cell.value = None
            cell.fill  = NO_FILL

def write_category_header(ws, row, label, num_cols):
    ws.cell(row, 1, label).fill = F_HDR
    ws.cell(row, 1).font = FN_HDR
    for c in range(2, num_cols + 1):
        ws.cell(row, c).fill = F_HDR

# ══════════════════════════════════════════════════════════════════════════════
# ACTUALIZARE SHEET-URI
# ══════════════════════════════════════════════════════════════════════════════

CATEGORII = [
    ("INDICI BURSIERI",  INDICI),
    ("ACTIUNI & ETF",    ACTIUNI),
    ("CRYPTOCURRENCY",   CRYPTO),
    ("VALUTE FOREX",     VALUTE),
    ("MATERII PRIME",    MATERII_PRIME),
]


def update_preturi_volume(ws, all_data):
    clear_rows(ws, 3)
    r = 3
    for cat_name, cat_dict in CATEGORII:
        write_category_header(ws, r, f"--- {cat_name} ---", 15)
        r += 1
        for name, ticker in cat_dict.items():
            d = all_data.get(ticker, {})
            ws.cell(r, 1,  d.get("data", datetime.now().strftime("%d.%m.%Y")))
            ws.cell(r, 2,  name).font = fnt(bold=True)
            ws.cell(r, 3,  ticker)
            if d:
                ws.cell(r, 4,  d["deschidere"]).number_format = "#,##0.0000"
                ws.cell(r, 5,  d["maxim"]).number_format      = "#,##0.0000"
                ws.cell(r, 6,  d["minim"]).number_format      = "#,##0.0000"
                ws.cell(r, 7,  d["inchidere"]).number_format  = "#,##0.0000"
                ws.cell(r, 8,  d["var_zi_pct"]   / 100).number_format = "0.00%"
                ws.cell(r, 9,  d["var_sapt_pct"] / 100).number_format = "0.00%"
                ws.cell(r, 10, d["var_luna_pct"] / 100).number_format = "0.00%"
                ws.cell(r, 11, d["volum"]).number_format      = "#,##0"
                ws.cell(r, 12, d["avg_vol_20"]).number_format = "#,##0"
                ws.cell(r, 13, d["rvol"]).number_format       = "0.00x"
                ws.cell(r, 14, d["semnal"])
                ws.cell(r, 15, d["trend"])
                color_pnl(ws.cell(r, 8),  d["var_zi_pct"])
                color_pnl(ws.cell(r, 9),  d["var_sapt_pct"])
                color_pnl(ws.cell(r, 10), d["var_luna_pct"])
                color_signal(ws.cell(r, 14), d["semnal"])
                color_trend(ws.cell(r, 15),  d["trend"])
            else:
                ws.cell(r, 7, "N/A").fill = F_GRY
            r += 1
    log.info(f"  OK Preturi Volume — {r - 3} randuri")


def update_indicatori_tehnici(ws, all_data):
    clear_rows(ws, 3)
    r = 3
    for cat_name, cat_dict in CATEGORII:
        write_category_header(ws, r, f"--- {cat_name} ---", 23)
        r += 1
        for name, ticker in cat_dict.items():
            d = all_data.get(ticker, {})
            ws.cell(r, 1, d.get("data", datetime.now().strftime("%d.%m.%Y")))
            ws.cell(r, 2, name).font = fnt(bold=True)
            ws.cell(r, 3, ticker)
            if d:
                nums = [
                    (4,  d["inchidere"],  "#,##0.0000"),
                    (5,  d["ma20"],       "#,##0.0000"),
                    (6,  d["ma50"],       "#,##0.0000"),
                    (7,  d["ma200"],      "#,##0.0000"),
                    (8,  d["rsi"],        "0.00"),
                    (9,  d["rsi_status"], None),
                    (10, d["macd"],       "0.000000"),
                    (11, d["macd_signal"],"0.000000"),
                    (12, d["macd_hist"],  "0.000000"),
                    (13, d["macd_cross"], None),
                    (14, d["bb_sup"],     "#,##0.0000"),
                    (15, d["bb_inf"],     "#,##0.0000"),
                    (16, d["bb_width"],   "#,##0.0000"),
                    (17, d["atr"],        "0.0000"),
                    (18, d["stoch_k"],    "0.00"),
                    (19, d["stoch_d"],    "0.00"),
                    (20, d["volum"],      "#,##0"),
                    (21, d["rvol"],       "0.00x"),
                    (22, d["trend"],      None),
                    (23, d["macross"],    None),
                ]
                for col, val, fmt in nums:
                    c = ws.cell(r, col, val)
                    if fmt and val is not None:
                        c.number_format = fmt
                color_rsi(ws.cell(r, 8),    d["rsi"])
                color_trend(ws.cell(r, 22), d["trend"])
                mc = ws.cell(r, 23)
                if d["macross"] == "Golden Cross": mc.fill = F_GRN; mc.font = FN_GRN
                elif d["macross"] == "Death Cross": mc.fill = F_RED; mc.font = FN_RED
                else:                               mc.fill = F_YLW; mc.font = FN_YLW
            r += 1
    log.info(f"  OK Indicatori Tehnici — {r - 3} randuri")


def update_semnale(ws, all_data):
    clear_rows(ws, 3)
    r = 3
    for cat_name, cat_dict in CATEGORII:
        write_category_header(ws, r, f"--- {cat_name} ---", 17)
        r += 1
        for name, ticker in cat_dict.items():
            d = all_data.get(ticker, {})
            if not d:
                r += 1
                continue
            conditie = (f"RSI={d['rsi']:.0f} | {d['macd_cross']} | "
                        f"{d['macross']} | RVOL={d['rvol']:.1f}x")
            vol_text = ("Crescut" if d["rvol"] > 1.3
                        else ("Scazut" if d["rvol"] < 0.8 else "Normal"))

            ws.cell(r, 1,  d["data"])
            ws.cell(r, 2,  name).font = fnt(bold=True)
            ws.cell(r, 3,  d["semnal"])
            ws.cell(r, 4,  conditie)
            ws.cell(r, 5,  d["rsi"]).number_format          = "0.00"
            ws.cell(r, 6,  d["macd_cross"])
            ws.cell(r, 7,  d["macross"])
            ws.cell(r, 8,  vol_text)
            ws.cell(r, 9,  d["momentum_10z"] / 100).number_format = "0.00%"
            ws.cell(r, 10, d["confluente"])
            ws.cell(r, 11, d["inchidere"]).number_format    = "#,##0.0000"
            ws.cell(r, 12, d["sl"]).number_format           = "#,##0.0000"
            ws.cell(r, 13, d["tp"]).number_format           = "#,##0.0000"
            ws.cell(r, 14).value         = f"=IFERROR((M{r}-K{r})/(K{r}-L{r}),\"N/A\")"
            ws.cell(r, 14).number_format = "0.00x"
            ws.cell(r, 15, d["probabilitate"]).number_format = "0%"
            ws.cell(r, 16, "Activ")
            ws.cell(r, 17, f"Auto {datetime.now().strftime('%H:%M')}")

            color_signal(ws.cell(r, 3),  d["semnal"])
            color_rsi(ws.cell(r, 5),     d["rsi"])
            prob_c = ws.cell(r, 15)
            if d["probabilitate"] >= 65:   prob_c.fill = F_GRN; prob_c.font = FN_GRN
            elif d["probabilitate"] >= 50: prob_c.fill = F_YLW; prob_c.font = FN_YLW
            else:                          prob_c.fill = F_RED; prob_c.font = FN_RED
            r += 1
    log.info(f"  OK Semnale Intrare — {r - 3} randuri")


def update_macro(ws, macro_live, fear_greed):
    today = datetime.now().strftime("%d.%m.%Y %H:%M")
    macro_map = {
        "VIX":          macro_live.get("VIX", {}).get("inchidere"),
        "Yield 10Y":    macro_live.get("Yield 10Y US", {}).get("inchidere"),
        "Yield 2Y":     macro_live.get("Yield 2Y US", {}).get("inchidere"),
        "USD Index":    macro_live.get("USD Index", {}).get("inchidere"),
        "Fear & Greed": fear_greed.get("value"),
    }
    if FRED_API_KEY:
        macro_map["Rata dobanzii"] = get_fred("FEDFUNDS")
        macro_map["CPI"]           = get_fred("CPIAUCSL")
        macro_map["Somaj"]         = get_fred("UNRATE")

    updated = 0
    for r in range(3, 50):
        ind = ws.cell(r, 1).value
        if not ind:
            break
        for key, val in macro_map.items():
            if key.lower() in str(ind).lower() and val is not None:
                prev = ws.cell(r, 2).value
                try:
                    if prev and float(str(prev).replace(",", ".")) != float(val):
                        ws.cell(r, 3).value = prev
                except (ValueError, TypeError):
                    pass
                ws.cell(r, 2).value = round(float(val), 4)
                ws.cell(r, 9).value = today
                updated += 1
                break
    log.info(f"  OK Indicatori Macro — {updated} valori")


def update_dashboard(ws, all_data, fear_greed):
    ws["D2"] = datetime.now().strftime("%d.%m.%Y %H:%M")
    buy = sell = wait = 0
    for d in all_data.values():
        if not d: continue
        s = d.get("semnal", "WAIT")
        if s == "BUY":    buy  += 1
        elif s == "SELL": sell += 1
        else:             wait += 1
    total = buy + sell + wait
    if total > 0:
        pct_buy  = round(buy  / total * 100, 1)
        pct_sell = round(sell / total * 100, 1)
        trend = ("Bullish" if pct_buy > 55
                 else ("Bearish" if pct_sell > 55 else "Mixt / Neutru"))
        log.info(f"  OK Dashboard — BUY:{buy}({pct_buy}%) "
                 f"SELL:{sell}({pct_sell}%) WAIT:{wait} | {trend}")


def update_historic(ws, all_data, fear_greed):
    luna_an = datetime.now().strftime("%b %Y")
    for row in ws.iter_rows(min_row=3, max_col=1):
        if row[0].value == luna_an:
            log.info(f"  OK Istoric — {luna_an} exista deja")
            return
    last_row = 2
    for row in ws.iter_rows(min_row=3, max_col=1):
        if row[0].value:
            last_row = row[0].row
    r = last_row + 1

    rsi_vals = [d["rsi"] for d in all_data.values() if d and d.get("rsi")]
    avg_rsi  = round(sum(rsi_vals) / len(rsi_vals), 1) if rsi_vals else 50
    buy  = sum(1 for d in all_data.values() if d and d.get("semnal") == "BUY")
    sell = sum(1 for d in all_data.values() if d and d.get("semnal") == "SELL")
    semnal_luna = "BUY" if buy > sell else ("SELL" if sell > buy else "WAIT")

    ws.cell(r, 1, luna_an)
    ws.cell(r, 2, avg_rsi)
    ws.cell(r, 8, semnal_luna)
    sp500 = all_data.get("^GSPC", {}).get("inchidere")
    if sp500: ws.cell(r, 4, sp500)
    if fear_greed.get("value"): ws.cell(r, 11, fear_greed["value"])
    color_signal(ws.cell(r, 8), semnal_luna)
    log.info(f"  OK Istoric — adaugat {luna_an} RSI={avg_rsi} Semnal={semnal_luna}")

# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    log.info("=" * 65)
    log.info("  PORNIRE ACTUALIZARE AUTOMATA")
    log.info(f"  {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}")
    log.info(f"  Active: {len(ACTIVE)} total  "
             f"({len(INDICI)} indici | {len(ACTIUNI)} actiuni | "
             f"{len(CRYPTO)} crypto | {len(VALUTE)} valute | "
             f"{len(MATERII_PRIME)} materii prime)")
    log.info("=" * 65)

    if not EXCEL_PATH.exists():
        log.error(f"Fisierul Excel nu a fost gasit: {EXCEL_PATH}")
        sys.exit(1)

    # ── 1. Preia date ─────────────────────────────────────────────
    log.info(f"\n  Preia date pentru {len(ACTIVE)} active (2-5 minute)...\n")
    all_data = {}
    for i, (name, ticker) in enumerate(ACTIVE.items(), 1):
        data = get_full_data(name, ticker)
        all_data[ticker] = data
        if data:
            v   = data["var_zi_pct"]
            log.info(f"  [{i:3d}/{len(ACTIVE)}] {name:25s} "
                     f"{data['inchidere']:>14.4f}  "
                     f"({'+' if v >= 0 else ''}{v:.2f}%)  "
                     f"RSI={data['rsi']:5.1f}  [{data['semnal']}]")
        else:
            log.warning(f"  [{i:3d}/{len(ACTIVE)}] {name:25s}  -- EROARE / N/A --")

    log.info("\n  Preia Fear & Greed Index...")
    fear_greed = get_fear_greed()
    log.info(f"  Fear & Greed: {fear_greed.get('display', 'N/A')}")

    log.info("\n  Preia date macro...")
    macro_live = {}
    for name, ticker in MACRO_TICKERS.items():
        d = get_full_data(name, ticker)
        if d:
            macro_live[name] = d
            log.info(f"  {name:25s} = {d['inchidere']:.4f}")

    # ── 2. Actualizeaza Excel ─────────────────────────────────────
    log.info("\n  Scriere Excel...\n")
    wb = load_workbook(str(EXCEL_PATH))

    tasks = [
        ("PRETURI VOLUME",     update_preturi_volume,     [all_data]),
        ("INDICATORI TEHNICI", update_indicatori_tehnici, [all_data]),
        ("SEMNALE INTRARE",    update_semnale,            [all_data]),
        ("INDICATORI MACRO",   update_macro,              [macro_live, fear_greed]),
        ("DASHBOARD",          update_dashboard,          [all_data, fear_greed]),
        ("ISTORIC TRENDING",   update_historic,           [all_data, fear_greed]),
    ]
    for sheet_name, fn, args in tasks:
        if sheet_name not in wb.sheetnames:
            log.warning(f"  Sheet '{sheet_name}' lipsa — skip")
            continue
        try:
            fn(wb[sheet_name], *args)
        except Exception as e:
            log.error(f"  EROARE {sheet_name}: {e}")

    # ── 3. Salveaza ───────────────────────────────────────────────
    wb.save(str(EXCEL_PATH))
    ok    = sum(1 for d in all_data.values() if d)
    fails = len(all_data) - ok
    log.info("\n" + "=" * 65)
    log.info(f"  SALVAT: {EXCEL_PATH.name}")
    log.info(f"  Procesate: {ok}/{len(ACTIVE)} active OK"
             + (f"  |  {fails} esuate" if fails else ""))
    log.info(f"  Finalizat: {datetime.now().strftime('%H:%M:%S')}")
    log.info("=" * 65)


if __name__ == "__main__":
    main()