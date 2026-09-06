import os
# ============================================================
# actualizare_unificat.py  –  Versiunea Finala Unificata
# Analiza_Piata_Profesionala.xlsx
# Python 3.10+ | yfinance>=0.2.18 | openpyxl>=3.1.2 | pandas>=2.0
# ============================================================

import sys
import logging
import requests
import pandas as pd
import yfinance as yf
from copy import copy
from datetime import datetime
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import StockChart, BarChart, Reference
from openpyxl.chart.axis import DateAxis, ChartLines
from openpyxl.chart.updown_bars import UpDownBars
from openpyxl.chart.data_source import NumData, NumVal

# ─────────────────────────────────────────────────────────────
# CONFIGURARE
# ─────────────────────────────────────────────────────────────
EXCEL_PATH   = Path(r"C:\Users\Marius\Desktop\Nu sterge\Analiza_Piata_Profesionala.xlsx")
FRED_API_KEY = os.environ.get("FRED_API_KEY", "")

# ─────────────────────────────────────────────────────────────
# LOGGING DUAL
# ─────────────────────────────────────────────────────────────
_fmt = logging.Formatter("%(asctime)s %(levelname)-8s %(message)s", datefmt="%H:%M:%S")
_sh  = logging.StreamHandler(sys.stdout); _sh.setFormatter(_fmt)
_fh  = logging.FileHandler("actualizare_unificat.log", encoding="utf-8"); _fh.setFormatter(_fmt)
log  = logging.getLogger("unificat")
log.setLevel(logging.DEBUG)
log.addHandler(_sh); log.addHandler(_fh)

# ─────────────────────────────────────────────────────────────
# DICȚIONARE ACTIVE
# ─────────────────────────────────────────────────────────────
INDICI = {
    "S&P 500":        "^GSPC",
    "NASDAQ 100":     "^NDX",
    "NASDAQ Comp.":   "^IXIC",
    "Dow Jones":      "^DJI",
    "Russell 2000":   "^RUT",
    "DAX Germany":    "^GDAXI",
    "FTSE 100":       "^FTSE",
    "CAC 40":         "^FCHI",
    "Nikkei 225":     "^N225",
    "Hang Seng":      "^HSI",
    "Shanghai":       "000001.SS",
    "MSCI World":     "URTH",
    "MSCI EM":        "EEM",
    "BET Romania":    "BET.RO",
}
ACTIUNI = {
    "Apple":          "AAPL",
    "Microsoft":      "MSFT",
    "NVIDIA":         "NVDA",
    "Alphabet":       "GOOGL",
    "Amazon":         "AMZN",
    "Meta":           "META",
    "Tesla":          "TSLA",
    "Berkshire B":    "BRK-B",
    "JPMorgan":       "JPM",
    "Visa":           "V",
    "UnitedHealth":   "UNH",
    "Exxon Mobil":    "XOM",
    "Johnson&Johnson":"JNJ",
    "Procter&Gamble": "PG",
    "ASML":           "ASML",
    "Samsung":        "005930.KS",
    "TSMC":           "TSM",
    "Netflix":        "NFLX",
    "Adobe":          "ADBE",
    "Salesforce":     "CRM",
    "Palantir":       "PLTR",
    "AMD":            "AMD",
    "Intel":          "INTC",
    "Broadcom":       "AVGO",
    "Qualcomm":       "QCOM",
    "PayPal":         "PYPL",
    "Coinbase":       "COIN",
    "Robinhood":      "HOOD",
    "Cathie Wood ARK":"ARKK",
    "SPY ETF":        "SPY",
}
CRYPTO = {
    "Bitcoin":        "BTC-USD",
    "Ethereum":       "ETH-USD",
    "BNB":            "BNB-USD",
    "Solana":         "SOL-USD",
    "XRP":            "XRP-USD",
    "Cardano":        "ADA-USD",
    "Avalanche":      "AVAX-USD",
    "Polkadot":       "DOT-USD",
    "Polygon":        "MATIC-USD",
    "Chainlink":      "LINK-USD",
    "Uniswap":        "UNI-USD",
    "Litecoin":       "LTC-USD",
    "Dogecoin":       "DOGE-USD",
    "Shiba Inu":      "SHIB-USD",
    "TRON":           "TRX-USD",
    "Stellar":        "XLM-USD",
    "Cosmos":         "ATOM-USD",
    "Monero":         "XMR-USD",
    "Filecoin":       "FIL-USD",
    "Internet Computer": "ICP-USD",
    "Hedera":         "HBAR-USD",
    "VeChain":        "VET-USD",
    "Algorand":       "ALGO-USD",
    "Fantom":         "FTM-USD",
    "NEAR Protocol":  "NEAR-USD",
}
VALUTE = {
    "EUR/USD":        "EURUSD=X",
    "GBP/USD":        "GBPUSD=X",
    "USD/JPY":        "USDJPY=X",
    "USD/CHF":        "USDCHF=X",
    "AUD/USD":        "AUDUSD=X",
    "USD/CAD":        "USDCAD=X",
    "NZD/USD":        "NZDUSD=X",
    "EUR/GBP":        "EURGBP=X",
    "EUR/JPY":        "EURJPY=X",
    "USD/CNY":        "USDCNY=X",
    "USD/HUF":        "USDHUF=X",
    "USD/TRY":        "USDTRY=X",
}
MATERII_PRIME = {
    "Gold":           "GC=F",
    "Silver":         "SI=F",
    "Oil WTI":        "CL=F",
    "Oil Brent":      "BZ=F",
    "Natural Gas":    "NG=F",
    "Copper":         "HG=F",
    "Platinum":       "PL=F",
    "Palladium":      "PA=F",
    "Corn":           "ZC=F",
    "Wheat":          "ZW=F",
    "Soybeans":       "ZS=F",
    "Coffee":         "KC=F",
    "Sugar":          "SB=F",
    "Cotton":         "CT=F",
}

ACTIVE = {**INDICI, **ACTIUNI, **CRYPTO, **VALUTE, **MATERII_PRIME}

MACRO_TICKERS = {
    "VIX":          "^VIX",
    "Yield 10Y US": "^TNX",
    "Yield 2Y US":  "^IRX",
    "Yield 30Y US": "^TYX",
    "USD Index":    "DX-Y.NYB",
}

CATEGORII = [
    ("INDICI",        INDICI),
    ("ACTIUNI",       ACTIUNI),
    ("CRYPTO",        CRYPTO),
    ("VALUTE",        VALUTE),
    ("MATERII_PRIME", MATERII_PRIME),
]

COMPETITORI_MAP = {
    "INDICI":  ["S&P 500","NASDAQ 100","Dow Jones","DAX Germany","FTSE 100","Nikkei 225"],
    "ACTIUNI": ["Apple","Microsoft","NVIDIA","Alphabet","Amazon","Meta"],
    "CRYPTO":  ["Bitcoin","Ethereum","BNB","Solana","XRP","Cardano"],
    "VALUTE":  ["EUR/USD","GBP/USD","USD/JPY","USD/CHF","AUD/USD","USD/CAD"],
    "MATERII": ["Gold","Silver","Oil WTI","Oil Brent","Natural Gas","Copper"],
}

RISK_LIBRARY = {
    "INDICI": [
        {"ID":"R-I-01","Tip":"Sistemic","Categorie":"INDICI","Descriere":"Recesiune globala / contractie PIB",         "Impact":5,"Probabilitate":30,"Orizont":"6-12 luni"},
        {"ID":"R-I-02","Tip":"Macro",   "Categorie":"INDICI","Descriere":"Crestere agresiva rate dobanda FED",         "Impact":4,"Probabilitate":35,"Orizont":"3-6 luni"},
        {"ID":"R-I-03","Tip":"Geopolit.","Categorie":"INDICI","Descriere":"Conflict armat major / tensiuni globale",   "Impact":4,"Probabilitate":25,"Orizont":"0-3 luni"},
        {"ID":"R-I-04","Tip":"Sectorial","Categorie":"INDICI","Descriere":"Criza bancara sistemica",                   "Impact":5,"Probabilitate":20,"Orizont":"3-12 luni"},
        {"ID":"R-I-05","Tip":"Tehnic",   "Categorie":"INDICI","Descriere":"Spargere suport major / Death Cross",       "Impact":3,"Probabilitate":40,"Orizont":"1-3 luni"},
        {"ID":"R-I-06","Tip":"Lichid.",  "Categorie":"INDICI","Descriere":"Criza lichiditate / credit crunch",         "Impact":4,"Probabilitate":20,"Orizont":"6-12 luni"},
    ],
    "ACTIUNI": [
        {"ID":"R-A-01","Tip":"Earnings", "Categorie":"ACTIUNI","Descriere":"Rezultate financiare sub asteptari",        "Impact":3,"Probabilitate":45,"Orizont":"0-1 luni"},
        {"ID":"R-A-02","Tip":"Macro",    "Categorie":"ACTIUNI","Descriere":"Stagflatie / crestere costuri operationale","Impact":4,"Probabilitate":30,"Orizont":"3-9 luni"},
        {"ID":"R-A-03","Tip":"Reglem.",  "Categorie":"ACTIUNI","Descriere":"Reglementari antitrust / investigatii",    "Impact":3,"Probabilitate":25,"Orizont":"6-18 luni"},
        {"ID":"R-A-04","Tip":"Tehnic",   "Categorie":"ACTIUNI","Descriere":"RSI supraextins / divergenta bearish",     "Impact":2,"Probabilitate":50,"Orizont":"0-1 luni"},
        {"ID":"R-A-05","Tip":"Sectorial","Categorie":"ACTIUNI","Descriere":"Disruptie tehnologica / obsolescenta",     "Impact":4,"Probabilitate":20,"Orizont":"12-36 luni"},
        {"ID":"R-A-06","Tip":"Macro",    "Categorie":"ACTIUNI","Descriere":"Dolar puternic — impact venituri externe", "Impact":3,"Probabilitate":35,"Orizont":"3-6 luni"},
    ],
    "CRYPTO": [
        {"ID":"R-C-01","Tip":"Reglementar","Categorie":"CRYPTO","Descriere":"Interdictie / restrictie legala crypto",  "Impact":5,"Probabilitate":20,"Orizont":"0-6 luni"},
        {"ID":"R-C-02","Tip":"Tehnic",    "Categorie":"CRYPTO","Descriere":"Spargere suport major / bear market",     "Impact":4,"Probabilitate":40,"Orizont":"1-3 luni"},
        {"ID":"R-C-03","Tip":"Hack",      "Categorie":"CRYPTO","Descriere":"Exploit exchange / protocol major",       "Impact":5,"Probabilitate":15,"Orizont":"0-1 luni"},
        {"ID":"R-C-04","Tip":"Macro",     "Categorie":"CRYPTO","Descriere":"Risk-off global / fuga spre siguranta",   "Impact":4,"Probabilitate":35,"Orizont":"0-3 luni"},
        {"ID":"R-C-05","Tip":"On-chain",  "Categorie":"CRYPTO","Descriere":"Whale dump / manipulare piata",           "Impact":3,"Probabilitate":30,"Orizont":"0-1 luni"},
        {"ID":"R-C-06","Tip":"Lichid.",   "Categorie":"CRYPTO","Descriere":"Criza stablecoin / de-peg major",         "Impact":5,"Probabilitate":10,"Orizont":"0-1 luni"},
    ],
    "VALUTE": [
        {"ID":"R-V-01","Tip":"Macro",     "Categorie":"VALUTE","Descriere":"Divergenta politici monetare FED/BCE",    "Impact":4,"Probabilitate":40,"Orizont":"3-6 luni"},
        {"ID":"R-V-02","Tip":"Geopolit.", "Categorie":"VALUTE","Descriere":"Criza geopolitica / sanctiuni comerciale","Impact":3,"Probabilitate":25,"Orizont":"0-3 luni"},
        {"ID":"R-V-03","Tip":"Lichid.",   "Categorie":"VALUTE","Descriere":"Volatilitate extrema weekend / gap",      "Impact":2,"Probabilitate":30,"Orizont":"0-1 saptamani"},
        {"ID":"R-V-04","Tip":"Tehnic",    "Categorie":"VALUTE","Descriere":"Interventie banca centrala la nivel cheie","Impact":3,"Probabilitate":20,"Orizont":"0-1 luni"},
        {"ID":"R-V-05","Tip":"Macro",     "Categorie":"VALUTE","Descriere":"Surpriza CPI / NFP semnificativa",        "Impact":3,"Probabilitate":35,"Orizont":"0-1 luni"},
        {"ID":"R-V-06","Tip":"Sistemic",  "Categorie":"VALUTE","Descriere":"Criza valutara piata emergenta",          "Impact":4,"Probabilitate":15,"Orizont":"3-12 luni"},
    ],
    "MATERII": [
        {"ID":"R-M-01","Tip":"Geopolit.", "Categorie":"MATERII","Descriere":"Conflict OPEC+ / embargo petrol",        "Impact":5,"Probabilitate":25,"Orizont":"0-3 luni"},
        {"ID":"R-M-02","Tip":"Macro",     "Categorie":"MATERII","Descriere":"Incetinire economica China",             "Impact":4,"Probabilitate":35,"Orizont":"3-12 luni"},
        {"ID":"R-M-03","Tip":"Meteo",     "Categorie":"MATERII","Descriere":"Fenomene climatice extreme / seceta",    "Impact":3,"Probabilitate":30,"Orizont":"0-6 luni"},
        {"ID":"R-M-04","Tip":"USD",       "Categorie":"MATERII","Descriere":"Apreciere USD puternica",                "Impact":3,"Probabilitate":35,"Orizont":"3-6 luni"},
        {"ID":"R-M-05","Tip":"Tehnic",    "Categorie":"MATERII","Descriere":"Supraoferta / stocuri in exces",         "Impact":3,"Probabilitate":25,"Orizont":"3-9 luni"},
        {"ID":"R-M-06","Tip":"Reglementar","Categorie":"MATERII","Descriere":"Reglementari energie verde / carbune",  "Impact":3,"Probabilitate":20,"Orizont":"12-36 luni"},
    ],
}

CALENDAR_LIBRARY = {
    "INDICI":  ["FOMC","NFP","CPI","GDP","PMI","Earnings Season"],
    "ACTIUNI": ["Earnings Report","FOMC","CPI","NFP","PCE","Retail Sales"],
    "CRYPTO":  ["Bitcoin Halving","FOMC","SEC Ruling","CPI","ETH Upgrade","Macro Risk"],
    "VALUTE":  ["FOMC","ECB","BOE","BOJ","CPI SUA","NFP"],
    "MATERII": ["OPEC+","EIA Crude","FOMC","China PMI","USD Index","Geopolitical Events"],
}

# ─────────────────────────────────────────────────────────────
# STILURI EXCEL
# ─────────────────────────────────────────────────────────────
def fill(hex_c):
    return PatternFill("solid", start_color=hex_c, fgColor=hex_c)

def fnt(bold=False, color="000000", size=10, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)

def aln(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

NO_FILL = PatternFill(fill_type=None)

F_GRN  = fill("C6EFCE");  FN_GRN  = fnt(color="375623", bold=True)
F_YLW  = fill("FFEB9C");  FN_YLW  = fnt(color="9C6500", bold=True)
F_RED  = fill("FFC7CE");  FN_RED  = fnt(color="9C0006", bold=True)
F_HDR  = fill("1F4E79");  FN_HDR  = fnt(bold=True, color="FFFFFF")
F_GRY  = fill("F2F2F2");  FN_GRY  = fnt(color="595959")
F_LBL  = fill("D6DCE4");  FN_LBL  = fnt(bold=True)
F_ORG  = fill("FCE4D6");  FN_ORG  = fnt(color="833C00")
F_BLU  = fill("DEEAF1");  FN_BLU  = fnt(color="1F4E79", size=10)
F_SEP  = fill("1F4E79")
F_TITLU= fill("0D2137");  FN_TITLU= fnt(bold=True, color="FFFFFF", size=15)
F_CAT  = fill("1F4E79");  FN_CAT  = fnt(bold=True, color="FFFFFF", size=11)

# Ghid invatare extra
F_ACTIV_BUY = fill("1E6B3C");  F_ACTIV_SEL = fill("8B0000");  F_ACTIV_WAI = fill("7D5A00")
FN_ACTIV    = fnt(bold=True, color="FFFFFF", size=11)
F_SEC_HDR   = fill("2E4057")
F_GHID_CAP  = fill("1F4E79");  F_GHID_SUB  = fill("D6DCE4");  F_GHID_TXT  = fill("F9F9F9")

# ─────────────────────────────────────────────────────────────
# COLORARE CONDITIONALA
# ─────────────────────────────────────────────────────────────
def color_pnl(cell, value):
    try:
        v = float(value)
        if v > 0:
            cell.fill = copy(F_GRN); cell.font = copy(FN_GRN)
        elif v < 0:
            cell.fill = copy(F_RED); cell.font = copy(FN_RED)
        else:
            cell.fill = copy(F_YLW); cell.font = copy(FN_YLW)
    except Exception:
        cell.fill = copy(F_GRY); cell.font = copy(FN_GRY)

def color_rsi(cell, rsi):
    try:
        r = float(rsi)
        if r < 30:
            cell.fill = copy(F_GRN); cell.font = copy(FN_GRN)
        elif r < 50:
            cell.fill = copy(F_YLW); cell.font = copy(FN_YLW)
        elif r < 70:
            cell.fill = copy(fill("E2EFDA")); cell.font = copy(fnt(color="375623"))
        else:
            cell.fill = copy(F_RED); cell.font = copy(FN_RED)
    except Exception:
        cell.fill = copy(F_GRY); cell.font = copy(FN_GRY)

def color_signal(cell, semnal):
    s = str(semnal).strip().upper()
    if s == "BUY":
        cell.fill = copy(F_GRN); cell.font = copy(FN_GRN)
    elif s == "SELL":
        cell.fill = copy(F_RED); cell.font = copy(FN_RED)
    else:
        cell.fill = copy(F_YLW); cell.font = copy(FN_YLW)

def color_trend(cell, trend):
    t = str(trend).strip().lower()
    if "bullish" in t:
        cell.fill = copy(F_GRN); cell.font = copy(FN_GRN)
    elif "bearish" in t:
        cell.fill = copy(F_RED); cell.font = copy(FN_RED)
    else:
        cell.fill = copy(F_YLW); cell.font = copy(FN_YLW)

def color_status_extended(cell, v):
    s = str(v).strip().lower()
    positive_kw = {"buy","pozitiv","bullish","golden cross","impuls pozitiv"}
    negative_kw = {"sell","negativ","bearish","death cross","impuls negativ"}
    if any(k in s for k in positive_kw):
        cell.fill = copy(F_GRN); cell.font = copy(FN_GRN)
    elif any(k in s for k in negative_kw):
        cell.fill = copy(F_RED); cell.font = copy(FN_RED)
    else:
        cell.fill = copy(F_YLW); cell.font = copy(FN_YLW)

# ─────────────────────────────────────────────────────────────
# HELPER FUNCTIONS
# ─────────────────────────────────────────────────────────────
def safe_write(ws, row, col, value=None):
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        return None
    cell.value = value
    return cell

def clear_rows(ws, start_row, end_row=5000):
    for row in ws.iter_rows(min_row=start_row, max_row=end_row):
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            cell.value = None
            cell.fill  = copy(NO_FILL)
            cell.font  = fnt()
            cell.number_format = "General"
            cell.alignment = Alignment(vertical="center")

def write_category_header(ws, row, label, num_cols):
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=c)
        if isinstance(cell, MergedCell):
            continue
        cell.fill  = copy(F_HDR)
        cell.font  = copy(FN_HDR)
        cell.alignment = aln(h="center")
    c1 = ws.cell(row=row, column=1)
    if not isinstance(c1, MergedCell):
        c1.value = label

def fmt_price(x, decimals=4):
    if x is None:
        return ""
    try:
        if pd.isna(x):
            return ""
        return f"{float(x):,.{decimals}f}"
    except Exception:
        return ""

def fmt_pct(val, decimals=2):
    if val is None:
        return "N/A"
    try:
        if pd.isna(val):
            return "N/A"
        v = float(val)
        return f"{v:+.{decimals}f}%"
    except Exception:
        return "N/A"

def safe(val, default=0.0):
    try:
        v = float(val)
        return default if pd.isna(v) else v
    except Exception:
        return default

def rr_value(entry, sl, tp):
    try:
        e, s, t = float(entry), float(sl), float(tp)
        risk   = abs(e - s)
        reward = abs(t - e)
        if risk == 0:
            return None
        return reward / risk
    except Exception:
        return None

def rr_text(entry, sl, tp):
    v = rr_value(entry, sl, tp)
    return f"{v:.2f}x" if v is not None else "N/A"

def map_rsi_status(rsi):
    try:
        r = float(rsi)
        if r < 30:   return "Presiune excesiva vanzare"
        if r < 45:   return "Presiune moderata vanzare"
        if r <= 55:  return "Echilibru"
        if r <= 70:  return "Momentum ascendent"
        return "Presiune excesiva cumparare"
    except Exception:
        return "N/A"

def signal_status_text(signal):
    s = str(signal).strip().upper()
    if s == "BUY":  return "Pozitiv"
    if s == "SELL": return "Negativ"
    return "Neutru"

# ─────────────────────────────────────────────────────────────
# INDICATORI TEHNICI
# ─────────────────────────────────────────────────────────────
def calc_rsi(prices: pd.Series, period=14) -> float:
    try:
        delta  = prices.diff()
        gain   = delta.clip(lower=0)
        loss   = (-delta).clip(lower=0)
        avg_g  = gain.rolling(window=period, min_periods=period).mean()
        avg_l  = loss.rolling(window=period, min_periods=period).mean()
        rs     = avg_g / avg_l.replace(0, 1e-10)
        rsi    = 100 - (100 / (1 + rs))
        result = rsi.iloc[-1]
        return 50.0 if pd.isna(result) else float(result)
    except Exception:
        return 50.0

def calc_macd(prices: pd.Series) -> dict:
    try:
        ema12  = prices.ewm(span=12, adjust=False).mean()
        ema26  = prices.ewm(span=26, adjust=False).mean()
        macd   = ema12 - ema26
        signal = macd.ewm(span=9, adjust=False).mean()
        hist   = macd - signal
        m, s, h = float(macd.iloc[-1]), float(signal.iloc[-1]), float(hist.iloc[-1])
        prev_h  = float(hist.iloc[-2]) if len(hist) >= 2 else h
        if m > s and prev_h < 0 <= h:
            cross = "Impuls pozitiv nou"
        elif m > s and h >= 0:
            cross = "Impuls pozitiv activ"
        elif m < s and prev_h > 0 >= h:
            cross = "Impuls negativ nou"
        else:
            cross = "Impuls negativ activ"
        return {"macd": m, "signal": s, "histogram": h, "cross": cross}
    except Exception:
        return {"macd": 0.0, "signal": 0.0, "histogram": 0.0, "cross": "N/A"}

def calc_ma(prices: pd.Series) -> dict:
    def _ma(n):
        if len(prices) < n:
            return None
        v = prices.rolling(n).mean().iloc[-1]
        return None if pd.isna(v) else float(v)
    ma20, ma50, ma200 = _ma(20), _ma(50), _ma(200)
    if ma50 is not None and ma200 is not None:
        if ma50 > ma200:   macross = "Golden Cross"
        elif ma50 < ma200: macross = "Death Cross"
        else:              macross = "Neutru"
    else:
        macross = "Neutru"
    return {"ma20": ma20, "ma50": ma50, "ma200": ma200, "macross": macross}

def calc_bollinger(prices: pd.Series, period=20) -> dict:
    try:
        m   = prices.rolling(period).mean()
        std = prices.rolling(period).std()
        sup = m + 2 * std
        inf = m - 2 * std
        bb_sup = float(sup.iloc[-1])   if pd.notna(sup.iloc[-1])   else None
        bb_inf = float(inf.iloc[-1])   if pd.notna(inf.iloc[-1])   else None
        bb_mid = float(m.iloc[-1])     if pd.notna(m.iloc[-1])     else None
        bb_width = (float(bb_sup) - float(bb_inf)) if bb_sup and bb_inf else None
        return {"bb_sup": bb_sup, "bb_inf": bb_inf, "bb_width": bb_width}
    except Exception:
        return {"bb_sup": None, "bb_inf": None, "bb_width": None}

def calc_atr(hist: pd.DataFrame, period=14) -> float:
    try:
        high  = hist["High"]
        low   = hist["Low"]
        close = hist["Close"].shift(1)
        tr    = pd.concat([
            high - low,
            (high - close).abs(),
            (low  - close).abs()
        ], axis=1).max(axis=1)
        atr   = tr.rolling(period).mean().iloc[-1]
        return 0.0 if pd.isna(atr) else float(atr)
    except Exception:
        return 0.0

def calc_stochastic(hist: pd.DataFrame, period=14) -> dict:
    try:
        low14  = hist["Low"].rolling(period).min()
        high14 = hist["High"].rolling(period).max()
        denom  = (high14 - low14).replace(0, 1e-10)
        k      = ((hist["Close"] - low14) / denom * 100)
        d      = k.rolling(3).mean()
        sk     = float(k.iloc[-1])  if pd.notna(k.iloc[-1])  else 50.0
        sd     = float(d.iloc[-1])  if pd.notna(d.iloc[-1])  else 50.0
        return {"stoch_k": sk, "stoch_d": sd}
    except Exception:
        return {"stoch_k": 50.0, "stoch_d": 50.0}

def calc_signal(rsi, macd_cross, ma_cross, rvol) -> tuple:
    score = 0
    r = safe(rsi, 50.0)
    if   r < 35: score += 2
    elif r < 45: score += 1
    elif r > 75: score -= 2
    elif r > 65: score -= 1

    mc = str(macd_cross).lower()
    if   "impuls pozitiv nou"    in mc: score += 2
    elif "impuls pozitiv activ"  in mc: score += 1
    elif "impuls negativ nou"    in mc: score -= 2
    elif "impuls negativ activ"  in mc: score -= 1

    mx = str(ma_cross).lower()
    if   "golden cross" in mx: score += 2
    elif "death cross"  in mx: score -= 2

    rv = safe(rvol, 1.0)
    if   rv > 1.5: score += 1
    elif rv < 0.6: score -= 1

    confluente = min(abs(score), 5)
    if   score >= 3:  semnal = "BUY"
    elif score <= -3: semnal = "SELL"
    else:             semnal = "WAIT"
    return semnal, confluente, score

# ─────────────────────────────────────────────────────────────
# PRELUARE DATE LIVE
# ─────────────────────────────────────────────────────────────
def get_full_data(name: str, ticker: str) -> dict:
    try:
        t    = yf.Ticker(ticker)
        hist = t.history(period="1y", interval="1d", auto_adjust=True, timeout=15)
        if hist is None or len(hist) < 5:
            return {}
        closes = hist["Close"].dropna()
        if len(closes) < 5:
            return {}

        price   = round(float(closes.iloc[-1]), 6)
        o_price = round(float(hist["Open"].iloc[-1]),  6)
        h_price = round(float(hist["High"].iloc[-1]),  6)
        l_price = round(float(hist["Low"].iloc[-1]),   6)

        def _pct(idx):
            try:
                prev = float(closes.iloc[idx])
                return round((price - prev) / prev * 100, 4) if prev else 0.0
            except Exception:
                return 0.0

        var_zi   = _pct(-2)
        var_sapt = _pct(-6)  if len(closes) > 5  else 0.0
        var_luna = _pct(-21) if len(closes) > 20 else 0.0

        vol_ser  = hist["Volume"].fillna(0)
        volum    = int(vol_ser.iloc[-1])
        avg_vol  = int(vol_ser.tail(20).mean()) if len(vol_ser) >= 20 else volum
        rvol     = round(volum / avg_vol, 2) if avg_vol > 0 else 1.0

        rsi        = calc_rsi(closes)
        rsi_status = map_rsi_status(rsi)
        macd_res   = calc_macd(closes)
        ma_res     = calc_ma(closes)
        bb_res     = calc_bollinger(closes)
        atr        = calc_atr(hist)
        stoch      = calc_stochastic(hist)

        try:
            mom10 = float(closes.pct_change(10).iloc[-1]) * 100
        except Exception:
            mom10 = 0.0

        ma50 = ma_res["ma50"]
        if ma50:
            if   price > ma50 * 1.01: trend = "Bullish"
            elif price < ma50 * 0.99: trend = "Bearish"
            else:                     trend = "Sideways"
        else:
            trend = "Sideways"

        semnal, confluente, score = calc_signal(
            rsi, macd_res["cross"], ma_res["macross"], rvol
        )

        if semnal == "BUY":
            sl = price - 1.5 * atr;  tp = price + 3.0 * atr
        elif semnal == "SELL":
            sl = price + 1.5 * atr;  tp = price - 3.0 * atr
        else:
            sl = None;               tp = None

        prob = min(90, 35 + confluente * 10 + (5 if rvol > 1.2 else 0))

        support    = float(hist["Low"].tail(20).min())
        resistance = float(hist["High"].tail(20).max())

        now = datetime.now()
        return {
            "name": name, "ticker": ticker,
            "data":      now.strftime("%d.%m.%Y"),
            "timestamp": now.strftime("%d.%m.%Y %H:%M"),
            "deschidere": o_price, "maxim": h_price,
            "minim": l_price,      "inchidere": price,
            "var_zi_pct":   var_zi, "var_sapt_pct": var_sapt, "var_luna_pct": var_luna,
            "volum": volum,  "avg_vol_20": avg_vol, "rvol": rvol,
            "rsi": rsi,      "rsi_status": rsi_status,
            "macd":          macd_res["macd"],
            "macd_signal":   macd_res["signal"],
            "macd_hist":     macd_res["histogram"],
            "macd_cross":    macd_res["cross"],
            "ma20": ma_res["ma20"], "ma50": ma_res["ma50"], "ma200": ma_res["ma200"],
            "macross": ma_res["macross"],
            "bb_sup":   bb_res["bb_sup"], "bb_inf": bb_res["bb_inf"],
            "bb_width": bb_res["bb_width"],
            "atr": atr,
            "stoch_k": stoch["stoch_k"], "stoch_d": stoch["stoch_d"],
            "momentum_10z": mom10,
            "trend": trend,
            "semnal": semnal, "score": score, "confluente": confluente,
            "sl": sl, "tp": tp,
            "probabilitate": prob,
            "support": support, "resistance": resistance,
        }
    except Exception as e:
        log.warning(f"get_full_data({ticker}): {e}")
        return {}

def get_fear_greed() -> dict:
    try:
        r    = requests.get("https://api.alternative.me/fng/?limit=1", timeout=10)
        r.raise_for_status()
        data = r.json()["data"][0]
        val  = int(data["value"])
        cls  = data.get("value_classification", "")
        if   val >= 55: status = "Pozitiv"
        elif val <= 45: status = "Negativ"
        else:           status = "Neutru"
        return {"value": val, "display": f"{val} - {cls}", "status": status}
    except Exception:
        return {"value": None, "display": "N/A", "status": "Neutru"}

def get_fred(series_id: str) -> tuple:
    if not FRED_API_KEY:
        return (None, None)
    try:
        url    = "https://api.stlouisfed.org/fred/series/observations"
        params = {
            "series_id":  series_id,
            "api_key":    FRED_API_KEY,
            "file_type":  "json",
            "sort_order": "desc",
            "limit":      2,
        }
        r   = requests.get(url, params=params, timeout=10)
        r.raise_for_status()
        obs = r.json().get("observations", [])
        def _parse(o):
            v = o.get("value", ".")
            if v == ".":
                return None
            try:
                return float(v)
            except Exception:
                return None
        current  = _parse(obs[0]) if len(obs) > 0 else None
        previous = _parse(obs[1]) if len(obs) > 1 else None
        return (current, previous)
    except Exception as e:
        log.warning(f"FRED {series_id}: {e}")
        return (None, None)

# ─────────────────────────────────────────────────────────────
# SHEET 1 — PRETURI VOLUME
# ─────────────────────────────────────────────────────────────
def update_preturi_volume(wb, all_data):
    ws = wb["PRETURI VOLUME"]
    clear_rows(ws, 3)
    row = 3
    count = 0
    for label, cat_dict in CATEGORII:
        write_category_header(ws, row, label, 15)
        row += 1
        for name, ticker in cat_dict.items():
            d = all_data.get(ticker, {})
            try:
                c = safe_write(ws, row, 1, d.get("data", ""))
                c2 = safe_write(ws, row, 2, d.get("name", name))
                if c2: c2.font = fnt(bold=True)
                safe_write(ws, row, 3, d.get("ticker", ticker))
                if not d:
                    c7 = safe_write(ws, row, 7, "N/A")
                    if c7: c7.fill = copy(F_GRY); c7.font = copy(FN_GRY)
                    row += 1; continue

                for col, key, fmt in [
                    (4, "deschidere", "#,##0.0000"),
                    (5, "maxim",      "#,##0.0000"),
                    (6, "minim",      "#,##0.0000"),
                    (7, "inchidere",  "#,##0.0000"),
                ]:
                    cell = safe_write(ws, row, col, d.get(key))
                    if cell: cell.number_format = fmt

                for col, key in [(8, "var_zi_pct"), (9, "var_sapt_pct"), (10, "var_luna_pct")]:
                    val  = d.get(key, 0)
                    cell = safe_write(ws, row, col, val / 100 if val is not None else None)
                    if cell:
                        cell.number_format = "0.00%"
                        color_pnl(cell, val)

                for col, key, fmt in [
                    (11, "volum",      "#,##0"),
                    (12, "avg_vol_20", "#,##0"),
                ]:
                    cell = safe_write(ws, row, col, d.get(key))
                    if cell: cell.number_format = fmt

                cell = safe_write(ws, row, 13, d.get("rvol"))
                if cell: cell.number_format = "0.00x"

                cell = safe_write(ws, row, 14, d.get("semnal"))
                if cell: color_signal(cell, d.get("semnal", "WAIT"))

                cell = safe_write(ws, row, 15, d.get("trend"))
                if cell: color_trend(cell, d.get("trend", "Sideways"))

                count += 1
            except Exception as ex:
                log.error(f"PRETURI VOLUME rând {row} ({name}): {ex}")
            row += 1
    log.info(f"OK PRETURI VOLUME — {count} rânduri actualizate")

# ─────────────────────────────────────────────────────────────
# SHEET 2 — INDICATORI TEHNICI
# ─────────────────────────────────────────────────────────────
def update_indicatori_tehnici(wb, all_data):
    ws = wb["INDICATORI TEHNICI"]
    clear_rows(ws, 3)
    row = 3
    count = 0
    for label, cat_dict in CATEGORII:
        write_category_header(ws, row, label, 23)
        row += 1
        for name, ticker in cat_dict.items():
            d = all_data.get(ticker, {})
            if not d:
                row += 1; continue
            try:
                safe_write(ws, row, 1, d.get("data"))
                c2 = safe_write(ws, row, 2, d.get("name", name))
                if c2: c2.font = fnt(bold=True)
                safe_write(ws, row, 3, d.get("ticker", ticker))

                cell = safe_write(ws, row, 4, d.get("inchidere")); 
                if cell: cell.number_format = "#,##0.0000"

                for col, key in [(5,"ma20"),(6,"ma50"),(7,"ma200")]:
                    v    = d.get(key)
                    cell = safe_write(ws, row, col, v if v is not None else "")
                    if cell and v is not None: cell.number_format = "#,##0.0000"

                cell = safe_write(ws, row, 8, d.get("rsi"))
                if cell: cell.number_format = "0.00"; color_rsi(cell, d.get("rsi"))

                safe_write(ws, row, 9, d.get("rsi_status"))

                for col, key, fmt in [
                    (10,"macd","0.000000"),(11,"macd_signal","0.000000"),
                    (12,"macd_hist","0.000000"),
                ]:
                    cell = safe_write(ws, row, col, d.get(key))
                    if cell: cell.number_format = fmt

                cell = safe_write(ws, row, 13, d.get("macd_cross"))
                if cell: color_status_extended(cell, d.get("macd_cross",""))

                for col, key in [(14,"bb_sup"),(15,"bb_inf"),(16,"bb_width")]:
                    cell = safe_write(ws, row, col, d.get(key))
                    if cell: cell.number_format = "#,##0.0000"

                cell = safe_write(ws, row, 17, d.get("atr"))
                if cell: cell.number_format = "0.0000"

                for col, key in [(18,"stoch_k"),(19,"stoch_d")]:
                    cell = safe_write(ws, row, col, d.get(key))
                    if cell: cell.number_format = "0.00"

                cell = safe_write(ws, row, 20, d.get("volum"))
                if cell: cell.number_format = "#,##0"

                cell = safe_write(ws, row, 21, d.get("rvol"))
                if cell: cell.number_format = "0.00x"

                cell = safe_write(ws, row, 22, d.get("trend"))
                if cell: color_trend(cell, d.get("trend","Sideways"))

                cell = safe_write(ws, row, 23, d.get("macross"))
                if cell: color_status_extended(cell, d.get("macross",""))

                count += 1
            except Exception as ex:
                log.error(f"INDICATORI rând {row} ({name}): {ex}")
            row += 1
    log.info(f"OK INDICATORI TEHNICI — {count} rânduri actualizate")

# ─────────────────────────────────────────────────────────────
# SHEET 3 — SEMNALE INTRARE
# ─────────────────────────────────────────────────────────────
def update_semnale(wb, all_data):
    ws = wb["SEMNALE INTRARE"]
    clear_rows(ws, 3)
    row = 3
    count = 0
    for label, cat_dict in CATEGORII:
        write_category_header(ws, row, label, 17)
        row += 1
        for name, ticker in cat_dict.items():
            d = all_data.get(ticker, {})
            if not d:
                continue
            try:
                rsi      = safe(d.get("rsi"), 50.0)
                rvol     = safe(d.get("rvol"), 1.0)
                score    = safe(d.get("score"), 0)
                mc       = d.get("macd_cross","")
                macross  = d.get("macross","")
                semnal   = d.get("semnal","WAIT")

                safe_write(ws, row, 1, d.get("data"))
                c2 = safe_write(ws, row, 2, d.get("name", name))
                if c2: c2.font = fnt(bold=True)

                cell = safe_write(ws, row, 3, semnal)
                if cell: color_signal(cell, semnal)

                motiv = f"RSI={rsi:.0f} | {mc} | {macross} | RVOL={rvol:.1f}x | Score={score:.0f}"
                safe_write(ws, row, 4, motiv)

                cell = safe_write(ws, row, 5, rsi)
                if cell: cell.number_format = "0.00"; color_rsi(cell, rsi)

                cell = safe_write(ws, row, 6, mc)
                if cell: color_status_extended(cell, mc)

                cell = safe_write(ws, row, 7, macross)
                if cell: color_status_extended(cell, macross)

                cell = safe_write(ws, row, 8, rvol)
                if cell: cell.number_format = "0.00x"

                cell = safe_write(ws, row, 9, safe(d.get("momentum_10z"),0)/100)
                if cell: cell.number_format = "0.00%"

                safe_write(ws, row, 10, d.get("confluente"))

                cell = safe_write(ws, row, 11, d.get("inchidere"))
                if cell: cell.number_format = "#,##0.0000"

                cell = safe_write(ws, row, 12, d.get("sl"))
                if cell: cell.number_format = "#,##0.0000"

                cell = safe_write(ws, row, 13, d.get("tp"))
                if cell: cell.number_format = "#,##0.0000"

                # R/R Excel formula
                cell = safe_write(ws, row, 14, f"=IFERROR((M{row}-K{row})/(K{row}-L{row}),\"N/A\")")
                if cell: cell.number_format = "0.00x"

                prob = safe(d.get("probabilitate"), 0)
                cell = safe_write(ws, row, 15, prob / 100)
                if cell:
                    cell.number_format = "0%"
                    if prob >= 65:
                        cell.fill = copy(F_GRN); cell.font = copy(FN_GRN)
                    elif prob >= 50:
                        cell.fill = copy(F_YLW); cell.font = copy(FN_YLW)
                    else:
                        cell.fill = copy(F_RED); cell.font = copy(FN_RED)

                safe_write(ws, row, 16, "Activ")
                safe_write(ws, row, 17, f"Auto {datetime.now().strftime('%H:%M')}")

                count += 1
            except Exception as ex:
                log.error(f"SEMNALE rând {row} ({name}): {ex}")
            row += 1
    log.info(f"OK SEMNALE INTRARE — {count} rânduri scrise")

# ─────────────────────────────────────────────────────────────
# SHEET 4 — INDICATORI MACRO
# ─────────────────────────────────────────────────────────────
def update_macro(wb, macro_live, fear_greed):
    ws = wb["INDICATORI MACRO"]
    fed,  _  = get_fred("FEDFUNDS")
    cpi,  _  = get_fred("CPIAUCSL")
    unem, _  = get_fred("UNRATE")

    macro_map = {
        "VIX":           macro_live.get("VIX",         {}).get("inchidere"),
        "Yield 10Y":     macro_live.get("Yield 10Y US", {}).get("inchidere"),
        "Yield 2Y":      macro_live.get("Yield 2Y US",  {}).get("inchidere"),
        "USD Index":     macro_live.get("USD Index",     {}).get("inchidere"),
        "Fear & Greed":  fear_greed.get("value"),
        "Rata dobanzii": fed,
        "CPI":           cpi,
        "Somaj":         unem,
    }

    ts = datetime.now().strftime("%d.%m.%Y %H:%M")
    updated = 0
    for row in range(3, 51):
        c1 = ws.cell(row=row, column=1)
        if isinstance(c1, MergedCell):
            continue
        if c1.value is None or str(c1.value).strip() == "":
            break
        indicator = str(c1.value).strip()
        for key, val in macro_map.items():
            if key.lower() in indicator.lower():
                try:
                    c2 = ws.cell(row=row, column=2)
                    c3 = ws.cell(row=row, column=3)
                    if not isinstance(c3, MergedCell) and not isinstance(c2, MergedCell):
                        old_val = c2.value
                        if not isinstance(c3, MergedCell):
                            c3.value = old_val
                        c2.value = round(float(val), 4) if val is not None else "N/A"
                    c9 = ws.cell(row=row, column=9)
                    if not isinstance(c9, MergedCell):
                        c9.value = ts
                    updated += 1
                except Exception as ex:
                    log.error(f"MACRO rând {row}: {ex}")
                break
    log.info(f"OK INDICATORI MACRO — {updated} valori")

# ─────────────────────────────────────────────────────────────
# SHEET 5 — DASHBOARD
# ─────────────────────────────────────────────────────────────
def update_dashboard(wb, all_data, fear_greed):
    ws = wb["DASHBOARD"]
    now_str = datetime.now().strftime("%d.%m.%Y %H:%M")

    # Timestamp actualizare
    c = ws["D2"]
    if not isinstance(c, MergedCell):
        c.value = now_str

    # ── Statistici globale ────────────────────────────────────
    buy_list  = [(t, d) for t, d in all_data.items() if d.get("semnal") == "BUY"]
    sell_list = [(t, d) for t, d in all_data.items() if d.get("semnal") == "SELL"]
    wait_list = [(t, d) for t, d in all_data.items() if d.get("semnal") == "WAIT"]
    total_v   = len(all_data) or 1
    nb, ns, nw = len(buy_list), len(sell_list), len(wait_list)
    pct_buy   = nb / total_v * 100
    pct_sell  = ns / total_v * 100

    if   pct_buy  > 55: trend_gen = "Bullish"
    elif pct_sell > 55: trend_gen = "Bearish"
    else:               trend_gen = "Mixt"

    # ── Semnal principal — cel mai bun BUY ───────────────────
    best = None
    best_score = -999
    for t, d in buy_list:
        sc = safe(d.get("score"), 0) + safe(d.get("probabilitate"), 0) * 0.1
        if sc > best_score:
            best_score = sc
            best = d
    if best is None and all_data:
        # Dacă nu e niciun BUY, ia cel cu cel mai mare score
        best = max(all_data.values(),
                   key=lambda x: safe(x.get("score"), -99) if x else -99,
                   default={})

    # ── Citește activul selectat din J2 ──────────────────────
    j2_cell = ws["J2"]
    if not isinstance(j2_cell, MergedCell) and j2_cell.value:
        selected_name = str(j2_cell.value).strip()
    else:
        selected_name = "NASDAQ Comp."

    ticker_sel = ACTIVE.get(selected_name)
    if not ticker_sel:
        # potrivire partiala
        sel_lower = selected_name.lower()
        for name, t in ACTIVE.items():
            if sel_lower in name.lower():
                ticker_sel = t
                selected_name = name
                break
    if not ticker_sel:
        ticker_sel = "^IXIC"
        selected_name = "NASDAQ Comp."

    sel_d = all_data.get(ticker_sel, {})

    def _sw(cell_addr, val):
        try:
            cell = ws[cell_addr]
            if not isinstance(cell, MergedCell):
                cell.value = val
        except Exception:
            pass

    # ── Statistici BUY/SELL/WAIT ─────────────────────────────
    _sw("E5", nb)
    _sw("E6", ns)
    _sw("E7", nw)
    _sw("E8", trend_gen)

    # ── Semnal principal (best BUY) ───────────────────────────
    if best:
        semnal_best = best.get("semnal", "WAIT")
        cell_b8 = ws["B8"]
        if not isinstance(cell_b8, MergedCell):
            cell_b8.value = semnal_best
            color_signal(cell_b8, semnal_best)
        _sw("C8", best.get("name", ""))
        _sw("D8", best.get("inchidere"))

        cell_d8 = ws["D8"]
        if not isinstance(cell_d8, MergedCell):
            cell_d8.number_format = "#,##0.0000"

        _sw("E9", best.get("sl"))
        _sw("F9", best.get("tp"))
        _sw("G9", rr_text(best.get("inchidere"), best.get("sl"), best.get("tp")))
        _sw("H9", best.get("confluente"))
        _sw("I9", best.get("probabilitate"))
        cond = (f"RSI={safe(best.get('rsi'), 50):.0f} | "
                f"{best.get('macd_cross', '')} | "
                f"{best.get('macross', '')} | "
                f"RVOL={safe(best.get('rvol'), 1):.1f}x")
        _sw("J9", cond)

    # ── REZUMAT INDICATORI — bazat pe activul SELECTAT din J2 ─
    if sel_d:
        rsi_sel   = safe(sel_d.get("rsi"), 50)
        rvol_sel  = safe(sel_d.get("rvol"), 1)
        trend_sel = sel_d.get("trend", "Sideways")
        mc_sel    = sel_d.get("macd_cross", "")
        macr_sel  = sel_d.get("macross", "")

        _sw("B15", round(rsi_sel, 2))
        _sw("B16", mc_sel)
        _sw("B17", macr_sel)
        _sw("B18", trend_sel)
        _sw("B21", round(rvol_sel, 2))

        # Colorare celule rezumat
        for addr, val, color_fn in [
            ("B15", rsi_sel,  color_rsi),
            ("B16", mc_sel,   color_status_extended),
            ("B17", macr_sel, color_status_extended),
            ("B18", trend_sel,color_trend),
        ]:
            try:
                cell = ws[addr]
                if not isinstance(cell, MergedCell):
                    color_fn(cell, val)
            except Exception:
                pass

    # VIX și Fear&Greed sunt globale, nu depind de activul selectat
    vix_d = all_data.get("^VIX", {})
    _sw("B19", vix_d.get("inchidere") if vix_d else None)
    _sw("B20", fear_greed.get("display"))

    # Titlu activ selectat în rezumat (dacă există celula)
    _sw("B14", f"Activ analizat: {selected_name}")

    log.info(f"OK DASHBOARD — BUY:{nb}({pct_buy:.1f}%) SELL:{ns}({pct_sell:.1f}%) "
             f"WAIT:{nw} | {trend_gen} | Activ rezumat: {selected_name}")

# ─────────────────────────────────────────────────────────────
# SHEET 6 — FISA ACTIV
# ─────────────────────────────────────────────────────────────
def _get_selected_activ(wb, all_data) -> tuple:
    """
    Citește activul selectat din DASHBOARD ws["J2"].
    Caută mai întâi după nume exact în ACTIVE, apoi după potrivire parțială.
    Returnează (dict_date, nume_activ).
    """
    selected_name = "NASDAQ Comp."  # fallback
    try:
        ws_dash = wb["DASHBOARD"]
        j2 = ws_dash["J2"]
        if not isinstance(j2, MergedCell) and j2.value and str(j2.value).strip():
            selected_name = str(j2.value).strip()
    except Exception:
        pass

    # Caută ticker exact după nume
    ticker = ACTIVE.get(selected_name)

    # Dacă nu găsit exact, caută case-insensitive parțial
    if not ticker:
        sel_lower = selected_name.lower()
        for name, t in ACTIVE.items():
            if sel_lower in name.lower() or name.lower() in sel_lower:
                ticker = t
                selected_name = name
                break

    # Fallback final la NASDAQ Comp.
    if not ticker:
        ticker = "^IXIC"
        selected_name = "NASDAQ Comp."

    d = all_data.get(ticker, {})
    return d, selected_name

def _detect_category(name: str) -> str:
    for k, cat_d in CATEGORII:
        if name in cat_d:
            return k
    return "INDICI"

def update_fisa_activ(wb, all_data, fear_greed, wb_ref):
    ws = wb["FISA ACTIV"]

    # ── Citește activul selectat ──────────────────────────────
    d, sel_name = _get_selected_activ(wb, all_data)

    if not d:
        log.warning(f"FISA ACTIV: date indisponibile pentru '{sel_name}'")
        return

    cat_key = _detect_category(sel_name)

    # ── Construiește mapările label → valori ──────────────────
    inchidere = d.get("inchidere")
    sl_v      = d.get("sl")
    tp_v      = d.get("tp")
    rsi_v     = safe(d.get("rsi"), 50)
    rvol_v    = safe(d.get("rvol"), 1)
    prob_v    = safe(d.get("probabilitate"), 0)
    semnal    = d.get("semnal", "WAIT")
    macross   = d.get("macross", "")
    macd_c    = d.get("macd_cross", "")
    trend_v   = d.get("trend", "Sideways")

    # Mapare completă: label col A → (valoare col B, valoare col E, tip_colorare)
    # tip_colorare: None | "signal" | "rsi" | "trend" | "status_ext" | "pct"
    label_map = {
        # ── Secțiunea 1 — Semnal ──────────────────────────────
        "Semnal (BUY/SELL/WAIT)": (semnal,    None,   "signal"),
        "Conditie declansare":    (
            f"RSI={rsi_v:.0f} | {macd_c} | {macross} | RVOL={rvol_v:.1f}x",
            None, None),
        "Confluentе (din 5)":     (d.get("confluente"), None, None),
        "Entry Price":            (inchidere,  None,  "price4"),
        "Stop Loss (SL)":         (sl_v,       None,  "price4"),
        "Take Profit (TP)":       (tp_v,       None,  "price4"),
        "Risk/Reward Ratio":      (rr_text(inchidere, sl_v, tp_v), None, None),
        "Probabilitate (%)":      (prob_v / 100, None, "pct"),
        "Status semnal":          (
            "Activ" if semnal != "WAIT" else "In asteptare",
            None, "signal_status"),
        "Ultima actualizare":     (d.get("timestamp"), None, None),
        # ── Secțiunea 2 — Indicatori tehnici ─────────────────
        "Pret curent":            (inchidere,          trend_v,     "price4",   "trend"),
        "MA20":                   (d.get("ma20"),       macross,     "price4",   "status_ext"),
        "MA50":                   (d.get("ma50"),       macd_c,      "price4",   "status_ext"),
        "MA200":                  (d.get("ma200"),      d.get("macd_hist"), "price4", None),
        "RSI(14)":                (rsi_v,               d.get("rsi_status"), "rsi", None),
        "RSI Status":             (d.get("rsi_status"), d.get("macd"),       None,  None),
        "MACD":                   (d.get("macd"),       d.get("volum"),      "price6", None),
        "MACD Signal":            (d.get("macd_signal"),rvol_v,              "price6", None),
        "MACD Histogram":         (d.get("macd_hist"),  trend_v,             "price6", "trend"),
        "MA Cross":               (macross,             d.get("support"),    "status_ext", "price4"),
        "MACD Cross":             (macd_c,              d.get("resistance"), "status_ext", "price4"),
        # ── Secțiunea 3 — Prețuri & Volume ───────────────────
        "Deschidere":             (d.get("deschidere"), inchidere,           "price4", "price4"),
        "Maxim":                  (d.get("maxim"),      d.get("support"),    "price4", "price4"),
        "Minim":                  (d.get("minim"),      d.get("avg_vol_20"), "price4", None),
        "Inchidere":              (inchidere,           rvol_v,              "price4", None),
        "Var. Zi":                (safe(d.get("var_zi_pct"), 0) / 100,
                                   semnal,  "pct_color", "signal"),
        "Var. Sapt":              (safe(d.get("var_sapt_pct"), 0) / 100,
                                   trend_v, "pct_color", "trend"),
    }

    def _apply_format(cell, val, fmt_type):
        """Aplică format și culoare pe celulă în funcție de tip."""
        if cell is None or isinstance(cell, MergedCell):
            return
        try:
            if fmt_type == "price4":
                cell.number_format = "#,##0.0000"
            elif fmt_type == "price6":
                cell.number_format = "0.000000"
            elif fmt_type in ("pct", "pct_color"):
                cell.number_format = "0.00%"
                if fmt_type == "pct_color":
                    color_pnl(cell, val * 100 if val else 0)
            elif fmt_type == "rsi":
                cell.number_format = "0.00"
                color_rsi(cell, val)
            elif fmt_type == "signal":
                color_signal(cell, val)
            elif fmt_type == "signal_status":
                if str(val).lower() == "activ":
                    cell.fill = copy(F_GRN); cell.font = copy(FN_GRN)
                else:
                    cell.fill = copy(F_YLW); cell.font = copy(FN_YLW)
            elif fmt_type == "trend":
                color_trend(cell, val)
            elif fmt_type == "status_ext":
                color_status_extended(cell, val)
        except Exception as ex:
            log.debug(f"_apply_format({fmt_type}): {ex}")

    # ── Iterare rânduri sheet ─────────────────────────────────
    macro_start = comp_start = risk_start = cal_start = None

    for row in range(1, 300):
        c1 = ws.cell(row=row, column=1)
        if isinstance(c1, MergedCell):
            continue
        label = str(c1.value).strip() if c1.value else ""

        # Marchează secțiunile speciale
        if "MACRO" in label.upper() and macro_start is None:
            macro_start = row + 1
        if "COMPET" in label.upper() and comp_start is None:
            comp_start = row + 1
        if "RISC" in label.upper() and risk_start is None:
            risk_start = row + 1
        if "CALENDAR" in label.upper() and cal_start is None:
            cal_start = row + 1

        if label not in label_map:
            continue

        entry = label_map[label]
        # entry poate fi (val_b, val_e, fmt_b, fmt_e) sau (val_b, val_e, fmt_b)
        val_b  = entry[0]
        val_e  = entry[1] if len(entry) > 1 else None
        fmt_b  = entry[2] if len(entry) > 2 else None
        fmt_e  = entry[3] if len(entry) > 3 else None

        try:
            cb = safe_write(ws, row, 2, val_b)
            _apply_format(cb, val_b, fmt_b)

            if val_e is not None:
                ce = safe_write(ws, row, 5, val_e)
                _apply_format(ce, val_e, fmt_e)
        except Exception as ex:
            log.error(f"FISA ACTIV rând {row} ({label}): {ex}")

    # ── Secțiunea 4 — MACRO ───────────────────────────────────
    if macro_start:
        fed_cur, fed_prv   = get_fred("FEDFUNDS")
        cpi_cur, cpi_prv   = get_fred("CPIAUCSL")
        unem_cur, unem_prv = get_fred("UNRATE")
        vix_d   = all_data.get("^VIX", {})
        tnx_d   = all_data.get("^TNX", {})
        irx_d   = all_data.get("^IRX", {})

        macro_rows = [
            ("VIX",           vix_d.get("inchidere") if vix_d else None,
                              None,   "Volatilitate"),
            ("Yield 10Y",     tnx_d.get("inchidere") if tnx_d else None,
                              None,   "Rata"),
            ("Yield 2Y",      irx_d.get("inchidere") if irx_d else None,
                              None,   "Rata"),
            ("Fear & Greed",  fear_greed.get("value"),
                              None,   fear_greed.get("status", "Neutru")),
            ("Rata dobanzii", fed_cur,  fed_prv,  "FED"),
            ("CPI",           cpi_cur,  cpi_prv,  "Inflatie"),
            ("Somaj",         unem_cur, unem_prv, "Piata muncii"),
        ]
        for offset, (ind, cur, prv, imp) in enumerate(macro_rows):
            mr = macro_start + offset
            safe_write(ws, mr, 1, ind)
            safe_write(ws, mr, 2, round(float(cur), 4) if cur is not None else "N/A")
            safe_write(ws, mr, 3, round(float(prv), 4) if prv is not None else "N/A")
            delta = None
            if cur and prv and float(prv) != 0:
                delta = round((float(cur) - float(prv)) / abs(float(prv)) * 100, 2)
            safe_write(ws, mr, 4, delta)
            safe_write(ws, mr, 5, imp)
            safe_write(ws, mr, 8, datetime.now().strftime("%d.%m.%Y"))

    # ── Secțiunea 5 — COMPETITORI (din activul selectat) ─────
    if comp_start:
        comp_key   = "MATERII" if cat_key == "MATERII_PRIME" else cat_key
        comp_names = COMPETITORI_MAP.get(comp_key, COMPETITORI_MAP["INDICI"])
        for offset, cname in enumerate(comp_names[:8]):
            cticker = ACTIVE.get(cname, "")
            cd = all_data.get(cticker, {})
            mr = comp_start + offset
            safe_write(ws, mr, 1, cname)
            safe_write(ws, mr, 2, cat_key)
            c5 = safe_write(ws, mr, 5, cd.get("inchidere") if cd else None)
            if c5: c5.number_format = "#,##0.0000"
            ct = safe_write(ws, mr, 6, cd.get("trend") if cd else None)
            if ct and cd: color_trend(ct, cd.get("trend", ""))
            cs = safe_write(ws, mr, 8, cd.get("semnal") if cd else None)
            if cs and cd: color_signal(cs, cd.get("semnal", "WAIT"))

    # ── Secțiunea 6 — RISCURI ────────────────────────────────
    if risk_start:
        risk_cat = "MATERII" if cat_key == "MATERII_PRIME" else cat_key
        risks    = RISK_LIBRARY.get(risk_cat, RISK_LIBRARY["INDICI"])
        for offset, r in enumerate(risks):
            mr = risk_start + offset
            safe_write(ws, mr, 1, r["ID"])
            safe_write(ws, mr, 2, r["Tip"])
            safe_write(ws, mr, 3, r["Categorie"])
            safe_write(ws, mr, 4, r["Descriere"])
            safe_write(ws, mr, 5, r["Impact"])
            safe_write(ws, mr, 6, r["Probabilitate"])
            safe_write(ws, mr, 7, round(r["Impact"] * r["Probabilitate"] / 100, 2))
            safe_write(ws, mr, 8, r["Orizont"])

    # ── Secțiunea 7 — CALENDAR ───────────────────────────────
    if cal_start:
        cal_cat  = "MATERII" if cat_key == "MATERII_PRIME" else cat_key
        cal_evts = CALENDAR_LIBRARY.get(cal_cat, [])
        for offset, evt in enumerate(cal_evts):
            mr = cal_start + offset
            safe_write(ws, mr, 1, "N/A")
            safe_write(ws, mr, 2, evt)
            safe_write(ws, mr, 8, d.get("name", sel_name))

    log.info(f"OK FISA ACTIV — activ selectat: {sel_name} [{cat_key}]")

# ─────────────────────────────────────────────────────────────
# SHEET 7 — REZUMAT EXECUTIV
# ─────────────────────────────────────────────────────────────
def update_rezumat_executiv(wb, all_data, fear_greed, macro_live):
    ws = wb["REZUMAT EXECUTIV"]
    ts = datetime.now().strftime("%d.%m.%Y %H:%M")

    def _sw(addr, val):
        cell = ws[addr]
        if not isinstance(cell, MergedCell):
            cell.value = val

    _sw("D3", ts)

    buy_n  = sum(1 for d in all_data.values() if d.get("semnal")=="BUY")
    sell_n = sum(1 for d in all_data.values() if d.get("semnal")=="SELL")
    wait_n = sum(1 for d in all_data.values() if d.get("semnal")=="WAIT")
    total  = len(all_data) or 1
    pct_b  = buy_n/total*100
    pct_s  = sell_n/total*100

    if   pct_b > 55:  trend_gen = "Bullish";  arrow = "↑"
    elif pct_s > 55:  trend_gen = "Bearish";  arrow = "↓"
    else:             trend_gen = "Mixt";     arrow = "→"

    vix_d   = all_data.get("^VIX", {})
    vix_val = safe(vix_d.get("inchidere"), 20) if vix_d else 20
    sp_d    = all_data.get("^GSPC", {})
    sp_rvol = safe(sp_d.get("rvol"), 1.0) if sp_d else 1.0

    if   vix_val > 30: vol_str = "Ridicata"
    elif vix_val > 20: vol_str = "Moderata"
    else:              vol_str = "Scazuta"

    if   sp_rvol > 1.3: vol_trx = "Crescut"
    elif sp_rvol > 0.7: vol_trx = "Normal"
    else:               vol_trx = "Scazut"

    if   vix_val > 30: risc_str = "Ridicat"
    elif vix_val > 20: risc_str = "Moderat"
    else:              risc_str = "Scazut"

    fg = fear_greed.get("display","N/A")

    sp_macross = sp_d.get("macross","") if sp_d else ""
    sp_semnal  = sp_d.get("semnal","WAIT") if sp_d else "WAIT"

    for row in range(4, 60):
        c1 = ws.cell(row=row, column=1)
        if isinstance(c1, MergedCell): continue
        label = str(c1.value).strip() if c1.value else ""
        try:
            if "Tendinta generala" in label or "Tendință generală" in label:
                safe_write(ws, row, 2, trend_gen)
                safe_write(ws, row, 4, arrow)
                safe_write(ws, row, 5, f"BUY:{buy_n} SELL:{sell_n} WAIT:{wait_n}")
            elif "Volatilitate" in label:
                safe_write(ws, row, 2, vol_str)
            elif "Volum" in label:
                safe_write(ws, row, 2, vol_trx)
            elif "Sentiment" in label:
                safe_write(ws, row, 2, fg)
            elif "Risc sistemic" in label or "Risc" in label:
                safe_write(ws, row, 2, risc_str)
        except Exception as ex:
            log.error(f"REZUMAT rând {row}: {ex}")

    # Concluzii — 5 bullets
    icon1 = "✅" if sp_semnal == "BUY" else "⚠️"
    icon3 = "✅" if "golden" in sp_macross.lower() else "⚠️"
    icon4 = "🛡️" if vix_val < 20 else "⚠️"
    bullets = [
        f"{icon1} Semnal general piata: {trend_gen} | S&P500: {sp_semnal}",
        f"🎯 VIX={vix_val:.1f} → Volatilitate {vol_str} | Risc sistemic: {risc_str}",
        f"{icon3} MA Cross S&P500: {sp_macross} | Momentum: {sp_d.get('momentum_10z',0):.1f}% (10z)" if sp_d else f"{icon3} Date S&P500 indisponibile",
        f"{icon4} Fear & Greed Index: {fg} | Sentiment: {fear_greed.get('status','N/A')}",
        f"📅 Actualizat: {ts} | Active analizate: {total} | BUY:{buy_n} SELL:{sell_n} WAIT:{wait_n}",
    ]
    for i, bullet in enumerate(bullets):
        for row in range(4, 60):
            c1 = ws.cell(row=row, column=1)
            if isinstance(c1, MergedCell): continue
            if str(c1.value or "").startswith(("✅","⚠️","🎯","🛡️","📅")):
                if i == 0 and "Semnal" in str(c1.value or ""):
                    safe_write(ws, row, 1, bullet); break
                elif i == 1 and "VIX" in str(c1.value or ""):
                    safe_write(ws, row, 1, bullet); break
                elif i == 2 and "MA Cross" in str(c1.value or ""):
                    safe_write(ws, row, 1, bullet); break
                elif i == 3 and "Fear" in str(c1.value or ""):
                    safe_write(ws, row, 1, bullet); break
                elif i == 4 and "Actualizat" in str(c1.value or ""):
                    safe_write(ws, row, 1, bullet); break

    log.info("OK REZUMAT EXECUTIV — actualizat")

# ─────────────────────────────────────────────────────────────
# SHEET 8 — ISTORIC TRENDING
# ─────────────────────────────────────────────────────────────
def update_historic(wb, all_data, fear_greed):
    ws = wb["ISTORIC TRENDING"]
    luna_an = datetime.now().strftime("%b %Y")

    last_row = 1
    for row in range(2, 5000):
        c1 = ws.cell(row=row, column=1)
        if isinstance(c1, MergedCell): continue
        if c1.value is None:
            last_row = row
            break
        if str(c1.value).strip() == luna_an:
            log.info(f"ISTORIC TRENDING — {luna_an} deja existent, skip")
            return
        last_row = row + 1

    all_rsi = [safe(d.get("rsi"),50) for d in all_data.values() if d]
    avg_rsi  = round(sum(all_rsi)/len(all_rsi),2) if all_rsi else 50.0
    sp_close = safe(all_data.get("^GSPC",{}).get("inchidere"),0)
    vix_close= safe(all_data.get("^VIX",{}).get("inchidere"),0)
    cpi_cur, _ = get_fred("CPIAUCSL")
    gdp_cur, _ = get_fred("GDP")

    buy_n  = sum(1 for d in all_data.values() if d.get("semnal")=="BUY")
    sell_n = sum(1 for d in all_data.values() if d.get("semnal")=="SELL")
    tot    = len(all_data) or 1
    if   buy_n/tot  > 0.5: semnal_luna = "BUY"
    elif sell_n/tot > 0.5: semnal_luna = "SELL"
    else:                   semnal_luna = "WAIT"

    r = last_row
    safe_write(ws, r, 1, luna_an)
    safe_write(ws, r, 2, avg_rsi)
    safe_write(ws, r, 3, sp_close)
    safe_write(ws, r, 4, gdp_cur)
    safe_write(ws, r, 5, cpi_cur)
    safe_write(ws, r, 6, vix_close)
    cell = safe_write(ws, r, 8, semnal_luna)
    if cell: color_signal(cell, semnal_luna)
    safe_write(ws, r, 11, fear_greed.get("value"))

    log.info(f"OK ISTORIC TRENDING — {luna_an} adaugat")

# ─────────────────────────────────────────────────────────────
# SHEET 9 — GHID INVATARE (text generation helpers)
# ─────────────────────────────────────────────────────────────
def explica_miscare(d: dict) -> str:
    name     = d.get("name","Activul")
    vz       = safe(d.get("var_zi_pct"),0)
    rvol     = safe(d.get("rvol"),1)
    rsi      = safe(d.get("rsi"),50)
    macd_c   = d.get("macd_cross","")
    macross  = d.get("macross","")
    bb_sup   = d.get("bb_sup"); bb_inf = d.get("bb_inf")
    price    = safe(d.get("inchidere"),0)
    ma50     = d.get("ma50"); ma200 = d.get("ma200")
    sl_v     = d.get("sl"); tp_v = d.get("tp")
    semnal   = d.get("semnal","WAIT")

    dir_txt = "crescut" if vz > 0 else "scazut"
    intens  = "semnificativ" if abs(vz) > 2 else "moderat" if abs(vz) > 0.5 else "marginal"
    vol_txt = "exceptionale" if rvol > 1.5 else "normale" if rvol > 0.7 else "scazute"
    rsi_txt = ("supravandut — potential rebound" if rsi < 30
               else "zona neutra" if rsi < 60
               else "zona supracumparare — prudenta recomandata")
    bb_txt = ""
    if bb_sup and bb_inf and price:
        if price > bb_sup * 0.99: bb_txt = "Pretul testeaza banda superioara Bollinger (potentiala rezistenta)."
        elif price < bb_inf * 1.01: bb_txt = "Pretul testeaza banda inferioara Bollinger (potential suport)."
        else: bb_txt = "Pretul se afla in interiorul benzilor Bollinger."
    ma_txt = ""
    if ma50 and ma200:
        ma_txt = f"Raport MA50/MA200: {macross}."
    sl_tp_txt = ""
    if sl_v and tp_v:
        sl_tp_txt = f"SL recomandat: {fmt_price(sl_v)} | TP: {fmt_price(tp_v)}."

    return (f"{name} a {dir_txt} {intens} cu {abs(vz):.2f}% in sedinta curenta. "
            f"Volumele sunt {vol_txt} (RVOL={rvol:.1f}x). "
            f"RSI({rsi:.1f}): {rsi_txt}. "
            f"MACD: {macd_c}. "
            f"{ma_txt} {bb_txt} "
            f"Semnal tehnic: {semnal}. {sl_tp_txt}")

def identifica_oportunitate(d: dict) -> str:
    semnal = d.get("semnal","WAIT")
    rsi    = safe(d.get("rsi"),50)
    rvol   = safe(d.get("rvol"),1)
    macross= d.get("macross","")
    name   = d.get("name","Activul")
    if semnal == "BUY" and rsi < 45:
        return f"✅ OPORTUNITATE: {name} prezinta semnal BUY cu RSI scazut ({rsi:.0f}) — potential entry atractiv."
    elif semnal == "BUY" and "golden" in macross.lower():
        return f"✅ OPORTUNITATE: Golden Cross confirmat pe {name} cu semnal BUY activ."
    elif semnal == "SELL" and rsi > 70:
        return f"⚠️ RISC: {name} supraevaluat (RSI={rsi:.0f}) cu semnal SELL — potential short sau exit long."
    elif rvol > 2.0:
        return f"🔥 ATENTIE: Volum exceptional pe {name} (RVOL={rvol:.1f}x) — miscare semnificativa posibila."
    else:
        return f"⏸ {name} in zona de asteptare (WAIT) — se monitorizeaza confirmare semnal."

def extrage_lectie(d: dict) -> str:
    rsi    = safe(d.get("rsi"),50)
    macross= d.get("macross","")
    macd_c = d.get("macd_cross","")
    semnal = d.get("semnal","WAIT")
    rvol   = safe(d.get("rvol"),1)
    if "golden" in macross.lower():
        return "Lectie: Golden Cross (MA50>MA200) este un semnal bullish pe termen lung, confirmat de MACD."
    elif "death" in macross.lower():
        return "Lectie: Death Cross (MA50<MA200) semnaleaza potentiala tendinta descendenta — fii prudent cu pozitiile long."
    elif rsi < 30:
        return "Lectie: RSI sub 30 indica supravanzare — pretul poate reveni, dar asteapta confirmare MACD."
    elif rsi > 70:
        return "Lectie: RSI peste 70 indica supracumparare — risc de corectie. NU intra in long la extrema RSI."
    elif "pozitiv nou" in macd_c.lower():
        return "Lectie: Crossover MACD pozitiv nou = potential inceput de tendinta ascendenta — oportunitate entry."
    elif rvol > 1.5:
        return "Lectie: Volum ridicat confirma miscarile de pret — un breakout cu volum este mai credibil."
    else:
        return "Lectie: In absenta unui semnal clar, cash-ul este o pozitie valida. Rabdarea este o virtute in trading."

def write_ghid_grafice(ws, start_row: int) -> int:
    row = start_row + 2

    def _hdr(r, text):
        cell = ws.cell(row=r, column=1)
        if not isinstance(cell, MergedCell):
            cell.value = text
            cell.fill  = copy(F_GHID_CAP)
            cell.font  = copy(FN_HDR)
        return r + 1

    def _sub(r, col, text):
        cell = ws.cell(row=r, column=col)
        if not isinstance(cell, MergedCell):
            cell.value = text
            cell.fill  = copy(F_GHID_SUB)
            cell.font  = fnt(bold=True)
        return r

    def _txt(r, col, text):
        cell = ws.cell(row=r, column=col)
        if not isinstance(cell, MergedCell):
            cell.value = text
            cell.fill  = copy(F_GHID_TXT)
            cell.font  = fnt(size=9)
            cell.alignment = aln(wrap=True)
        return r

    sections = [
        ("CITIRE CANDLESTICK", [
            ("Corp verde (bullish):", "Pretul de inchidere > deschidere — cumparatorii au dominat sedinta."),
            ("Corp rosu (bearish):", "Pretul de inchidere < deschidere — vanzatorii au dominat sedinta."),
            ("Wick superior lung:", "Respingere la rezistenta — semnal de slabiciune bullish."),
            ("Wick inferior lung:", "Respingere la suport — semnal de forta bullish potential."),
        ]),
        ("RSI — Relative Strength Index", [
            ("RSI < 30:", "Zona supravanzare — potential rebound, dar confirma cu alt indicator."),
            ("RSI 30-70:", "Zona neutra — tendinta dominanta in vigoare."),
            ("RSI > 70:", "Zona supracumparare — risc de corectie, nu intra in long la extrema."),
        ]),
        ("MACD — Moving Average Convergence Divergence", [
            ("Cross pozitiv:", "Linia MACD traverseaza semnalul de jos in sus — semnal BUY potential."),
            ("Cross negativ:", "Linia MACD traverseaza semnalul de sus in jos — semnal SELL potential."),
            ("Divergenta bullish:", "Pretul scade dar MACD creste — inversare bullish posibila."),
        ]),
        ("BENZILE BOLLINGER", [
            ("Pret langa banda sup:", "Activ in zona supracumparare relativa — potential reversal."),
            ("Pret langa banda inf:", "Activ in zona supravanzare relativa — potential rebound."),
            ("Banda ingusta:", "Volatilitate scazuta — breakout iminent, pregateste-te."),
        ]),
        ("MA CROSSES", [
            ("Golden Cross:", "MA50 trece peste MA200 — semnal bullish pe termen lung, confirmat de volum."),
            ("Death Cross:", "MA50 trece sub MA200 — semnal bearish pe termen lung."),
        ]),
        ("STOCHASTIC OSCILLATOR", [
            ("%K sub 20:", "Supravanzat — posibil entry long daca %K incepe sa urce."),
            ("%K peste 80:", "Supracumparat — risc de corectie, evita intrarea in long."),
        ]),
        ("ATR — VOLATILITATE", [
            ("ATR mare:", "Miscare mare asteptata — ajusteaza SL mai larg."),
            ("ATR mic:", "Piata lenta — spread-urile sunt proportional mai mari ca % din ATR."),
        ]),
        ("SEMNALE BUY / SELL / WAIT", [
            ("BUY (Score >= 3):", "Minimum 3 confluente pozitive: RSI+MACD+MA+RVOL."),
            ("SELL (Score <= -3):", "Minimum 3 confluente negative."),
            ("WAIT:", "Confluente insuficiente — pastreaza cash, asteapta confirmare."),
        ]),
        ("RISK MANAGEMENT — SL / TP / R/R", [
            ("SL = 1.5x ATR:", "Stop Loss plasat la 1.5 multipli ATR fata de entry."),
            ("TP = 3.0x ATR:", "Take Profit la 3x ATR — R/R minim 2:1 recomandat."),
            ("R/R >= 2:", "Accepta tranzactia doar daca recompensa e de 2x riscul."),
        ]),
        ("PSIHOLOGIE TRADING", [
            ("FOMO:", "Fear Of Missing Out — cel mai frecvent motiv pentru intrari proaste."),
            ("Disciplina:", "Respecta SL-ul INTOTDEAUNA. O pierdere mica e mai buna decat una mare."),
            ("Jurnalul de tranzactii:", "Noteaza fiecare trade — pattern-urile se repeta, invata din ele."),
        ]),
    ]

    for sec_title, items in sections:
        row = _hdr(row, sec_title)
        for sub, txt in items:
            _sub(row, 1, sub)
            _txt(row, 2, txt)
            row += 1
        row += 1

    return row

def update_ghid_invatare(wb, all_data):
    sheet_name = "GHID INVATARE"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]          # ← șterge direct după nume string
    ws = wb.create_sheet(sheet_name)

    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 30

    buy_n  = sum(1 for d in all_data.values() if d.get("semnal")=="BUY")
    sell_n = sum(1 for d in all_data.values() if d.get("semnal")=="SELL")
    wait_n = sum(1 for d in all_data.values() if d.get("semnal")=="WAIT")
    n_total= len(all_data)
    ts     = datetime.now().strftime("%d.%m.%Y %H:%M")

    # Titlu principal
    c1 = ws.cell(row=1, column=1)
    c1.value = (f"GHID DE INVATARE ZILNIC — ANALIZA DE PIATA PROFESIONALA\n"
                f"Generat: {ts} | {n_total} active analizate | BUY:{buy_n} SELL:{sell_n} WAIT:{wait_n}")
    c1.fill  = copy(F_TITLU)
    c1.font  = copy(FN_TITLU)
    c1.alignment = aln(h="center", wrap=True)
    ws.row_dimensions[1].height = 40

    row = 3
    bloc_count = 0

    for label, cat_dict in CATEGORII:
        # Header categorie
        hdr_cell = ws.cell(row=row, column=1)
        hdr_cell.value = f"══ {label} ══"
        hdr_cell.fill  = copy(F_CAT)
        hdr_cell.font  = copy(FN_CAT)
        row += 1

        for name, ticker in cat_dict.items():
            d = all_data.get(ticker, {})
            if not d:
                continue
            try:
                semnal  = d.get("semnal","WAIT")
                price   = safe(d.get("inchidere"),0)
                vz      = safe(d.get("var_zi_pct"),0)

                # ── Rând HEADER ACTIV ──────────────────────────────────
                if   semnal == "BUY":  hdr_fill = copy(F_ACTIV_BUY)
                elif semnal == "SELL": hdr_fill = copy(F_ACTIV_SEL)
                else:                  hdr_fill = copy(F_ACTIV_WAI)

                hdr_text = (f"▶ {name} ({ticker}) | {semnal} | "
                            f"Pret: {fmt_price(price)} | Var: {fmt_pct(vz)}")
                for col in range(1, 4):
                    cell = ws.cell(row=row, column=col)
                    if not isinstance(cell, MergedCell):
                        cell.fill = copy(hdr_fill)
                        cell.font = copy(FN_ACTIV)
                        cell.alignment = aln(h="left", wrap=False)
                        if col == 1:
                            cell.value = hdr_text
                row += 1

                # ── Rând PREȚURI / INDICATORI ─────────────────────────
                rsi_v    = safe(d.get("rsi"),50)
                stk      = safe(d.get("stoch_k"),50)
                std      = safe(d.get("stoch_d"),50)
                atr_v    = safe(d.get("atr"),0)
                rvol_v   = safe(d.get("rvol"),1)
                volum_v  = d.get("volum",0)
                vs       = safe(d.get("var_sapt_pct"),0)
                vl       = safe(d.get("var_luna_pct"),0)

                row_data = [
                    f"Pret: {fmt_price(price)} | Zi: {fmt_pct(vz)} | Sapt: {fmt_pct(vs)} | Luna: {fmt_pct(vl)}",
                    f"RSI: {rsi_v:.1f} [{d.get('rsi_status','')}] | Stoch: K={stk:.1f} D={std:.1f}",
                    f"ATR: {fmt_price(atr_v,4)} | RVOL: {rvol_v:.1f}x | Vol: {volum_v:,}",
                ]
                for col, txt in enumerate(row_data, start=1):
                    cell = ws.cell(row=row, column=col)
                    if not isinstance(cell, MergedCell):
                        cell.value = txt
                        cell.fill  = copy(F_SEC_HDR)
                        cell.font  = fnt(color="FFFFFF", size=9)
                        cell.alignment = aln(wrap=True)
                row += 1

                # ── Rând MEDII MOBILE + SEMNALE ───────────────────────
                ma20_v   = d.get("ma20"); ma50_v = d.get("ma50"); ma200_v = d.get("ma200")
                macross  = d.get("macross","")
                macd_v   = safe(d.get("macd"),0)
                msig_v   = safe(d.get("macd_signal"),0)
                mhst_v   = safe(d.get("macd_hist"),0)
                macd_c   = d.get("macd_cross","")
                bb_sup_v = d.get("bb_sup"); bb_inf_v = d.get("bb_inf")
                sl_v     = d.get("sl"); tp_v = d.get("tp")

                ma_txt   = (f"MA20:{fmt_price(ma20_v,2)} MA50:{fmt_price(ma50_v,2)} "
                            f"MA200:{fmt_price(ma200_v,2)} → {macross}")
                macd_txt = (f"MACD:{macd_v:.4f} Signal:{msig_v:.4f} "
                            f"Hist:{mhst_v:.4f} → {macd_c}")
                bb_txt   = (f"BB Sup:{fmt_price(bb_sup_v,2)} BB Inf:{fmt_price(bb_inf_v,2)} "
                            f"| SL:{fmt_price(sl_v)} TP:{fmt_price(tp_v)}")

                row_ind = [ma_txt, macd_txt, bb_txt]
                for col, txt in enumerate(row_ind, start=1):
                    cell = ws.cell(row=row, column=col)
                    if not isinstance(cell, MergedCell):
                        cell.value = txt
                        cell.fill  = copy(F_BLU)
                        cell.font  = copy(FN_BLU)
                        cell.alignment = aln(wrap=True)
                        if col == 1:
                            if "golden" in macross.lower():
                                cell.fill = copy(F_GRN); cell.font = copy(FN_GRN)
                            elif "death" in macross.lower():
                                cell.fill = copy(F_RED); cell.font = copy(FN_RED)
                row += 1

                # ── Rând EXPLICAȚIE MIȘCARE ───────────────────────────
                expl_txt = explica_miscare(d)
                prob_v   = safe(d.get("probabilitate"),0)
                conf_v   = safe(d.get("confluente"),0)
                score_v  = safe(d.get("score"),0)
                rr_v     = rr_text(d.get("inchidere"), sl_v, tp_v)

                cell_expl = ws.cell(row=row, column=1)
                if not isinstance(cell_expl, MergedCell):
                    cell_expl.value = expl_txt
                    cell_expl.fill  = copy(F_BLU)
                    cell_expl.font  = fnt(color="1F4E79", size=9, italic=True)
                    cell_expl.alignment = aln(wrap=True)

                # col 2 — gol (vizual merge simulat)
                cell_b = ws.cell(row=row, column=2)
                if not isinstance(cell_b, MergedCell):
                    cell_b.fill = copy(F_BLU)

                cell_rr = ws.cell(row=row, column=3)
                if not isinstance(cell_rr, MergedCell):
                    cell_rr.value = (f"R/R: {rr_v}\n"
                                     f"Prob: {prob_v:.0f}%\n"
                                     f"Conf: {conf_v:.0f}/5\n"
                                     f"Score: {score_v:.0f}")
                    cell_rr.fill  = copy(F_LBL)
                    cell_rr.font  = fnt(bold=True, size=9)
                    cell_rr.alignment = aln(h="center", wrap=True)
                ws.row_dimensions[row].height = 60
                row += 1

                # ── Rând OPORTUNITATE / LECȚIE ────────────────────────
                op_txt  = identifica_oportunitate(d)
                lec_txt = extrage_lectie(d)

                cell_op = ws.cell(row=row, column=1)
                if not isinstance(cell_op, MergedCell):
                    cell_op.value = op_txt
                    cell_op.fill  = copy(F_GRN if semnal=="BUY" else F_RED if semnal=="SELL" else F_YLW)
                    cell_op.font  = copy(FN_GRN if semnal=="BUY" else FN_RED if semnal=="SELL" else FN_YLW)
                    cell_op.alignment = aln(wrap=True)

                cell_op2 = ws.cell(row=row, column=2)
                if not isinstance(cell_op2, MergedCell):
                    cell_op2.fill = copy(F_GRN if semnal=="BUY" else F_RED if semnal=="SELL" else F_YLW)

                cell_lec = ws.cell(row=row, column=3)
                if not isinstance(cell_lec, MergedCell):
                    cell_lec.value = lec_txt
                    cell_lec.fill  = copy(F_ORG)
                    cell_lec.font  = copy(FN_ORG)
                    cell_lec.alignment = aln(wrap=True)
                ws.row_dimensions[row].height = 40
                row += 1

                # ── Separator ─────────────────────────────────────────
                for col in range(1, 4):
                    cell = ws.cell(row=row, column=col)
                    if not isinstance(cell, MergedCell):
                        cell.fill = copy(F_SEP)
                        cell.value = None
                ws.row_dimensions[row].height = 4
                row += 1

                bloc_count += 1

            except Exception as ex:
                log.error(f"GHID INVATARE bloc {name}: {ex}")
                row += 1

    # ── Ghid grafice la final ──────────────────────────────────────────
    row = write_ghid_grafice(ws, row)

    log.info(f"OK GHID INVATARE — {bloc_count} blocuri scrise + ghid grafice")


# ─────────────────────────────────────────────────────────────
# MAIN ORCHESTRATOR
# ─────────────────────────────────────────────────────────────
def main():
    sep = "=" * 44
    log.info(sep)
    log.info(" PORNIRE ACTUALIZARE UNIFICATA COMPLETA")

    n_indici  = len(INDICI)
    n_actiuni = len(ACTIUNI)
    n_crypto  = len(CRYPTO)
    n_valute  = len(VALUTE)
    n_materii = len(MATERII_PRIME)
    n_total   = len(ACTIVE)
    log.info(f" Active: {n_total} total ({n_indici} indici | {n_actiuni} actiuni | "
             f"{n_crypto} crypto | {n_valute} valute | {n_materii} materii prime)")
    log.info(sep)

    # ── Verificare fișier ─────────────────────────────────────
    if not EXCEL_PATH.exists():
        log.error(f"Fisierul nu exista: {EXCEL_PATH}")
        sys.exit(1)

    # ── Preluare date ACTIVE (paralel) ────────────────────────
    all_data   = {}
    failed     = []
    items_list = list(ACTIVE.items())

    log.info("Preluare date live...")
    with ThreadPoolExecutor(max_workers=10) as executor:
        futures = {
            executor.submit(get_full_data, name, ticker): (name, ticker)
            for name, ticker in items_list
        }
        for i, future in enumerate(as_completed(futures), start=1):
            name, ticker = futures[future]
            try:
                result = future.result(timeout=30)
                all_data[ticker] = result
                if result:
                    price  = result.get("inchidere", 0)
                    vz     = result.get("var_zi_pct", 0)
                    rsi    = result.get("rsi", 50)
                    semnal = result.get("semnal", "WAIT")
                    log.info(f"[{i:3d}/{n_total}] {name:<22} "
                             f"{price:>14,.4f}  ({vz:+.2f}%)  "
                             f"RSI={rsi:.1f}  [{semnal}]")
                else:
                    log.warning(f"[{i:3d}/{n_total}] {name:<22} — date indisponibile")
                    failed.append(name)
            except Exception as ex:
                log.error(f"[{i:3d}/{n_total}] {name}: {ex}")
                all_data[ticker] = {}
                failed.append(name)

    # ── Fear & Greed ──────────────────────────────────────────
    fear_greed = get_fear_greed()
    log.info(f"Fear & Greed: {fear_greed.get('display','N/A')}")

    # ── Date MACRO ────────────────────────────────────────────
    macro_live = {}
    for mname, mticker in MACRO_TICKERS.items():
        md = get_full_data(mname, mticker)
        macro_live[mname] = md
        if md:
            log.info(f"MACRO {mname}: {md.get('inchidere','N/A')}")

    # ── Deschide workbook ─────────────────────────────────────
    log.info(f"Deschid: {EXCEL_PATH}")
    wb = load_workbook(str(EXCEL_PATH))

    # ── Execuție tasks ────────────────────────────────────────
    tasks = [
        ("PRETURI VOLUME",     update_preturi_volume,     [all_data]),
        ("INDICATORI TEHNICI", update_indicatori_tehnici, [all_data]),
        ("SEMNALE INTRARE",    update_semnale,            [all_data]),
        ("INDICATORI MACRO",   update_macro,              [macro_live, fear_greed]),
        ("DASHBOARD",          update_dashboard,          [all_data, fear_greed]),
        ("FISA ACTIV",         update_fisa_activ,         [all_data, fear_greed, wb]),
        ("REZUMAT EXECUTIV",   update_rezumat_executiv,   [all_data, fear_greed, macro_live]),
        ("ISTORIC TRENDING",   update_historic,           [all_data, fear_greed]),
        ("GHID INVATARE",      update_ghid_invatare,      [all_data]),
    ]

    ok_tasks = 0
    for sheet_name, func, args in tasks:
        # GHID INVATARE poate fi creat — nu verificam existenta
        if sheet_name != "GHID INVATARE" and sheet_name not in wb.sheetnames:
            log.warning(f"Sheet '{sheet_name}' nu exista in workbook — skip")
            continue
        try:
            func(wb, *args)
            ok_tasks += 1
        except Exception as ex:
            log.error(f"EROARE la task '{sheet_name}': {ex}", exc_info=True)

    # ── Salvare ───────────────────────────────────────────────
    wb.save(str(EXCEL_PATH))
    log.info(f"SALVAT: {EXCEL_PATH.name}")

    ok_active = n_total - len(failed)
    ts_final  = datetime.now().strftime("%H:%M:%S")
    log.info(sep)
    log.info(f"Procesate: {ok_active}/{n_total} active OK | {len(failed)} esuate")
    if failed:
        log.warning(f"Esuate: {', '.join(failed[:10])}{'...' if len(failed)>10 else ''}")
    log.info(f"Finalizat: {ts_final}")
    log.info(sep)


if __name__ == "__main__":
    main()