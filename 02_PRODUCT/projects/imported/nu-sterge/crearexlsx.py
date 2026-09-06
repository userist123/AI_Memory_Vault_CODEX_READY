import os
"""
╔══════════════════════════════════════════════════════════════════════════════╗
║  ANALIZA PIATA PROFESIONALA — Script Complet v3.1                          ║
║  80+ active | Dashboard dinamic | Ghid Invatare | Office 2021              ║
╚══════════════════════════════════════════════════════════════════════════════╝
INSTALARE:  pip install yfinance openpyxl requests pandas
RULARE:     python piata_v3.py
"""
import sys, shutil, logging
from datetime import datetime, timedelta
from pathlib import Path
import requests, pandas as pd, yfinance as yf
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter as gcl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule

# ═══════════════════════════════════════════════════════════════════
# CONFIGURARE
# ═══════════════════════════════════════════════════════════════════
EXCEL_TEMPLATE = Path(r"C:\Users\Marius\Desktop\Analiza_Piata_Profesionala.xlsx")
FRED_API_KEY   = os.environ.get("FRED_API_KEY", "")

INDICI = {
    "S&P 500":"^GSPC","NASDAQ 100":"^NDX","NASDAQ Comp.":"^IXIC",
    "Dow Jones":"^DJI","Russell 2000":"^RUT","DAX Germany":"^GDAXI",
    "FTSE 100":"^FTSE","CAC 40":"^FCHI","Nikkei 225":"^N225",
    "Hang Seng":"^HSI","Shanghai Comp.":"000001.SS",
    "MSCI World ETF":"URTH","MSCI EM ETF":"EEM","BET Romania":"BET.RO",
}
ACTIUNI = {
    "Apple":"AAPL","Microsoft":"MSFT","NVIDIA":"NVDA","Alphabet":"GOOGL",
    "Amazon":"AMZN","Meta":"META","Tesla":"TSLA","AMD":"AMD","Intel":"INTC",
    "Broadcom":"AVGO","ASML":"ASML","Taiwan Semi":"TSM","Palantir":"PLTR",
    "Salesforce":"CRM","Oracle":"ORCL","JPMorgan":"JPM","Goldman Sachs":"GS",
    "Berkshire B":"BRK-B","Visa":"V","Mastercard":"MA","ExxonMobil":"XOM",
    "Chevron":"CVX","Shell":"SHEL","Caterpillar":"CAT","Boeing":"BA",
    "SPY":"SPY","QQQ":"QQQ","GLD ETF":"GLD","TLT Bond ETF":"TLT","ARKK":"ARKK",
}
CRYPTO = {
    "Bitcoin":"BTC-USD","Ethereum":"ETH-USD","BNB":"BNB-USD","Solana":"SOL-USD",
    "XRP":"XRP-USD","Cardano":"ADA-USD","Avalanche":"AVAX-USD","Dogecoin":"DOGE-USD",
    "Chainlink":"LINK-USD","Polkadot":"DOT-USD","Litecoin":"LTC-USD",
    "Shiba Inu":"SHIB-USD","Polygon":"MATIC-USD","Uniswap":"UNI-USD",
    "Cosmos":"ATOM-USD","Stellar":"XLM-USD","Monero":"XMR-USD","Tron":"TRX-USD",
    "Filecoin":"FIL-USD","Aave":"AAVE-USD","Arbitrum":"ARB-USD",
    "Optimism":"OP-USD","Render":"RNDR-USD","Sui":"SUI-USD","Near Protocol":"NEAR-USD",
}
VALUTE = {
    "EUR/USD":"EURUSD=X","GBP/USD":"GBPUSD=X","USD/JPY":"USDJPY=X",
    "USD/CHF":"USDCHF=X","AUD/USD":"AUDUSD=X","USD/CAD":"USDCAD=X",
    "EUR/RON":"EURRON=X","USD/RON":"USDRON=X","GBP/RON":"GBPRON=X",
    "EUR/GBP":"EURGBP=X","USD/CNY":"USDCNY=X","USD/TRY":"USDTRY=X",
}
MATERII = {
    "Gold":"GC=F","Silver":"SI=F","Platinum":"PL=F","Palladium":"PA=F",
    "Oil WTI":"CL=F","Oil Brent":"BZ=F","Natural Gas":"NG=F","Copper":"HG=F",
    "Corn":"ZC=F","Wheat":"ZW=F","Soybean":"ZS=F","Coffee":"KC=F",
    "Sugar":"SB=F","Cotton":"CT=F",
}
MACRO_T = {"VIX":"^VIX","Yield 10Y US":"^TNX","Yield 2Y US":"^IRX","USD Index":"DX-Y.NYB"}
ACTIVE    = {**INDICI,**ACTIUNI,**CRYPTO,**VALUTE,**MATERII}
CATEGORII = [
    ("INDICI BURSIERI",INDICI),("ACTIUNI & ETF",ACTIUNI),
    ("CRYPTOCURRENCY",CRYPTO),("VALUTE FOREX",VALUTE),("MATERII PRIME",MATERII),
]

def categorie_activ(name):
    if name in INDICI:  return "INDICI"
    if name in ACTIUNI: return "ACTIUNI"
    if name in CRYPTO:  return "CRYPTO"
    if name in VALUTE:  return "VALUTE"
    return "MATERII"

# ═══════════════════════════════════════════════════════════════════
# LOGGING
# ═══════════════════════════════════════════════════════════════════
logging.basicConfig(
    level=logging.INFO, format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
    handlers=[logging.StreamHandler(sys.stdout),
              logging.FileHandler(Path(__file__).parent/"analiza.log",encoding="utf-8")],
)
log = logging.getLogger(__name__)

# ═══════════════════════════════════════════════════════════════════
# STILURI
# ═══════════════════════════════════════════════════════════════════
def fill(h): return PatternFill("solid",start_color=h,fgColor=h)
def fnt(bold=False,color="000000",size=10,italic=False):
    return Font(name="Arial",bold=bold,color=color,size=size,italic=italic)
def aln(h="left",v="center",wrap=True):
    return Alignment(horizontal=h,vertical=v,wrap_text=wrap)
def tb(color="BFBFBF"):
    s=Side(style="thin",color=color); return Border(left=s,right=s,top=s,bottom=s)

NO_FILL=PatternFill(fill_type=None)
F_H =fill("1F4E79");FN_H =fnt(bold=True,color="FFFFFF",size=10)
F_RH=fill("D6DCE4");FN_RH=fnt(bold=True,color="000000",size=10)
F_IN=fill("D9E1F2");FN_IN=fnt(color="1F4E79",size=10)
F_CA=fill("F2F2F2");FN_CA=fnt(color="595959",size=10)
F_GN=fill("C6EFCE");FN_GN=fnt(color="375623",bold=True)
F_YL=fill("FFEB9C");FN_YL=fnt(color="9C6500",bold=True)
F_RD=fill("FFC7CE");FN_RD=fnt(color="9C0006",bold=True)
F_OR=fill("FCE4D6");FN_OR=fnt(color="833C00")
F_BL=fill("DEEAF1");FN_BL=fnt(color="1F4E79",size=10)
F_WH=fill("FFFFFF")
F_TI=fill("0D2137");FN_TI=fnt(bold=True,color="FFFFFF",size=14)
F_CT=fill("1F4E79");FN_CT=fnt(bold=True,color="FFFFFF",size=11)
F_GT=fill("0A1628");FN_GT=fnt(bold=True,color="FFFFFF",size=14)
F_GC=fill("154360");FN_GC=fnt(bold=True,color="FFFFFF",size=11)
F_GS=fill("1A5276");FN_GS=fnt(bold=True,color="FFFFFF",size=10)
F_GX=fill("EBF5FB");FN_GX=fnt(color="1A252F",size=10)
F_SL=fill("FFF2CC");FN_SL=fnt(bold=True,color="7D5A00",size=13)
F_SH=fill("ED7D31");FN_SH=fnt(bold=True,color="FFFFFF",size=11)

def set_h(ws,row,h): ws.row_dimensions[row].height=h

def mc(ws,r1,c1,r2,c2,val=None,f=None,fn=None,ha="left",va="center"):
    cell=ws.cell(r1,c1)
    if val is not None: cell.value=val
    if f: cell.fill=f
    if fn: cell.font=fn
    cell.alignment=aln(ha,va)
    if r2>r1 or c2>c1:
        ws.merge_cells(start_row=r1,start_column=c1,end_row=r2,end_column=c2)
    return cell

def hdr_row(ws,row,headers,col=1):
    for i,h in enumerate(headers):
        c=ws.cell(row,col+i,h)
        c.fill=F_H;c.font=FN_H;c.alignment=aln("center");c.border=tb()

def clr(ws,start=3,end=1000):
    for row in ws.iter_rows(min_row=start,max_row=end):
        for c in row: c.value=None;c.fill=NO_FILL

def unmerge_row(ws,row,c1,c2):
    to_rm=[str(m) for m in ws.merged_cells.ranges
           if m.min_row==row and m.min_col>=c1 and m.max_col<=c2]
    for mr in to_rm:
        try: ws.merged_cells.remove(mr)
        except: pass

def cf_signal(ws,rng):
    ws.conditional_formatting.add(rng,CellIsRule("equal",['"BUY"'], fill=F_GN,font=FN_GN))
    ws.conditional_formatting.add(rng,CellIsRule("equal",['"SELL"'],fill=F_RD,font=FN_RD))
    ws.conditional_formatting.add(rng,CellIsRule("equal",['"WAIT"'],fill=F_YL,font=FN_YL))

def cf_status(ws,rng):
    cf_signal(ws,rng)
    for v,f_,fn_ in [("Pozitiv",F_GN,FN_GN),("Negativ",F_RD,FN_RD),("Neutru",F_YL,FN_YL),
                      ("Bullish",F_GN,FN_GN),("Bearish",F_RD,FN_RD),("Sideways",F_YL,FN_YL)]:
        ws.conditional_formatting.add(rng,CellIsRule("equal",[f'"{v}"'],fill=f_,font=fn_))

# ═══════════════════════════════════════════════════════════════════
# FORMAT HELPERS
# ═══════════════════════════════════════════════════════════════════
def fp(val,d=4):
    if val is None: return "N/A"
    try:
        f=float(val)
        if f!=f: return "N/A"
        return f"{{:,.{d}f}}".format(f)
    except: return "N/A"

def fpc(val,d=2):
    if val is None: return "N/A"
    try:
        f=float(val)
        return f"{'+' if f>=0 else ''}{f:.{d}f}%"
    except: return "N/A"

def safe(val,default=0.0):
    if val is None: return default
    try:
        f=float(val); return default if f!=f else f
    except: return default

# ═══════════════════════════════════════════════════════════════════
# INDICATORI TEHNICI
# ═══════════════════════════════════════════════════════════════════
def calc_rsi(prices,p=14):
    d=prices.diff()
    g=d.clip(lower=0).rolling(p).mean()
    l=(-d.clip(upper=0)).rolling(p).mean()
    rs=g/l.replace(0,1e-10)
    v=(100-100/(1+rs)).iloc[-1]
    return round(float(v),2) if pd.notna(v) else 50.0

def calc_macd(prices):
    e12=prices.ewm(span=12,adjust=False).mean()
    e26=prices.ewm(span=26,adjust=False).mean()
    ml=e12-e26; sl=ml.ewm(span=9,adjust=False).mean(); hs=ml-sl
    m,s,h=float(ml.iloc[-1]),float(sl.iloc[-1]),float(hs.iloc[-1])
    ph=float(hs.iloc[-2]) if len(hs)>1 else 0
    if m>s:
        cross="Crossover bullish" if ph<0 else "Momentum ascendent"
    else:
        cross="Crossover bearish" if ph>0 else "Momentum descendent"
    return {"macd":round(m,6),"signal":round(s,6),"hist":round(h,6),"cross":cross}

def calc_ma(prices):
    n=len(prices)
    def ma(p): return round(float(prices.tail(p).mean()),6) if n>=p else None
    m20,m50,m200=ma(20),ma(50),ma(200)
    cross=("Golden Cross" if m50 and m200 and m50>m200
           else "Death Cross" if m50 and m200 and m50<m200
           else "Neutru")
    return {"ma20":m20,"ma50":m50,"ma200":m200,"macross":cross}

def calc_boll(prices,p=20):
    if len(prices)<p: return {"bb_sup":None,"bb_inf":None,"bb_w":None}
    m=prices.rolling(p).mean(); sd=prices.rolling(p).std()
    s=float((m+2*sd).iloc[-1]); i=float((m-2*sd).iloc[-1])
    return {"bb_sup":round(s,6),"bb_inf":round(i,6),"bb_w":round(s-i,6)}

def calc_atr(hist,p=14):
    hi,lo,cl=hist["High"],hist["Low"],hist["Close"]
    tr=pd.concat([(hi-lo),(hi-cl.shift()).abs(),(lo-cl.shift()).abs()],axis=1).max(axis=1)
    v=tr.rolling(p).mean().iloc[-1]
    return round(float(v),6) if pd.notna(v) else 0.0

def calc_stoch(hist,p=14):
    if len(hist)<p: return {"k":50.0,"d":50.0}
    lo,hi,cl=hist["Low"],hist["High"],hist["Close"]
    l14=lo.rolling(p).min(); h14=hi.rolling(p).max()
    k=(cl-l14)/(h14-l14).replace(0,1e-10)*100; d=k.rolling(3).mean()
    return {"k":round(float(k.iloc[-1]),2) if pd.notna(k.iloc[-1]) else 50.0,
            "d":round(float(d.iloc[-1]),2) if pd.notna(d.iloc[-1]) else 50.0}

def calc_semnal(rsi,macd_cross,macross,rvol):
    sc=0
    if rsi<35:   sc+=2
    elif rsi<45: sc+=1
    elif rsi>75: sc-=2
    elif rsi>65: sc-=1
    if "bullish" in macd_cross.lower():  sc+=2
    elif "ascendent" in macd_cross.lower(): sc+=1
    elif "bearish" in macd_cross.lower():  sc-=2
    elif "descendent" in macd_cross.lower(): sc-=1
    if macross=="Golden Cross": sc+=2
    elif macross=="Death Cross": sc-=2
    if rvol>1.5: sc+=1
    elif rvol<0.6: sc-=1
    conf=min(abs(sc),5)
    if sc>=3:    return "BUY",conf
    elif sc<=-3: return "SELL",conf
    else:        return "WAIT",conf

# ═══════════════════════════════════════════════════════════════════
# PRELUARE DATE
# ═══════════════════════════════════════════════════════════════════
def get_data(name,ticker):
    try:
        t=yf.Ticker(ticker)
        hist=t.history(period="1y",auto_adjust=True)
        if hist.empty or len(hist)<5:
            log.warning(f"  {name} — date insuficiente"); return {}
        cl=hist["Close"]; n=len(cl)
        lt=hist.iloc[-1]; pv=hist.iloc[-2] if n>1 else lt
        now=float(lt["Close"]); prev=float(pv["Close"])
        c5=float(hist.iloc[-min(5,n)]["Close"]); c20=float(hist.iloc[-min(20,n)]["Close"])
        vzi  =(now-prev)/prev*100 if prev else 0
        vsapt=(now-c5)/c5*100    if c5   else 0
        vluna=(now-c20)/c20*100  if c20  else 0
        vol=int(lt.get("Volume",0))
        avg=int(hist["Volume"].tail(min(20,n)).mean()) if "Volume" in hist else 0
        rvol=round(vol/avg,2) if avg>0 else 1.0
        rsi=calc_rsi(cl); macd=calc_macd(cl); ma=calc_ma(cl)
        boll=calc_boll(cl); atr=calc_atr(hist); stoch=calc_stoch(hist)
        mom10=round(float(cl.pct_change(10).iloc[-1]*100),2) if n>10 else 0.0
        if ma["ma50"] and now>ma["ma50"]*1.01:   trend="Bullish"
        elif ma["ma50"] and now<ma["ma50"]*0.99: trend="Bearish"
        else:                                     trend="Sideways"
        if rsi<30:   rsi_st="Zona de acumulare"
        elif rsi<45: rsi_st="Presiune vanzatori"
        elif rsi<55: rsi_st="Echilibru"
        elif rsi<70: rsi_st="Momentum pozitiv"
        else:        rsi_st="Zona de distributie"
        semnal,conf=calc_semnal(rsi,macd["cross"],ma["macross"],rvol)
        sl=(round(now-1.5*atr,6) if semnal=="BUY" else
            round(now+1.5*atr,6) if semnal=="SELL" else round(now-2.0*atr,6))
        tp=(round(now+3.0*atr,6) if semnal=="BUY" else
            round(now-3.0*atr,6) if semnal=="SELL" else round(now+2.0*atr,6))
        prob=min(90,35+conf*10+(5 if rvol>1.2 else 0))
        return {
            "name":name,"ticker":ticker,"data":datetime.now().strftime("%d.%m.%Y"),
            "now":round(now,6),"open":round(float(lt.get("Open",now)),6),
            "high":round(float(lt.get("High",now)),6),"low":round(float(lt.get("Low",now)),6),
            "vzi":round(vzi,4),"vsapt":round(vsapt,4),"vluna":round(vluna,4),
            "vol":vol,"avg_vol":avg,"rvol":rvol,
            "rsi":rsi,"rsi_st":rsi_st,
            "macd":macd["macd"],"macd_sig":macd["signal"],"macd_hist":macd["hist"],
            "macd_cross":macd["cross"],
            "ma20":ma["ma20"],"ma50":ma["ma50"],"ma200":ma["ma200"],"macross":ma["macross"],
            "bb_sup":boll["bb_sup"],"bb_inf":boll["bb_inf"],"bb_w":boll["bb_w"],
            "atr":atr,"stoch_k":stoch["k"],"stoch_d":stoch["d"],
            "mom10":mom10,"trend":trend,
            "semnal":semnal,"conf":conf,"sl":sl,"tp":tp,"prob":prob,"n":n,
        }
    except Exception as e:
        log.error(f"  {name} ({ticker}): {e}"); return {}

def get_fear_greed():
    try:
        r=requests.get("https://api.alternative.me/fng/?limit=1",timeout=10)
        d=r.json()["data"][0]; v=int(d["value"])
        return {"value":v,"display":f"{v} — {d['value_classification']}",
                "status":"Pozitiv" if v>60 else ("Negativ" if v<40 else "Neutru")}
    except: return {"value":None,"display":"N/A","status":"Neutru"}

def get_fred(sid):
    if not FRED_API_KEY: return None
    try:
        url=(f"https://api.stlouisfed.org/fred/series/observations"
             f"?series_id={sid}&api_key={FRED_API_KEY}&file_type=json&sort_order=desc&limit=1")
        obs=requests.get(url,timeout=10).json().get("observations",[])
        if obs and obs[0]["value"]!=".": return float(obs[0]["value"])
    except: pass
    return None

# ═══════════════════════════════════════════════════════════════════
# DATE CONTEXTUALE PER CATEGORIE
# ═══════════════════════════════════════════════════════════════════
COMPETITORI_MAP = {
    "INDICI": [("S&P 500","^GSPC","SUA","Piata actiuni"),("NASDAQ 100","^NDX","SUA","Tech"),
               ("Dow Jones","^DJI","SUA","Blue chip"),("DAX Germany","^GDAXI","Germania","Europa"),
               ("FTSE 100","^FTSE","UK","Europa"),("Nikkei 225","^N225","Japonia","Asia")],
    "ACTIUNI": [("Apple","AAPL","Technology","Tech consumer"),("Microsoft","MSFT","Technology","Cloud/AI"),
                ("NVIDIA","NVDA","Semiconductors","AI/GPU"),("Alphabet","GOOGL","Technology","Search/Cloud"),
                ("Amazon","AMZN","E-Commerce","Cloud/Retail"),("Meta","META","Social Media","Digital Ads")],
    "CRYPTO": [("Bitcoin","BTC-USD","Crypto","Store of value"),("Ethereum","ETH-USD","Crypto","Smart contracts"),
               ("BNB","BNB-USD","Crypto","Exchange token"),("Solana","SOL-USD","Crypto","Layer 1"),
               ("XRP","XRP-USD","Crypto","Payments"),("Cardano","ADA-USD","Crypto","Layer 1")],
    "VALUTE": [("EUR/USD","EURUSD=X","FX Major","Euro vs Dollar"),("GBP/USD","GBPUSD=X","FX Major","Pound"),
               ("USD/JPY","USDJPY=X","FX Major","Yen"),("USD/CHF","USDCHF=X","FX Major","Franc"),
               ("AUD/USD","AUDUSD=X","FX Major","Aussie"),("USD/CAD","USDCAD=X","FX Major","Loonie")],
    "MATERII": [("Gold","GC=F","Metale","Refugiu"),("Silver","SI=F","Metale","Industrial/refugiu"),
                ("Oil WTI","CL=F","Energie","Petrol SUA"),("Oil Brent","BZ=F","Energie","Petrol global"),
                ("Natural Gas","NG=F","Energie","Gaze"),("Copper","HG=F","Metale","Industrial")],
}
RISCURI_MAP = {
    "INDICI": [
        ("R01","Risc","Macro","Recesiune SUA — scadere PIB doua trimestre consecutive",4,35,"3-12 luni","Defensive positioning; cash buffer"),
        ("R02","Risc","Macro","Inflatie persistenta — rate ridicate mai mult timp",4,55,"3-12 luni","Reducere expunere rate-sensitive"),
        ("R03","Risc","Geopolitic","Escaladare tensiuni geopolitice — impact sentiment",3,40,"1-3 luni","Gold hedge; diversificare"),
        ("R04","Risc","Tehnic","Breakdown MA200 — inversare trend major",4,25,"Imediat","SL stricte; reducere sizing"),
        ("O01","Oportunitate","Macro","Ciclu relaxare Fed — rate in scadere",4,60,"3-12 luni","Growth stocks; REIT exposure"),
        ("O02","Oportunitate","Tehnic","Golden Cross confirmat — semnal bullish pe termen lung",5,70,"Imediat","BUY la pullback la MA50"),
    ],
    "ACTIUNI": [
        ("R01","Risc","Sector","Reglementari antitrust Tech — impact valuation",3,50,"3-12 luni","Monitorizare legislativa"),
        ("R02","Risc","Macro","Comprimarea marjelor — costuri mai mari",3,45,"3-12 luni","Selectivitate sectoriala"),
        ("R03","Risc","Tehnic","RSI extins — risc corectie pe termen scurt",3,55,"Imediat","Trailing stop; luare profit partial"),
        ("R04","Risc","Operationale","Miss earnings — dezamagire vs asteptari",4,30,"Imediat","Atentie la data de raportare"),
        ("O01","Oportunitate","Sector","AI buildout — CAPEX data centers",5,75,"3-12 luni","Long NVDA, MSFT, AMD"),
        ("O02","Oportunitate","Tehnic","Pullback la MA50 in uptrend — entry optim",4,65,"Imediat","BUY cu SL sub MA50"),
    ],
    "CRYPTO": [
        ("R01","Risc","Regulatoriu","Reglementari SEC/UE — impact adoptie",4,60,"3-12 luni","Monitorizare cadru legal"),
        ("R02","Risc","Tehnic","Volatilitate extrema — miscare -30% posibila",5,50,"Imediat","Position sizing redus; SL larg"),
        ("R03","Risc","Macro","Sentiment risk-off — vanzare active riscante",4,45,"1-3 luni","Reducere expunere in scadere VIX"),
        ("R04","Risc","Operationale","Hack exchange/protocol — pierdere capital",3,20,"Imediat","Cold wallet; diversificare exchange"),
        ("O01","Oportunitate","Sector","Adoptie institutionala — ETF spot aprobate",5,70,"3-12 luni","Acumulare Bitcoin, Ethereum"),
        ("O02","Oportunitate","Tehnic","Post-halving rally Bitcoin — ciclu istoric",4,65,"3-12 luni","Expunere BTC, altcoin majori"),
    ],
    "VALUTE": [
        ("R01","Risc","Macro","Divergenta politici monetare — volatilitate FX",3,55,"1-3 luni","Hedging pozitii FX"),
        ("R02","Risc","Geopolitic","Criza geopolitica — flight to safety USD",4,35,"Imediat","Monitorizare USD Index"),
        ("R03","Risc","Macro","Interventie banca centrala — schimbare brusca directie",4,25,"Imediat","SL stricte; atentie la meeting-uri"),
        ("O01","Oportunitate","Macro","Carry trade — diferential rate dobanda",3,60,"3-12 luni","Long valute high-yield vs low-yield"),
        ("O02","Oportunitate","Tehnic","Breakout din range — trend nou in formare",4,55,"Imediat","Entry la confirmare cu volum"),
    ],
    "MATERII": [
        ("R01","Risc","Macro","Incetinire economica globala — scadere cerere",4,40,"3-12 luni","Reducere expunere ciclice"),
        ("R02","Risc","Geopolitic","Perturbari lantul de aprovizionare",3,45,"1-3 luni","Monitorizare rute comerciale"),
        ("R03","Risc","Macro","USD puternic — presiune pe materii prime",3,50,"1-3 luni","Hedging USD; monitorizare DXY"),
        ("O01","Oportunitate","Macro","Tranzitie energetica — cerere metale industriale",4,70,"1+ an","Long Copper, Silver, Lithium"),
        ("O02","Oportunitate","Geopolitic","Tensiuni geopolitice — crestere Gold",4,60,"1-3 luni","Alocare 5-10% Gold ca hedge"),
    ],
}
CALENDAR_MAP = {
    "INDICI": [("FOMC Rate Decision","SUA","Fed","US Equities, Bonds"),("Non-Farm Payrolls","SUA","BLS","USD, S&P500"),
               ("CPI Inflation","SUA","BLS","Bonds, USD, Equities"),("GDP Growth Rate","SUA","BEA","USD, S&P500"),
               ("PMI Manufacturing","SUA","ISM","Industrials, Equities"),("Earnings Season Q","SUA","Companii","Sectorial")],
    "ACTIUNI": [("Earnings Report Q","SUA","Companie","Actiune"),("FOMC Rate Decision","SUA","Fed","Growth stocks"),
                ("CPI Inflation","SUA","BLS","Tech valuations"),("Non-Farm Payrolls","SUA","BLS","Consumer stocks"),
                ("PCE Price Index","SUA","BEA","Fed policy"),("Retail Sales MoM","SUA","Census","Consumer stocks")],
    "CRYPTO": [("Bitcoin Halving","Global","Protocol","BTC, Altcoins"),("FOMC Rate Decision","SUA","Fed","Risk assets, BTC"),
               ("SEC Crypto Ruling","SUA","SEC","Crypto general"),("CPI Inflation","SUA","BLS","BTC hedge narrative"),
               ("Ethereum Network Upgrade","Global","Protocol","ETH, DeFi"),("Macro Risk Events","Global","Multiple","Risk-off sentiment")],
    "VALUTE": [("FOMC Rate Decision","SUA","Fed","USD pairs"),("ECB Rate Decision","UE","ECB","EUR pairs"),
               ("BOE Rate Decision","UK","BOE","GBP pairs"),("BOJ Rate Decision","Japonia","BOJ","JPY pairs"),
               ("CPI Inflation SUA","SUA","BLS","USD pairs"),("Non-Farm Payrolls","SUA","BLS","USD general")],
    "MATERII": [("OPEC+ Production Decision","Global","OPEC","Oil WTI, Brent"),("EIA Crude Oil Inventories","SUA","EIA","Oil prices"),
                ("FOMC Rate Decision","SUA","Fed","Gold, commodities"),("China PMI Manufacturing","China","NBS","Copper, Iron ore"),
                ("USD Index Movement","SUA","Fed","All commodities"),("Geopolitical Events","Global","Multiple","Gold, Oil")],
}

# ═══════════════════════════════════════════════════════════════════
# TEXTE EDUCATIVE — GHID INVATARE
# ═══════════════════════════════════════════════════════════════════
def text_miscare(d):
    name=d["name"]; pret=d["now"]; vzi=d["vzi"]; vsapt=d["vsapt"]; vluna=d["vluna"]
    rsi=d["rsi"]; mc_=d["macd_cross"]; mh=d["macd_hist"]; macross=d["macross"]
    rvol=d["rvol"]; ma50=d["ma50"]; ma200=d["ma200"]
    bb_sup=d["bb_sup"]; bb_inf=d["bb_inf"]; bb_w=d["bb_w"]
    stoch_k=d["stoch_k"]; semnal=d["semnal"]; n=d["n"]
    ma50_s=fp(ma50,2); ma200_s=fp(ma200,2)
    L=[]
    dir_="crescut" if vzi>0 else "scazut"
    int_="semnificativ" if abs(vzi)>3 else "moderat" if abs(vzi)>1 else "usor"
    L.append(f"{name} a {dir_} {int_} cu {abs(vzi):.2f}% astazi (pret: {fp(pret)}).\nSaptamana: {fpc(vsapt)} | Luna: {fpc(vluna)}.")
    if rvol>1.5: L.append(f"VOLUM EXCEPTIONAL: {rvol:.1f}x media 20 zile — miscarea este confirmata institutional.")
    elif rvol<0.7: L.append(f"VOLUM SCAZUT: {rvol:.1f}x media — miscare neconfirmata, fara participare institutionala.")
    else: L.append(f"Volum normal ({rvol:.1f}x) — participare obisnuita a pietei.")
    if rsi<30: L.append(f"RSI = {rsi:.1f} — ZONA DE ACUMULARE (sub 30).\nActivul a coborat rapid si acumuleaza potential de revenire.")
    elif rsi<45: L.append(f"RSI = {rsi:.1f} — Presiune din partea vanzatorilor. Momentum descendent.")
    elif rsi<55: L.append(f"RSI = {rsi:.1f} — Zona de echilibru. Cumparatori si vanzatori aproximativ egali.")
    elif rsi<70: L.append(f"RSI = {rsi:.1f} — Momentum pozitiv solid. Zona 55-70 este optima pentru trend-following.")
    else: L.append(f"RSI = {rsi:.1f} — ZONA DE DISTRIBUTIE (peste 70).\nActivul a crescut rapid. Probabilitatea unei consolidari creste.")
    if "bullish" in mc_.lower(): L.append(f"MACD: Crossover bullish (histogram: {mh:+.4f}) — schimbare de directie ascendenta confirmata.")
    elif "ascendent" in mc_.lower(): L.append(f"MACD: Momentum ascendent (histogram: {mh:+.4f}) — trendul ascendent se mentine.")
    elif "bearish" in mc_.lower(): L.append(f"MACD: Crossover bearish (histogram: {mh:+.4f}) — momentum ascendent s-a epuizat.")
    else: L.append(f"MACD: Momentum descendent (histogram: {mh:+.4f}) — presiunea descendenta domina.")
    if macross=="Golden Cross" and ma50 and ma200: L.append(f"GOLDEN CROSS ACTIV: MA50 ({ma50_s}) > MA200 ({ma200_s}). Cel mai puternic semnal bullish.")
    elif macross=="Death Cross" and ma50 and ma200: L.append(f"DEATH CROSS ACTIV: MA50 ({ma50_s}) < MA200 ({ma200_s}). Semnal bearish major.")
    elif ma50 and not ma200: L.append(f"MA50 disponibila ({ma50_s}), MA200 indisponibila ({n} zile date).")
    else: L.append(f"MA50 ({ma50_s}) si MA200 ({ma200_s}) — piata in tranzitie.")
    if bb_sup and bb_inf and bb_w:
        pf=float(pret)
        if pf>=float(bb_sup)*0.99: L.append(f"Pret la Banda Bollinger SUPERIOARA ({fp(bb_sup)}) — potential supraextindere.")
        elif pf<=float(bb_inf)*1.01: L.append(f"Pret la Banda Bollinger INFERIOARA ({fp(bb_inf)}) — zona suport.")
        elif float(bb_w)/pf<0.03: L.append(f"BOLLINGER SQUEEZE (latime {fp(bb_w)}) — volatilitate comprimata, explozie iminenta.")
    if stoch_k<20: L.append(f"Stochastic %K = {stoch_k:.1f} — Zona de acumulare pe termen scurt.")
    elif stoch_k>80: L.append(f"Stochastic %K = {stoch_k:.1f} — Zona de distributie pe termen scurt.")
    if semnal=="BUY": L.append("CONCLUZIE: Configuratie FAVORABILA CUMPARARII. Respecta strict stop loss-ul!")
    elif semnal=="SELL": L.append("CONCLUZIE: Presiune de VANZARE dominanta. Prudenta cu pozitiile long.")
    else: L.append("CONCLUZIE: Semnale MIXTE — ASTEPTARE recomandata. Nu forta o tranzactie.")
    return "\n\n".join(L)

def text_oportunitate(d):
    semnal=d["semnal"]; pret=d["now"]; sl=d["sl"]; tp=d["tp"]
    atr=d["atr"]; rvol=d["rvol"]; conf=d["conf"]; prob=d["prob"]
    rr=abs(float(tp)-float(pret))/abs(float(sl)-float(pret)) if abs(float(sl)-float(pret))>0 else 0
    risc_pct=abs(float(pret)-float(sl))/float(pret)*100 if float(pret)>0 else 0
    rr_e="EXCELENT" if rr>=3 else "BUN" if rr>=2 else "ACCEPTABIL" if rr>=1.5 else "SLAB"
    L=[]
    if semnal in ("BUY","SELL"):
        tip="CUMPARARE" if semnal=="BUY" else "SHORT/VANZARE"
        L.append(f"OPORTUNITATE DE {tip} — {conf} confluente | Probabilitate: {prob}%\n\n"
                 f"  Entry               : {fp(pret,4)}\n  Stop Loss (SL)      : {fp(sl,4)}   risc {risc_pct:.1f}%\n"
                 f"  Take Profit (TP)    : {fp(tp,4)}   castig {abs(float(tp)-float(pret))/float(pret)*100:.1f}%\n"
                 f"  Risk/Reward Ratio   : {rr:.2f}x   {rr_e}\n  ATR (volatilitate)  : {fp(atr,4)}")
        if rr>=2: L.append(f"RR de {rr:.1f}x — profitabil matematic chiar cu win rate de 34%.")
        else: L.append(f"RR de {rr:.1f}x — sub pragul ideal de 2x. Evalueaza ajustarea SL/TP.")
    else:
        L.append(f"NU EXISTA OPORTUNITATE CLARA — Confluente: {conf}/5\n\nUrmareste:\n"
                 f"  RSI sub 40 (BUY) sau peste 70 (SELL)\n  Crossover MACD bullish/bearish\n"
                 f"  Volum peste 1.2x medie\n  Golden/Death Cross confirmat")
    if rvol>1.5 and semnal!="WAIT":
        L.append(f"CONFIRMARE VOLUM: {rvol:.1f}x medie — semnal de calitate superioara.")
    return "\n\n".join(L)

def text_pattern(d):
    pret=d["now"]; rsi=d["rsi"]; mc_=d["macd_cross"]; mh=d["macd_hist"]
    macross=d["macross"]; sk=d["stoch_k"]; sd_v=d["stoch_d"]
    bb_sup=d["bb_sup"]; bb_inf=d["bb_inf"]; bb_w=d["bb_w"]
    ma50=d["ma50"]; ma200=d["ma200"]; vzi=d["vzi"]; rvol=d["rvol"]
    ma50_s=fp(ma50,2); ma200_s=fp(ma200,2)
    P=[]
    if macross=="Golden Cross" and ma50 and ma200:
        dp=(float(ma50)-float(ma200))/float(ma200)*100
        P.append(f"PATTERN: GOLDEN CROSS\nMA50 ({ma50_s}) > MA200 ({ma200_s}) cu {dp:.2f}%.\nCel mai urmarit semnal bullish institutional.")
    elif macross=="Death Cross" and ma50 and ma200:
        dp=(float(ma200)-float(ma50))/float(ma200)*100
        P.append(f"PATTERN: DEATH CROSS\nMA50 ({ma50_s}) < MA200 ({ma200_s}) cu {dp:.2f}%.\nSemnal bearish major.")
    if bb_w and bb_sup and bb_inf and float(bb_w)/float(pret)<0.04:
        P.append(f"PATTERN: BOLLINGER SQUEEZE\nLatime benzi {fp(bb_w,4)} — compresie extrema. Explozie de volatilitate iminenta.")
    if "bullish" in mc_.lower():
        P.append(f"PATTERN: MACD CROSSOVER BULLISH (histogram: {mh:+.6f})\nHistograma trece din negativ in pozitiv — schimbare de momentum.")
    elif "bearish" in mc_.lower():
        P.append(f"PATTERN: MACD CROSSOVER BEARISH (histogram: {mh:+.6f})\nHistograma trece din pozitiv in negativ — momentum ascendent s-a epuizat.")
    if rsi<32 and vzi<-2 and rvol>1.3:
        P.append(f"PATTERN POSIBIL: CLIMAX DE VANZARE\nScadere {vzi:.1f}% + RSI {rsi:.1f} + Volum {rvol:.1f}x — potential punct de inversare.")
    elif rsi>75 and vzi>2 and rvol>1.3:
        P.append(f"PATTERN POSIBIL: CLIMAX DE CUMPARARE\nCrestere {vzi:.1f}% + RSI {rsi:.1f} + Volum {rvol:.1f}x — potential distributie smart money.")
    if sk<25 and sk>sd_v:
        P.append(f"STOCHASTIC: Zona acumulare cu inversare (%K {sk:.1f} > %D). Semnal scurt termen BUY.")
    elif sk>75 and sk<sd_v:
        P.append(f"STOCHASTIC: Zona distributie cu inversare (%K {sk:.1f} < %D). Semnal scurt termen SELL.")
    if ma50:
        dist=(float(pret)-float(ma50))/float(ma50)*100
        if abs(dist)<0.8:
            P.append(f"PATTERN: TEST MA50\nPret la {dist:+.2f}% fata de MA50 ({ma50_s}). Zona cheie suport/rezistenta.")
    if not P: P.append("Niciun pattern tehnic dominant azi. Activ in consolidare — urmareste zilnic.")
    return "\n\n".join(P)

def text_lectie(d):
    semnal=d["semnal"]; rsi=d["rsi"]; macross=d["macross"]
    rvol=d["rvol"]; conf=d["conf"]; pret=d["now"]; sl=d["sl"]; tp=d["tp"]
    rr=abs(float(tp)-float(pret))/abs(float(sl)-float(pret)) if abs(float(sl)-float(pret))>0 else 0
    L=[]
    if semnal=="WAIT" and conf<2:
        L.append("LECTIE — RABDAREA CA STRATEGIE:\n\nProfesioniistii stau in cash 60-70% din timp si ataca DOAR cand setup-ul e perfect. "
                 "Selectivitatea extrema este un avantaj competitiv. Daca nu exista semnal clar — NU TRANZACTIONA.")
    if 0<rr<1.5:
        L.append(f"LECTIE — MATEMATICA RISK/REWARD:\n\nRR curent: {rr:.1f}x — sub pragul profesionist de 2x.\n\n"
                 f"  RR=1.0x + win rate 50% = breakeven\n  RR=1.5x + win rate 40% = profitabil\n"
                 f"  RR=2.0x + win rate 34% = profitabil\n  RR=3.0x + win rate 26% = profitabil")
    if rvol<0.7:
        L.append("LECTIE — VOLUMUL CONFIRMA:\n\nO miscare pe volum sub 70% din medie este neconvingatoare. "
                 "Regula de aur: 'Nu cumpara breakout-uri pe volum mic.'")
    if rsi<30:
        L.append("LECTIE — ZONA DE ACUMULARE NU INSEAMNA CUMPARARE IMEDIATA:\n\nRSI sub 30 este conditie necesara, nu suficienta. "
                 "Ai nevoie de confirmare: candle bullish + MACD crossover + volum crescut.")
    elif rsi>70:
        L.append("LECTIE — ZONA DE DISTRIBUTIE NU INSEAMNA VINDE IMEDIAT:\n\nActivele pot ramane in zona de distributie luni intregi in bull market. "
                 "NVIDIA: RSI peste 70 timp de 8 luni consecutive, +400%.")
    if macross=="Golden Cross":
        L.append("LECTIE — GOLDEN CROSS: ANTICIPEAZA, NU REACTIONA:\n\nGolden Cross apare DUPA ce pretul a urcat deja 10-30% (indicator tardiv).")
    elif macross=="Death Cross":
        L.append("LECTIE — DEATH CROSS: EROAREA CLASICA:\n\nDeath Cross apare la MIJLOCUL corectiei, nu la inceput. "
                 "Strategie corecta: reduce expunerea la PRIMELE semnale.")
    L.append("LECTIE PSIHOLOGIE — LOSS AVERSION:\n\nDurerea pierderii de 100 lei este 2x mai intensa decat bucuria castigului de 100 lei. "
             "(Kahneman, Nobel 2002)\n\nEfecte in trading:\n  Tii pozitiile perdante prea mult\n  Vinzi castigatorii prea devreme\n"
             "  Faci revenge trading dupa pierdere\n\nSolutia: PLAN SCRIS inainte de intrare.")
    return "\n\n" + ("\n\n".join(L[:2]))

GHID_GRAFICE=[
    {"titlu":"CANDLESTICK — Lumanari japoneze","sectiuni":[
        ("Ce este o lumanare","Fiecare lumanare = o perioada de timp (zi, ora) cu 4 preturi: Open, High, Low, Close.\n"
         "VERDE: Inchidere > Deschidere — pret in crestere.\nROSU: Inchidere < Deschidere — pret in scadere.\nMarimea corpului = forta miscarii."),
        ("Umbrele (fitilele)","Umbra lunga SUS: pretul a urcat dar Bears l-au respins.\nUmbra lunga JOS: pretul a coborat dar Bulls l-au aparat.\nUmbra scurta = directie clara."),
        ("Pattern-uri esentiale","HAMMER: corp mic sus, umbra lunga jos — potential revenire bullish\nSHOOTING STAR: corp mic jos, umbra lunga sus — potential inversare bearish\n"
         "DOJI: corp inexistent — indecizie, potential inversare\nBULLISH ENGULFING: verde mare inghite rosu — semnal BUY puternic\nBEARISH ENGULFING: rosu mare inghite verde — semnal SELL puternic"),
        ("Cum aplici","1. Cauta pattern-uri LA NIVELURI CHEIE (suport, rezistenta, MA50, MA200)\n2. Confirma cu VOLUM crescut\n3. Asteapta confirmarea: urmatorul candle sa confirme directia\n4. Pattern-urile pe timeframe ZILNIC sau SAPTAMANAL sunt cele mai fiabile"),
    ]},
    {"titlu":"RSI — Relative Strength Index (0-100)","sectiuni":[
        ("Ce masoara","Viteza si amplitudinea miscarilor din ultimele 14 perioade.\nRaspunde la: A urcat/coborat pretul prea repede fata de norma sa?"),
        ("Zonele RSI","0-30   = Zona de acumulare: pret coborat rapid, potential revenire\n30-45  = Presiune vanzatori\n45-55  = Echilibru, piata cauta directie\n55-70  = Momentum pozitiv, zona optima trend-following\n70-100 = Zona de distributie: pret urcat rapid, potential consolidare"),
        ("RSI Divergenta","BULLISH: Pretul face un nou minim, RSI face un minim MAI MARE — presiunea de vanzare slabeste\nBEARISH: Pretul face un nou maxim, RSI face un maxim MAI MIC — momentum ascendent slabeste"),
        ("Greseli frecvente","Nu vinde DOAR pe RSI>70 (poate ramane acolo luni in bull market)\nNu cumpara DOAR pe RSI<30 (poate scadea la 10 in bear market)\nNu folosi RSI izolat fara confirmare\nRSI pe 5 minute = zgomot; zilnic/saptamanal = semnal"),
    ]},
    {"titlu":"MACD — Moving Average Convergence/Divergence","sectiuni":[
        ("Componentele","Linia MACD  = EMA(12) - EMA(26)  [linia rapida]\nLinia Signal = EMA(9) a MACD   [linia lenta]\nHistograma   = MACD - Signal   [barele verzi/rosii]\nRaspunde la: Care este directia si forta momentum-ului?"),
        ("Semnale principale","CROSSOVER BULLISH: MACD > Signal, histograma din negativ in pozitiv — BUY\nCROSSOVER BEARISH: MACD < Signal, histograma din pozitiv in negativ — SELL\nZERO LINE CROSS: MACD din negativ in pozitiv = confirmare trend bullish major\nHISTOGRAMA CRESTE = momentum se accelereaza | SCADE = momentum slabeste"),
        ("Cel mai puternic semnal","MACD crossover bullish LA LINIA ZERO\n+ Golden Cross MA50/MA200\n+ Volum crescut (RVOL > 1.5x)\n+ RSI in zona 40-55\n= Unul dintre cele mai fiabile setup-uri in analiza tehnica."),
        ("Limitari","Indicator tardiv: confirma, nu prezice.\nCel mai eficient pe timeframe ZILNIC si SAPTAMANAL.\nGenereaza zgomot pe timeframe sub 1 ora.\nFunctioneaza mai bine in trenduri clare."),
    ]},
    {"titlu":"MEDII MOBILE — MA20, MA50, MA200","sectiuni":[
        ("Ce reprezinta","MA20  (~1 luna)   = tendinta pe termen scurt\nMA50  (~2.5 luni) = tendinta pe termen mediu\nMA200 (~10 luni)  = tendinta pe termen lung"),
        ("Regulile esentiale","1. Pret DEASUPRA MA200 = Bull Market\n2. Pret DEDESUBT MA200 = Bear Market\n3. GOLDEN CROSS (MA50 > MA200) = semnal bullish major\n4. DEATH CROSS (MA50 < MA200) = bearish major\n5. MA50 = suport dinamic in uptrend si rezistenta in downtrend"),
        ("Utilizare practica","UPTREND: Cumpara la pullback-uri la MA50 cu candle de confirmare\nDOWNTREND: Short la retestul MA50 cu candle bearish\nTRANZITIE (MA50 langa MA200): Evita tranzactiile — risc maxim de whipsaw."),
    ]},
    {"titlu":"BOLLINGER BANDS","sectiuni":[
        ("Structura","Banda Superioara = MA20 + (2 x deviatie standard)\nBanda Mijlocie  = MA20\nBanda Inferioara = MA20 - (2 x deviatie standard)\n95% din timp pretul se afla INTRE benzile externe."),
        ("Strategia Bounce","In piete laterale: cumpara la banda inferioara, vinde la superioara.\nNU aplica in trenduri puternice."),
        ("Strategia Squeeze","Benzi inguste (sub 4% din pret) = compresie = explozie iminenta.\nBUY STOP deasupra benzii superioare + SELL STOP sub cea inferioara."),
    ]},
    {"titlu":"VOLUM & RVOL","sectiuni":[
        ("De ce conteaza","Pretul spune UNDE s-a dus piata. Volumul spune CAT DE CONVINSA era.\nAxioma: Volumul preceda pretul."),
        ("Scala RVOL","RVOL > 2.0x = Exceptional: eveniment major\nRVOL 1.5-2.0x = Ridicat: miscare de calitate\nRVOL 1.0-1.5x = Normal\nRVOL 0.7-1.0x = Normal spre scazut\nRVOL < 0.7x = Scazut: miscare neconfirmata"),
        ("Regulile de aur","UPTREND SANATOS: zile UP volum mare + zile DOWN volum mic\nDOWNTREND SANATOS: zile DOWN volum mare + zile UP volum mic\nBREAKOUT VOLUM MIC = capcana | BREAKOUT VOLUM MARE = real"),
    ]},
    {"titlu":"SUPORT & REZISTENTA","sectiuni":[
        ("Definitie","SUPORT: nivel unde cumparatorii opresc si inverseaza scaderea.\nREZISTENTA: nivel unde vanzatorii opresc si inverseaza cresterea."),
        ("Cum identifici","1. Maxime si minime anterioare semnificative (min. 2-3 atingeri)\n2. Zone de consolidare\n3. Niveluri psihologice rotunde: 1000, 50000, 100\n4. Medii mobile (MA50, MA200) — dinamice\n5. Fibonacci 38.2%, 50%, 61.8%"),
        ("Role Reversal","SUPORTUL SPART devine REZISTENTA\nREZISTENTA SPARTA devine SUPORT\nPunct de intrare cu risc redus si potential ridicat."),
        ("Stop Loss optim","NU plasa SL exact la nivel — piata il testeaza cu cateva procente.\nREGULA: SL = suport MINUS 1-3% (sau 1-2x ATR). Spatiu de respiratie = esential."),
    ]},
    {"titlu":"MANAGEMENT RISC","sectiuni":[
        ("Regula 1-2%","NU risca mai mult de 1-2% din capital per tranzactie.\nMarime pozitie = Risc acceptat / (Entry - Stop Loss)\nExemplu: 10.000 EUR, risc 1% (100 EUR), Entry 50, SL 48 = 100/(50-48) = 50 actiuni"),
        ("RR — matematica supravietuirii","RR = (TP - Entry) / (Entry - SL)\nWin rate minim: RR=1.5x >40% | RR=2.0x >34% | RR=3.0x >26%\nCu RR bun, pierzi mai multe tranzactii si totusi esti profitabil."),
        ("Cele 10 reguli de aur","1. Planifica tranzactia, tranzactioneaza planul.\n2. Niciodata RR sub 1.5x.\n3. Niciodata risc peste 2% din capital.\n4. Confirma breakout-urile cu VOLUMUL.\n5. Nu face averaging down neselectiv.\n6. Lasa castigatorii sa curga, taie perdantii rapid.\n7. Nu tranzactiona din plictiseala sau FOMO.\n8. Nu revenge trade dupa pierdere — ia o pauza.\n9. Tine jurnal de tranzactii.\n10. Piata exista si maine — capitalul protejat = oportunitati viitoare."),
    ]},
]

# ═══════════════════════════════════════════════════════════════════
# TEMPLATE EXCEL GOL
# ═══════════════════════════════════════════════════════════════════
SHEET_TABS = {
    "DASHBOARD":"1F4E79","REZUMAT EXECUTIV":"1F4E79","SEMNALE INTRARE":"375623",
    "INDICATORI TEHNICI":"595959","INDICATORI MACRO":"595959","COMPETITORI SECTOR":"595959",
    "PRETURI VOLUME":"595959","RISCURI OPORTUNITATI":"9C0006","CALENDAR ECONOMIC":"595959",
    "JURNAL TRANZACTII":"595959","ISTORIC TRENDING":"595959","GHID INVATARE":"154360",
    "LEGENDA":"595959","FISA ACTIV":"ED7D31",
}

def create_template(path):
    wb=Workbook(); wb.remove(wb.active)
    for sname,tab_color in SHEET_TABS.items():
        ws=wb.create_sheet(sname); ws.sheet_properties.tabColor=tab_color
        builders={"DASHBOARD":_tmpl_dashboard,"REZUMAT EXECUTIV":_tmpl_rezumat,
                  "SEMNALE INTRARE":_tmpl_semnale,"INDICATORI TEHNICI":_tmpl_tehnic,
                  "INDICATORI MACRO":_tmpl_macro,"COMPETITORI SECTOR":_tmpl_comp,
                  "PRETURI VOLUME":_tmpl_preturi,"RISCURI OPORTUNITATI":_tmpl_riscuri,
                  "CALENDAR ECONOMIC":_tmpl_calendar,"JURNAL TRANZACTII":_tmpl_jurnal,
                  "ISTORIC TRENDING":_tmpl_istoric,"GHID INVATARE":_tmpl_ghid,
                  "LEGENDA":_tmpl_legenda,"FISA ACTIV":_tmpl_fisa}
        if sname in builders: builders[sname](ws)
    wb.save(str(path))
    log.info(f"  Template creat: {path.name}")

def _tmpl_dashboard(ws):
    for col,w in {"A":18,"B":14,"C":12,"D":12,"E":12,"F":12,"G":14,"H":14,"I":14,"J":14,"K":12,"L":12,"M":14,"N":12,"O":12,"P":12}.items():
        ws.column_dimensions[col].width=w
    ws.freeze_panes="A3"
    set_h(ws,1,40); mc(ws,1,1,1,16,"DASHBOARD — ANALIZA DE PIATA",F_TI,FN_TI,"center")
    set_h(ws,2,28)
    mc(ws,2,1,2,3,"Data actualizare:",F_H,fnt(bold=True,color="FFFFFF",size=10),"right")
    mc(ws,2,4,2,6,"",F_IN,FN_IN,"center")
    mc(ws,2,8,2,8,"Activ selectat:",F_SH,FN_SH,"right")
    mc(ws,2,9,2,13,"",F_SL,FN_SL,"center")
    for r in [4,5,6]: set_h(ws,r,22 if r==5 else 16)
    for ci,(lbl,bg) in enumerate([("TREND GENERAL","C6EFCE"),("VOLATILITATE","FFEB9C"),
                                   ("VOLUM RELATIV","C6EFCE"),("MOMENT INTRARE","C6EFCE")]):
        c=ci*4+1
        fg_="375623" if bg=="C6EFCE" else "9C6500"
        mc(ws,4,c,4,c+3,lbl,fill(bg),fnt(bold=True,color=fg_,size=9),"center")
        mc(ws,5,c,5,c+3,"—",fill(bg),fnt(bold=True,color=fg_,size=14),"center")
        mc(ws,6,c,6,c+3,"—",fill(bg),fnt(italic=True,color=fg_,size=9),"center")
    set_h(ws,8,20); mc(ws,8,1,8,16,"SEMNAL PRINCIPAL DE TRADING",F_H,FN_H,"center")
    for r,lbl in {9:"SEMNAL ACTIV",10:"Activ analizat",11:"Entry Price",12:"Stop Loss (SL)",
                   13:"Take Profit (TP)",14:"Risk/Reward Ratio",15:"Confluente aliniate",
                   16:"Probabilitate (%)",17:"Conditie declansare"}.items():
        set_h(ws,r,24)
        mc(ws,r,1,r,6,lbl,F_RH,FN_RH,"right"); mc(ws,r,7,r,16,"—",F_CA,fnt(size=11),"left")
    set_h(ws,19,20); mc(ws,19,1,19,16,"REZUMAT INDICATORI (per activ selectat)",F_H,FN_H,"left")
    hdr_row(ws,20,["Indicator","Valoare","","Status","","Observatie","","","","","","","","","",""],1)
    for i,lbl in enumerate(["RSI(14)","MACD Status","MA50 vs MA200","Trend","Bollinger","Stochastic %K","RVOL activ"]):
        r=21+i; set_h(ws,r,20)
        ws.cell(r,1,lbl).fill=F_RH; ws.cell(r,1).font=FN_RH; ws.cell(r,1).alignment=aln("left")
        mc(ws,r,2,r,4,"—",F_CA,fnt(size=10),"center")
        mc(ws,r,5,r,8,"—",F_CA,fnt(bold=True),"center")
        mc(ws,r,9,r,16,"",F_WH,fnt(italic=True,color="595959",size=9),"left")

def _tmpl_rezumat(ws):
    for col,w in {"A":30,"B":16,"C":12,"D":10,"E":42,"F":12,"G":12,"H":12}.items():
        ws.column_dimensions[col].width=w
    ws.freeze_panes="A3"
    set_h(ws,1,40); mc(ws,1,1,1,8,"REZUMAT EXECUTIV — MARKET OVERVIEW",F_TI,FN_TI,"center")
    set_h(ws,2,22); ws.cell(2,1,"Activ selectat:").fill=F_RH; ws.cell(2,1).font=FN_RH
    mc(ws,2,4,2,6,"",F_SL,FN_SL,"center")
    set_h(ws,4,20); mc(ws,4,1,4,8,"TENDINTA GENERALA PIATA",F_H,FN_H,"left")
    hdr_row(ws,5,["Indicator","Valoare","","Trending","Observatii","","",""],1)
    for i,lbl in enumerate(["Tendinta generala","Volatilitate","Volum","Sentiment","Risc sistemic"]):
        r=6+i; set_h(ws,r,22)
        ws.cell(r,1,lbl).fill=F_RH; ws.cell(r,1).font=FN_RH
        mc(ws,r,2,r,3,"",F_IN,FN_IN,"center")
        ws.cell(r,4,"").fill=F_CA; ws.cell(r,4).alignment=aln("center")
        mc(ws,r,5,r,8,"",F_WH,fnt(italic=True,color="595959"),"left")
    set_h(ws,13,20); mc(ws,13,1,13,8,"CONCLUZII & RECOMANDARI",F_H,FN_H,"left")
    for i in range(5):
        r=14+i; set_h(ws,r,22)
        ws.cell(r,1,"").fill=fill("F2F2F2" if i%2==0 else "FFFFFF"); ws.merge_cells(f"A{r}:H{r}")

def _tmpl_semnale(ws):
    for c,w in {"A":12,"B":20,"C":8,"D":30,"E":8,"F":16,"G":16,"H":14,"I":10,"J":10,"K":12,"L":12,"M":12,"N":10,"O":10,"P":10,"Q":25}.items():
        ws.column_dimensions[c].width=w
    ws.freeze_panes="B3"
    set_h(ws,1,35); mc(ws,1,1,1,17,"SEMNALE DE INTRARE PE PIATA",F_TI,FN_TI,"center")
    hdr_row(ws,2,["Data","Activ","Semnal","Conditie","RSI(14)","MACD Status","MA Cross","Volum vs Medie",
                  "Momentum 10z","Confluente","Entry","Stop Loss","Take Profit","RR Ratio","Probabilitate","Status","Note"])
    cf_signal(ws,"C3:C2000"); cf_status(ws,"P3:P2000")

def _tmpl_tehnic(ws):
    for c,w in {"A":12,"B":20,"C":12,"D":12,"E":12,"F":12,"G":8,"H":18,"I":12,"J":12,"K":12,"L":12,"M":12,"N":10,"O":10,"P":8,"Q":8,"R":14,"S":8,"T":10,"U":12,"V":12,"W":16}.items():
        ws.column_dimensions[c].width=w
    ws.freeze_panes="C3"
    set_h(ws,1,35); mc(ws,1,1,1,23,"INDICATORI TEHNICI",F_TI,FN_TI,"center")
    hdr_row(ws,2,["Data","Activ","Pret","MA20","MA50","MA200","RSI(14)","RSI Status","MACD","MACD Signal","MACD Hist",
                  "BB Superior","BB Inferior","BB Latime","ATR","Stoch %K","Stoch %D","Volum","RVOL","Trend","Suport cheie","Rezistenta","MA Cross"])
    rng="G3:G2000"
    ws.conditional_formatting.add(rng,CellIsRule("lessThan",["30"],fill=F_GN,font=FN_GN))
    ws.conditional_formatting.add(rng,CellIsRule("between",["30","50"],fill=F_YL,font=FN_YL))
    ws.conditional_formatting.add(rng,CellIsRule("between",["50","70"],fill=fill("E2EFDA"),font=fnt(color="375623")))
    ws.conditional_formatting.add(rng,CellIsRule("greaterThan",["70"],fill=F_RD,font=FN_RD))
    cf_status(ws,"T3:T2000"); cf_status(ws,"W3:W2000")

def _tmpl_macro(ws):
    for c,w in {"A":28,"B":14,"C":14,"D":10,"E":8,"F":12,"G":14,"H":14,"I":14,"J":12,"K":10,"L":10,"M":30}.items():
        ws.column_dimensions[c].width=w
    ws.freeze_panes="B3"
    set_h(ws,1,35); mc(ws,1,1,1,13,"INDICATORI MACROECONOMICI",F_TI,FN_TI,"center")
    hdr_row(ws,2,["Indicator","Valoare curenta","Valoare anterioara","Delta Abs","Delta %",
                  "Consens","Dev vs Estimare","Impact piata","Data publicare","Frecventa","Trending","Status","Note"])
    macro_lista=["PIB YoY (%)","CPI YoY (%)","Core CPI (%)","Rata dobanzii (%)","Rata somajului (%)",
                 "PMI Manufacturing","PMI Services","Retail Sales MoM (%)","Balanta comerciala ($B)",
                 "USD Index (DXY)","EUR/USD","Petrol Brent ($)","Petrol WTI ($)","Yield 10Y US (%)","VIX","Fear & Greed (0-100)"]
    for i,ind in enumerate(macro_lista):
        r=3+i; set_h(ws,r,18)
        ws.cell(r,1,ind).fill=F_RH; ws.cell(r,1).font=FN_RH; ws.cell(r,1).alignment=aln("left")
        for c in range(2,14): ws.cell(r,c).fill=F_IN if c in [2,3,6] else F_CA
    cf_status(ws,"L3:L50"); cf_status(ws,"K3:K50")

def _tmpl_comp(ws):
    for c,w in {"A":22,"B":20,"C":10,"D":10,"E":12,"F":12,"G":12,"H":10,"I":30,"J":30,"K":12,"L":8,"M":14,"N":12,"O":20}.items():
        ws.column_dimensions[c].width=w
    ws.freeze_panes="B3"
    set_h(ws,1,35); mc(ws,1,1,1,15,"ANALIZA COMPETITORI & SECTOR (per activ selectat)",F_TI,FN_TI,"center")
    set_h(ws,2,20); mc(ws,2,1,2,15,"Sectorul si competitorii se actualizeaza automat la selectia activului din Dashboard",F_BL,FN_BL,"center")
    hdr_row(ws,3,["Entitate","Sector","Cota %","YoY","Pret mediu","Trending","Revenue ($B)","Marja (%)","Puncte forte","Puncte slabe","Scor (1-10)","Risc (1-5)","Status","Data","Note"])
    cf_status(ws,"F4:F100")

def _tmpl_preturi(ws):
    for c,w in {"A":12,"B":20,"C":12,"D":12,"E":12,"F":12,"G":10,"H":10,"I":10,"J":14,"K":14,"L":8,"M":14,"N":12,"O":10}.items():
        ws.column_dimensions[c].width=w
    ws.freeze_panes="C3"
    set_h(ws,1,35); mc(ws,1,1,1,15,"PRETURI, VOLUME & SEZONALITATE",F_TI,FN_TI,"center")
    hdr_row(ws,2,["Data","Activ","Deschidere","Maxim","Minim","Inchidere","Var Zi (%)","Var Sapt (%)","Var Luna (%)",
                  "Volum","Medie Vol 20z","RVOL","Sezon","Factor sezonier (%)","Trend"])
    ws.conditional_formatting.add("G3:I2000",CellIsRule("greaterThan",["0"],fill=F_GN,font=FN_GN))
    ws.conditional_formatting.add("G3:I2000",CellIsRule("lessThan",["0"],fill=F_RD,font=FN_RD))
    cf_status(ws,"O3:O2000")

def _tmpl_riscuri(ws):
    for c,w in {"A":6,"B":14,"C":16,"D":45,"E":10,"F":12,"G":12,"H":14,"I":38,"J":16,"K":12,"L":12,"M":25}.items():
        ws.column_dimensions[c].width=w
    ws.freeze_panes="C3"
    set_h(ws,1,35); mc(ws,1,1,1,13,"RISCURI & OPORTUNITATI (per activ selectat)",F_TI,FN_TI,"center")
    set_h(ws,2,20); mc(ws,2,1,2,13,"Se actualizeaza automat la selectia activului din Dashboard",F_BL,FN_BL,"center")
    hdr_row(ws,3,["ID","Tip","Categorie","Descriere","Impact (1-5)","Probabilitate (%)","Scor prioritate","Orizont","Actiuni","Owner","Status","Data","Note"])
    ws.conditional_formatting.add("B4:B100",CellIsRule("equal",['"Risc"'],fill=F_RD,font=FN_RD))
    ws.conditional_formatting.add("B4:B100",CellIsRule("equal",['"Oportunitate"'],fill=F_GN,font=FN_GN))
    cf_status(ws,"K4:K100")

def _tmpl_calendar(ws):
    for c,w in {"A":22,"B":36,"C":8,"D":14,"E":10,"F":10,"G":10,"H":10,"I":16,"J":18,"K":30}.items():
        ws.column_dimensions[c].width=w
    ws.freeze_panes="B3"
    set_h(ws,1,35); mc(ws,1,1,1,11,"CALENDAR ECONOMIC (per activ selectat)",F_TI,FN_TI,"center")
    set_h(ws,2,20); mc(ws,2,1,2,11,"Evenimentele relevante se filtreaza automat la selectia activului",F_BL,FN_BL,"center")
    hdr_row(ws,3,["Data & Ora","Eveniment","Tara","Impact","Anterior","Estimare","Actual","Deviere","Impact real","Activ afectat","Note"])
    ws.conditional_formatting.add("D4:D100",CellIsRule("equal",['"Ridicat"'],fill=F_OR,font=FN_OR))
    ws.conditional_formatting.add("D4:D100",CellIsRule("equal",['"Mediu"'],fill=F_YL,font=FN_YL))
    cf_status(ws,"I4:I100")

def _tmpl_jurnal(ws):
    for c,w in {"A":6,"B":12,"C":8,"D":14,"E":7,"F":14,"G":12,"H":12,"I":12,"J":10,"K":10,"L":12,"M":12,"N":12,"O":8,"P":10,"Q":10,"R":14,"S":14,"T":35,"U":15}.items():
        ws.column_dimensions[c].width=w
    ws.freeze_panes="D3"
    set_h(ws,1,35); mc(ws,1,1,1,21,"JURNAL TRANZACTII — HEDGE FUND STYLE",F_TI,FN_TI,"center")
    hdr_row(ws,2,["ID","Data","Ora","Activ","L/S","Setup","Entry","SL","TP","Marime poz","Risc $","Exit","Data iesire","P&L $","P&L %","RR Realizat","Cal exec (1-10)","Emotie","Plan?","Lectie","Link/SS"])
    ws.conditional_formatting.add("N3:N500",CellIsRule("greaterThan",["0"],fill=F_GN,font=FN_GN))
    ws.conditional_formatting.add("N3:N500",CellIsRule("lessThan",["0"],fill=F_RD,font=FN_RD))
    ws.conditional_formatting.add("E3:E500",CellIsRule("equal",['"LONG"'],fill=F_GN,font=FN_GN))
    ws.conditional_formatting.add("E3:E500",CellIsRule("equal",['"SHORT"'],fill=F_RD,font=FN_RD))
    set_h(ws,25,22); mc(ws,25,1,25,10,"STATISTICI PORTOFOLIU",F_H,FN_H,"left")
    stats=[("Nr tranzactii",'=COUNTA(A3:A22)'),("Win Rate",'=IFERROR(COUNTIF(N3:N22,">0")/COUNTA(N3:N22),0)'),
           ("P&L Total $",'=IFERROR(SUM(N3:N22),0)'),("P&L Mediu",'=IFERROR(AVERAGE(N3:N22),0)'),
           ("Best trade",'=IFERROR(MAX(N3:N22),0)'),("Worst trade",'=IFERROR(MIN(N3:N22),0)')]
    hdr_row(ws,26,["Metric","Valoare"],1)
    for i,(lbl,frm) in enumerate(stats):
        r=27+i; set_h(ws,r,20)
        ws.cell(r,1,lbl).fill=F_RH; ws.cell(r,1).font=FN_RH; ws.cell(r,1).alignment=aln("left")
        ws.cell(r,2,frm).fill=F_CA; ws.cell(r,2).alignment=aln("center")

def _tmpl_istoric(ws):
    for c,w in {"A":12,"B":8,"C":14,"D":14,"E":8,"F":8,"G":8,"H":10,"I":12,"J":12,"K":10,"L":10,"M":10,"N":10,"O":10,"P":10,"Q":20}.items():
        ws.column_dimensions[c].width=w
    ws.freeze_panes="B3"
    set_h(ws,1,35); mc(ws,1,1,1,17,"ISTORIC & TRENDING — SNAPSHOT LUNAR 24 LUNI",F_TI,FN_TI,"center")
    hdr_row(ws,2,["Luna/An","RSI Medie","Pret activ selectat","Pret S&P500","PIB YoY","CPI","VIX","Semnal Luna","Vol Mediu (B)","Trend dominant","Fear&Greed","Rate Dobanda","Yield 10Y","Gold","Oil WTI","EUR/USD","Note"])
    cf_signal(ws,"H3:H100"); cf_status(ws,"J3:J100")

def _tmpl_ghid(ws):
    for col,w in {"A":60,"B":50,"C":28}.items(): ws.column_dimensions[col].width=w
    ws.freeze_panes="A2"
    set_h(ws,1,55); mc(ws,1,1,1,3,"GHID DE INVATARE ZILNIC — ANALIZA DE PIATA PROFESIONALA",F_TI,FN_TI,"center")

def _tmpl_legenda(ws):
    for col,w in {"A":30,"B":12,"C":10,"D":45}.items(): ws.column_dimensions[col].width=w
    set_h(ws,1,35); mc(ws,1,1,1,4,"LEGENDA — SISTEM CULORI & ABREVIERI",F_TI,FN_TI,"center")
    rows=[("Celule editabile (input)","D9E1F2","000000","Date introduse manual"),
          ("Formule/calcule","FFFFFF","000000","Calculat automat"),
          ("Header coloane","1F4E79","FFFFFF","Titlu coloane principale"),
          ("Header randuri","D6DCE4","000000","Titlu rand/categorie"),
          ("BUY / Pozitiv","C6EFCE","375623","Semnal favorabil, trend bullish"),
          ("WAIT / Neutru","FFEB9C","9C6500","Semnale mixte, asteptare"),
          ("SELL / Negativ","FFC7CE","9C0006","Semnal negativ, trend bearish"),
          ("Alerta","FCE4D6","833C00","Atentie — actiune necesara")]
    hdr_row(ws,3,["Tip celula","Fundal","Text","Semnificatie"],1)
    for i,(tip,bg,fg_,sem) in enumerate(rows):
        r=4+i
        ws.cell(r,1,tip).fill=fill(bg); ws.cell(r,1).font=fnt(color=fg_); ws.cell(r,1).border=tb()
        ws.cell(r,2,"█████").fill=fill(bg); ws.cell(r,2).alignment=aln("center"); ws.cell(r,2).border=tb()
        ws.cell(r,3,"Abc").fill=fill(bg); ws.cell(r,3).font=fnt(color=fg_,bold=True); ws.cell(r,3).alignment=aln("center"); ws.cell(r,3).border=tb()
        ws.cell(r,4,sem).border=tb()
    set_h(ws,14,20); mc(ws,14,1,14,4,"ABREVIERI",F_H,FN_H,"left")
    abr=[("RSI","Relative Strength Index — momentum 0-100"),("MACD","Moving Average Convergence/Divergence"),
         ("MA","Medie Mobila (MA20=scurt, MA50=mediu, MA200=lung)"),("ATR","Average True Range — volatilitate medie"),
         ("RVOL","Relative Volume — volum vs medie 20z"),("RR","Risk/Reward Ratio — raport risc/castig"),
         ("SL","Stop Loss — nivel de iesire la pierdere"),("TP","Take Profit — nivel de iesire la profit"),
         ("MoM","Month over Month"),("YoY","Year over Year")]
    hdr_row(ws,15,["Abreviere","Semnificatie"],1); ws.merge_cells("B15:D15")
    for i,(ab,sig) in enumerate(abr):
        r=16+i; ws.cell(r,1,ab).fill=F_BL; ws.cell(r,1).font=fnt(bold=True,color="1F4E79"); ws.cell(r,1).border=tb()
        ws.cell(r,2,sig).border=tb(); ws.merge_cells(f"B{r}:D{r}")

def _tmpl_fisa(ws):
    for col,w in {"A":22,"B":16,"C":14,"D":14,"E":14,"F":14,"G":14,"H":14}.items():
        ws.column_dimensions[col].width=w
    ws.freeze_panes="A3"
    set_h(ws,1,45); mc(ws,1,1,1,8,"FISA COMPLETA ACTIV",F_TI,FN_TI,"center")
    set_h(ws,2,28); mc(ws,2,1,2,3,"Activ analizat:",F_SH,FN_SH,"right"); mc(ws,2,4,2,8,"",F_SL,FN_SL,"center")
    set_h(ws,3,18)
    mc(ws,3,1,3,8,"Schimba activul in DASHBOARD → celula portocalie → toate sectiunile se actualizeaza automat",
       fill("FFF9E6"),fnt(italic=True,color="9C6500",size=9),"center")

# ═══════════════════════════════════════════════════════════════════
# ACTUALIZARE SHEET-URI
# ═══════════════════════════════════════════════════════════════════
REF = "DASHBOARD!I2"

def _xlk(search_col, return_col, sheet, default='""'):
    """XLOOKUP wrapper — Office 2021 compatible."""
    return f'=IFERROR(XLOOKUP({REF},\'{sheet}\'!{search_col}:{search_col},\'{sheet}\'!{return_col}:{return_col},{default},0,-1),{default})'

def upd_dashboard(wb, all_data, fg):
    ws=wb["DASHBOARD"]
    if "LIST_ACTIVE" not in wb.sheetnames:
        wl=wb.create_sheet("LIST_ACTIVE"); wl.sheet_state="hidden"
    else: wl=wb["LIST_ACTIVE"]
    assets=[d["name"] for d in all_data.values() if d]
    for i,a in enumerate(assets,1): wl.cell(i,1,a)
    ws.cell(2,4).value=datetime.now().strftime("%d.%m.%Y %H:%M")
    unmerge_row(ws,2,9,13)
    if ws["I2"].value in (None,"","—"): ws["I2"].value=assets[0] if assets else "S&P 500"
    ws["I2"].fill=F_SL; ws["I2"].font=FN_SL; ws["I2"].alignment=aln("center","center")
    ws.merge_cells("I2:M2")
    dv=DataValidation(type="list",formula1=f"LIST_ACTIVE!$A$1:$A${len(assets)}",allow_blank=False,showDropDown=False)
    dv.sqref="I2:M2"; ws.add_data_validation(dv)
    # KPI Cards
    kpi_defs=[
        (1, _xlk("B","T","INDICATORI TEHNICI"), _xlk("B","W","INDICATORI TEHNICI"), "C6EFCE","375623"),
        (5, f'=IFERROR("ATR: "&TEXT(XLOOKUP({REF},\'INDICATORI TEHNICI\'!B:B,\'INDICATORI TEHNICI\'!O:O,"",0,-1),"0.0000"),"N/A")',
            f'=IFERROR("BB: "&TEXT(XLOOKUP({REF},\'INDICATORI TEHNICI\'!B:B,\'INDICATORI TEHNICI\'!N:N,"",0,-1),"0.0000"),"")',
            "FFEB9C","9C6500"),
        (9, f'=IFERROR(TEXT(XLOOKUP({REF},\'INDICATORI TEHNICI\'!B:B,\'INDICATORI TEHNICI\'!S:S,"",0,-1),"0.00")&"x medie","N/A")',
            f'=IFERROR("Vol: "&TEXT(XLOOKUP({REF},\'INDICATORI TEHNICI\'!B:B,\'INDICATORI TEHNICI\'!R:R,"",0,-1),"#,##0"),"")',
            "C6EFCE","375623"),
        (13,_xlk("B","C","SEMNALE INTRARE",'"N/A"'),
            f'=IFERROR(XLOOKUP({REF},\'SEMNALE INTRARE\'!B:B,\'SEMNALE INTRARE\'!J:J,0,0,-1)&"/5 conf","")',
            "C6EFCE","375623"),
    ]
    for cs,val_f,sub_f,bg,fg_ in kpi_defs:
        unmerge_row(ws,5,cs,cs+3); unmerge_row(ws,6,cs,cs+3)
        for r,frm,sz in [(5,val_f,14),(6,sub_f,9)]:
            mc(ws,r,cs,r,cs+3,frm,fill(bg),fnt(bold=(r==5),color=fg_,size=sz,italic=(r==6)),"center")
    for rng in ["A5:D5","M5:P5"]:
        for v,bg,fg_ in [("Bullish","C6EFCE","375623"),("Bearish","FFC7CE","9C0006"),("Sideways","FFEB9C","9C6500"),
                          ("BUY","C6EFCE","375623"),("SELL","FFC7CE","9C0006"),("WAIT","FFEB9C","9C6500")]:
            ws.conditional_formatting.add(rng,CellIsRule("equal",[f'"{v}"'],fill=fill(bg),font=fnt(bold=True,color=fg_,size=14)))
    # Signal Box
    sig={
        9: ("SEMNAL ACTIV",  _xlk("B","C","SEMNALE INTRARE",'"N/A"'), None),
        10:("Activ analizat",f"={REF}",None),
        11:("Entry Price",   _xlk("B","K","SEMNALE INTRARE"),"#,##0.0000"),
        12:("Stop Loss (SL)",_xlk("B","L","SEMNALE INTRARE"),"#,##0.0000"),
        13:("Take Profit (TP)",_xlk("B","M","SEMNALE INTRARE"),"#,##0.0000"),
        14:("Risk/Reward Ratio",f"=IFERROR((G13-G11)/(G11-G12),\"N/A\")","0.00\"x\""),
        15:("Confluente",    f'=IFERROR(XLOOKUP({REF},\'SEMNALE INTRARE\'!B:B,\'SEMNALE INTRARE\'!J:J,"",0,-1)&"/5","")',""),
        16:("Probabilitate",_xlk("B","O","SEMNALE INTRARE"),"0%"),
        17:("Conditie",      _xlk("B","D","SEMNALE INTRARE"),None),
    }
    fill_v={9:F_GN,12:fill("FFC7CE"),13:F_GN}
    font_v={9:fnt(bold=True,color="375623",size=13),12:fnt(bold=True,color="9C0006",size=11),13:fnt(bold=True,color="375623",size=11)}
    for row,(lbl,frm,fmt) in sig.items():
        set_h(ws,row,24); unmerge_row(ws,row,1,6); unmerge_row(ws,row,7,16)
        lc=ws.cell(row,1,lbl); lc.fill=F_RH; lc.font=FN_RH; lc.alignment=aln("right","center")
        ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=6)
        vc=ws.cell(row,7,frm); vc.fill=fill_v.get(row,F_CA); vc.font=font_v.get(row,fnt(size=11)); vc.alignment=aln("left","center")
        ws.merge_cells(start_row=row,start_column=7,end_row=row,end_column=16)
        if fmt: vc.number_format=fmt
    for v,bg,fg_ in [("BUY","C6EFCE","375623"),("SELL","FFC7CE","9C0006"),("WAIT","FFEB9C","9C6500")]:
        ws.conditional_formatting.add("G9:P9",CellIsRule("equal",[f'"{v}"'],fill=fill(bg),font=fnt(bold=True,color=fg_,size=13)))
    # Rezumat Indicatori
    ind={
        21:("RSI(14)",
            f'=IFERROR(TEXT(XLOOKUP({REF},\'INDICATORI TEHNICI\'!B:B,\'INDICATORI TEHNICI\'!G:G,"",0,-1),"0.0"),"")',
            _xlk("B","H","INDICATORI TEHNICI"),
            f'=IFERROR(IF(VALUE(XLOOKUP({REF},\'INDICATORI TEHNICI\'!B:B,\'INDICATORI TEHNICI\'!G:G,50,0,-1))<30,"Zona de acumulare",IF(VALUE(XLOOKUP({REF},\'INDICATORI TEHNICI\'!B:B,\'INDICATORI TEHNICI\'!G:G,50,0,-1))>70,"Zona de distributie","Zona echilibru")),"")'),
        22:("MACD Status", _xlk("B","W","INDICATORI TEHNICI"),
            f'=IFERROR(IF(ISNUMBER(SEARCH("bullish",XLOOKUP({REF},\'INDICATORI TEHNICI\'!B:B,\'INDICATORI TEHNICI\'!W:W,"",0,-1))),"BUY",IF(ISNUMBER(SEARCH("bearish",XLOOKUP({REF},\'INDICATORI TEHNICI\'!B:B,\'INDICATORI TEHNICI\'!W:W,"",0,-1))),"SELL","WAIT")),"")',
            f'=IFERROR("Hist: "&TEXT(XLOOKUP({REF},\'INDICATORI TEHNICI\'!B:B,\'INDICATORI TEHNICI\'!K:K,"",0,-1),"0.000"),"")'),
        23:("MA50 vs MA200", _xlk("B","V","INDICATORI TEHNICI"),
            f'=IFERROR(IF(XLOOKUP({REF},\'INDICATORI TEHNICI\'!B:B,\'INDICATORI TEHNICI\'!V:V,"",0,-1)="Golden Cross","BUY",IF(XLOOKUP({REF},\'INDICATORI TEHNICI\'!B:B,\'INDICATORI TEHNICI\'!V:V,"",0,-1)="Death Cross","SELL","WAIT")),"")',
            f'=IFERROR("MA50: "&TEXT(XLOOKUP({REF},\'INDICATORI TEHNICI\'!B:B,\'INDICATORI TEHNICI\'!E:E,"",0,-1),"#,##0.00")&" | MA200: "&TEXT(XLOOKUP({REF},\'INDICATORI TEHNICI\'!B:B,\'INDICATORI TEHNICI\'!F:F,"",0,-1),"#,##0.00"),"")'),
        24:("Trend activ", _xlk("B","T","INDICATORI TEHNICI"),
            f'=IFERROR(XLOOKUP({REF},\'INDICATORI TEHNICI\'!B:B,\'INDICATORI TEHNICI\'!T:T,"WAIT",0,-1),"WAIT")',
            f'=IFERROR("RVOL: "&TEXT(XLOOKUP({REF},\'INDICATORI TEHNICI\'!B:B,\'INDICATORI TEHNICI\'!S:S,"",0,-1),"0.00")&"x | ATR: "&TEXT(XLOOKUP({REF},\'INDICATORI TEHNICI\'!B:B,\'INDICATORI TEHNICI\'!O:O,"",0,-1),"0.0000"),"")'),
        25:("VIX",
            '=IFERROR(XLOOKUP("VIX",\'INDICATORI MACRO\'!A:A,\'INDICATORI MACRO\'!B:B,""),"")',
            '=IFERROR(IF(XLOOKUP("VIX",\'INDICATORI MACRO\'!A:A,\'INDICATORI MACRO\'!B:B,0)<15,"Pozitiv",IF(XLOOKUP("VIX",\'INDICATORI MACRO\'!A:A,\'INDICATORI MACRO\'!B:B,0)<25,"Neutru","Negativ")),"")',
            '=IFERROR(IF(XLOOKUP("VIX",\'INDICATORI MACRO\'!A:A,\'INDICATORI MACRO\'!B:B,0)<15,"Volatilitate scazuta",IF(XLOOKUP("VIX",\'INDICATORI MACRO\'!A:A,\'INDICATORI MACRO\'!B:B,0)<25,"Volatilitate moderata","Volatilitate ridicata")),"")'),
        26:("Fear & Greed",
            '=IFERROR(XLOOKUP("Fear & Greed (0-100)",\'INDICATORI MACRO\'!A:A,\'INDICATORI MACRO\'!B:B,""),"")',
            '=IFERROR(IF(XLOOKUP("Fear & Greed (0-100)",\'INDICATORI MACRO\'!A:A,\'INDICATORI MACRO\'!B:B,0)>60,"BUY",IF(XLOOKUP("Fear & Greed (0-100)",\'INDICATORI MACRO\'!A:A,\'INDICATORI MACRO\'!B:B,0)<40,"SELL","WAIT")),"")',
            '=IFERROR(IF(XLOOKUP("Fear & Greed (0-100)",\'INDICATORI MACRO\'!A:A,\'INDICATORI MACRO\'!B:B,0)>60,"Greed — sentiment favorabil",IF(XLOOKUP("Fear & Greed (0-100)",\'INDICATORI MACRO\'!A:A,\'INDICATORI MACRO\'!B:B,0)<40,"Fear — sentiment negativ","Neutru")),"")'),
        27:("RVOL activ",
            f'=IFERROR(TEXT(XLOOKUP({REF},\'INDICATORI TEHNICI\'!B:B,\'INDICATORI TEHNICI\'!S:S,"",0,-1),"0.00")&"x","")',
            f'=IFERROR(IF(XLOOKUP({REF},\'INDICATORI TEHNICI\'!B:B,\'INDICATORI TEHNICI\'!S:S,0,0,-1)>1.3,"Pozitiv",IF(XLOOKUP({REF},\'INDICATORI TEHNICI\'!B:B,\'INDICATORI TEHNICI\'!S:S,0,0,-1)<0.7,"Negativ","Neutru")),"")',
            f'=IFERROR("Vol: "&TEXT(XLOOKUP({REF},\'INDICATORI TEHNICI\'!B:B,\'INDICATORI TEHNICI\'!R:R,"",0,-1),"#,##0"),"")'),
    }
    for row,(lbl,val,status,obs) in ind.items():
        set_h(ws,row,20); unmerge_row(ws,row,2,4); unmerge_row(ws,row,5,8); unmerge_row(ws,row,9,16)
        ws.cell(row,1,lbl).fill=F_RH; ws.cell(row,1).font=FN_RH; ws.cell(row,1).alignment=aln("left"); ws.cell(row,1).border=tb()
        cv=ws.cell(row,2,val); cv.fill=F_CA; cv.font=fnt(bold=True,size=10); cv.alignment=aln("center"); cv.border=tb()
        ws.merge_cells(start_row=row,start_column=2,end_row=row,end_column=4)
        cs=ws.cell(row,5,status); cs.fill=F_CA; cs.font=fnt(bold=True,size=10); cs.alignment=aln("center"); cs.border=tb()
        ws.merge_cells(start_row=row,start_column=5,end_row=row,end_column=8)
        co=ws.cell(row,9,obs); co.fill=F_WH; co.font=fnt(size=9,italic=True,color="595959"); co.alignment=aln("left"); co.border=tb()
        ws.merge_cells(start_row=row,start_column=9,end_row=row,end_column=16)
    for v,bg,fg_ in [("BUY","C6EFCE","375623"),("SELL","FFC7CE","9C0006"),("WAIT","FFEB9C","9C6500"),
                      ("Pozitiv","C6EFCE","375623"),("Negativ","FFC7CE","9C0006"),("Neutru","FFEB9C","9C6500"),
                      ("Bullish","C6EFCE","375623"),("Bearish","FFC7CE","9C0006"),("Sideways","FFEB9C","9C6500")]:
        ws.conditional_formatting.add("E21:H27",CellIsRule("equal",[f'"{v}"'],fill=fill(bg),font=fnt(bold=True,color=fg_)))
    log.info("  DASHBOARD — complet dinamic")

def upd_semnale(ws, all_data):
    clr(ws); r=3
    today=datetime.now().strftime("%d.%m.%Y")
    for cat_name,cat_dict in CATEGORII:
        set_h(ws,r,18); mc(ws,r,1,r,17,f"— {cat_name} —",F_CT,FN_CT,"left"); r+=1
        for name,ticker in cat_dict.items():
            d=all_data.get(ticker,{}); 
            if not d: continue
            set_h(ws,r,18)
            vals=[today,name,d["semnal"],
                  f"RSI={d['rsi']:.0f} | {d['macd_cross']} | {d['macross']} | RVOL={d['rvol']:.1f}x",
                  d["rsi"],d["macd_cross"],d["macross"],
                  "Crescut" if d["rvol"]>1.3 else ("Scazut" if d["rvol"]<0.8 else "Normal"),
                  d["mom10"]/100,d["conf"],d["now"],d["sl"],d["tp"],None,d["prob"]/100,"Activ",f"Auto {datetime.now().strftime('%H:%M')}"]
            fmts=[None,None,None,None,"0.00",None,None,None,"0.00%","0","#,##0.0000","#,##0.0000","#,##0.0000","0.00\"x\"","0%",None,None]
            for j,(v,fmt) in enumerate(zip(vals,fmts)):
                c=ws.cell(r,j+1)
                if v is not None: c.value=v
                if fmt: c.number_format=fmt
                c.alignment=aln("left" if j in [1,3,16] else "center")
            ws.cell(r,14).value=f"=IFERROR((M{r}-K{r})/(K{r}-L{r}),\"N/A\")"; ws.cell(r,14).number_format="0.00\"x\""
            sc=ws.cell(r,3)
            if d["semnal"]=="BUY": sc.fill=F_GN;sc.font=FN_GN
            elif d["semnal"]=="SELL": sc.fill=F_RD;sc.font=FN_RD
            else: sc.fill=F_YL;sc.font=FN_YL
            ws.cell(r,2).font=fnt(bold=True); r+=1
    log.info(f"  SEMNALE INTRARE — {r-3} randuri (toate activele)")

def upd_tehnic(ws, all_data):
    clr(ws); r=3
    for cat_name,cat_dict in CATEGORII:
        set_h(ws,r,18); mc(ws,r,1,r,23,f"— {cat_name} —",F_CT,FN_CT,"left"); r+=1
        for name,ticker in cat_dict.items():
            d=all_data.get(ticker,{}); 
            if not d: continue
            set_h(ws,r,18)
            vals=[d["data"],name,d["now"],d["ma20"],d["ma50"],d["ma200"],d["rsi"],d["rsi_st"],
                  d["macd"],d["macd_sig"],d["macd_hist"],d["bb_sup"],d["bb_inf"],d["bb_w"],d["atr"],
                  d["stoch_k"],d["stoch_d"],d["vol"],d["rvol"],d["trend"],None,None,d["macross"]]
            fmts=[None,None,"#,##0.0000","#,##0.0000","#,##0.0000","#,##0.0000","0.00",None,
                  "0.000000","0.000000","0.000000","#,##0.0000","#,##0.0000","#,##0.0000","0.0000",
                  "0.00","0.00","#,##0","0.00\"x\"",None,None,None,None]
            for j,(v,fmt) in enumerate(zip(vals,fmts)):
                c=ws.cell(r,j+1)
                if v is not None: c.value=v
                if fmt: c.number_format=fmt
                c.alignment=aln("left" if j==1 else "center")
            ws.cell(r,2).font=fnt(bold=True)
            rsi_c=ws.cell(r,7)
            if d["rsi"]<30: rsi_c.fill=F_GN;rsi_c.font=FN_GN
            elif d["rsi"]<50: rsi_c.fill=F_YL;rsi_c.font=FN_YL
            elif d["rsi"]<70: rsi_c.fill=fill("E2EFDA");rsi_c.font=fnt(color="375623")
            else: rsi_c.fill=F_RD;rsi_c.font=FN_RD
            r+=1
    log.info(f"  INDICATORI TEHNICI — {r-3} randuri")

def upd_preturi(ws, all_data):
    clr(ws); r=3
    for cat_name,cat_dict in CATEGORII:
        set_h(ws,r,18); mc(ws,r,1,r,15,f"— {cat_name} —",F_CT,FN_CT,"left"); r+=1
        for name,ticker in cat_dict.items():
            d=all_data.get(ticker,{}); 
            if not d: continue
            set_h(ws,r,18)
            vals=[d["data"],name,d["open"],d["high"],d["low"],d["now"],
                  d["vzi"]/100,d["vsapt"]/100,d["vluna"]/100,d["vol"],d["avg_vol"],d["rvol"],None,None,d["trend"]]
            fmts=[None,None,"#,##0.0000","#,##0.0000","#,##0.0000","#,##0.0000",
                  "0.00%","0.00%","0.00%","#,##0","#,##0","0.00\"x\"",None,None,None]
            for j,(v,fmt) in enumerate(zip(vals,fmts)):
                c=ws.cell(r,j+1)
                if v is not None: c.value=v
                if fmt: c.number_format=fmt
            ws.cell(r,2).font=fnt(bold=True)
            for ci,val in [(7,d["vzi"]),(8,d["vsapt"]),(9,d["vluna"])]:
                cc=ws.cell(r,ci)
                if val>0: cc.fill=F_GN;cc.font=FN_GN
                elif val<0: cc.fill=F_RD;cc.font=FN_RD
            r+=1
    log.info(f"  PRETURI VOLUME — {r-3} randuri")

def upd_macro(ws, macro_live, fg):
    today=datetime.now().strftime("%d.%m.%Y %H:%M")
    macro_map={"VIX":macro_live.get("VIX",{}).get("now"),"Yield 10Y":macro_live.get("Yield 10Y US",{}).get("now"),
               "Yield 2Y":macro_live.get("Yield 2Y US",{}).get("now"),"USD Index":macro_live.get("USD Index",{}).get("now"),
               "Fear & Greed (0-100)":fg.get("value")}
    if FRED_API_KEY:
        macro_map["Rata dobanzii (%)"]=get_fred("FEDFUNDS")
        macro_map["CPI YoY (%)"]=get_fred("CPIAUCSL")
        macro_map["Rata somajului (%)"]=get_fred("UNRATE")
    upd=0
    for r in range(3,25):
        ind=ws.cell(r,1).value
        if not ind: break
        for k,v in macro_map.items():
            if k.lower() in str(ind).lower() and v is not None:
                prev=ws.cell(r,2).value
                try:
                    if prev and float(str(prev).replace(",","."))!=float(v): ws.cell(r,3).value=prev
                except: pass
                ws.cell(r,2).value=round(float(v),4); ws.cell(r,9).value=today; upd+=1; break
    log.info(f"  INDICATORI MACRO — {upd} valori")

def upd_comp(ws, all_data):
    clr(ws,4); r=4; today=datetime.now().strftime("%d.%m.%Y")
    for cat_name,cat_dict in CATEGORII:
        cat_key=list(cat_dict.keys())[0] if cat_dict else ""
        cat_id=categorie_activ(cat_key)
        competitori=COMPETITORI_MAP.get(cat_id,[])
        set_h(ws,r,18); mc(ws,r,1,r,15,f"— {cat_name} —",F_CT,FN_CT,"left"); r+=1
        for comp_name,comp_ticker,sector,tip in competitori:
            d=all_data.get(comp_ticker,{})
            set_h(ws,r,18)
            ws.cell(r,1,comp_name).font=fnt(bold=True)
            ws.cell(r,2,sector); ws.cell(r,5,d.get("now","") if d else "").number_format="#,##0.0000"
            trend_v=d.get("trend","") if d else ""
            ws.cell(r,6,trend_v).alignment=aln("center")
            ws.cell(r,9,tip); ws.cell(r,14,today)
            if trend_v=="Bullish": ws.cell(r,6).fill=F_GN;ws.cell(r,6).font=FN_GN
            elif trend_v=="Bearish": ws.cell(r,6).fill=F_RD;ws.cell(r,6).font=FN_RD
            elif trend_v: ws.cell(r,6).fill=F_YL;ws.cell(r,6).font=FN_YL
            r+=1
    log.info(f"  COMPETITORI SECTOR — {r-4} randuri")

def upd_riscuri(ws, all_data):
    clr(ws,4); r=4; today=datetime.now().strftime("%d.%m.%Y")
    for cat_name,cat_dict in CATEGORII:
        cat_key=list(cat_dict.keys())[0] if cat_dict else ""
        cat_id=categorie_activ(cat_key)
        riscuri=RISCURI_MAP.get(cat_id,[])
        set_h(ws,r,18); mc(ws,r,1,r,13,f"— {cat_name} —",F_CT,FN_CT,"left"); r+=1
        for id_r,tip,cat,desc,impact,prob,orizont,actiuni in riscuri:
            set_h(ws,r,22); scor=impact*prob/100
            vals=[id_r,tip,cat,desc,impact,prob,round(scor,1),orizont,actiuni,"Analyst","Monitorizat",today,""]
            for j,v in enumerate(vals):
                c=ws.cell(r,j+1,v); c.alignment=aln("left" if j in [3,8] else "center")
            tc=ws.cell(r,2)
            if tip=="Risc": tc.fill=F_RD;tc.font=FN_RD
            else: tc.fill=F_GN;tc.font=FN_GN
            r+=1
    log.info(f"  RISCURI OPORTUNITATI — {r-4} randuri")

def upd_calendar(ws, all_data):
    clr(ws,4); r=4; today=datetime.now()
    for cat_name,cat_dict in CATEGORII:
        cat_key=list(cat_dict.keys())[0] if cat_dict else ""
        cat_id=categorie_activ(cat_key)
        events=CALENDAR_MAP.get(cat_id,[])
        set_h(ws,r,18); mc(ws,r,1,r,11,f"— {cat_name} —",F_CT,FN_CT,"left"); r+=1
        for i,(evt,tara,sursa,activ) in enumerate(events):
            set_h(ws,r,18)
            data_evt=(today+timedelta(days=i*3+1)).strftime("%d.%m.%Y 14:30")
            impact="Ridicat" if i<3 else "Mediu"
            vals=[data_evt,evt,tara,impact,"","","","","",activ,""]
            for j,v in enumerate(vals): ws.cell(r,j+1,v).alignment=aln("left" if j in [1,9,10] else "center")
            ic=ws.cell(r,4)
            if impact=="Ridicat": ic.fill=F_OR;ic.font=FN_OR
            else: ic.fill=F_YL;ic.font=FN_YL
            r+=1
    log.info(f"  CALENDAR ECONOMIC — {r-4} randuri")

def upd_jurnal(ws, all_data):
    log.info("  JURNAL TRANZACTII — pastrat editabil")

def upd_istoric(ws, all_data, fg):
    luna=datetime.now().strftime("%b %Y")
    for row in ws.iter_rows(min_row=3,max_col=1):
        if row[0].value==luna: log.info(f"  ISTORIC — {luna} exista"); return
    last=2
    for row in ws.iter_rows(min_row=3,max_col=1):
        if row[0].value: last=row[0].row
    r=last+1
    rsi_vals=[d["rsi"] for d in all_data.values() if d and d.get("rsi")]
    avg_rsi=round(sum(rsi_vals)/len(rsi_vals),1) if rsi_vals else 50
    buy=sum(1 for d in all_data.values() if d and d.get("semnal")=="BUY")
    sell=sum(1 for d in all_data.values() if d and d.get("semnal")=="SELL")
    sm="BUY" if buy>sell else ("SELL" if sell>buy else "WAIT")
    sp500=all_data.get("^GSPC",{}).get("now")
    ws.cell(r,1,luna); ws.cell(r,2,avg_rsi); ws.cell(r,8,sm)
    if sp500: ws.cell(r,4,sp500)
    if fg.get("value"): ws.cell(r,11,fg["value"])
    sc=ws.cell(r,8)
    if sm=="BUY": sc.fill=F_GN;sc.font=FN_GN
    elif sm=="SELL": sc.fill=F_RD;sc.font=FN_RD
    else: sc.fill=F_YL;sc.font=FN_YL
    log.info(f"  ISTORIC — {luna} RSI={avg_rsi} Semnal={sm}")

def upd_rezumat(ws, all_data, fg):
    today=datetime.now().strftime("%d.%m.%Y %H:%M")
    ws.cell(2,4).value=f"={REF}"
    buy=sum(1 for d in all_data.values() if d and d.get("semnal")=="BUY")
    sell=sum(1 for d in all_data.values() if d and d.get("semnal")=="SELL")
    tot=len([d for d in all_data.values() if d])
    trend="Bullish" if tot and buy/tot>0.55 else ("Bearish" if tot and sell/tot>0.55 else "Mixt")
    rezumat_data=[
        ("Tendinta generala",trend,"↑" if trend=="Bullish" else ("↓" if trend=="Bearish" else "→"),f"BUY:{buy} SELL:{sell} din {tot} active"),
        ("VIX (volatilitate)",str(fg.get("display","N/A")),"→","Sentiment global"),
        ("Fear & Greed",str(fg.get("display","N/A")),"↑" if fg.get("status")=="Pozitiv" else "↓","Index 0-100"),
        ("Volume (general)","Normal","→","Fara anomalii majore"),
        ("Risc sistemic","Moderat","→","Inflatie si rate principale riscuri"),
    ]
    for i,(ind,val,trnd,obs) in enumerate(rezumat_data):
        r=6+i; set_h(ws,r,22)
        ws.cell(r,1,ind).fill=F_RH; ws.cell(r,1).font=FN_RH
        mc(ws,r,2,r,3,val,F_IN,FN_IN,"center")
        ws.cell(r,4,trnd).fill=F_CA; ws.cell(r,4).alignment=aln("center")
        mc(ws,r,5,r,8,obs,F_WH,fnt(italic=True,size=9,color="595959"),"left")
    concluzii=[
        f"Trend {'bullish' if trend=='Bullish' else 'bearish' if trend=='Bearish' else 'mixt'} — {buy} BUY, {sell} SELL din {tot}.",
        f"Fear & Greed: {fg.get('display','N/A')}.",
        "Selecteaza un activ in Dashboard pentru analiza detaliata.",
        "Respecta managementul riscului — max 2% din capital per tranzactie.",
        "Consulta FISA ACTIV pentru detalii complete pe activul selectat.",
    ]
    for i,c_ in enumerate(concluzii):
        r=14+i; set_h(ws,r,22)
        ws.cell(r,1,c_).fill=fill("F2F2F2" if i%2==0 else "FFFFFF"); ws.cell(r,1).font=fnt(size=10)
        ws.cell(r,1).alignment=aln("left"); ws.merge_cells(f"A{r}:H{r}")
    log.info("  REZUMAT EXECUTIV")

def upd_fisa_activ(wb, assets):
    SHEET="FISA ACTIV"
    if SHEET not in wb.sheetnames: wb.create_sheet(SHEET).sheet_properties.tabColor="ED7D31"
    ws=wb[SHEET]
    for col,w in {"A":22,"B":16,"C":14,"D":14,"E":14,"F":14,"G":14,"H":14}.items(): ws.column_dimensions[col].width=w
    ws.freeze_panes="A3"
    for row in ws.iter_rows(): [setattr(c,'value',None) or setattr(c,'fill',NO_FILL) for c in row]
    def sec2(row,label):
        set_h(ws,row,20); mc(ws,row,1,row,8,f"  {label}",fill("1F4E79"),fnt(bold=True,color="FFFFFF",size=10),"left")
    def kv1(row,lbl,frm,fmt=None):
        set_h(ws,row,20)
        ws.cell(row,1,lbl).fill=F_RH; ws.cell(row,1).font=FN_RH; ws.cell(row,1).border=tb()
        cv=ws.cell(row,2,frm); cv.fill=F_CA; cv.font=fnt(size=10); cv.border=tb()
        ws.merge_cells(start_row=row,start_column=2,end_row=row,end_column=8)
        if fmt: cv.number_format=fmt
    def kv2(row,l1,f1,fmt1,l2,f2,fmt2):
        set_h(ws,row,20)
        for col,lbl,frm,fmt in [(1,l1,f1,fmt1),(5,l2,f2,fmt2)]:
            ws.cell(row,col,lbl).fill=F_RH; ws.cell(row,col).font=FN_RH; ws.cell(row,col).border=tb()
            cv=ws.cell(row,col+1,frm); cv.fill=F_CA; cv.font=fnt(size=10); cv.border=tb()
            ws.merge_cells(start_row=row,start_column=col+1,end_row=row,end_column=col+3)
            if fmt: cv.number_format=fmt
    r=1
    set_h(ws,r,45); mc(ws,r,1,r,8,"FISA COMPLETA ACTIV",F_TI,FN_TI,"center"); r+=1
    set_h(ws,r,28); mc(ws,r,1,r,3,"Activ analizat:",F_SH,FN_SH,"right"); mc(ws,r,4,r,8,f"={REF}",F_SL,FN_SL,"center"); r+=1
    set_h(ws,r,18); mc(ws,r,1,r,8,"Schimba activul in DASHBOARD → celula portocalie",fill("FFF9E6"),fnt(italic=True,color="9C6500",size=9),"center"); r+=2
    sec2(r,"1 — SEMNAL DE INTRARE"); r+=1
    kv1(r,"Semnal",_xlk("B","C","SEMNALE INTRARE",'"N/A"')); r+=1
    kv1(r,"Conditie",_xlk("B","D","SEMNALE INTRARE")); r+=1
    kv2(r,"Entry",_xlk("B","K","SEMNALE INTRARE"),"#,##0.0000","Confluente",f'=IFERROR(XLOOKUP({REF},\'SEMNALE INTRARE\'!B:B,\'SEMNALE INTRARE\'!J:J,0,0,-1)&"/5","")',''); r+=1
    kv2(r,"Stop Loss",_xlk("B","L","SEMNALE INTRARE"),"#,##0.0000","Take Profit",_xlk("B","M","SEMNALE INTRARE"),"#,##0.0000"); r+=2
    sec2(r,"2 — INDICATORI TEHNICI"); r+=1
    for l1,c1,f1,l2,c2,f2 in [("Pret","C","#,##0.0000","Trend","T",""),("MA50","E","#,##0.0000","MA Cross","W",""),
                                ("RSI","G","0.00","RSI Status","H",""),("MACD Hist","K","0.000000","ATR","O","0.0000"),
                                ("Stoch %K","P","0.00","RVOL","S","0.00\"x\"")]:
        kv2(r,l1,f'=IFERROR(XLOOKUP({REF},\'INDICATORI TEHNICI\'!B:B,\'INDICATORI TEHNICI\'!{c1}:{c1},"",0,-1),"")',f1,
            l2,f'=IFERROR(XLOOKUP({REF},\'INDICATORI TEHNICI\'!B:B,\'INDICATORI TEHNICI\'!{c2}:{c2},"",0,-1),"")',f2); r+=1
    r+=1; sec2(r,"3 — PRETURI & VOLUME"); r+=1
    for l1,c1,f1,l2,c2,f2 in [("Deschidere","C","#,##0.0000","Inchidere","F","#,##0.0000"),
                                ("Var Zi","G","0.00%","Var Sapt","H","0.00%")]:
        kv2(r,l1,f'=IFERROR(XLOOKUP({REF},\'PRETURI VOLUME\'!B:B,\'PRETURI VOLUME\'!{c1}:{c1},"",0,-1),"")',f1,
            l2,f'=IFERROR(XLOOKUP({REF},\'PRETURI VOLUME\'!B:B,\'PRETURI VOLUME\'!{c2}:{c2},"",0,-1),"")',f2); r+=1
    r+=1
    set_h(ws,r,22)
    mc(ws,r,1,r,8,"Generata automat | Schimba activul in DASHBOARD",fill("E2EFDA"),fnt(italic=True,color="375623",size=9),"center")
    log.info("  FISA ACTIV — complet")

def upd_ghid(wb, all_data):
    SHEET="GHID INVATARE"
    if SHEET not in wb.sheetnames: ws=wb.create_sheet(SHEET); ws.sheet_properties.tabColor="154360"
    else: ws=wb[SHEET]
    for row in ws.iter_rows(): [setattr(c,'value',None) or setattr(c,'fill',NO_FILL) for c in row]
    for col,w in {"A":60,"B":50,"C":28}.items(): ws.column_dimensions[col].width=w
    ws.freeze_panes="A2"
    buy=sum(1 for d in all_data.values() if d and d.get("semnal")=="BUY")
    sell=sum(1 for d in all_data.values() if d and d.get("semnal")=="SELL")
    wait=sum(1 for d in all_data.values() if d and d.get("semnal")=="WAIT")
    ok=len([d for d in all_data.values() if d])
    r=1
    set_h(ws,r,55)
    mc(ws,r,1,r,3,f"GHID DE INVATARE ZILNIC — ANALIZA DE PIATA PROFESIONALA\nGenerat: {datetime.now().strftime('%d.%m.%Y  %H:%M')}   |   {ok} active   |   BUY: {buy}   SELL: {sell}   WAIT: {wait}",F_TI,FN_TI,"center"); r+=2
    for cat_name,cat_dict in CATEGORII:
        activi=[(n,t) for n,t in cat_dict.items() if t in all_data and all_data[t]]
        if not activi: continue
        set_h(ws,r,28); mc(ws,r,1,r,3,f"  {cat_name}",F_CT,FN_CT,"center"); r+=2
        for name,ticker in activi:
            d=all_data[ticker]; semnal=d["semnal"]
            f_head=(fill("1E6B3C") if semnal=="BUY" else fill("8B0000") if semnal=="SELL" else fill("7D5A00"))
            set_h(ws,r,24)
            mc(ws,r,1,r,3,f"  {name}   |   Pret: {fp(d['now'],4)}   |   Zi: {fpc(d['vzi'])}   |   RSI: {d['rsi']:.1f}   |   Conf: {d['conf']}/5   |   SEMNAL: {semnal}  ",f_head,FN_CT,"left"); r+=1
            for titlu,continut in [("  DE CE S-A MISCAT ASTAZI",text_miscare(d)),("  OPORTUNITATE DE TRADING",text_oportunitate(d)),
                                    ("  PATTERN GRAFIC DETECTAT",text_pattern(d)),("  LECTIA ZILEI",text_lectie(d))]:
                set_h(ws,r,18)
                h=ws.cell(r,1,titlu); h.fill=fill("D6DCE4"); h.font=fnt(bold=True,color="1F4E79",size=10); h.alignment=aln("left","center")
                ws.merge_cells(f"A{r}:C{r}"); r+=1
                nl=max(continut.count("\n")+1,3); set_h(ws,r,max(nl*13+16,50))
                cv=ws.cell(r,1,continut); cv.fill=F_WH; cv.font=fnt(size=10)
                cv.alignment=Alignment(horizontal="left",vertical="top",wrap_text=True)
                ws.merge_cells(f"A{r}:C{r}"); r+=1
            set_h(ws,r,6)
            for col in range(1,4): ws.cell(r,col).fill=fill("1F4E79")
            r+=2
        r+=1
    r+=2; set_h(ws,r,55)
    mc(ws,r,1,r,3,"GHID COMPLET DE CITIRE GRAFICE & INDICATORI TEHNICI\nReferinta permanenta — citeste zilnic pentru a-ti forma ochiul de trader",F_GT,FN_GT,"center"); r+=2
    for cap in GHID_GRAFICE:
        set_h(ws,r,28); mc(ws,r,1,r,3,f"  {cap['titlu']}",F_GC,FN_GC,"left"); r+=1
        for sub,txt in cap["sectiuni"]:
            set_h(ws,r,20)
            h=ws.cell(r,1,f"    {sub}"); h.fill=F_GS; h.font=FN_GS; h.alignment=aln("left","center")
            ws.merge_cells(f"A{r}:C{r}"); r+=1
            nl=max(txt.count("\n")+1,3); set_h(ws,r,max(nl*13+16,60))
            cv=ws.cell(r,1,txt); cv.fill=F_GX; cv.font=F_GX.copy() if hasattr(F_GX,'copy') else FN_GX
            cv.font=FN_GX; cv.alignment=Alignment(horizontal="left",vertical="top",wrap_text=True)
            ws.merge_cells(f"A{r}:C{r}"); r+=1
        set_h(ws,r,6)
        for col in range(1,4): ws.cell(r,col).fill=fill("2E75B6")
        r+=2
    log.info(f"  GHID INVATARE — {r} randuri")

# ═══════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════
def main():
    start=datetime.now()
    stamp=start.strftime("%Y-%m-%d_%H-%M")
    EXCEL_PATH=EXCEL_TEMPLATE.parent/f"Analiza_Piata_{stamp}.xlsx"
    log.info("="*65)
    log.info("  ANALIZA PIATA v3.1 — Script Complet")
    log.info(f"  {start.strftime('%d.%m.%Y  %H:%M:%S')}")
    log.info(f"  Active: {len(ACTIVE)} | Output: {EXCEL_PATH.name}")
    log.info("="*65)
    if not EXCEL_TEMPLATE.exists():
        log.info("\n  Template inexistent — creare template gol...")
        create_template(EXCEL_TEMPLATE)
    shutil.copy2(str(EXCEL_TEMPLATE),str(EXCEL_PATH))
    log.info(f"  Template copiat: {EXCEL_PATH.name}")
    log.info(f"\n  [1/3] Preia date pentru {len(ACTIVE)} active...\n")
    all_data={}
    for i,(name,ticker) in enumerate(ACTIVE.items(),1):
        d=get_data(name,ticker); all_data[ticker]=d
        if d: log.info(f"  [{i:3d}/{len(ACTIVE)}] {name:25s} {fp(d['now'],4):>14}  {fpc(d['vzi']):>8}  RSI={d['rsi']:5.1f}  [{d['semnal']}]")
        else: log.warning(f"  [{i:3d}/{len(ACTIVE)}] {name:25s}  -- N/A --")
    fg=get_fear_greed(); log.info(f"\n  Fear & Greed: {fg.get('display','N/A')}")
    macro_live={}
    for name,ticker in MACRO_T.items():
        d=get_data(name,ticker)
        if d: macro_live[name]=d; log.info(f"  {name:25s} = {fp(d['now'],4)}")
    ok=sum(1 for d in all_data.values() if d)
    log.info(f"\n  Procesate: {ok}/{len(ACTIVE)}")
    log.info("\n  [2/3] Actualizeaza Excel...\n")
    wb=load_workbook(str(EXCEL_PATH))
    steps=[("PRETURI VOLUME",upd_preturi,[all_data]),("INDICATORI TEHNICI",upd_tehnic,[all_data]),
           ("SEMNALE INTRARE",upd_semnale,[all_data]),("INDICATORI MACRO",upd_macro,[macro_live,fg]),
           ("COMPETITORI SECTOR",upd_comp,[all_data]),("RISCURI OPORTUNITATI",upd_riscuri,[all_data]),
           ("CALENDAR ECONOMIC",upd_calendar,[all_data]),("JURNAL TRANZACTII",upd_jurnal,[all_data]),
           ("ISTORIC TRENDING",upd_istoric,[all_data,fg]),("REZUMAT EXECUTIV",upd_rezumat,[all_data,fg]),
           ("DASHBOARD",upd_dashboard,[all_data,fg])]
    for sname,fn,args in steps:
        if sname not in wb.sheetnames: log.warning(f"  Sheet '{sname}' lipsa — skip"); continue
        try: fn(wb[sname],*args)
        except Exception as e: log.error(f"  EROARE {sname}: {e}")
    try: upd_fisa_activ(wb,[d["name"] for d in all_data.values() if d])
    except Exception as e: log.error(f"  EROARE FISA ACTIV: {e}")
    log.info("\n  [3/3] Ghid Invatare...\n")
    try: upd_ghid(wb,all_data)
    except Exception as e: log.error(f"  EROARE GHID: {e}")
    wb.save(str(EXCEL_PATH))
    fisiere=sorted(EXCEL_TEMPLATE.parent.glob("Analiza_Piata_*.xlsx"),key=lambda f:f.stat().st_mtime)
    for f in fisiere[:-30]:
        try: f.unlink()
        except: pass
    elapsed=(datetime.now()-start).seconds
    log.info("\n"+"="*65)
    log.info(f"  SALVAT: {EXCEL_PATH.name}")
    log.info(f"  Active OK: {ok}/{len(ACTIVE)}")
    log.info(f"  Durata: {elapsed//60}m {elapsed%60}s")
    log.info("="*65)

if __name__=="__main__":
    main()