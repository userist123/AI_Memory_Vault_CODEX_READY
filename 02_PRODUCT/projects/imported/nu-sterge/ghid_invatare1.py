"""
╔══════════════════════════════════════════════════════════════════════════════╗
║         GHID DE ÎNVĂȚARE ZILNIC — Analiză Piață Profesională                ║
║         Explică fiecare mișcare, oportunitate și pattern grafic             ║
╚══════════════════════════════════════════════════════════════════════════════╝

INSTALARE (o singură dată):
    pip install yfinance openpyxl requests pandas

RULARE (după actualizare_piata.py):
    python ghid_invatare.py

Ce face:
    • Preia date live pentru toate activele
    • Generează explicații detaliate: DE CE s-a mișcat, CE pattern există,
      CE oportunitate oferă, CE lecție practică se poate extrage
    • Adaugă sheet-ul "GHID INVATARE" în Excel
    • La finalul sheet-ului: ghid complet permanent de citire grafice
"""

import sys
import logging
from datetime import datetime
from pathlib import Path

import pandas as pd
import yfinance as yf
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ══════════════════════════════════════════════════════════════════════════════
# CONFIGURARE
# ══════════════════════════════════════════════════════════════════════════════

EXCEL_PATH = Path(r"C:\Users\Marius\Desktop\Nu sterge\Analiza_Piata_Profesionala.xlsx")

# ── ACTIVE ────────────────────────────────────────────────────────────────────
INDICI = {
    "S&P 500": "^GSPC", "NASDAQ 100": "^NDX", "NASDAQ Comp.": "^IXIC",
    "Dow Jones": "^DJI", "Russell 2000": "^RUT", "DAX Germany": "^GDAXI",
    "FTSE 100": "^FTSE", "CAC 40": "^FCHI", "Nikkei 225": "^N225",
    "Hang Seng": "^HSI", "Shanghai Comp.": "000001.SS",
    "MSCI World ETF": "URTH", "MSCI EM ETF": "EEM", "BET Romania": "BET.RO",
}
ACTIUNI = {
    "Apple": "AAPL", "Microsoft": "MSFT", "NVIDIA": "NVDA",
    "Alphabet": "GOOGL", "Amazon": "AMZN", "Meta": "META",
    "Tesla": "TSLA", "AMD": "AMD", "Intel": "INTC",
    "Broadcom": "AVGO", "ASML": "ASML", "Taiwan Semi": "TSM",
    "Palantir": "PLTR", "Salesforce": "CRM", "Oracle": "ORCL",
    "JPMorgan": "JPM", "Goldman Sachs": "GS", "Berkshire B": "BRK-B",
    "Visa": "V", "Mastercard": "MA", "ExxonMobil": "XOM",
    "Chevron": "CVX", "Shell": "SHEL", "Caterpillar": "CAT", "Boeing": "BA",
    "SPY": "SPY", "QQQ": "QQQ", "GLD ETF": "GLD",
    "TLT Bond ETF": "TLT", "ARKK": "ARKK",
}
CRYPTO = {
    "Bitcoin": "BTC-USD", "Ethereum": "ETH-USD", "BNB": "BNB-USD",
    "Solana": "SOL-USD", "XRP": "XRP-USD", "Cardano": "ADA-USD",
    "Avalanche": "AVAX-USD", "Dogecoin": "DOGE-USD", "Chainlink": "LINK-USD",
    "Polkadot": "DOT-USD", "Litecoin": "LTC-USD", "Shiba Inu": "SHIB-USD",
    "Polygon": "MATIC-USD", "Uniswap": "UNI-USD", "Cosmos": "ATOM-USD",
    "Stellar": "XLM-USD", "Monero": "XMR-USD", "Tron": "TRX-USD",
    "Filecoin": "FIL-USD", "Aave": "AAVE-USD", "Arbitrum": "ARB-USD",
    "Optimism": "OP-USD", "Render": "RNDR-USD", "Sui": "SUI-USD",
    "Near Protocol": "NEAR-USD",
}
VALUTE = {
    "EUR/USD": "EURUSD=X", "GBP/USD": "GBPUSD=X", "USD/JPY": "USDJPY=X",
    "USD/CHF": "USDCHF=X", "AUD/USD": "AUDUSD=X", "USD/CAD": "USDCAD=X",
    "EUR/RON": "EURRON=X", "USD/RON": "USDRON=X", "GBP/RON": "GBPRON=X",
    "EUR/GBP": "EURGBP=X", "USD/CNY": "USDCNY=X", "USD/TRY": "USDTRY=X",
}
MATERII_PRIME = {
    "Gold": "GC=F", "Silver": "SI=F", "Platinum": "PL=F", "Palladium": "PA=F",
    "Oil WTI": "CL=F", "Oil Brent": "BZ=F", "Natural Gas": "NG=F",
    "Copper": "HG=F", "Corn": "ZC=F", "Wheat": "ZW=F",
    "Soybean": "ZS=F", "Coffee": "KC=F", "Sugar": "SB=F", "Cotton": "CT=F",
}
ACTIVE = {**INDICI, **ACTIUNI, **CRYPTO, **VALUTE, **MATERII_PRIME}
CATEGORII = [
    ("INDICI BURSIERI", INDICI),
    ("ACTIUNI & ETF",   ACTIUNI),
    ("CRYPTOCURRENCY",  CRYPTO),
    ("VALUTE FOREX",    VALUTE),
    ("MATERII PRIME",   MATERII_PRIME),
]

# ── LOGGING ───────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(Path(__file__).parent / "ghid.log", encoding="utf-8"),
    ],
)
log = logging.getLogger(__name__)

# ══════════════════════════════════════════════════════════════════════════════
# STILURI EXCEL
# ══════════════════════════════════════════════════════════════════════════════

def fill(hex_c):
    return PatternFill("solid", start_color=hex_c, fgColor=hex_c)

def fnt(bold=False, color="000000", size=10, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)

def aln(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

NO_FILL = PatternFill(fill_type=None)

# Paleta culori
F_TITLU     = fill("0D2137"); FN_TITLU     = fnt(bold=True,  color="FFFFFF", size=15)
F_CAT       = fill("1F4E79"); FN_CAT       = fnt(bold=True,  color="FFFFFF", size=11)
F_ACTIV_BUY = fill("1E6B3C"); FN_ACTIV     = fnt(bold=True,  color="FFFFFF", size=11)
F_ACTIV_SEL = fill("8B0000")
F_ACTIV_WAI = fill("7D5A00")
F_SEC_HDR   = fill("D6DCE4"); FN_SEC_HDR   = fnt(bold=True,  color="1F4E79", size=10)
F_GRN       = fill("C6EFCE"); FN_GRN       = fnt(color="375623", bold=True)
F_YLW       = fill("FFEB9C"); FN_YLW       = fnt(color="9C6500", bold=True)
F_RED       = fill("FFC7CE"); FN_RED       = fnt(color="9C0006", bold=True)
F_ORG       = fill("FCE4D6"); FN_ORG       = fnt(color="833C00")
F_BLU       = fill("DEEAF1"); FN_BLU       = fnt(color="1F4E79", size=10)
F_GRY       = fill("F2F2F2"); FN_GRY       = fnt(color="404040", size=10)
F_WHITE     = fill("FFFFFF")
F_SEP       = fill("1F4E79")
# Ghid grafice
F_GHID_T    = fill("0A1628"); FN_GHID_T    = fnt(bold=True,  color="FFFFFF", size=14)
F_GHID_CAP  = fill("154360"); FN_GHID_CAP  = fnt(bold=True,  color="FFFFFF", size=11)
F_GHID_SUB  = fill("1A5276"); FN_GHID_SUB  = fnt(bold=True,  color="FFFFFF", size=10)
F_GHID_TXT  = fill("EBF5FB"); FN_GHID_TXT  = fnt(color="1A252F", size=10)

# ══════════════════════════════════════════════════════════════════════════════
# HELPER: formatare sigură (evită erori cu None)
# ══════════════════════════════════════════════════════════════════════════════

def fmt_price(val, decimals=4):
    """Formatează un preț în mod sigur, returnează 'N/A' dacă e None."""
    if val is None or (isinstance(val, float) and (val != val)):
        return "N/A"
    try:
        fmt = f"{{:,.{decimals}f}}"
        return fmt.format(float(val))
    except Exception:
        return "N/A"

def fmt_pct(val, decimals=2):
    """Formatează un procent în mod sigur."""
    if val is None:
        return "N/A"
    try:
        sign = "+" if float(val) >= 0 else ""
        return f"{sign}{float(val):.{decimals}f}%"
    except Exception:
        return "N/A"

def safe(val, default=0.0):
    """Returnează valoarea sau default dacă e None/NaN."""
    if val is None:
        return default
    try:
        f = float(val)
        return default if f != f else f  # NaN check
    except Exception:
        return default

# ══════════════════════════════════════════════════════════════════════════════
# PRELUARE DATE
# ══════════════════════════════════════════════════════════════════════════════

def get_data(name: str, ticker: str) -> dict:
    """Preia date istorice și calculează toți indicatorii tehnici."""
    try:
        t    = yf.Ticker(ticker)
        hist = t.history(period="1y", auto_adjust=True)
        if hist.empty or len(hist) < 5:
            log.warning(f"  ⚠  {name} — date insuficiente")
            return {}

        closes = hist["Close"]
        n      = len(closes)
        latest = hist.iloc[-1]
        prev   = hist.iloc[-2] if n > 1 else latest

        close_now  = float(latest["Close"])
        close_prev = float(prev["Close"])
        close_5d   = float(hist.iloc[-min(5,  n)]["Close"])
        close_20d  = float(hist.iloc[-min(20, n)]["Close"])

        var_zi   = (close_now - close_prev) / close_prev * 100 if close_prev else 0
        var_sapt = (close_now - close_5d)   / close_5d   * 100 if close_5d   else 0
        var_luna = (close_now - close_20d)  / close_20d  * 100 if close_20d  else 0

        volume  = int(latest.get("Volume", 0))
        avg_vol = int(hist["Volume"].tail(min(20, n)).mean()) if "Volume" in hist else 0
        rvol    = round(volume / avg_vol, 2) if avg_vol > 0 else 1.0

        # RSI(14)
        delta = closes.diff()
        gain  = delta.clip(lower=0).rolling(14).mean()
        loss  = (-delta.clip(upper=0)).rolling(14).mean()
        rs    = gain / loss.replace(0, 1e-10)
        rsi_s = (100 - 100 / (1 + rs)).iloc[-1]
        rsi   = round(float(rsi_s), 2) if pd.notna(rsi_s) else 50.0

        # MACD(12,26,9)
        ema12    = closes.ewm(span=12, adjust=False).mean()
        ema26    = closes.ewm(span=26, adjust=False).mean()
        ml       = ema12 - ema26
        sl       = ml.ewm(span=9, adjust=False).mean()
        hist_m   = ml - sl
        m_v      = float(ml.iloc[-1])
        s_v      = float(sl.iloc[-1])
        h_v      = float(hist_m.iloc[-1])
        prev_h   = float(hist_m.iloc[-2]) if n > 1 else 0
        if m_v > s_v:
            macd_cross = "Bullish Cross" if prev_h < 0 else "Pozitiv"
        else:
            macd_cross = "Bearish Cross" if prev_h > 0 else "Negativ"

        # Medii mobile (None dacă date insuficiente)
        def ma(p):
            if n >= p:
                return round(float(closes.tail(p).mean()), 6)
            return None

        ma20  = ma(20)
        ma50  = ma(50)
        ma200 = ma(200)

        if ma50 is not None and ma200 is not None:
            macross = "Golden Cross" if ma50 > ma200 else "Death Cross"
        elif ma50 is not None:
            macross = "Sub MA200 (insuf. date)"
        else:
            macross = "Date insuficiente"

        # Bollinger Bands(20,2)
        if n >= 20:
            ma20_s  = closes.rolling(20).mean()
            std20   = closes.rolling(20).std()
            bb_sup  = float((ma20_s + 2 * std20).iloc[-1])
            bb_inf  = float((ma20_s - 2 * std20).iloc[-1])
            bb_w    = bb_sup - bb_inf
        else:
            bb_sup = bb_inf = bb_w = None

        # ATR(14)
        hi, lo, cl = hist["High"], hist["Low"], hist["Close"]
        tr  = pd.concat([(hi - lo),
                          (hi - cl.shift()).abs(),
                          (lo - cl.shift()).abs()], axis=1).max(axis=1)
        atr_raw = tr.rolling(14).mean().iloc[-1]
        atr = round(float(atr_raw), 6) if pd.notna(atr_raw) else close_now * 0.02

        # Stochastic(14,3)
        if n >= 14:
            lo14   = lo.rolling(14).min()
            hi14   = hi.rolling(14).max()
            diff14 = (hi14 - lo14).replace(0, 1e-10)
            k_s    = (cl - lo14) / diff14 * 100
            d_s    = k_s.rolling(3).mean()
            stoch_k = round(float(k_s.iloc[-1]), 2)  if pd.notna(k_s.iloc[-1]) else 50.0
            stoch_d = round(float(d_s.iloc[-1]), 2)  if pd.notna(d_s.iloc[-1]) else 50.0
        else:
            stoch_k = stoch_d = 50.0

        # Momentum 10 zile
        mom10 = round(float(closes.pct_change(10).iloc[-1] * 100), 2) if n > 10 else 0.0

        # Trend
        if ma50 is not None and close_now > ma50 * 1.01:
            trend = "Bullish"
        elif ma50 is not None and close_now < ma50 * 0.99:
            trend = "Bearish"
        else:
            trend = "Sideways"

        # Semnal
        score = 0
        if rsi < 35:                          score += 2
        elif rsi < 45:                        score += 1
        elif rsi > 75:                        score -= 2
        elif rsi > 65:                        score -= 1
        if "Bullish Cross" in macd_cross:     score += 2
        elif "Pozitiv" in macd_cross:         score += 1
        elif "Bearish Cross" in macd_cross:   score -= 2
        elif "Negativ" in macd_cross:         score -= 1
        if macross == "Golden Cross":         score += 2
        elif macross == "Death Cross":        score -= 2
        if rvol > 1.5:                        score += 1
        elif rvol < 0.6:                      score -= 1
        conf = min(abs(score), 5)
        if score >= 3:    semnal = "BUY"
        elif score <= -3: semnal = "SELL"
        else:             semnal = "WAIT"

        price = close_now
        sl_v = round(price - 1.5 * atr, 6) if semnal == "BUY"  else \
               round(price + 1.5 * atr, 6) if semnal == "SELL" else \
               round(price - 2.0 * atr, 6)
        tp_v = round(price + 3.0 * atr, 6) if semnal == "BUY"  else \
               round(price - 3.0 * atr, 6) if semnal == "SELL" else \
               round(price + 2.0 * atr, 6)
        prob = min(90, 35 + conf * 10 + (5 if rvol > 1.2 else 0))

        return {
            "name":        name,
            "ticker":      ticker,
            "data":        datetime.now().strftime("%d.%m.%Y"),
            "inchidere":   round(close_now, 6),
            "var_zi":      round(var_zi,   4),
            "var_sapt":    round(var_sapt, 4),
            "var_luna":    round(var_luna, 4),
            "volum":       volume,
            "avg_vol":     avg_vol,
            "rvol":        rvol,
            "rsi":         rsi,
            "macd":        round(m_v, 6),
            "macd_signal": round(s_v, 6),
            "macd_hist":   round(h_v, 6),
            "macd_cross":  macd_cross,
            "ma20":        ma20,
            "ma50":        ma50,
            "ma200":       ma200,
            "macross":     macross,
            "bb_sup":      bb_sup,
            "bb_inf":      bb_inf,
            "bb_width":    bb_w,
            "atr":         atr,
            "stoch_k":     stoch_k,
            "stoch_d":     stoch_d,
            "mom10":       mom10,
            "trend":       trend,
            "semnal":      semnal,
            "confluente":  conf,
            "sl":          sl_v,
            "tp":          tp_v,
            "prob":        prob,
            "n_zile":      n,
        }
    except Exception as e:
        log.error(f"  ✗  {name} ({ticker}): {e}")
        return {}

# ══════════════════════════════════════════════════════════════════════════════
# MOTOR EXPLICAȚII — tot textul educativ generat din date reale
# ══════════════════════════════════════════════════════════════════════════════

def explica_miscare(d: dict) -> str:
    """Explicație detaliată DE CE s-a mișcat activul azi."""
    name      = d["name"]
    pret      = d["inchidere"]
    var_zi    = d["var_zi"]
    var_sapt  = d["var_sapt"]
    var_luna  = d["var_luna"]
    rsi       = d["rsi"]
    macd_c    = d["macd_cross"]
    macd_h    = d["macd_hist"]
    macross   = d["macross"]
    rvol      = d["rvol"]
    ma50      = d["ma50"]
    ma200     = d["ma200"]
    bb_sup    = d["bb_sup"]
    bb_inf    = d["bb_inf"]
    bb_w      = d["bb_width"]
    stoch_k   = d["stoch_k"]
    mom10     = d["mom10"]
    semnal    = d["semnal"]
    n_zile    = d["n_zile"]

    # formate sigure
    p_str    = fmt_price(pret)
    ma50_s   = fmt_price(ma50,  2)
    ma200_s  = fmt_price(ma200, 2)
    bb_sup_s = fmt_price(bb_sup, 4)
    bb_inf_s = fmt_price(bb_inf, 4)
    bb_w_s   = fmt_price(bb_w,  4)

    linii = []

    # 1. Mișcarea zilei
    directie   = "crescut" if var_zi > 0 else "scăzut"
    intensitate = ("semnificativ" if abs(var_zi) > 3
                   else "moderat" if abs(var_zi) > 1
                   else "ușor")
    linii.append(
        f"{name} a {directie} {intensitate} cu {abs(var_zi):.2f}% astăzi "
        f"(preț curent: {p_str}).\n"
        f"Pe săptămână: {fmt_pct(var_sapt)} | Pe lună: {fmt_pct(var_luna)}."
    )

    # 2. Volum
    if rvol > 1.5:
        linii.append(
            f"VOLUM EXCEPȚIONAL: {rvol:.1f}x față de media pe 20 de zile.\n"
            f"Mișcarea de astăzi este CONFIRMATĂ de participare instituțională. "
            f"Volumul ridicat validează direcția — nu este o mișcare întâmplătoare. "
            f"Fondurile, băncile și algos tranzacționează activ acest activ azi."
        )
    elif rvol > 1.2:
        linii.append(
            f"Volum ușor peste medie ({rvol:.1f}x) — mișcarea are suport real "
            f"din piață, fără anomalii. Participare normală spre ridicată."
        )
    elif rvol < 0.7:
        linii.append(
            f"VOLUM SCĂZUT: {rvol:.1f}x față de medie — ATENȚIE!\n"
            f"Mișcările pe volum mic sunt adesea false (fake moves). "
            f"Marii jucători nu participă activ azi. Probabilitate crescută ca "
            f"mișcarea să se inverseze sau să nu continue."
        )
    else:
        linii.append(
            f"Volum în linie cu media ({rvol:.1f}x) — nicio anomalie de participare."
        )

    # 3. RSI
    if rsi < 30:
        linii.append(
            f"RSI = {rsi:.1f} — ZONĂ SUPRAVÂNDUT (sub 30).\n"
            f"Activul a scăzut prea rapid și există presiune acumulată de cumpărare. "
            f"Istoric, RSI sub 30 semnalează că prețul este 'ieftin' față de media sa. "
            f"Nu înseamnă revenire imediată — poate rămâne sub 30 zile sau săptămâni — "
            f"dar crește semnificativ probabilitatea unui bounce (recuperare) în zilele imediat următoare."
        )
    elif rsi < 40:
        linii.append(
            f"RSI = {rsi:.1f} — Zonă slabă dar fără extremă.\n"
            f"Vânzătorii domină încă, dar presiunea începe să scadă. "
            f"Urmărește o stabilizare a prețului și un candle bullish de confirmare "
            f"înainte de a considera o intrare long."
        )
    elif rsi < 55:
        linii.append(
            f"RSI = {rsi:.1f} — Zonă NEUTRĂ.\n"
            f"Echilibru între cumpărători și vânzători. Piața caută direcție. "
            f"Evita pozițiile mari până la o confirmare clară de trend."
        )
    elif rsi < 70:
        linii.append(
            f"RSI = {rsi:.1f} — Momentum POZITIV.\n"
            f"Cumpărătorii au controlul, dar activul nu este încă supraevaluat. "
            f"Zona 55-70 este adesea CEA MAI PROFITABILĂ pentru trend-following. "
            f"Trendul bullish se menține solid."
        )
    else:
        linii.append(
            f"RSI = {rsi:.1f} — ZONĂ SUPRAEVALUAT (peste 70).\n"
            f"Activul a crescut prea rapid. Probabilitatea unei corecții sau "
            f"consolidări crește. Cei cu poziții long ar trebui să ia în considerare "
            f"protejarea profiturilor (trailing stop sau reducere parțială a expunerii)."
        )

    # 4. MACD
    if "Bullish Cross" in macd_c:
        linii.append(
            f"MACD: BULLISH CROSSOVER CONFIRMAT (histogram: {macd_h:+.4f}).\n"
            f"Linia MACD a trecut DEASUPRA liniei de semnal — unul dintre cele mai "
            f"puternice semnale de schimbare a direcției spre bullish. "
            f"Momentumul câștigă forță în favoarea cumpărătorilor. "
            f"Confirmă cu volumul crescut pentru un semnal de calitate."
        )
    elif "Bearish Cross" in macd_c:
        linii.append(
            f"MACD: BEARISH CROSSOVER CONFIRMAT (histogram: {macd_h:+.4f}).\n"
            f"Linia MACD a trecut DEDESUBT liniei de semnal. "
            f"Momentumul bullish s-a epuizat — vânzătorii preiau controlul. "
            f"Semnal de prudență pentru deținătorii de poziții long."
        )
    elif "Pozitiv" in macd_c:
        linii.append(
            f"MACD pozitiv și crescător (histogram: {macd_h:+.4f}).\n"
            f"Momentumul bullish se menține, fără un crossover proaspăt. "
            f"Trendul continuă dar fără accelerare nouă."
        )
    else:
        linii.append(
            f"MACD negativ (histogram: {macd_h:+.4f}).\n"
            f"Presiunea de vânzare domină pe termen scurt. "
            f"Fii precaut cu pozițiile long în acest context."
        )

    # 5. Medii mobile
    if macross == "Golden Cross" and ma50 is not None and ma200 is not None:
        linii.append(
            f"GOLDEN CROSS ACTIV: MA50 ({ma50_s}) > MA200 ({ma200_s}).\n"
            f"Cea mai importantă configurație bullish pe termen mediu-lung. "
            f"Fondurile de pensii, ETF-urile și robo-advisors cumpără automat la Golden Cross. "
            f"Trendul major este ascendent — scăderile sunt oportunități de cumpărare, "
            f"nu semnale de panică."
        )
    elif macross == "Death Cross" and ma50 is not None and ma200 is not None:
        linii.append(
            f"DEATH CROSS ACTIV: MA50 ({ma50_s}) < MA200 ({ma200_s}).\n"
            f"Semnal bearish major pe termen mediu-lung. Mulți investitori "
            f"instituționali reduc expunerea automat la Death Cross. "
            f"Creșterile temporare pot fi oportunități de vânzare (short) "
            f"sau ieșire din poziții long."
        )
    elif ma50 is not None and ma200 is None:
        linii.append(
            f"MA50 disponibilă ({ma50_s}), dar MA200 nu poate fi calculată "
            f"(activ cu istoric sub 200 zile — {n_zile} zile disponibile). "
            f"Urmărește relația prețului față de MA50 ca referință principală."
        )
    else:
        linii.append(
            f"MA50 ({ma50_s}) și MA200 ({ma200_s}) aproape la același nivel — "
            f"piața în tranziție, fără un trend major clar stabilit pe termen lung."
        )

    # 6. Bollinger Bands
    if bb_sup is not None and bb_inf is not None and bb_w is not None:
        pret_f = float(pret)
        if pret_f >= float(bb_sup) * 0.99:
            linii.append(
                f"PREȚ LA BANDA SUPERIOARĂ Bollinger ({bb_sup_s}).\n"
                f"Statistic, prețul revine spre medie în ~80% din cazuri după atingerea "
                f"benzii superioare. Semnal: posibilă supraextindere pe termen scurt — "
                f"risc crescut de corecție iminentă sau consolidare orizontală."
            )
        elif pret_f <= float(bb_inf) * 1.01:
            linii.append(
                f"PREȚ LA BANDA INFERIOARĂ Bollinger ({bb_inf_s}).\n"
                f"Zonă de potențial suport tehnic. Statistic, 80% din oscilații "
                f"rămân în interiorul benzilor. Posibilă revenire spre MA20 "
                f"(mijlocul benzilor) în zilele imediat următoare."
            )
        elif float(bb_w) / pret_f < 0.03:
            linii.append(
                f"BOLLINGER SQUEEZE detectat (lățime: {bb_w_s}).\n"
                f"Volatilitate extrem de scăzută — benzile sunt comprimate. "
                f"Istoric, perioadele de compresie sunt urmate INVARIABIL de "
                f"mișcări explozive. Direcția nu este clară, dar amplitudinea va fi mare. "
                f"Pregătește ordine pending în ambele direcții."
            )
        else:
            linii.append(
                f"Prețul în interiorul Bollinger ({bb_inf_s} — {bb_sup_s}). "
                f"Nicio extremă pentru moment, volatilitate normală."
            )
    else:
        linii.append(
            f"Benzile Bollinger nu pot fi calculate (date insuficiente: {n_zile} zile)."
        )

    # 7. Stochastic
    if stoch_k < 20:
        linii.append(
            f"Stochastic %K = {stoch_k:.1f} — OVERSOLD (sub 20).\n"
            f"Presiune de vânzare excesivă pe termen scurt. Când %K depășește %D "
            f"din jos în sus sub nivelul 20, este un semnal clasic de cumpărare pe termen scurt."
        )
    elif stoch_k > 80:
        linii.append(
            f"Stochastic %K = {stoch_k:.1f} — OVERBOUGHT (peste 80).\n"
            f"Activul s-a apreciat rapid pe termen scurt. Când %K coboară sub %D "
            f"din zona >80, este semnal de vânzare / luare profit pe termen scurt."
        )

    # 8. Momentum
    if abs(mom10) > 5:
        dir_m = "PUTERNIC POZITIV" if mom10 > 0 else "PUTERNIC NEGATIV"
        linii.append(
            f"Momentum 10 zile: {fmt_pct(mom10)} — {dir_m}.\n"
            f"Activul {'a acumulat câștiguri consistente' if mom10 > 0 else 'a pierdut teren consistent'} "
            f"în ultimele 2 săptămâni. Trendul pe termen scurt este {'bullish' if mom10 > 0 else 'bearish'} solid."
        )

    # 9. Concluzie
    if semnal == "BUY":
        linii.append(
            f"CONCLUZIE: Configurația tehnică este FAVORABILĂ CUMPĂRĂRII.\n"
            f"Momentul tehnic este bun. Riscul principal: o deteriorare bruscă a "
            f"sentimentului global (VIX spike, știri macro negative) poate invalida setup-ul. "
            f"Respectă strict stop loss-ul!"
        )
    elif semnal == "SELL":
        linii.append(
            f"CONCLUZIE: Configurația tehnică indică PRESIUNE DE VÂNZARE.\n"
            f"Prudență cu pozițiile long existente. Dacă ești short, "
            f"respectă stop loss-ul pentru a nu fi prins în rally-uri de recuperare (bear rallies)."
        )
    else:
        linii.append(
            f"CONCLUZIE: Semnalele sunt MIXTE — AȘTEPTARE recomandată.\n"
            f"Nu forța o tranzacție când piața nu are direcție clară. "
            f"Cel mai bun trade uneori este să nu tranzacționezi."
        )

    return "\n\n".join(linii)


def explica_oportunitate(d: dict) -> str:
    """Explică oportunitatea de tranzacționare cu parametrii completi."""
    name   = d["name"]
    semnal = d["semnal"]
    rsi    = d["rsi"]
    pret   = d["inchidere"]
    sl     = d["sl"]
    tp     = d["tp"]
    atr    = d["atr"]
    rvol   = d["rvol"]
    conf   = d["confluente"]
    prob   = d["prob"]
    macross= d["macross"]

    p_str  = fmt_price(pret, 4)
    sl_str = fmt_price(sl,   4)
    tp_str = fmt_price(tp,   4)
    atr_s  = fmt_price(atr,  4)

    dist_sl = abs(float(pret) - float(sl))
    dist_tp = abs(float(tp)   - float(pret))
    rr      = round(dist_tp / dist_sl, 2) if dist_sl > 0 else 0
    risc_pct = dist_sl / float(pret) * 100 if float(pret) > 0 else 0

    linii = []

    if semnal == "BUY":
        rr_eval = ("EXCELENT ★★★" if rr >= 3
                   else "BUN ★★" if rr >= 2
                   else "ACCEPTABIL ★" if rr >= 1.5
                   else "SLAB — reconsideră")
        linii.append(
            f"OPORTUNITATE DE CUMPĂRARE — {conf} confluențe tehnice aliniate.\n"
            f"Probabilitate estimată de succes: {prob}%\n\n"
            f"PARAMETRI TRANZACȚIE:\n"
            f"  Entry (intrare recomandata) :  {p_str}\n"
            f"  Stop Loss (SL)             :  {sl_str}   "
            f"  ← risc {risc_pct:.1f}% din pozitie\n"
            f"  Take Profit (TP)           :  {tp_str}   "
            f"  ← recompensa {dist_tp/float(pret)*100:.1f}%\n"
            f"  Risk/Reward Ratio          :  {rr:.2f}x   {rr_eval}\n"
            f"  ATR (volatilitate zilnica) :  {atr_s}"
        )
        if rr >= 3:
            linii.append(
                f"RR de {rr:.1f}x este excepțional.\n"
                f"La fiecare 1 leu riscat, potențialul câștig este {rr:.1f} lei. "
                f"Chiar cu un win rate de 33%, acest tip de setup este profitabil pe termen lung. "
                f"Acesta este tipul de setup pe care fondurile de hedging îl caută."
            )
        elif rr >= 2:
            linii.append(
                f"RR de {rr:.1f}x este solid și profesionist.\n"
                f"Cu un win rate de 40% și un RR constant de 2x, sistemul "
                f"este matematic profitabil pe termen lung. Nu ceda tentației "
                f"de a muta TP-ul prea devreme."
            )
        else:
            linii.append(
                f"RR de {rr:.1f}x este sub pragul ideal de 2x.\n"
                f"Ia în considerare: lărgirea TP-ului (dacă rezistența permite) "
                f"sau reducerea SL-ului (cu atenție la zgomotul prețului)."
            )

    elif semnal == "SELL":
        rr_eval = ("EXCELENT ★★★" if rr >= 3
                   else "BUN ★★" if rr >= 2
                   else "ACCEPTABIL ★" if rr >= 1.5
                   else "SLAB — reconsideră")
        linii.append(
            f"OPORTUNITATE SHORT / VÂNZARE — {conf} confluențe tehnice aliniate.\n"
            f"Probabilitate estimată: {prob}%\n\n"
            f"PARAMETRI SHORT:\n"
            f"  Entry short                :  {p_str}\n"
            f"  Stop Loss (SL)             :  {sl_str}   "
            f"  ← risc {risc_pct:.1f}% din pozitie\n"
            f"  Take Profit (TP)           :  {tp_str}   "
            f"  ← target scadere {dist_tp/float(pret)*100:.1f}%\n"
            f"  Risk/Reward Ratio          :  {rr:.2f}x   {rr_eval}\n"
            f"  ATR (volatilitate zilnica) :  {atr_s}"
        )
        linii.append(
            "IMPORTANT — SHORT-UL este o strategie avansată:\n"
            "  • Short-ul implică a vinde ceva ce nu deții (împrumutat)\n"
            "  • Pierderea teoretică este NELIMITATĂ (prețul poate urca infinit)\n"
            "  • Necesită cont de marjă și experiență prealabilă\n"
            "  • Alternativa mai sigură: cumpără un ETF inversat sau opțiuni PUT"
        )
    else:
        linii.append(
            f"NU EXISTĂ OPORTUNITATE CLARĂ în acest moment.\n"
            f"Confluențe tehnice: {conf}/5 — sub pragul minim de 3.\n\n"
            f"CE SĂ URMĂREȘTI pentru a identifica o intrare viitoare:\n"
            f"  • RSI să coboare sub 40 (pentru BUY) sau să depășească 70 (SELL)\n"
            f"  • MACD să confirme cu un crossover în direcția dorită\n"
            f"  • Volumul să fie peste 1.2x medie la momentul mișcării\n"
            f"  • MA50 deasupra MA200 (Golden Cross) pentru semnale BUY\n"
            f"  • Un candle bullish/bearish puternic la un nivel de suport/rezistență"
        )

    if rvol > 1.5 and semnal != "WAIT":
        linii.append(
            f"BONUS CONFIRMARE: Volum excepțional ({rvol:.1f}x medie).\n"
            f"Volumul ridicat crește semnificativ credibilitatea semnalului. "
            f"Mișcările cu volum mare sunt mult mai puțin susceptibile de a fi false. "
            f"Acesta este un semnal de calitate superioară."
        )

    return "\n\n".join(linii)


def explica_pattern(d: dict) -> str:
    """Detectează și explică pattern-urile tehnice vizibile în date."""
    pret    = d["inchidere"]
    rsi     = d["rsi"]
    macd_c  = d["macd_cross"]
    macd_h  = d["macd_hist"]
    macross = d["macross"]
    stoch_k = d["stoch_k"]
    stoch_d = d["stoch_d"]
    bb_sup  = d["bb_sup"]
    bb_inf  = d["bb_inf"]
    bb_w    = d["bb_width"]
    ma50    = d["ma50"]
    ma200   = d["ma200"]
    var_zi  = d["var_zi"]
    rvol    = d["rvol"]
    n_zile  = d["n_zile"]

    p_f      = float(pret)
    ma50_s   = fmt_price(ma50,  2)
    ma200_s  = fmt_price(ma200, 2)
    bb_sup_s = fmt_price(bb_sup, 4)
    bb_inf_s = fmt_price(bb_inf, 4)
    bb_w_s   = fmt_price(bb_w,  4)

    patterns = []

    # Golden / Death Cross
    if macross == "Golden Cross" and ma50 is not None and ma200 is not None:
        diff_pct = (float(ma50) - float(ma200)) / float(ma200) * 100
        patterns.append(
            f"PATTERN ACTIV: GOLDEN CROSS\n"
            f"MA50 ({ma50_s}) > MA200 ({ma200_s}) cu {diff_pct:.2f}%.\n\n"
            f"Semnificație completă:\n"
            f"Este cel mai urmărit semnal bullish pe termen mediu-lung. "
            f"Instituțiile, fondurile de indici și robo-advisors CUMPĂRĂ automat la Golden Cross. "
            f"Istoric, S&P 500 a generat un randament mediu de +15% în primele "
            f"12 luni după un Golden Cross.\n\n"
            f"CUM ARATĂ PE GRAFIC:\n"
            f"Linia MA50 (albastru) taie MA200 (roșu) de jos în sus. "
            f"Caută un candle bullish (verde) care să confirme încrucișarea. "
            f"Volumul crescut la momentul crossover-ului este esențial pentru validare."
        )
    elif macross == "Death Cross" and ma50 is not None and ma200 is not None:
        diff_pct = (float(ma200) - float(ma50)) / float(ma200) * 100
        patterns.append(
            f"PATTERN ACTIV: DEATH CROSS\n"
            f"MA50 ({ma50_s}) < MA200 ({ma200_s}) cu {diff_pct:.2f}%.\n\n"
            f"Semnificație completă:\n"
            f"Semnal bearish major pe termen mediu-lung. Mulți algoritmi "
            f"de trailing stop se activează automat. Poate semnala începutul "
            f"unui bear market prelungit.\n\n"
            f"CAPCANA FRECVENTĂ:\n"
            f"Death Cross-ul apare DUPĂ ce prețul a coborât deja 15-25%. "
            f"Nu vinde în panică exact la semnal. Caută un retest (pull-back) "
            f"al MA50 (care a devenit rezistență) pentru un entry short mai bun."
        )

    # Bollinger Squeeze
    if bb_w is not None and bb_sup is not None and bb_inf is not None:
        if float(bb_w) / p_f < 0.04:
            patterns.append(
                f"PATTERN ACTIV: BOLLINGER BAND SQUEEZE\n"
                f"Lățime benzi: {bb_w_s} (sub 4% din preț) — compresie extremă.\n\n"
                f"Semnificație:\n"
                f"Perioadele de volatilitate scăzută sunt urmate INVARIABIL de explozii "
                f"de volatilitate. Piața 'se încolăcește' înainte de o mișcare mare.\n\n"
                f"STRATEGIE PRACTICĂ:\n"
                f"Plasează un ordin BUY STOP cu {fmt_price(float(bb_sup) * 1.005, 4)} "
                f"(deasupra benzii superioare {bb_sup_s}) și un SELL STOP cu "
                f"{fmt_price(float(bb_inf) * 0.995, 4)} (sub banda inferioară {bb_inf_s}). "
                f"Cel care se activează primul = direcția breakout-ului. "
                f"Anulează imediat celălalt ordin."
            )

    # MACD Crossover detaliat
    if "Bullish Cross" in macd_c:
        patterns.append(
            f"PATTERN ACTIV: MACD BULLISH CROSSOVER\n"
            f"Histogramă: {macd_h:+.6f} (trecere în teritoriu pozitiv).\n\n"
            f"CE ÎNSEAMNĂ PE GRAFIC:\n"
            f"Barele histogramei au trecut de la ROȘU (negativ) la VERDE (pozitiv). "
            f"Linia MACD (rapidă) a depășit linia Signal (lentă).\n\n"
            f"CEL MAI PUTERNIC SEMNAL:\n"
            f"Crossover bullish care apare EXACT LA LINIA ZERO este considerat "
            f"cel mai fiabil. Un crossover la -0.5 sau -1.0 (departe de zero) "
            f"este mai slab.\n\n"
            f"CAPCANA:\n"
            f"MACD crossover pe VOLUM MIC poate fi fals. "
            f"Confirmă întotdeauna cu volumul crescut! "
            f"Rvol curent: {d['rvol']:.1f}x {'✓ Confirmat' if d['rvol'] > 1.2 else '⚠ Slab'}."
        )
    elif "Bearish Cross" in macd_c:
        patterns.append(
            f"PATTERN ACTIV: MACD BEARISH CROSSOVER\n"
            f"Histogramă: {macd_h:+.6f} (trecere în teritoriu negativ).\n\n"
            f"CE ÎNSEAMNĂ PE GRAFIC:\n"
            f"Barele histogramei trec din VERDE în ROȘU. "
            f"Linia MACD a coborât sub linia Signal — momentumul pozitiv s-a epuizat.\n\n"
            f"IMPLICAȚII:\n"
            f"Semnal de prudență pentru deținătorii de poziții long. "
            f"Nu înseamnă obligatoriu că prețul va scădea mult, "
            f"dar probabilitatea de continuare bullish scade semnificativ."
        )

    # Selling / Buying Climax
    if rsi < 32 and var_zi < -2 and rvol > 1.3:
        patterns.append(
            f"PATTERN POSIBIL: SELLING CLIMAX\n"
            f"Scădere {var_zi:.1f}% + RSI {rsi:.1f} supravândut + Volum {d['rvol']:.1f}x.\n\n"
            f"CE ESTE:\n"
            f"Selling climax = momentul în care TOȚI cei care voiau să vândă "
            f"AU VÂNDUT. Volumul mare la un minim + RSI extrem = potențial punct de inversare.\n\n"
            f"CUM ARATĂ PE GRAFIC:\n"
            f"Caută un candle cu FITIL LUNG în jos și ÎNCHIDERE APROAPE DE MAXIM. "
            f"Acesta se numește 'Hammer' sau 'Pin Bar' și semnifică că bulls "
            f"au preluat controlul în cursul zilei respective.\n\n"
            f"ATENȚIE: Nu este garantat. Confirmă cu volumul scăzut a doua zi "
            f"(vânzătorii s-au epuizat) și un candle bullish ulterior."
        )
    elif rsi > 75 and var_zi > 2 and rvol > 1.3:
        patterns.append(
            f"PATTERN POSIBIL: BUYING CLIMAX\n"
            f"Creștere {var_zi:.1f}% + RSI {rsi:.1f} supraevaluat + Volum {d['rvol']:.1f}x.\n\n"
            f"CE ESTE:\n"
            f"Buying climax = momentul maxim al euforiei. Smart money (marii "
            f"jucători) VINDE în climax-urile de cumpărare, distribuind pozițiile "
            f"lor investitorilor retail entuziaști (Distribution Phase).\n\n"
            f"CUM ARATĂ PE GRAFIC:\n"
            f"Candle cu umbră lungă în sus și închidere aproape de MINIM = "
            f"'Shooting Star' sau 'Bearish Engulfing' — semnal de inversare bearish.\n\n"
            f"Regula: când toată lumea e entuziastă și 'știe' că prețul va urca — "
            f"fii cel mai precaut."
        )

    # Stochastic Crossover
    if stoch_k < 25 and stoch_k > stoch_d:
        patterns.append(
            f"PATTERN: STOCHASTIC OVERSOLD CROSSOVER\n"
            f"%K ({stoch_k:.1f}) a depășit %D ({stoch_d:.1f}) din zona sub 25.\n\n"
            f"Semnal clasic de cumpărare pe termen scurt (1-5 zile). "
            f"Cel mai eficient în piețele laterale (sideways range). "
            f"Mai puțin fiabil în downtrend-uri puternice."
        )
    elif stoch_k > 75 and stoch_k < stoch_d:
        patterns.append(
            f"PATTERN: STOCHASTIC OVERBOUGHT CROSSOVER\n"
            f"%K ({stoch_k:.1f}) a coborât sub %D ({stoch_d:.1f}) din zona peste 75.\n\n"
            f"Semnal de luare a profitului / vânzare pe termen scurt."
        )

    # Test MA50
    if ma50 is not None:
        dist_ma50 = (p_f - float(ma50)) / float(ma50) * 100
        if abs(dist_ma50) < 0.8:
            patterns.append(
                f"PATTERN: TEST MA50\n"
                f"Prețul ({fmt_price(pret, 4)}) se află la {dist_ma50:+.2f}% față de MA50 ({ma50_s}).\n\n"
                f"MA50 este CEL MAI URMĂRIT nivel dinamic de suport/rezistență. "
                f"Un test al MA50 poate fi:\n"
                f"  BOUNCE (revenire): prețul respinge MA50 și continuă trendul → confirmă trendul\n"
                f"  BREAKDOWN: prețul sparge MA50 pe volum → semnal bearish\n"
                f"  FAKE: prețul taie MA50 dar revine rapid (whipsaw) → semn de piață slabă\n\n"
                f"Fii FOARTE ATENT la candle-ul de închidere față de MA50 azi!"
            )

    if not patterns:
        patterns.append(
            f"Nu există un pattern tehnic dominant clar astăzi.\n"
            f"Activul se află în faza de consolidare sau așteptare. "
            f"Monitorizează zilnic pentru apariția unui setup clar. "
            f"Răbdarea este o strategie validă."
        )

    return "\n\n".join(patterns)


def lectia_zilei(d: dict) -> str:
    """Generează 1-2 lecții educative adaptate situației activului."""
    semnal  = d["semnal"]
    rsi     = d["rsi"]
    macross = d["macross"]
    rvol    = d["rvol"]
    conf    = d["confluente"]
    pret    = d["inchidere"]
    sl      = d["sl"]
    tp      = d["tp"]

    dist_sl = abs(float(pret) - float(sl))
    dist_tp = abs(float(tp)   - float(pret))
    rr      = round(dist_tp / dist_sl, 2) if dist_sl > 0 else 0

    lectii = []

    # Lecție despre răbdare
    if semnal == "WAIT" and conf < 2:
        lectii.append(
            "LECȚIE — PUTEREA 'NU TRANZACȚIONEZ' :\n\n"
            "Una dintre cele mai greu de aplicat lecții în trading este să nu faci nimic. "
            "Mulți traderi simt nevoia permanentă de a fi 'în piață'. "
            "Realitatea: marii profesioniști stau în cash sau neutral 60-70% din timp "
            "și atacă DOAR când setup-ul este perfect.\n\n"
            "Jack Schwager, în cartea 'Market Wizards', a intervievat cei mai buni "
            "traderi din lume. Toți aveau în comun un lucru: SELECTIVITATE EXTREMĂ. "
            "Nu numărul de tranzacții te îmbogățește, ci CALITATEA lor.\n\n"
            "Aplicare practică: stabilește-ți un checklist cu 5 condiții. "
            "Intri DOAR dacă toate 5 sunt bifate. Astăzi acest activ nu le bifează."
        )

    # Lecție despre RR
    if 0 < rr < 1.5:
        lectii.append(
            f"LECȚIE — MATEMTATICA RISK/REWARD RATIO:\n\n"
            f"RR-ul curent pentru acest setup este {rr:.1f}x — sub pragul profesionist de 2x.\n\n"
            f"De ce contează matematic:\n"
            f"  Cu RR = 1.0x și win rate 50% → breakeven (zero profit după costuri)\n"
            f"  Cu RR = 1.5x și win rate 40% → PROFITABIL pe termen lung\n"
            f"  Cu RR = 2.0x și win rate 34% → PROFITABIL pe termen lung\n"
            f"  Cu RR = 3.0x și win rate 26% → PROFITABIL pe termen lung\n\n"
            f"Concluzie: cu un RR bun, poți câștiga mai mult decât pierzi "
            f"chiar dacă ai mai multe tranzacții pierzătoare decât câștigătoare. "
            f"Ajustează SL și TP pentru a obține minim 2x înainte de a intra."
        )

    # Lecție despre volum
    if rvol < 0.7:
        lectii.append(
            "LECȚIE — 'FOLLOW THE VOLUME':\n\n"
            "Prețul poate fi manipulat, volumul nu minte.\n\n"
            "O mișcare de preț pe volum sub 70% din medie este un semnal roșu. "
            "Poate fi manipulare de piață, poate fi o mișcare de 'thin market' "
            "(piață subțire în afara orelor principale), sau pur și simplu lipsă "
            "de interes real.\n\n"
            "Regula de aur nr. 1 din trading: "
            "'Nu cumpăra breakout-uri pe volum mic.'\n\n"
            "Regula de aur nr. 2: "
            "'Dacă prețul urcă pe volum scăzut, vânzătorii NU sunt convinși că e scump — "
            "cumpărătorii sunt cei care ridică prețul singuri.'"
        )

    # Lecție despre RSI extrem
    if rsi < 30:
        lectii.append(
            "LECȚIE — 'FALLING KNIFE': CUM NU TE RĂNEȘTI:\n\n"
            "RSI sub 30 atrage mulți traderi cu gândul 'e ieftin, cumpăr acum'. "
            "Aceasta este CAPCANA clasică a prețului în scădere.\n\n"
            "Regula: RSI sub 30 este o condiție NECESARĂ pentru un BUY, "
            "nu una SUFICIENTĂ. Ai nevoie de CONFIRMARE:\n"
            "  1. Un candle bullish solid (Hammer, Pin Bar, Engulfing verde)\n"
            "  2. MACD care formează un crossover bullish\n"
            "  3. Un nivel clar de suport ținut (prețul respinge un nivel cheie)\n"
            "  4. Volum scăzut pe continuarea scăderii (vânzătorii se epuizează)\n\n"
            "Fără aceste confirmări, 'ieftin poate deveni și mai ieftin'."
        )
    elif rsi > 70:
        lectii.append(
            "LECȚIE — NU SHORT-UI ÎNTR-UN BULL MARKET:\n\n"
            "RSI peste 70 nu înseamnă automat 'vinde'. "
            "NVIDIA în 2023-2024 a avut RSI peste 70 timp de 8 luni consecutive "
            "și a crescut cu 400% în acea perioadă.\n\n"
            "Activele pot rămâne 'supraevaluate' luni întregi în bull market-uri puternice. "
            "Regula: short-ul pe RSI ridicat funcționează în piețe laterale sau bearish. "
            "În uptrend puternic (Golden Cross + volum crescut), "
            "short-ul pe RSI ridicat = rețetă pentru pierderi mari.\n\n"
            "Alternativa corectă: dacă ești long, folosești trailing stop "
            "pentru a lăsa profitul să curgă și a te proteja simultan."
        )

    # Lecție despre Golden/Death Cross
    if macross == "Golden Cross":
        lectii.append(
            "LECȚIE — GOLDEN CROSS: ÎNTÂRZIERE VS. ANTICIPARE:\n\n"
            "Golden Cross este real și puternic, dar are un defect major: "
            "este un indicator LAGGING (cu întârziere). "
            "Apare DUPĂ ce prețul a urcat deja 10-30%.\n\n"
            "Strategia amatorilor: cumpără exact la Golden Cross (prea târziu).\n"
            "Strategia profesioniștilor: anticipează Golden Cross-ul.\n\n"
            "Cum anticipezi:\n"
            "  1. MA200 încetinește scăderea (se aplatizează)\n"
            "  2. MA50 oprește scăderea și inversează\n"
            "  3. RSI revine din sub 30 → zona 40-50\n"
            "  4. MACD formează un bullish crossover la zero line\n"
            "  5. Volumul crește pe zilele UP\n\n"
            "Dacă bifezi 4 din 5 și Golden Cross nu s-a format încă — "
            "ACELA este momentul optim de cumpărare."
        )
    elif macross == "Death Cross":
        lectii.append(
            "LECȚIE — DEATH CROSS: GREȘEALA CLASICĂ:\n\n"
            "Death Cross apare adesea la MIJLOCUL unei corecții, nu la început. "
            "Dacă vinzi în panică la Death Cross, vinzi la jumătatea scăderii.\n\n"
            "Strategia corectă:\n"
            "  1. Reducerea expunerii la PRIMELE semnale (MA50 se aplatizează, "
            "     MACD bearish cross, volum crescut pe zile negative)\n"
            "  2. La Death Cross complet format: evaluezi dacă teza investițională "
            "     mai este validă pe termen lung\n"
            "  3. Dacă fundamentele sunt bune: Death Cross = oportunitate de "
            "     acumulare în etape (Dollar Cost Averaging)\n"
            "  4. Dacă fundamentele s-au deteriorat: Death Cross = confirmare "
            "     că ai dreptate să ieși\n\n"
            "Nu există un răspuns universal — contextul fundamental contează enorm."
        )

    # Lecție psihologie (mereu inclusă)
    lectii.append(
        "LECȚIE PSIHOLOGIE — EMOȚIILE ȘI PIERDEREA BANILOR:\n\n"
        "Studiile arată că durerea unei pierderi de 100 lei este psihologic de "
        "2x mai intensă decât bucuria unui câștig de 100 lei (Loss Aversion, "
        "Daniel Kahneman, Nobel 2002).\n\n"
        "Cum afectează trading-ul:\n"
        "  1. Ții pozițiile perdante prea mult ('o să revină')\n"
        "  2. Vinzi pozițiile câștigătoare prea repede (frică să nu pierzi profitul)\n"
        "  3. Dublezi pe pierdere pentru a 'recupera' (averaging down neselectiv)\n"
        "  4. Faci revenge trading după o pierdere (emoție, nu analiză)\n\n"
        "Soluția: PLAN SCRIS înainte de intrare. Entry, SL, TP, sizing — "
        "scrise pe hârtie sau în jurnal ÎNAINTE de a deschide poziția. "
        "Odată în poziție, lăsă planul să lucreze, nu emoțiile."
    )

    return "\n\n" + ("─" * 45) + "\n\n".join(lectii[:2])


# ══════════════════════════════════════════════════════════════════════════════
# GHID COMPLET CITIRE GRAFICE (secțiunea permanentă de jos)
# ══════════════════════════════════════════════════════════════════════════════

GHID_GRAFICE = [
    {
        "titlu": "CANDLESTICK (LUMÂNĂRI JAPONEZE) — Fundația analizei tehnice",
        "sectiuni": [
            ("Ce este o lumânare",
             "Fiecare lumânare reprezintă o perioadă de timp (zi, oră, 5 minute etc.) "
             "și conține 4 informații esențiale:\n"
             "  • OPEN (Deschidere): prețul la începutul perioadei\n"
             "  • HIGH (Maxim): cel mai mare preț atins\n"
             "  • LOW (Minim): cel mai mic preț atins\n"
             "  • CLOSE (Închidere): prețul la sfârșitul perioadei"),
            ("Corpul lumânării",
             "VERDE (bullish): Închidere > Deschidere → prețul a URCAT în acea perioadă.\n"
             "ROȘU (bearish): Închidere < Deschidere → prețul a COBORÂT în acea perioadă.\n\n"
             "Mărimea corpului = forța mișcării. Corp mare verde = bulls puternici. "
             "Corp mic = indeczie, echilibru între cumpărători și vânzători."),
            ("Umbrele (fitilele / shadows)",
             "Liniile subțiri deasupra și dedesubt corpului:\n"
             "  • Umbră LUNGĂ sus: prețul a urcat la maxim dar a revenit jos → Bears au respins mișcarea\n"
             "  • Umbră LUNGĂ jos: prețul a coborât la minim dar a recuperat → Bulls au apărat nivelul\n"
             "  • Umbră scurtă sau absentă: direcție clară, presiune uniformă"),
            ("Cele mai importante pattern-uri de lumânări",
             "DOJI: corp inexistent sau minimal, umbre similare → indeczie pură, potențial punct de inversare\n\n"
             "HAMMER (Ciocan): corp mic la vârf, umbră lungă în jos (min. 2x corpul) → reversal BULLISH\n"
             "  → Semnificație: Bears au coborât prețul mult dar Bulls au preluat controlul complet\n\n"
             "SHOOTING STAR (Stea căzătoare): corp mic la bază, umbră lungă în sus → reversal BEARISH\n"
             "  → Semnificație: Bulls au împins prețul sus dar Bears au preluat controlul\n\n"
             "BULLISH ENGULFING: candle verde mare 'înghite' complet candle roșu precedent → BUY puternic\n\n"
             "BEARISH ENGULFING: candle roșu mare 'înghite' complet candle verde precedent → SELL puternic\n\n"
             "MARUBOZU: corp mare fără umbre → presiune unidirecțională extremă (trend foarte puternic)"),
            ("Cum aplici în practică",
             "1. Caută pattern-urile LA NIVELURI CHEIE (suport, rezistență, MA50, MA200)\n"
             "2. Confirmă cu VOLUMUL: un Hammer pe volum mare = semnal puternic\n"
             "3. Așteptă CONFIRMAREA: candle-ul următor să confirme direcția\n"
             "4. Un singur pattern nu e suficient — combină cu RSI, MACD și volum\n"
             "5. Pattern-urile pe timeframe-uri mai mari (zilnic, săptămânal) sunt mai fiabile"),
        ]
    },
    {
        "titlu": "RSI — Relative Strength Index (0-100)",
        "sectiuni": [
            ("Ce măsoară",
             "RSI măsoară VITEZA și AMPLITUDINEA mișcărilor de preț din ultimele 14 perioade. "
             "Răspunde la întrebarea: 'A urcat sau coborât prețul prea repede față de norma sa?'\n\n"
             "Formula simplificată: RSI = 100 - (100 / (1 + medie_câștiguri/medie_pierderi_14z))"),
            ("Zonele RSI și semnificația lor",
             "0 - 30   → SUPRAVÂNDUT (Oversold): prețul a scăzut prea mult, prea repede\n"
             "           Probabilitate crescută de revenire, dar nu garantată\n\n"
             "30 - 45  → Zonă slabă: Bears în control, dar presiunea scade\n\n"
             "45 - 55  → Neutru: echilibru, piața caută direcție\n\n"
             "55 - 70  → Zonă puternică: Bulls în control, trend sănătos\n"
             "           CEA MAI PROFITABILĂ ZONĂ pentru trend-following\n\n"
             "70 - 100 → SUPRAEVALUAT (Overbought): prețul a crescut prea mult, prea repede\n"
             "           Probabilitate crescută de corecție sau consolidare"),
            ("RSI Divergence — semnalul cel mai puternic",
             "BULLISH DIVERGENCE (semnificație majoră):\n"
             "  Prețul face un nou MINIM (lower low)\n"
             "  dar RSI face un minim MAI MARE (higher low)\n"
             "  → Presiunea de vânzare SLĂBEȘTE deși prețul coboară\n"
             "  → Semnal de inversare bullish iminent (unul dintre cele mai fiabile)\n\n"
             "BEARISH DIVERGENCE:\n"
             "  Prețul face un nou MAXIM (higher high)\n"
             "  dar RSI face un maxim MAI MIC (lower high)\n"
             "  → Momentumul bullish SLĂBEȘTE deși prețul urcă\n"
             "  → Semnal de inversare bearish sau corecție majoră"),
            ("Greșeli frecvente cu RSI",
             "✗ A vinde DOAR pentru că RSI > 70 (poate rămâne acolo LUNI în bull market puternic)\n"
             "✗ A cumpăra DOAR pentru că RSI < 30 (poate scădea la 15-10 în bear market)\n"
             "✗ Folosirea RSI izolat fără confirmare de la alt indicator\n"
             "✗ Ignorarea timeframe-ului: RSI pe 5 minute e zgomot, pe zilnic e semnal"),
        ]
    },
    {
        "titlu": "MACD — Moving Average Convergence/Divergence",
        "sectiuni": [
            ("Ce este și cum funcționează",
             "MACD arată relația dintre două medii mobile exponențiale.\n\n"
             "Calculul:\n"
             "  Linia MACD  = EMA(12 zile) - EMA(26 zile)  [linia rapidă]\n"
             "  Linia Signal = EMA(9 zile) a liniei MACD     [linia lentă]\n"
             "  Histogramă  = Linia MACD - Linia Signal      [barele verzi/roșii]\n\n"
             "Răspunde la: 'Care este direcția și forța momentumului pe termen scurt vs mediu?'"),
            ("Semnalele principale MACD",
             "BULLISH CROSSOVER: Linia MACD trece DEASUPRA Signal → semnal BUY\n"
             "  → Histograma trece de la roșu la verde\n\n"
             "BEARISH CROSSOVER: Linia MACD trece DEDESUBT Signal → semnal SELL\n"
             "  → Histograma trece de la verde la roșu\n\n"
             "ZERO LINE CROSS (cel mai puternic):\n"
             "  MACD trece de la negativ la pozitiv = confirmare trend bullish major\n"
             "  MACD trece de la pozitiv la negativ = confirmare trend bearish major\n\n"
             "HISTOGRAMA CREȘTE = momentum se accelerează\n"
             "HISTOGRAMA SCADE = momentum se slăbește (atenție: prețul poate urca încă)"),
            ("Cel mai puternic semnal MACD combinat",
             "MACD Bullish Crossover LA LINIA ZERO\n"
             "+ Confirmat de Golden Cross MA50/MA200\n"
             "+ Volum crescut (RVOL > 1.5x)\n"
             "+ RSI în zona 40-55 (nu supraevaluat)\n\n"
             "= Unul dintre cele mai fiabile setup-uri din analiza tehnică clasică.\n"
             "Apare de câteva ori pe an pe activele majore — merită așteptat."),
            ("Limitări importante",
             "LAGGING indicator: confirmă tendințele, nu le prezice.\n"
             "Cel mai eficient pe timeframe-uri ZILNICE și SĂPTĂMÂNALE.\n"
             "Pe timeframe-uri mici (sub 1 oră) generează foarte mult zgomot.\n"
             "Funcționează cel mai bine în trenduri clare, mai puțin în piețe laterale."),
        ]
    },
    {
        "titlu": "MEDII MOBILE — MA20, MA50, MA200",
        "sectiuni": [
            ("Ce sunt și ce reprezintă",
             "Mediilele mobile 'netezesc' zgomotul prețului și arată DIRECȚIA TENDINTEI.\n\n"
             "  MA20  = media ultimelor 20 zile  (~1 lună)\n"
             "          Urmărită de: day traderi, scalpers\n\n"
             "  MA50  = media ultimelor 50 zile  (~2.5 luni)\n"
             "          Urmărită de: swing traderi, fonduri\n\n"
             "  MA200 = media ultimelor 200 zile (~10 luni)\n"
             "          Urmărită de: investitori pe termen lung, fonduri de pensii"),
            ("Regulile esențiale",
             "1. Preț DEASUPRA MA200 = Bull Market (piață primară bullish)\n"
             "2. Preț DEDESUBT MA200 = Bear Market (piață primară bearish)\n"
             "3. GOLDEN CROSS (MA50 > MA200) = semnal bullish major pe termen lung\n"
             "4. DEATH CROSS  (MA50 < MA200) = semnal bearish major pe termen lung\n"
             "5. MA50 = SUPORT DINAMIC în uptrend și REZISTENȚĂ DINAMICĂ în downtrend\n"
             "6. Cu cât mai mulți jucători urmăresc un nivel, cu atât e mai puternic\n"
             "   (Profecie care se autoîmplinește — Self-fulfilling prophecy)"),
            ("Cum utilizezi pe grafic",
             "UPTREND: Prețul tinde să revină la MA50 (pullback = oportunitate de cumpărare)\n"
             "  → Entry: la atingerea MA50 cu candle bullish de confirmare\n"
             "  → Stop Loss: sub MA50 (sau sub ultimul minim semnificativ)\n\n"
             "DOWNTREND: Prețul tinde să respingă MA50 (bear rally = oportunitate de vânzare)\n"
             "  → Short Entry: la retestul MA50 cu candle bearish de confirmare\n"
             "  → Stop Loss: deasupra MA50\n\n"
             "TRANZIȚIE (MA50 lângă MA200): Evită tranzacțiile în această zonă — risc maxim."),
            ("EMA vs SMA — care e mai bună",
             "SMA (Simple Moving Average): medie simplă — toate zilele au greutate egală\n"
             "EMA (Exponential Moving Average): zilele recente au greutate mai mare\n\n"
             "EMA reacționează MAI RAPID la schimbări de preț → mai puține semnale întârziate\n"
             "SMA este mai STABILĂ și generează mai puțin zgomot\n\n"
             "Recomandare practică: folosește EMA20, EMA50 pentru semnale, SMA200 pentru tendința majoră."),
        ]
    },
    {
        "titlu": "BOLLINGER BANDS — Benzile lui Bollinger",
        "sectiuni": [
            ("Structura benzilor",
             "Trei linii calculate pe baza MA20 și deviației standard:\n\n"
             "  Banda Superioară = MA20 + (2 × deviație standard)\n"
             "  Banda Mijlocie   = MA20 (media mobilă simplă 20 zile)\n"
             "  Banda Inferioară = MA20 - (2 × deviație standard)\n\n"
             "Statistic: 95% din timp prețul se află ÎNTRE cele două benzi externe.\n"
             "Benzile se LĂRGESC când volatilitatea crește și se ÎNGUSTEAZĂ când scade."),
            ("Strategia Bollinger Bounce",
             "Funcționează CEL MAI BINE în piețe LATERALE (sideways/range):\n\n"
             "  BUY: când prețul atinge banda INFERIOARĂ → țintă banda mijlocie sau superioară\n"
             "  SELL: când prețul atinge banda SUPERIOARĂ → țintă banda mijlocie sau inferioară\n\n"
             "ATENȚIE CRITICĂ: NU aplica în trenduri puternice!\n"
             "În uptrend, prețul poate 'plimba' banda superioară săptămâni întregi.\n"
             "Vânzarea la banda superioară în uptrend = cea mai costisitoare greșeală."),
            ("Strategia Bollinger Squeeze (Breakout)",
             "Detectează momentul ÎNAINTEA unei mișcări mari:\n\n"
             "SEMNALUL: Benzile devin foarte înguste (lățime sub 4% din preț)\n"
             "CE URMEAZĂ: O explozie de volatilitate — nu știm direcția, dar mișcarea va fi mare.\n\n"
             "EXECUȚIE:\n"
             "  1. Identifică squeeze-ul (vizual: benzile aproape paralele)\n"
             "  2. Plasează BUY STOP puțin deasupra benzii superioare\n"
             "  3. Plasează SELL STOP puțin sub banda inferioară\n"
             "  4. Cel activat = direcția breakout-ului → anulează imediat celălalt\n\n"
             "Avantaj: intri DUPĂ confirmare, nu ghicești direcția."),
            ("Benzile ca filtru de trend",
             "PREȚ DEASUPRA benzii superioare constant = trend bullish puternic (pozitiv)\n"
             "PREȚ DEDESUBT benzii inferioare constant = trend bearish puternic (negativ)\n"
             "PREȚ OSCILÂND ÎNTRE BENZI = piață laterală, folosește bounce strategy\n\n"
             "Lățimea benzilor ca indicator de piață:\n"
             "Benzi LARGI = volatilitate ridicată, mișcări mari, risc crescut\n"
             "Benzi ÎNGUSTE = volatilitate scăzută, piața se 'odihnește' → explozie iminentă"),
        ]
    },
    {
        "titlu": "VOLUM & RVOL — Combustibilul mișcărilor de piață",
        "sectiuni": [
            ("De ce volumul este fundamental",
             "PREȚUL spune unde s-a dus piața.\n"
             "VOLUMUL spune CÂT DE CONVINSĂ era piața de acea mișcare.\n\n"
             "O mișcare de preț fără volum = neconvingătoare, probabilitate mare de inversare.\n"
             "O mișcare de preț CU volum = participare reală, probabilitate mare de continuare.\n\n"
             "Axioma nr. 1 a analizei tehnice: 'Volumul precedă prețul.'"),
            ("RVOL — Relative Volume",
             "RVOL = Volum curent / Media volumului pe 20 de zile\n\n"
             "  RVOL > 2.0x → Volum EXCEPȚIONAL: eveniment major, instituții active\n"
             "  RVOL 1.5-2.0x → Volum RIDICAT: mișcare de calitate superioară\n"
             "  RVOL 1.0-1.5x → Volum NORMAL-RIDICAT: mișcare validă\n"
             "  RVOL 0.7-1.0x → Volum NORMAL: nimic special\n"
             "  RVOL < 0.7x  → Volum SCĂZUT: mișcare suspectă, probabilitate fake-out"),
            ("Regulile de aur ale volumului",
             "1. UPTREND SĂNĂTOS:\n"
             "   Zilele UP → volum CRESCUT (bulls convinși)\n"
             "   Zilele DOWN → volum SCĂZUT (pullback normal, bears nu convinși)\n\n"
             "2. DOWNTREND SĂNĂTOS:\n"
             "   Zilele DOWN → volum CRESCUT (bears convinși)\n"
             "   Zilele UP → volum SCĂZUT (bear rally, nu inversare reală)\n\n"
             "3. SEMN DE SLĂBIRE A TRENDULUI:\n"
             "   Prețul face noi maxime dar VOLUMUL SCADE → divergență negativă\n"
             "   = Distribuție (smart money vinde în timp ce retail cumpără)\n\n"
             "4. BREAKOUT PE VOLUM MIC = CAPCANĂ (fake breakout)\n"
             "   BREAKOUT PE VOLUM MARE = REAL, urmărește continuation"),
            ("Cum recunoști distribuția și acumularea",
             "ACUMULARE (Smart Money cumpără discret):\n"
             "  Prețul lateral sau ușor descendent + Zile de creștere cu volum mare\n"
             "  + Zile de scădere cu volum mic → Bulls acumulează treptat\n\n"
             "DISTRIBUȚIE (Smart Money vinde discret):\n"
             "  Prețul lateral sau ușor ascendent + Zile de scădere cu volum mare\n"
             "  + Zile de creștere cu volum mic → Bears distribuie pozițiile\n\n"
             "Metodologia Wyckoff (1930, valabilă și azi) descrie aceste faze în detaliu."),
        ]
    },
    {
        "titlu": "SUPORT & REZISTENȚĂ — Nivelurile care controlează prețul",
        "sectiuni": [
            ("Definiție și logică",
             "SUPORT: Nivel de preț unde cumpărătorii intervin suficient de puternic "
             "pentru a OPRI și INVERSA scăderea.\n\n"
             "REZISTENȚĂ: Nivel de preț unde vânzătorii intervin suficient de puternic "
             "pentru a OPRI și INVERSA creșterea.\n\n"
             "De ce există aceste niveluri?\n"
             "  → Memorie colectivă a pieței: traders care au cumpărat la un nivel "
             "     vor cumpăra din nou dacă prețul revine (confirmare psihologică)\n"
             "  → Algoritmi programați să cumpere/vândă la niveluri istorice cheie\n"
             "  → Ordine limită (limit orders) plasate în masă la niveluri rotunde"),
            ("Cum identifici nivelurile cheie",
             "1. MAXIME și MINIME ANTERIOARE semnificative (pivots)\n"
             "   → Cu cât mai recent și mai pronunțat, cu atât mai puternic\n\n"
             "2. ZONE DE CONSOLIDARE (range): niveluri unde prețul a petrecut timp\n"
             "   → Volumul tranzacționat în acea zonă = forța nivelului\n\n"
             "3. NIVELURI PSIHOLOGICE ROTUNDE: 1000, 50.000, 100, 1.00\n"
             "   → Oamenii plasează ordine la numere rotunde → self-fulfilling\n\n"
             "4. MEDII MOBILE (dinamice): MA50, MA200\n"
             "5. RETRACEMENT FIBONACCI: 38.2%, 50%, 61.8% din mișcarea anterioară\n\n"
             "Regula: un nivel este PUTERNIC dacă a fost testat DE MULTIPLE ORI (min. 2-3x)."),
            ("Role Reversal — Regula de aur",
             "SUPORTUL SPART devine REZISTENȚĂ.\n"
             "REZISTENȚA SPARTĂ devine SUPORT.\n\n"
             "Aceasta este una dintre cele mai puternice concepte din analiza tehnică.\n\n"
             "Exemplu practic:\n"
             "Prețul era susținut la 100 (suport). Sparge 100 → coboară la 85.\n"
             "Prețul revine la 100 (bear rally) → acum 100 este REZISTENȚĂ.\n"
             "Traderii care au cumpărat la 100 și sunt în pierdere VOR VINDE "
             "când prețul revine la 100 (breakeven) → generează presiune de vânzare la 100.\n\n"
             "Role reversal = punct de intrare cu RISC REDUS și POTENȚIAL RIDICAT."),
            ("Stop Loss optim față de suport/rezistență",
             "NU plasa stop loss EXACT la nivelul de suport/rezistență.\n"
             "Piața testează adesea nivelurile cu câteva procente înainte de inversare.\n\n"
             "REGULA: Stop loss = nivelul de suport MINUS 1-3% (sau 1-2x ATR).\n\n"
             "Exemplu: Suport la 100 → SL la 97-98 (lasă 'breathing room').\n\n"
             "Un SL prea strâns = ești scos din tranzacție de 'zgomotul' normal al prețului, "
             "exact înainte ca prețul să revină în direcția ta. Frustrant și costisitor."),
        ]
    },
    {
        "titlu": "MANAGEMENTUL RISCULUI — Regulile care separă profesioniștii de amatori",
        "sectiuni": [
            ("Regula 1-2% per tranzacție",
             "NU risca mai mult de 1-2% din capitalul total pe o singură tranzacție.\n\n"
             "Exemplu cu capital 10.000 EUR:\n"
             "  Risc per tranzacție (1%): 100 EUR\n"
             "  Risc per tranzacție (2%): 200 EUR\n\n"
             "Calculul MĂRIMII POZIȚIEI (Position Sizing):\n"
             "  Mărime = Risc acceptat ÷ (Entry Price - Stop Loss Price)\n\n"
             "Exemplu: Capital 10.000 EUR, Entry la 50 EUR, SL la 48 EUR, risc 1% (100 EUR)\n"
             "  Mărime = 100 EUR ÷ (50 - 48) = 50 acțiuni\n\n"
             "De ce contează: 10 pierderi consecutive la 2% = -18% capital (nu -20%)\n"
             "Cu poziții de 10% per tranzacție: 10 pierderi = -65% (greu de recuperat)"),
            ("Risk/Reward Ratio (RR) — Matematica supraviețuirii",
             "RR = (Take Profit - Entry) ÷ (Entry - Stop Loss)\n\n"
             "Tabel matematic — win rate minim necesar pentru profitabilitate:\n"
             "  RR = 1.0x → Win rate necesar > 50% pentru profit\n"
             "  RR = 1.5x → Win rate necesar > 40% pentru profit\n"
             "  RR = 2.0x → Win rate necesar > 34% pentru profit\n"
             "  RR = 3.0x → Win rate necesar > 26% pentru profit\n\n"
             "Concluzie: un RR bun îți permite să PIERZI MAI MULTE tranzacții decât câștigi "
             "și să rămâi profitabil. Acesta este secretul longevității în trading.\n\n"
             "REGULA: Nu intra niciodată cu RR sub 1.5x. Ideal 2x sau mai mult."),
            ("Diversificarea și corelația",
             "Nu aloca mai mult de 10-15% din portofoliu pe un singur activ.\n"
             "Nu aloca mai mult de 25-30% pe o singură clasă de active.\n\n"
             "ATENȚIE LA CORELAȚIE:\n"
             "Bitcoin și Ethereum sunt corelate 85-90% → nu sunt cu adevărat diversificate.\n"
             "Acțiunile tech (NVIDIA, AMD, Intel) sunt corelate 70-80%.\n"
             "Gold și S&P500 sunt NEGATIV corelate în criză → gold = hedge real.\n\n"
             "Diversificare adevărată înseamnă active care NU se mișcă în același timp:\n"
             "  Exemple: Acțiuni + Obligațiuni + Gold + Valute + Materii prime"),
            ("Jurnalul de tranzacții — arma secretă",
             "Cel mai subevaluat instrument în trading. Notează OBLIGATORIU:\n\n"
             "  ÎNAINTE de tranzacție:\n"
             "    → De ce intru? (setup, confluențe)\n"
             "    → Care este teza? (ce trebuie să se întâmple pentru ca trade-ul să funcționeze)\n"
             "    → Care este planul dacă mă înșel? (SL clar)\n\n"
             "  DUPĂ tranzacție:\n"
             "    → Ce s-a întâmplat? (a funcționat teza?)\n"
             "    → Ce am simțit? (frică, FOMO, lăcomie, răbdare)\n"
             "    → Am respectat planul?\n"
             "    → Ce lecție extrag?\n\n"
             "Traderii profesioniști analizează PIERDERILE mai mult decât câștigurile. "
             "Pierderile conțin informația cea mai valoroasă despre punctele slabe."),
            ("Cele 10 reguli de aur ale traderului profesionist",
             "1. Planifică tranzacția, tranzacționează planul.\n"
             "2. Niciodată RR sub 1.5x — ideal 2x sau mai mult.\n"
             "3. Niciodată mai mult de 2% risc din capital pe o tranzacție.\n"
             "4. Confirmă breakout-urile cu VOLUMUL — fake breakouts costă scump.\n"
             "5. Nu face averaging down neselectiv — 'ieftin' poate deveni 'mai ieftin'.\n"
             "6. Lasă câștigătoarele să curgă, taie pierdătoarele rapid (contrar instinctului).\n"
             "7. Nu tranzacționa din plictiseală sau din dorința de 'a fi activ'.\n"
             "8. Nu revenge trade după o pierdere — ia o pauză.\n"
             "9. Jurnalul de tranzacții este obligatoriu — dacă nu îl măsori, nu îl îmbunătățești.\n"
             "10. Piața există și mâine — capitalul protejat = oportunități viitoare."),
        ]
    },
]

# ══════════════════════════════════════════════════════════════════════════════
# SCRIERE EXCEL
# ══════════════════════════════════════════════════════════════════════════════

def set_h(ws, row, height):
    ws.row_dimensions[row].height = height

def write_row(ws, row, col1, col2, val, f_fill=None, f_font=None,
              h_align="left", height=None):
    c = ws.cell(row, col1, val)
    if f_fill: c.fill      = f_fill
    if f_font: c.font      = f_font
    c.alignment = aln(h_align, "center")
    if col2 > col1:
        ws.merge_cells(start_row=row, start_column=col1,
                       end_row=row, end_column=col2)
    if height:
        set_h(ws, row, height)
    return row + 1


def write_bloc_activ(ws, d: dict, r: int) -> int:
    """Scrie blocul complet pentru un activ. Returnează noul rând."""
    name   = d["name"]
    pret   = d["inchidere"]
    var    = d["var_zi"]
    semnal = d["semnal"]
    rsi    = d["rsi"]
    conf   = d["confluente"]

    # Culoare header activ
    if semnal == "BUY":
        f_head = F_ACTIV_BUY
    elif semnal == "SELL":
        f_head = F_ACTIV_SEL
    else:
        f_head = F_ACTIV_WAI

    # ── Header activ ─────────────────────────────────────────────────────
    set_h(ws, r, 24)
    header_text = (
        f"  {name}   |   Preț: {fmt_price(pret, 4)}   |   "
        f"Zi: {fmt_pct(var)}   |   RSI: {rsi:.1f}   |   "
        f"Confluențe: {conf}/5   |   SEMNAL: {semnal}  "
    )
    c = ws.cell(r, 1, header_text)
    c.fill = f_head
    c.font = FN_ACTIV
    c.alignment = aln("left", "center")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    r += 1

    # ── Cele 4 secțiuni ──────────────────────────────────────────────────
    sectiuni = [
        ("  DE CE S-A MIȘCAT ASTĂZI",   explica_miscare(d),    F_BLU,  fnt(color="1F4E79", size=10)),
        ("  OPORTUNITATE DE TRADING",    explica_oportunitate(d), F_GRN if semnal=="BUY" else F_RED if semnal=="SELL" else F_YLW,
                                         FN_GRN if semnal=="BUY" else FN_RED if semnal=="SELL" else FN_YLW),
        ("  PATTERN GRAFIC DETECTAT",    explica_pattern(d),    F_GRY,  fnt(color="303030", size=10)),
        ("  LECȚIA ZILEI",               lectia_zilei(d),       F_ORG,  FN_ORG),
    ]

    for titlu, continut, f_hdr_s, fn_hdr_s in sectiuni:
        # Sub-header
        set_h(ws, r, 18)
        h = ws.cell(r, 1, titlu)
        h.fill = F_SEC_HDR
        h.font = fnt(bold=True, color="1F4E79", size=10)
        h.alignment = aln("left", "center")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        r += 1

        # Conținut
        n_linii = max(continut.count("\n") + 1, 3)
        inaltimea = max(n_linii * 13 + 16, 50)
        set_h(ws, r, inaltimea)
        c = ws.cell(r, 1, continut)
        c.fill = F_WHITE
        c.font = fnt(size=10)
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        r += 1

    # Separator
    set_h(ws, r, 6)
    for col in range(1, 4):
        ws.cell(r, col).fill = F_SEP
    r += 2

    return r


def write_ghid_grafice(ws, r: int) -> int:
    """Scrie ghidul complet de citire grafice la finalul sheet-ului."""
    # Titlu ghid
    set_h(ws, r, 55)
    c = ws.cell(r, 1,
                "GHID COMPLET DE CITIRE GRAFICE & INDICATORI TEHNICI\n"
                "Referință permanentă — citește zilnic pentru a-ți forma ochiul de trader")
    c.fill = F_GHID_T; c.font = FN_GHID_T
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    r += 2

    for capitol in GHID_GRAFICE:
        # Titlu capitol
        set_h(ws, r, 28)
        c = ws.cell(r, 1, f"  {capitol['titlu']}")
        c.fill = F_GHID_CAP; c.font = FN_GHID_CAP
        c.alignment = aln("left", "center")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        r += 1

        for subtitlu, text in capitol["sectiuni"]:
            # Sub-titlu
            set_h(ws, r, 20)
            h = ws.cell(r, 1, f"    ▸  {subtitlu}")
            h.fill = F_GHID_SUB; h.font = FN_GHID_SUB
            h.alignment = aln("left", "center")
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
            r += 1

            # Text
            n_linii = max(text.count("\n") + 1, 3)
            inaltimea = max(n_linii * 13 + 16, 60)
            set_h(ws, r, inaltimea)
            c = ws.cell(r, 1, text)
            c.fill = F_GHID_TXT; c.font = FN_GHID_TXT
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
            r += 1

        # Separator capitol
        set_h(ws, r, 6)
        for col in range(1, 4):
            ws.cell(r, col).fill = fill("2E75B6")
        r += 2

    return r


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    log.info("=" * 65)
    log.info("  GENERARE GHID DE INVATARE ZILNIC")
    log.info(f"  {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}")
    log.info(f"  Active: {len(ACTIVE)}")
    log.info("=" * 65)

    if not EXCEL_PATH.exists():
        log.error(f"Fisierul Excel nu a fost gasit: {EXCEL_PATH}")
        sys.exit(1)

    # ── 1. Preia datele ───────────────────────────────────────────
    log.info(f"\n  Preia date pentru {len(ACTIVE)} active (2-5 minute)...\n")
    all_data = {}
    total = len(ACTIVE)
    for i, (name, ticker) in enumerate(ACTIVE.items(), 1):
        log.info(f"  [{i:3d}/{total}] {name:25s} ({ticker})")
        d = get_data(name, ticker)
        if d:
            all_data[ticker] = d
            log.info(f"             Pret={fmt_price(d['inchidere'],4)}  "
                     f"Zi={fmt_pct(d['var_zi'])}  "
                     f"RSI={d['rsi']:.1f}  [{d['semnal']}]  "
                     f"({d['n_zile']}z date)")
        else:
            log.warning(f"             -- EROARE / DATE INSUFICIENTE --")

    ok    = len(all_data)
    fails = total - ok
    buy   = sum(1 for d in all_data.values() if d.get("semnal") == "BUY")
    sell  = sum(1 for d in all_data.values() if d.get("semnal") == "SELL")
    wait  = sum(1 for d in all_data.values() if d.get("semnal") == "WAIT")

    log.info(f"\n  Rezultat: {ok}/{total} active OK  "
             f"| BUY:{buy}  SELL:{sell}  WAIT:{wait}")

    # ── 2. Deschide Excel ─────────────────────────────────────────
    log.info("\n  Scrie Ghid Invatare in Excel...")
    wb = load_workbook(str(EXCEL_PATH))

    SHEET = "GHID INVATARE"
    if SHEET in wb.sheetnames:
        del wb[SHEET]
    ws = wb.create_sheet(SHEET)
    ws.sheet_properties.tabColor = "154360"

    # Lățimi coloane
    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 28
    ws.freeze_panes = "A2"

    # ── 3. Titlu principal ────────────────────────────────────────
    r = 1
    set_h(ws, r, 55)
    c = ws.cell(r, 1,
                f"GHID DE INVATARE ZILNIC — ANALIZA DE PIATA PROFESIONALA\n"
                f"Generat: {datetime.now().strftime('%d.%m.%Y  %H:%M')}   "
                f"|   {ok} active analizate   "
                f"|   BUY: {buy}   SELL: {sell}   WAIT: {wait}")
    c.fill = F_TITLU; c.font = FN_TITLU
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    r += 2

    # ── 4. Scrie blocuri pe categorii ─────────────────────────────
    for cat_name, cat_dict in CATEGORII:
        activi_cat = [(name, ticker) for name, ticker in cat_dict.items()
                      if ticker in all_data]
        if not activi_cat:
            continue

        # Header categorie
        set_h(ws, r, 28)
        c = ws.cell(r, 1, f"  ══════  {cat_name}  ══════")
        c.fill = F_CAT; c.font = FN_CAT
        c.alignment = aln("center", "center")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        r += 2

        for name, ticker in activi_cat:
            d = all_data[ticker]
            log.info(f"  Scrie bloc: {name:25s} [{d['semnal']}]")
            r = write_bloc_activ(ws, d, r)

        r += 1

    # ── 5. Ghid de citire grafice ─────────────────────────────────
    r += 2
    r = write_ghid_grafice(ws, r)

    # ── 6. Salvează ───────────────────────────────────────────────
    wb.save(str(EXCEL_PATH))

    log.info("\n" + "=" * 65)
    log.info(f"  SALVAT CU SUCCES")
    log.info(f"  Sheet: '{SHEET}' in {EXCEL_PATH.name}")
    log.info(f"  Active procesate: {ok}/{total}"
             + (f"  ({fails} erori)" if fails else "  (toate OK)"))
    log.info(f"  Finalizat: {datetime.now().strftime('%H:%M:%S')}")
    log.info("=" * 65)


if __name__ == "__main__":
    main()
