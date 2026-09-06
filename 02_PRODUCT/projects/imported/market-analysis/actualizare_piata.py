import os
import sys
import logging
from copy import copy
from datetime import datetime, timedelta
from pathlib import Path
import requests
import pandas as pd
import yfinance as yf
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.chart import StockChart, BarChart, Reference
from openpyxl.chart.axis import DateAxis, ChartLines
from openpyxl.chart.updown_bars import UpDownBars
from openpyxl.chart.data_source import NumData, NumVal
from copy import deepcopy

EXCEL_PATH = Path(r"C:\\Users\\Marius\\Desktop\\Nu sterge\\Analiza_Piata_Profesionala.xlsx")
FRED_API_KEY = os.environ.get("FRED_API_KEY", "")

INDICI = {
    "S&P 500": "^GSPC", "NASDAQ 100": "^NDX", "NASDAQ Comp.": "^IXIC",
    "Dow Jones": "^DJI", "Russell 2000": "^RUT", "DAX Germany": "^GDAXI",
    "FTSE 100": "^FTSE", "CAC 40": "^FCHI", "Nikkei 225": "^N225",
    "Hang Seng": "^HSI", "Shanghai Comp.": "000001.SS", "MSCI World ETF": "URTH",
    "MSCI EM ETF": "EEM", "BET Romania": "BET.RO",
}

ACTIUNI = {
    "Apple": "AAPL", "Microsoft": "MSFT", "NVIDIA": "NVDA", "Alphabet": "GOOGL",
    "Amazon": "AMZN", "Meta": "META", "Tesla": "TSLA", "AMD": "AMD",
    "Intel": "INTC", "Broadcom": "AVGO", "ASML": "ASML", "Taiwan Semi": "TSM",
    "Palantir": "PLTR", "Salesforce": "CRM", "Oracle": "ORCL", "JPMorgan": "JPM",
    "Goldman Sachs": "GS", "Berkshire B": "BRK-B", "Visa": "V", "Mastercard": "MA",
    "ExxonMobil": "XOM", "Chevron": "CVX", "Shell": "SHEL", "Caterpillar": "CAT",
    "Boeing": "BA", "SPY": "SPY", "QQQ": "QQQ", "GLD ETF": "GLD",
    "TLT Bond ETF": "TLT", "ARKK": "ARKK",
}

CRYPTO = {
    "Bitcoin": "BTC-USD", "Ethereum": "ETH-USD", "BNB": "BNB-USD", "Solana": "SOL-USD",
    "XRP": "XRP-USD", "Cardano": "ADA-USD", "Avalanche": "AVAX-USD", "Dogecoin": "DOGE-USD",
    "Chainlink": "LINK-USD", "Polkadot": "DOT-USD", "Litecoin": "LTC-USD", "Shiba Inu": "SHIB-USD",
    "Polygon": "MATIC-USD", "Uniswap": "UNI-USD", "Cosmos": "ATOM-USD", "Stellar": "XLM-USD",
    "Monero": "XMR-USD", "Tron": "TRX-USD", "Filecoin": "FIL-USD", "Aave": "AAVE-USD",
    "Arbitrum": "ARB-USD", "Optimism": "OP-USD", "Render": "RNDR-USD", "Sui": "SUI-USD",
    "Near Protocol": "NEAR-USD",
}

VALUTE = {
    "EUR/USD": "EURUSD=X", "GBP/USD": "GBPUSD=X", "USD/JPY": "USDJPY=X", "USD/CHF": "USDCHF=X",
    "AUD/USD": "AUDUSD=X", "USD/CAD": "USDCAD=X", "EUR/RON": "EURRON=X", "USD/RON": "USDRON=X",
    "GBP/RON": "GBPRON=X", "EUR/GBP": "EURGBP=X", "USD/CNY": "USDCNY=X", "USD/TRY": "USDTRY=X",
}

MATERII_PRIME = {
    "Gold": "GC=F", "Silver": "SI=F", "Platinum": "PL=F", "Palladium": "PA=F",
    "Oil WTI": "CL=F", "Oil Brent": "BZ=F", "Natural Gas": "NG=F", "Copper": "HG=F",
    "Corn": "ZC=F", "Wheat": "ZW=F", "Soybean": "ZS=F", "Coffee": "KC=F",
    "Sugar": "SB=F", "Cotton": "CT=F",
}

ACTIVE = {**INDICI, **ACTIUNI, **CRYPTO, **VALUTE, **MATERII_PRIME}

MACRO_TICKERS = {
    "VIX": "^VIX",
    "Yield 10Y US": "^TNX",
    "Yield 2Y US": "^IRX",
    "Yield 30Y US": "^TYX",
    "USD Index": "DX-Y.NYB",
}

CATEGORII = [
    ("INDICI BURSIERI", INDICI),
    ("ACTIUNI & ETF", ACTIUNI),
    ("CRYPTOCURRENCY", CRYPTO),
    ("VALUTE FOREX", VALUTE),
    ("MATERII PRIME", MATERII_PRIME),
]

COMPETITORI_MAP = {
    "INDICI": ["S&P 500", "NASDAQ 100", "Dow Jones", "DAX Germany", "FTSE 100", "Nikkei 225"],
    "ACTIUNI": ["Apple", "Microsoft", "NVIDIA", "Alphabet", "Amazon", "Meta"],
    "CRYPTO": ["Bitcoin", "Ethereum", "BNB", "Solana", "XRP", "Cardano"],
    "VALUTE": ["EUR/USD", "GBP/USD", "USD/JPY", "USD/CHF", "AUD/USD", "USD/CAD"],
    "MATERII": ["Gold", "Silver", "Oil WTI", "Oil Brent", "Natural Gas", "Copper"],
}

RISK_LIBRARY = {
    "INDICI": [
        ("R01", "Risc", "Recesiune SUA", 5, 70, "1-3 luni"),
        ("R02", "Risc", "Inflație persistentă", 4, 65, "1-3 luni"),
        ("R03", "Risc", "Breakdown sub MA200", 4, 55, "2-6 săpt."),
        ("R04", "Risc", "Tensiuni geopolitice", 5, 45, "Imediat"),
        ("O01", "Oportunitate", "Revenire pe sentiment extrem negativ", 4, 50, "1-4 săpt."),
        ("O02", "Oportunitate", "Breakout peste rezistențe majore", 4, 40, "1-8 săpt."),
    ],
    "ACTIUNI": [
        ("R01", "Risc", "Reglementări antitrust", 4, 45, "1-6 luni"),
        ("R02", "Risc", "Comprimarea marjelor", 4, 55, "1-2 trimestre"),
        ("R03", "Risc", "Miss la earnings", 5, 40, "Eveniment"),
        ("R04", "Risc", "RSI extins / pullback", 3, 60, "1-3 săpt."),
        ("O01", "Oportunitate", "Rezultate peste estimări", 4, 35, "Eveniment"),
        ("O02", "Oportunitate", "Reluare trend peste MA50", 4, 45, "2-6 săpt."),
    ],
    "CRYPTO": [
        ("R01", "Risc", "Reglementări SEC", 5, 55, "1-6 luni"),
        ("R02", "Risc", "Volatilitate extremă", 5, 75, "Imediat"),
        ("R03", "Risc", "Hack exchange / protocol", 4, 30, "Imediat"),
        ("R04", "Risc", "Sentiment risk-off macro", 4, 60, "1-4 săpt."),
        ("O01", "Oportunitate", "Short squeeze cu RVOL ridicat", 4, 35, "1-10 zile"),
        ("O02", "Oportunitate", "Breakout post-consolidare", 4, 40, "1-4 săpt."),
    ],
    "VALUTE": [
        ("R01", "Risc", "Divergență politici monetare", 4, 65, "1-3 luni"),
        ("R02", "Risc", "Intervenție bancă centrală", 5, 30, "Imediat"),
        ("R03", "Risc", "Flight to safety USD", 4, 50, "1-6 săpt."),
        ("R04", "Risc", "Date macro sub așteptări", 3, 60, "Eveniment"),
        ("O01", "Oportunitate", "Carry trade pe trend stabil", 3, 45, "1-8 săpt."),
        ("O02", "Oportunitate", "Breakout după ședință CB", 4, 35, "1-10 zile"),
    ],
    "MATERII": [
        ("R01", "Risc", "Încetinire economică globală", 4, 55, "1-3 luni"),
        ("R02", "Risc", "Supply chain disruptions", 4, 40, "1-8 săpt."),
        ("R03", "Risc", "USD puternic", 4, 65, "1-6 săpt."),
        ("R04", "Risc", "Volatilitate geopolitică", 5, 45, "Imediat"),
        ("O01", "Oportunitate", "Rebalansare pe inflație", 4, 35, "1-3 luni"),
        ("O02", "Oportunitate", "Breakout pe ofertă restrânsă", 4, 40, "2-8 săpt."),
    ],
}

CALENDAR_LIBRARY = {
    "INDICI": ["FOMC", "NFP", "CPI", "GDP", "PMI", "Earnings Season"],
    "ACTIUNI": ["Earnings Report", "FOMC", "CPI", "NFP", "PCE", "Retail Sales"],
    "CRYPTO": ["Bitcoin Halving", "FOMC", "SEC Ruling", "CPI", "ETH Upgrade", "Macro Risk"],
    "VALUTE": ["FOMC", "ECB", "BOE", "BOJ", "CPI SUA", "NFP"],
    "MATERII": ["OPEC+", "EIA Crude", "FOMC", "China PMI", "USD Index", "Geopolitical Events"],
}

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(Path(__file__).with_name("actualizare.log"), encoding="utf-8"),
    ],
)
log = logging.getLogger(__name__)


def fill(hex_c):
    return PatternFill("solid", fgColor=hex_c)


def fnt(bold=False, color="000000", size=10):
    return Font(name="Arial", bold=bold, color=color, size=size)


F_GRN = fill("C6EFCE")
FN_GRN = fnt(color="375623")
F_YLW = fill("FFEB9C")
FN_YLW = fnt(color="9C6500")
F_RED = fill("FFC7CE")
FN_RED = fnt(color="9C0006")
F_HDR = fill("1F4E79")
FN_HDR = fnt(bold=True, color="FFFFFF")
F_GRY = fill("F2F2F2")
FN_GRY = fnt(color="595959")
F_LBL = fill("D6DCE4")
FN_LBL = fnt(bold=True)
NO_FILL = PatternFill(fill_type=None)


def calc_rsi(prices, period=14):
    delta = prices.diff()
    gain = delta.clip(lower=0).rolling(period).mean()
    loss = (-delta.clip(upper=0)).rolling(period).mean()
    rs = gain / loss.replace(0, 1e-10)
    rsi = 100 - (100 / (1 + rs))
    val = rsi.iloc[-1]
    return round(float(val), 2) if pd.notna(val) else 50.0


def calc_macd(prices):
    ema12 = prices.ewm(span=12, adjust=False).mean()
    ema26 = prices.ewm(span=26, adjust=False).mean()
    ml = ema12 - ema26
    sl = ml.ewm(span=9, adjust=False).mean()
    hist = ml - sl

    m = float(ml.iloc[-1])
    s = float(sl.iloc[-1])
    h = float(hist.iloc[-1])
    prev_h = float(hist.iloc[-2]) if len(hist) > 1 and pd.notna(hist.iloc[-2]) else 0.0

    if m > s and prev_h < 0 <= h:
        cross = "Impuls pozitiv nou"
    elif m > s and h >= 0:
        cross = "Impuls pozitiv activ"
    elif m < s and prev_h > 0 >= h:
        cross = "Impuls negativ nou"
    else:
        cross = "Impuls negativ activ"

    return {"macd": round(m, 6), "signal": round(s, 6), "histogram": round(h, 6), "cross": cross}


def calc_ma(prices):
    def ma(n):
        return round(float(prices.tail(n).mean()), 6) if len(prices) >= n else None

    ma20 = ma(20)
    ma50 = ma(50)
    ma200 = ma(200)
    cross = "Neutru"
    if ma50 is not None and ma200 is not None:
        cross = "Golden Cross" if ma50 > ma200 else "Death Cross"
    return {"ma20": ma20, "ma50": ma50, "ma200": ma200, "macross": cross}


def calc_bollinger(prices, period=20):
    m = prices.rolling(period).mean()
    std = prices.rolling(period).std()
    sup = float((m + 2 * std).iloc[-1]) if pd.notna((m + 2 * std).iloc[-1]) else 0.0
    inf = float((m - 2 * std).iloc[-1]) if pd.notna((m - 2 * std).iloc[-1]) else 0.0
    return {"bb_sup": round(sup, 6), "bb_inf": round(inf, 6), "bb_width": round(sup - inf, 6)}


def calc_atr(hist, period=14):
    hi, lo, cl = hist["High"], hist["Low"], hist["Close"]
    tr = pd.concat([(hi - lo), (hi - cl.shift()).abs(), (lo - cl.shift()).abs()], axis=1).max(axis=1)
    atr = tr.rolling(period).mean().iloc[-1]
    return round(float(atr), 6) if pd.notna(atr) else 0.0


def calc_stochastic(hist, period=14):
    lo14 = hist["Low"].rolling(period).min()
    hi14 = hist["High"].rolling(period).max()
    diff = (hi14 - lo14).replace(0, 1e-10)
    k = (hist["Close"] - lo14) / diff * 100
    d = k.rolling(3).mean()
    return {
        "stoch_k": round(float(k.iloc[-1]), 2) if pd.notna(k.iloc[-1]) else 50.0,
        "stoch_d": round(float(d.iloc[-1]), 2) if pd.notna(d.iloc[-1]) else 50.0,
    }


def map_rsi_status(rsi):
    if rsi < 30:
        return "Presiune excesivă vânzare"
    if rsi < 45:
        return "Presiune moderată vânzare"
    if rsi <= 55:
        return "Echilibru"
    if rsi <= 70:
        return "Momentum ascendent"
    return "Presiune excesivă cumpărare"


def calc_signal(rsi, macd_cross, ma_cross, rvol):
    score = 0

    if rsi < 35:
        score += 2
    elif rsi < 45:
        score += 1
    elif rsi > 75:
        score -= 2
    elif rsi > 65:
        score -= 1

    if macd_cross == "Impuls pozitiv nou":
        score += 2
    elif macd_cross == "Impuls pozitiv activ":
        score += 1
    elif macd_cross == "Impuls negativ nou":
        score -= 2
    elif macd_cross == "Impuls negativ activ":
        score -= 1

    if ma_cross == "Golden Cross":
        score += 2
    elif ma_cross == "Death Cross":
        score -= 2

    if rvol > 1.5:
        score += 1
    elif rvol < 0.6:
        score -= 1

    conf = min(abs(score), 5)

    if score >= 3:
        return "BUY", conf, score
    if score <= -3:
        return "SELL", conf, score
    return "WAIT", conf, score


def color_pnl(cell, value):
    if value > 0:
        cell.fill = copy(F_GRN)
        cell.font = copy(FN_GRN)
    elif value < 0:
        cell.fill = copy(F_RED)
        cell.font = copy(FN_RED)
    else:
        cell.fill = copy(F_YLW)
        cell.font = copy(FN_YLW)


def color_rsi(cell, rsi):
    if rsi < 30:
        cell.fill = copy(F_GRN)
        cell.font = copy(FN_GRN)
    elif rsi < 50:
        cell.fill = copy(F_YLW)
        cell.font = copy(FN_YLW)
    elif rsi < 70:
        cell.fill = fill("E2EFDA")
        cell.font = fnt(color="375623")
    else:
        cell.fill = copy(F_RED)
        cell.font = copy(FN_RED)


def color_signal(cell, semnal):
    if semnal == "BUY":
        cell.fill = copy(F_GRN)
        cell.font = copy(FN_GRN)
    elif semnal == "SELL":
        cell.fill = copy(F_RED)
        cell.font = copy(FN_RED)
    else:
        cell.fill = copy(F_YLW)
        cell.font = copy(FN_YLW)


def color_trend(cell, trend):
    if trend == "Bullish":
        cell.fill = copy(F_GRN)
        cell.font = copy(FN_GRN)
    elif trend == "Bearish":
        cell.fill = copy(F_RED)
        cell.font = copy(FN_RED)
    else:
        cell.fill = copy(F_YLW)
        cell.font = copy(FN_YLW)


def color_status_extended(cell, value):
    txt = str(value).strip().lower()
    if txt in ("buy", "pozitiv", "bullish", "golden cross", "impuls pozitiv nou", "impuls pozitiv activ"):
        cell.fill = copy(F_GRN)
        cell.font = copy(FN_GRN)
    elif txt in ("sell", "negativ", "bearish", "death cross", "impuls negativ nou", "impuls negativ activ"):
        cell.fill = copy(F_RED)
        cell.font = copy(FN_RED)
    else:
        cell.fill = copy(F_YLW)
        cell.font = copy(FN_YLW)


def safe_write(ws, row, col, value=None):
    cell = ws.cell(row, col)
    if isinstance(cell, MergedCell):
        return None
    cell.value = value
    return cell


def clear_rows(ws, start_row, end_row=5000):
    for row in ws.iter_rows(min_row=start_row, max_row=end_row):
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            cell.value = None
            cell.fill = copy(NO_FILL)
            cell.font = fnt()
            cell.number_format = "General"
            cell.alignment = Alignment(vertical="center")


def write_category_header(ws, row, label, num_cols):
    for col in range(1, num_cols + 1):
        c = safe_write(ws, row, col, label if col == 1 else None)
        if c:
            c.fill = copy(F_HDR)
            c.font = copy(FN_HDR)


def fmt_price(x):
    if x is None or x == "":
        return ""
    x = float(x)
    if abs(x) >= 1000:
        return f"{x:,.4f}"
    if abs(x) >= 1:
        return f"{x:.4f}"
    return f"{x:.6f}"


def rr_value(entry, sl, tp):
    if entry in (None, 0) or sl is None or tp is None or entry == sl:
        return None
    return abs((tp - entry) / (entry - sl))


def rr_text(entry, sl, tp):
    rr = rr_value(entry, sl, tp)
    return f"{rr:.2f}x" if rr is not None else "N/A"


def signal_status_text(signal):
    if signal == "BUY":
        return "Pozitiv"
    if signal == "SELL":
        return "Negativ"
    return "Neutru"


def get_full_data(name, ticker):
    try:
        hist = yf.Ticker(ticker).history(period="1y", interval="1d", auto_adjust=True, timeout=15)
        if hist.empty or len(hist) < 5:
            log.warning(f"  ⚠ Date insuficiente: {name} ({ticker})")
            return {}

        closes = hist["Close"]
        latest = hist.iloc[-1]
        prev = hist.iloc[-2]

        close_now = float(latest["Close"])
        close_prev = float(prev["Close"])
        close_5d = float(hist.iloc[-5]["Close"]) if len(hist) >= 5 else close_prev
        close_20d = float(hist.iloc[-20]["Close"]) if len(hist) >= 20 else close_prev

        var_zi = (close_now - close_prev) / close_prev * 100 if close_prev else 0
        var_sapt = (close_now - close_5d) / close_5d * 100 if close_5d else 0
        var_luna = (close_now - close_20d) / close_20d * 100 if close_20d else 0

        volume = int(latest.get("Volume", 0))
        avg_vol = int(hist["Volume"].tail(20).mean()) if "Volume" in hist else 0
        rvol = round(volume / avg_vol, 2) if avg_vol > 0 else 1.0

        rsi = calc_rsi(closes)
        macd = calc_macd(closes)
        ma = calc_ma(closes)
        boll = calc_bollinger(closes)
        atr = calc_atr(hist)
        stoch = calc_stochastic(hist)
        mom10 = round(float(closes.pct_change(10).iloc[-1] * 100), 2) if len(closes) > 10 else 0.0

        price = close_now
        if ma["ma50"] is not None and price > ma["ma50"] * 1.01:
            trend = "Bullish"
        elif ma["ma50"] is not None and price < ma["ma50"] * 0.99:
            trend = "Bearish"
        else:
            trend = "Sideways"

        rsi_status = map_rsi_status(rsi)
        semnal, confluente, score = calc_signal(rsi, macd["cross"], ma["macross"], rvol)

        if semnal == "BUY":
            sl = round(price - 1.5 * atr, 6)
            tp = round(price + 3.0 * atr, 6)
        elif semnal == "SELL":
            sl = round(price + 1.5 * atr, 6)
            tp = round(price - 3.0 * atr, 6)
        else:
            sl = None
            tp = None

        prob = min(90, 35 + confluente * 10 + (5 if rvol > 1.2 else 0))
        support = round(float(hist["Low"].tail(20).min()), 6)
        resistance = round(float(hist["High"].tail(20).max()), 6)

        return {
            "name": name,
            "ticker": ticker,
            "data": datetime.now().strftime("%d.%m.%Y"),
            "timestamp": datetime.now().strftime("%d.%m.%Y %H:%M"),
            "deschidere": round(float(latest.get("Open", close_now)), 6),
            "maxim": round(float(latest.get("High", close_now)), 6),
            "minim": round(float(latest.get("Low", close_now)), 6),
            "inchidere": round(close_now, 6),
            "var_zi_pct": round(var_zi, 4),
            "var_sapt_pct": round(var_sapt, 4),
            "var_luna_pct": round(var_luna, 4),
            "volum": volume,
            "avg_vol_20": avg_vol,
            "rvol": rvol,
            "rsi": rsi,
            "rsi_status": rsi_status,
            "macd": macd["macd"],
            "macd_signal": macd["signal"],
            "macd_hist": macd["histogram"],
            "macd_cross": macd["cross"],
            "ma20": ma["ma20"],
            "ma50": ma["ma50"],
            "ma200": ma["ma200"],
            "macross": ma["macross"],
            "bb_sup": boll["bb_sup"],
            "bb_inf": boll["bb_inf"],
            "bb_width": boll["bb_width"],
            "atr": atr,
            "stoch_k": stoch["stoch_k"],
            "stoch_d": stoch["stoch_d"],
            "momentum_10z": mom10,
            "trend": trend,
            "semnal": semnal,
            "score": score,
            "confluente": confluente,
            "sl": sl,
            "tp": tp,
            "probabilitate": prob,
            "support": support,
            "resistance": resistance,
        }
    except Exception as e:
        log.error(f"  X Eroare {name} ({ticker}): {e}")
        return {}


def get_fear_greed():
    try:
        r = requests.get("https://api.alternative.me/fng/?limit=1", timeout=10)
        r.raise_for_status()
        d = r.json()["data"][0]
        val = int(d["value"])
        label = d["value_classification"]
        return {
            "value": val,
            "display": f"{val} - {label}",
            "status": "Pozitiv" if val > 60 else ("Negativ" if val < 40 else "Neutru"),
        }
    except Exception as e:
        log.warning(f"  Fear & Greed indisponibil: {e}")
        return {"value": None, "display": "N/A", "status": "Neutru"}


def get_fred(series_id):
    if not FRED_API_KEY:
        return None
    try:
        url = (
            "https://api.stlouisfed.org/fred/series/observations"
            f"?series_id={series_id}&api_key={FRED_API_KEY}&file_type=json&sort_order=desc&limit=2"
        )
        resp = requests.get(url, timeout=10)
        resp.raise_for_status()
        obs = resp.json().get("observations", [])
        vals = [float(x["value"]) for x in obs if x.get("value") not in (None, ".")]
        if vals:
            return vals[0], vals[1] if len(vals) > 1 else None
    except Exception as e:
        log.warning(f"  FRED {series_id}: {e}")
    return None, None


def update_preturi_volume(ws, all_data):
    clear_rows(ws, 3)
    r = 3

    for cat_name, cat_dict in CATEGORII:
        write_category_header(ws, r, f"--- {cat_name} ---", 15)
        r += 1

        for name, ticker in cat_dict.items():
            d = all_data.get(ticker, {})
            safe_write(ws, r, 1, d.get("data", datetime.now().strftime("%d.%m.%Y")))
            c = safe_write(ws, r, 2, name)
            if c:
                c.font = fnt(bold=True)

            if d:
                vals = {
                    3: d["deschidere"],
                    4: d["maxim"],
                    5: d["minim"],
                    6: d["inchidere"],
                    7: d["var_zi_pct"] / 100,
                    8: d["var_sapt_pct"] / 100,
                    9: d["var_luna_pct"] / 100,
                    10: d["volum"],
                    11: d["avg_vol_20"],
                    12: d["rvol"],
                    13: "",
                    14: "",
                    15: d["trend"],
                }
                for col, val in vals.items():
                    c = safe_write(ws, r, col, val)
                    if not c:
                        continue
                    if col in (3, 4, 5, 6):
                        c.number_format = "#,##0.0000"
                    elif col in (7, 8, 9):
                        c.number_format = "0.00%"
                        color_pnl(c, d["var_zi_pct"] if col == 7 else d["var_sapt_pct"] if col == 8 else d["var_luna_pct"])
                    elif col in (10, 11):
                        c.number_format = "#,##0"
                    elif col == 12:
                        c.number_format = "0.00x"
                    elif col == 15:
                        color_trend(c, d["trend"])
            else:
                c = safe_write(ws, r, 6, "N/A")
                if c:
                    c.fill = copy(F_GRY)
                    c.font = copy(FN_GRY)
            r += 1

    log.info(f"  OK PRETURI VOLUME — {r - 3} randuri actualizate")


def update_indicatori_tehnici(ws, all_data):
    clear_rows(ws, 3)
    r = 3

    for cat_name, cat_dict in CATEGORII:
        write_category_header(ws, r, f"--- {cat_name} ---", 23)
        r += 1

        for name, ticker in cat_dict.items():
            d = all_data.get(ticker, {})
            safe_write(ws, r, 1, d.get("data", datetime.now().strftime("%d.%m.%Y")))
            c = safe_write(ws, r, 2, name)
            if c:
                c.font = fnt(bold=True)

            if d:
                vals = {
                    3: d["inchidere"],
                    4: d["ma20"],
                    5: d["ma50"],
                    6: d["ma200"],
                    7: d["rsi"],
                    8: d["rsi_status"],
                    9: d["macd"],
                    10: d["macd_signal"],
                    11: d["macd_hist"],
                    12: d["bb_sup"],
                    13: d["bb_inf"],
                    14: d["bb_width"],
                    15: d["atr"],
                    16: d["stoch_k"],
                    17: d["stoch_d"],
                    18: d["volum"],
                    19: d["rvol"],
                    20: d["trend"],
                    21: d["support"],
                    22: d["resistance"],
                    23: d["macross"],
                }
                for col, val in vals.items():
                    c = safe_write(ws, r, col, val)
                    if not c:
                        continue
                    if col in (3, 4, 5, 6, 12, 13, 14, 15, 21, 22):
                        c.number_format = "#,##0.0000"
                    elif col == 7:
                        c.number_format = "0.00"
                        color_rsi(c, d["rsi"])
                    elif col in (9, 10, 11):
                        c.number_format = "0.000000"
                    elif col in (16, 17):
                        c.number_format = "0.00"
                    elif col == 18:
                        c.number_format = "#,##0"
                    elif col == 19:
                        c.number_format = "0.00x"
                    elif col == 20:
                        color_trend(c, d["trend"])
                    elif col == 23:
                        color_status_extended(c, d["macross"])
            r += 1

    log.info(f"  OK INDICATORI TEHNICI — {r - 3} randuri actualizate")


def update_semnale(ws, all_data):
    clear_rows(ws, 3)
    r = 3

    for cat_name, cat_dict in CATEGORII:
        write_category_header(ws, r, f"--- {cat_name} ---", 17)
        r += 1

        for name, ticker in cat_dict.items():
            d = all_data.get(ticker, {})
            if not d:
                r += 1
                continue

            conditie = (
                f"RSI={d['rsi']:.0f} | {d['macd_cross']} | "
                f"{d['macross']} | RVOL={d['rvol']:.1f}x | Score={d['score']}"
            )

            safe_write(ws, r, 1, d["data"])
            c = safe_write(ws, r, 2, name)
            if c:
                c.font = fnt(bold=True)

            safe_write(ws, r, 3, d["semnal"])
            safe_write(ws, r, 4, conditie)
            c = safe_write(ws, r, 5, d["rsi"])
            if c:
                c.number_format = "0.00"

            safe_write(ws, r, 6, d["macd_cross"])
            safe_write(ws, r, 7, d["macross"])
            c = safe_write(ws, r, 8, d["rvol"])
            if c:
                c.number_format = "0.00x"

            c = safe_write(ws, r, 9, d["momentum_10z"] / 100)
            if c:
                c.number_format = "0.00%"

            safe_write(ws, r, 10, d["confluente"])
            c = safe_write(ws, r, 11, d["inchidere"])
            if c:
                c.number_format = "#,##0.0000"

            c = safe_write(ws, r, 12, d["sl"] if d["semnal"] != "WAIT" else None)
            if c and d["sl"] is not None:
                c.number_format = "#,##0.0000"

            c = safe_write(ws, r, 13, d["tp"] if d["semnal"] != "WAIT" else None)
            if c and d["tp"] is not None:
                c.number_format = "#,##0.0000"

            c = safe_write(ws, r, 14)
            if c:
                c.value = f'=IFERROR(IF(OR(C{r}="WAIT",L{r}="",M{r}=""),"N/A",(M{r}-K{r})/(K{r}-L{r})),"N/A")'
                c.number_format = "0.00x"

            c = safe_write(ws, r, 15, d["probabilitate"] / 100)
            if c:
                c.number_format = "0%"

            safe_write(ws, r, 16, "Activ")
            safe_write(ws, r, 17, f"Auto {datetime.now().strftime('%H:%M')}")

            color_signal(ws.cell(r, 3), d["semnal"])
            color_rsi(ws.cell(r, 5), d["rsi"])
            color_status_extended(ws.cell(r, 6), d["macd_cross"])
            color_status_extended(ws.cell(r, 7), d["macross"])

            prob_c = ws.cell(r, 15)
            if d["probabilitate"] >= 65:
                prob_c.fill = copy(F_GRN)
                prob_c.font = copy(FN_GRN)
            elif d["probabilitate"] >= 50:
                prob_c.fill = copy(F_YLW)
                prob_c.font = copy(FN_YLW)
            else:
                prob_c.fill = copy(F_RED)
                prob_c.font = copy(FN_RED)

            r += 1

    log.info(f"  OK SEMNALE INTRARE — {r - 3} randuri actualizate")


def update_macro(ws, macro_live, fear_greed):
    clear_rows(ws, 3)

    fred_specs = {
        "PIB YoY (%)": ("A191RL1Q225SBEA", "Trimestrial"),
        "CPI YoY (%)": ("CPIAUCSL", "Lunar"),
        "Core CPI (%)": ("CPILFESL", "Lunar"),
        "Rata dobânzii (%)": ("FEDFUNDS", "Lunar"),
        "Rata șomajului (%)": ("UNRATE", "Lunar"),
        "Retail Sales MoM (%)": ("RSAFS", "Lunar"),
    }

    rows = [
        "PIB YoY (%)", "CPI YoY (%)", "Core CPI (%)", "Rata dobânzii (%)", "Rata șomajului (%)",
        "PMI Manufacturing", "PMI Services", "Retail Sales MoM (%)", "Balanță comercială ($B)",
        "USD Index (DXY)", "EUR/USD", "Petrol Brent ($)", "Petrol WTI ($)", "Yield 10Y US (%)",
        "VIX", "Fear & Greed (0-100)",
    ]

    prev_map = {}
    if ws.max_row >= 3:
        for rr in range(3, min(ws.max_row, 50) + 1):
            name = ws.cell(rr, 1).value
            prev = ws.cell(rr, 2).value
            if name:
                prev_map[str(name).strip()] = prev

    eurusd = macro_live.get("EUR/USD", None)
    brent = macro_live.get("Oil Brent", None)
    wti = macro_live.get("Oil WTI", None)

    values = {
        "USD Index (DXY)": macro_live.get("USD Index"),
        "Yield 10Y US (%)": macro_live.get("Yield 10Y US"),
        "VIX": macro_live.get("VIX"),
        "Fear & Greed (0-100)": fear_greed.get("value"),
        "EUR/USD": eurusd,
        "Petrol Brent ($)": brent,
        "Petrol WTI ($)": wti,
    }

    r = 3
    for ind in rows:
        curr = values.get(ind)
        prev = prev_map.get(ind)

        if ind in fred_specs:
            curr, fred_prev = get_fred(fred_specs[ind][0])
            if prev in (None, "") and fred_prev is not None:
                prev = fred_prev

        safe_write(ws, r, 1, ind)
        c = safe_write(ws, r, 2, curr)
        if c and isinstance(curr, (int, float)):
            c.number_format = "0.00"

        c = safe_write(ws, r, 3, prev)
        if c and isinstance(prev, (int, float)):
            c.number_format = "0.00"

        c = safe_write(ws, r, 4)
        if c:
            c.value = f'=IFERROR(B{r}-C{r},"")'
            c.number_format = "0.00"

        c = safe_write(ws, r, 5)
        if c:
            c.value = f'=IFERROR((B{r}-C{r})/C{r},"")'
            c.number_format = "0.00%"

        safe_write(ws, r, 6, "")
        safe_write(ws, r, 7, "")
        safe_write(ws, r, 8, "Impact market")
        safe_write(ws, r, 9, datetime.now().strftime("%d.%m.%Y %H:%M"))
        safe_write(ws, r, 10, fred_specs[ind][1] if ind in fred_specs else "Live")

        trend = "Creștere" if isinstance(curr, (int, float)) and isinstance(prev, (int, float)) and curr > prev else \
                "Scădere" if isinstance(curr, (int, float)) and isinstance(prev, (int, float)) and curr < prev else "Stabil"
        status = "Pozitiv" if ind == "Fear & Greed (0-100)" and fear_greed.get("status") == "Pozitiv" else \
                 "Negativ" if ind == "Fear & Greed (0-100)" and fear_greed.get("status") == "Negativ" else \
                 "Neutru"

        if ind == "VIX" and isinstance(curr, (int, float)):
            status = "Negativ" if curr > 25 else "Neutru" if curr > 15 else "Pozitiv"

        if ind == "Yield 10Y US (%)" and isinstance(curr, (int, float)):
            status = "Negativ" if curr > 4.5 else "Neutru" if curr > 3.5 else "Pozitiv"

        safe_write(ws, r, 11, trend)
        safe_write(ws, r, 12, status)
        safe_write(ws, r, 13, "")

        color_status_extended(ws.cell(r, 12), status)
        r += 1

    log.info("  OK INDICATORI MACRO actualizat")


def update_competitori_sector(ws, all_data):
    clear_rows(ws, 3)
    r = 3

    for cat, names in COMPETITORI_MAP.items():
        write_category_header(ws, r, f"--- {cat} ---", 15)
        r += 1

        for name in names:
            ticker = ACTIVE.get(name)
            d = all_data.get(ticker, {}) if ticker else {}

            safe_write(ws, r, 1, name)
            safe_write(ws, r, 2, cat)
            safe_write(ws, r, 3, None)
            safe_write(ws, r, 4, None)

            c = safe_write(ws, r, 5, d.get("inchidere") if d else None)
            if c and d:
                c.number_format = "#,##0.0000"

            safe_write(ws, r, 6, d.get("trend") if d else "N/A")
            safe_write(ws, r, 7, None)
            safe_write(ws, r, 8, None)
            safe_write(ws, r, 9, "")
            safe_write(ws, r, 10, "")
            safe_write(ws, r, 11, None)
            safe_write(ws, r, 12, None)
            safe_write(ws, r, 13, signal_status_text(d.get("semnal")) if d else "Neutru")
            safe_write(ws, r, 14, datetime.now().strftime("%d.%m.%Y"))
            safe_write(ws, r, 15, f"Var zi {d.get('var_zi_pct', 0):+.2f}% | RVOL {d.get('rvol', 0):.2f}x" if d else "N/A")

            if d:
                color_trend(ws.cell(r, 6), d["trend"])
                color_status_extended(ws.cell(r, 13), ws.cell(r, 13).value)

            r += 1

    log.info("  OK COMPETITORI SECTOR actualizat")


def update_riscuri_oportunitati(ws):
    clear_rows(ws, 3)
    r = 3

    for cat, items in RISK_LIBRARY.items():
        write_category_header(ws, r, f"--- {cat} ---", 13)
        r += 1

        for rid, tip, desc, impact, prob, horizon in items:
            score = round(impact * prob / 100, 2)
            safe_write(ws, r, 1, rid)
            safe_write(ws, r, 2, tip)
            safe_write(ws, r, 3, cat)
            safe_write(ws, r, 4, desc)
            safe_write(ws, r, 5, impact)

            c = safe_write(ws, r, 6, prob / 100)
            if c:
                c.number_format = "0%"

            safe_write(ws, r, 7, score)
            safe_write(ws, r, 8, horizon)
            safe_write(ws, r, 9, "Monitorizare / hedge / confirmare tehnică")
            safe_write(ws, r, 10, "Analist")
            safe_write(ws, r, 11, "Deschis")
            safe_write(ws, r, 12, datetime.now().strftime("%d.%m.%Y"))
            safe_write(ws, r, 13, "")

            if tip == "Risc":
                ws.cell(r, 2).fill = copy(F_RED)
                ws.cell(r, 2).font = copy(FN_RED)
            else:
                ws.cell(r, 2).fill = copy(F_GRN)
                ws.cell(r, 2).font = copy(FN_GRN)

            r += 1

    log.info("  OK RISCURI OPORTUNITATI actualizat")


def update_calendar_economic(ws):
    clear_rows(ws, 3)
    r = 3
    base = datetime.now()

    for cat, events in CALENDAR_LIBRARY.items():
        write_category_header(ws, r, f"--- {cat} ---", 11)
        r += 1

        for i, ev in enumerate(events, start=1):
            dt = base + timedelta(days=i * 2, hours=9 + (i % 5))
            impact = "Ridicat" if i <= 3 else "Mediu"

            safe_write(ws, r, 1, dt.strftime("%d.%m.%Y %H:%M"))
            safe_write(ws, r, 2, ev)
            safe_write(ws, r, 3, "US" if cat in ("INDICI", "ACTIUNI", "CRYPTO") else "Global")
            safe_write(ws, r, 4, impact)
            safe_write(ws, r, 5, "")
            safe_write(ws, r, 6, "")
            safe_write(ws, r, 7, "")
            safe_write(ws, r, 8, "")
            safe_write(ws, r, 9, "")
            safe_write(ws, r, 10, cat)
            safe_write(ws, r, 11, "Generat automat")

            color_status_extended(ws.cell(r, 4), "BUY" if impact == "Ridicat" else "WAIT")
            r += 1

    log.info("  OK CALENDAR ECONOMIC actualizat")


def update_jurnal_stats(ws):
    stats = {
        25: ("Nr tranzacții", '=COUNTA(A3:A22)'),
        26: ("Win Rate", '=IFERROR(COUNTIF(N3:N22,">0")/COUNTA(N3:N22),0)'),
        27: ("P&L Total", '=IFERROR(SUM(N3:N22),0)'),
        28: ("P&L Mediu", '=IFERROR(AVERAGE(N3:N22),0)'),
        29: ("Best trade", '=IFERROR(MAX(N3:N22),0)'),
        30: ("Worst trade", '=IFERROR(MIN(N3:N22),0)'),
    }
    for row, (label, formula) in stats.items():
        safe_write(ws, row, 1, label)
        c = safe_write(ws, row, 2)
        if c:
            c.value = formula
            if row == 26:
                c.number_format = "0.00%"
            else:
                c.number_format = "#,##0.00"
    log.info("  OK JURNAL TRANZACTII statistici actualizate")


def update_historic(ws, all_data, fear_greed):
    luna_curenta = datetime.now().strftime("%b %Y")

    for row in range(2, 500):
        if ws.cell(row, 1).value == luna_curenta:
            log.info(f"  OK ISTORIC TRENDING — {luna_curenta} exista deja")
            return

    r = 2
    while ws.cell(r, 1).value:
        r += 1

    avg_rsi = sum(d.get("rsi", 50) for d in all_data.values()) / len(all_data) if all_data else 50
    sp500 = all_data.get("^GSPC", {})
    vix_d = all_data.get("^VIX", {})
    total_buy = sum(1 for d in all_data.values() if d.get("semnal") == "BUY")
    total_sell = sum(1 for d in all_data.values() if d.get("semnal") == "SELL")
    semnal_luna = "BUY" if total_buy > total_sell else "SELL" if total_sell > total_buy else "WAIT"
    trend_dom = "Bullish" if total_buy > total_sell else "Bearish" if total_sell > total_buy else "Sideways"

    safe_write(ws, r, 1, luna_curenta)
    safe_write(ws, r, 2, round(avg_rsi, 2))
    safe_write(ws, r, 3, sp500.get("inchidere"))
    safe_write(ws, r, 4, sp500.get("inchidere"))
    safe_write(ws, r, 5, "")
    safe_write(ws, r, 6, "")
    safe_write(ws, r, 7, vix_d.get("inchidere"))
    safe_write(ws, r, 8, semnal_luna)
    safe_write(ws, r, 9, "")
    safe_write(ws, r, 10, trend_dom)
    safe_write(ws, r, 11, fear_greed.get("value"))
    safe_write(ws, r, 12, "")
    safe_write(ws, r, 13, all_data.get("^TNX", {}).get("inchidere"))
    safe_write(ws, r, 14, all_data.get("GC=F", {}).get("inchidere"))
    safe_write(ws, r, 15, all_data.get("CL=F", {}).get("inchidere"))
    safe_write(ws, r, 16, all_data.get("EURUSD=X", {}).get("inchidere"))
    safe_write(ws, r, 17, "")

    color_signal(ws.cell(r, 8), semnal_luna)
    color_trend(ws.cell(r, 10), trend_dom)
    log.info(f"  OK ISTORIC TRENDING — adăugat {luna_curenta}")


def get_selected_asset_from_dashboard(wb):
    if "DASHBOARD" not in wb.sheetnames:
        return "S&P 500"
    ws = wb["DASHBOARD"]
    for row in range(1, 12):
        for col in range(1, 16):
            v = ws.cell(row, col).value
            if v and "Activ selectat" in str(v):
                cand = ws.cell(row, min(col + 2, ws.max_column)).value
                if cand:
                    return str(cand).strip()
    return "S&P 500"


def update_dashboard(ws, all_data, fear_greed, activ_sel):
    ticker_sel = ACTIVE.get(activ_sel)
    d = all_data.get(ticker_sel, {}) if ticker_sel else {}

    safe_write(ws, 2, 4, datetime.now().strftime("%d.%m.%Y %H:%M"))

    if not d:
        log.warning("  DASHBOARD fără date pentru activul selectat")
        return

    safe_write(ws, 5, 1, d.get("trend"))
    safe_write(ws, 6, 1, d.get("macross"))
    safe_write(ws, 5, 5, f"ATR: {fmt_price(d.get('atr'))}")
    safe_write(ws, 6, 5, f"BB: {fmt_price(d.get('bb_width'))}")
    safe_write(ws, 5, 9, f"{d.get('rvol', 0):.2f}x medie")
    safe_write(ws, 6, 9, f"Vol: {d.get('volum', 0):,}")
    safe_write(ws, 5, 13, d.get("semnal"))
    safe_write(ws, 6, 13, f"{d.get('confluente', 0)}/5 conf")

    color_trend(ws.cell(5, 1), d.get("trend"))
    color_status_extended(ws.cell(6, 1), d.get("macross"))
    color_signal(ws.cell(5, 13), d.get("semnal"))

    safe_write(ws, 9, 2, d.get("semnal"))
    safe_write(ws, 10, 2, activ_sel)
    safe_write(ws, 11, 2, d.get("inchidere"))
    safe_write(ws, 12, 2, d.get("sl") if d.get("semnal") != "WAIT" else "N/A")
    safe_write(ws, 13, 2, d.get("tp") if d.get("semnal") != "WAIT" else "N/A")
    safe_write(ws, 14, 2, rr_text(d.get("inchidere"), d.get("sl"), d.get("tp")))
    safe_write(ws, 15, 2, f"{d.get('confluente', 0)}/5")
    safe_write(ws, 16, 2, f"{d.get('probabilitate', 0)}%")
    safe_write(ws, 17, 2, f"RSI={d['rsi']:.0f} | {d['macd_cross']} | {d['macross']} | RVOL={d['rvol']:.1f}x")
    color_signal(ws.cell(9, 2), d["semnal"])

    indicators = {
        21: ("RSI(14)", d.get("rsi"), d.get("rsi_status")),
        22: ("Impuls MACD", d.get("macd_cross"), signal_status_text(d.get("semnal"))),
        23: ("MA50 vs MA200", d.get("macross"), signal_status_text(d.get("semnal"))),
        24: ("Trend activ", d.get("trend"), d.get("trend")),
        25: ("VIX (global)", all_data.get("^VIX", {}).get("inchidere", "N/A"), "Negativ" if all_data.get("^VIX", {}).get("inchidere", 0) > 25 else "Neutru"),
        26: ("Fear & Greed", fear_greed.get("display", "N/A"), "BUY" if fear_greed.get("status") == "Pozitiv" else "SELL" if fear_greed.get("status") == "Negativ" else "WAIT"),
        27: ("RVOL activ", f"{d.get('rvol', 0):.2f}x", "Pozitiv" if d.get("rvol", 0) > 1.2 else "Negativ" if d.get("rvol", 1) < 0.7 else "Neutru"),
    }

    for row, (label, val, status) in indicators.items():
        safe_write(ws, row, 1, label)
        safe_write(ws, row, 2, val)
        safe_write(ws, row, 5, status)
        color_status_extended(ws.cell(row, 5), status)

    log.info("  OK DASHBOARD actualizat")


def update_fisa_activ(ws, all_data, activ_selectat):
    ticker_sel = ACTIVE.get(activ_selectat)
    d = all_data.get(ticker_sel, {}) if ticker_sel else {}
    safe_write(ws, 2, 2, activ_selectat)

    if not d:
        log.warning(f"  Nu există date pentru {activ_selectat}")
        return

    labels = {
        "Semnal (BUY/SELL/WAIT)": d["semnal"],
        "Condiție declanșare": f"RSI={d['rsi']:.0f} | {d['macd_cross']} | {d['macross']} | RVOL={d['rvol']:.1f}x",
        "Entry Price": d["inchidere"],
        "Stop Loss (SL)": d["sl"] if d["semnal"] != "WAIT" else "N/A",
        "Take Profit (TP)": d["tp"] if d["semnal"] != "WAIT" else "N/A",
        "Risk/Reward Ratio": rr_text(d["inchidere"], d["sl"], d["tp"]),
        "Status semnal": "Activ",
        "Preț curent": d["inchidere"],
        "MA20": d["ma20"],
        "MA50": d["ma50"],
        "MA200": d["ma200"],
        "RSI(14)": d["rsi"],
        "RSI Status": d["rsi_status"],
        "BB Superior": d["bb_sup"],
        "BB Lățime": d["bb_width"],
        "Stoch %K": d["stoch_k"],
        "Volum": d["volum"],
        "Deschidere": d["deschidere"],
        "Maxim": d["maxim"],
        "Var. Zi (%)": d["var_zi_pct"],
        "Var. Lună (%)": d["var_luna_pct"],
        "Medie Vol.20z": d["avg_vol_20"],
    }

    pairs = {
        "Entry Price": d["confluente"],
        "Stop Loss (SL)": d["tp"] if d["semnal"] != "WAIT" else "N/A",
        "Risk/Reward Ratio": f"{d['probabilitate']}%",
        "Status semnal": d["timestamp"],
        "Preț curent": d["trend"],
        "MA20": d["macross"],
        "MA50": d["macd_cross"],
        "MA200": d["macd_hist"],
        "RSI(14)": d["rsi_status"],
        "RSI Status": d["macd_signal"],
        "BB Superior": d["bb_inf"],
        "BB Lățime": d["atr"],
        "Stoch %K": d["stoch_d"],
        "Volum": d["rvol"],
        "Deschidere": d["inchidere"],
        "Maxim": d["minim"],
        "Var. Zi (%)": d["var_sapt_pct"],
        "Var. Lună (%)": d["semnal"],
        "Medie Vol.20z": d["trend"],
    }

    for row in range(4, 140):
        label = ws.cell(row, 1).value
        if not label:
            continue
        label = str(label).strip()
        if label in labels:
            safe_write(ws, row, 2, labels[label])
        if label in pairs:
            safe_write(ws, row, 5, pairs[label])

    log.info(f"  OK FISA ACTIV pentru {activ_selectat}")


def update_rezumat_executiv(ws, all_data, fear_greed):
    safe_write(ws, 3, 4, datetime.now().strftime("%d.%m.%Y %H:%M"))

    total_buy = sum(1 for d in all_data.values() if d.get("semnal") == "BUY")
    total_sell = sum(1 for d in all_data.values() if d.get("semnal") == "SELL")
    total_wait = sum(1 for d in all_data.values() if d.get("semnal") == "WAIT")
    avg_rsi = sum(d.get("rsi", 50) for d in all_data.values()) / len(all_data) if all_data else 50

    sp500 = all_data.get("^GSPC", {})
    vix_d = all_data.get("^VIX", {})

    trend_general = "Bullish" if total_buy > total_sell else "Bearish" if total_sell > total_buy else "Sideways"
    volatilitate = "Ridicată" if vix_d.get("inchidere", 20) > 25 else "Moderată" if vix_d.get("inchidere", 20) > 15 else "Scăzută"
    volum_tranz = "Crescut" if sp500.get("rvol", 1) > 1.2 else "Scăzut" if sp500.get("rvol", 1) < 0.8 else "Normal"
    risc = "Ridicat" if vix_d.get("inchidere", 20) > 30 else "Moderat" if vix_d.get("inchidere", 20) > 20 else "Scăzut"

    vals = {
        "Tendință generală": (trend_general, f"BUY: {total_buy} | SELL: {total_sell} | WAIT: {total_wait}"),
        "Volatilitate": (volatilitate, f"VIX: {vix_d.get('inchidere', 'N/A')}"),
        "Volum tranzacții": (volum_tranz, f"RVOL S&P500: {sp500.get('rvol', 'N/A')}x"),
        "Sentiment piață": (fear_greed.get("display", "N/A"), f"RSI mediu: {avg_rsi:.1f}"),
        "Risc sistemic": (risc, f"Status Fear&Greed: {fear_greed.get('status', 'Neutru')}"),
    }

    for row in range(6, 20):
        key = ws.cell(row, 1).value
        if key in vals:
            safe_write(ws, row, 2, vals[key][0])
            safe_write(ws, row, 5, vals[key][1])

    log.info("  OK REZUMAT EXECUTIV actualizat")


def ghid_miscare_text(d):
    dir_txt = "a crescut" if d["var_zi_pct"] > 0 else "a scăzut" if d["var_zi_pct"] < 0 else "a închis lateral"
    intens = "ușor" if abs(d["var_zi_pct"]) < 1 else "moderat" if abs(d["var_zi_pct"]) < 2.5 else "semnificativ"
    vol = "excepțional" if d["rvol"] > 1.5 else "normal" if d["rvol"] >= 0.7 else "scăzut"
    return (
        f"Prețul {dir_txt} {intens}. Var zi {d['var_zi_pct']:+.2f}%, săptămână {d['var_sapt_pct']:+.2f}%, "
        f"lună {d['var_luna_pct']:+.2f}%. Volumul este {vol} (RVOL {d['rvol']:.2f}x), RSI indică "
        f"{d['rsi_status']}, iar MACD arată {d['macd_cross']}. Structura MA este {d['macross']}, "
        f"trendul este {d['trend']}."
    )


def ghid_trade_text(d):
    if d["semnal"] == "WAIT":
        return "Nu există semnal complet. Urmărește confirmare prin volum, menținere peste/sub MA50 și impuls MACD."
    rr = rr_text(d["inchidere"], d["sl"], d["tp"])
    warn = " Atenție: short-ul are risc teoretic nelimitat." if d["semnal"] == "SELL" else ""
    return (
        f"Semnal {d['semnal']}. Entry {fmt_price(d['inchidere'])}, SL {fmt_price(d['sl'])}, "
        f"TP {fmt_price(d['tp'])}, RR {rr}, probabilitate {d['probabilitate']}%.{warn}"
    )


def ghid_pattern_text(d):
    pats = []
    if d["macross"] in ("Golden Cross", "Death Cross"):
        pats.append(d["macross"])
    if d["bb_width"] and d["inchidere"] and (d["bb_width"] / d["inchidere"]) < 0.04:
        pats.append("Bollinger Squeeze")
    if d["macd_cross"] in ("Impuls pozitiv nou", "Impuls negativ nou"):
        pats.append(d["macd_cross"])
    if d["rsi"] < 32 and d["var_zi_pct"] < -2 and d["rvol"] > 1.3:
        pats.append("Climax de vânzare")
    if d["rsi"] > 75 and d["var_zi_pct"] > 2 and d["rvol"] > 1.3:
        pats.append("Climax de cumpărare")
    if d["ma50"] and abs(d["inchidere"] - d["ma50"]) / d["ma50"] <= 0.008:
        pats.append("Test MA50")
    return " | ".join(pats) if pats else "Consolidare, monitorizează zilnic."


def ghid_lectie_text(d):
    rr = rr_value(d["inchidere"], d["sl"], d["tp"])
    if d["semnal"] == "WAIT" and d["confluente"] < 2:
        return "Lecția zilei: răbdarea este strategie. Lipsa confluenței înseamnă neacțiune disciplinată."
    if rr is not None and rr < 1.5:
        return "Lecția zilei: RR mic cere win rate mai mare; nu forța setup-uri mediocre."
    if d["rvol"] < 0.7:
        return "Lecția zilei: volumul este filtru de validare; mișcarea fără participare rămâne fragilă."
    if d["rsi"] < 30:
        return "Lecția zilei: presiunea excesivă de vânzare nu înseamnă cumpărare imediată."
    if d["rsi"] > 70:
        return "Lecția zilei: o zonă extinsă poate persista în trend, dar crește riscul de pullback."
    if d["macross"] == "Golden Cross":
        return "Lecția zilei: Golden Cross confirmă, nu anticipează; este indicator tardiv."
    if d["macross"] == "Death Cross":
        return "Lecția zilei: Death Cross apare adesea târziu; evită vânzarea panicată."
    return "Lecția zilei: Loss Aversion distorsionează decizia; execută planul, nu emoția."


def update_ghid_invatre(ws, all_data):
    clear_rows(ws, 2)
    r = 2

    total_buy = sum(1 for d in all_data.values() if d.get("semnal") == "BUY")
    total_sell = sum(1 for d in all_data.values() if d.get("semnal") == "SELL")
    total_wait = sum(1 for d in all_data.values() if d.get("semnal") == "WAIT")

    title = (
        "GHID DE ÎNVĂȚARE ZILNIC — ANALIZĂ DE PIAȚĂ PROFESIONALĂ | "
        f"{datetime.now().strftime('%d.%m.%Y %H:%M')} | BUY={total_buy} | SELL={total_sell} | WAIT={total_wait}"
    )
    safe_write(ws, r, 1, title)
    ws.cell(r, 1).fill = fill("0D2137")
    ws.cell(r, 1).font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    r += 2

    for cat_name, cat_dict in CATEGORII:
        safe_write(ws, r, 1, f"=== {cat_name} ===")
        ws.cell(r, 1).fill = copy(F_HDR)
        ws.cell(r, 1).font = copy(FN_HDR)
        r += 1

        for name, ticker in cat_dict.items():
            d = all_data.get(ticker, {})
            if not d:
                continue

            safe_write(ws, r, 1, f"{name} | Preț {fmt_price(d['inchidere'])} | Var zi {d['var_zi_pct']:+.2f}% | RSI {d['rsi']:.1f} | Conf {d['confluente']}/5 | {d['semnal']}")
            color_signal(ws.cell(r, 1), d["semnal"])
            r += 1

            blocks = [
                ("A. DE CE S-A MIȘCAT ASTĂZI", ghid_miscare_text(d)),
                ("B. OPORTUNITATE DE TRADING", ghid_trade_text(d)),
                ("C. PATTERN GRAFIC DETECTAT", ghid_pattern_text(d)),
                ("D. LECȚIA ZILEI", ghid_lectie_text(d)),
            ]

            for hdr, txt in blocks:
                safe_write(ws, r, 1, hdr)
                ws.cell(r, 1).fill = copy(F_LBL)
                ws.cell(r, 1).font = Font(name="Arial", size=10, bold=True, color="1F4E79")
                r += 1
                safe_write(ws, r, 1, txt)
                ws.cell(r, 1).alignment = Alignment(wrap_text=True, vertical="top")
                r += 2

        r += 1

    safe_write(ws, r, 1, "GHID COMPLET DE CITIRE GRAFICE & INDICATORI TEHNICI")
    ws.cell(r, 1).fill = fill("0A1628")
    ws.cell(r, 1).font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    r += 1

    guide = [
        "1. CANDLESTICK — umbre, hammer, shooting star, doji, engulfing, marubozu.",
        "2. RSI — zone, divergențe, interpretare contextuală, greșeli frecvente.",
        "3. MACD — linie, semnal, histogramă, impuls pozitiv/negativ nou/activ.",
        "4. MEDII MOBILE — MA20/50/200, trend, pullback, întârziere inerentă.",
        "5. BOLLINGER BANDS — bounce, squeeze breakout, volatilitate.",
        "6. VOLUM & RVOL — confirmare, acumulare, distribuție, participare.",
        "7. SUPORT & REZISTENȚĂ — role reversal, reacții, stop-uri plasate logic.",
        "8. MANAGEMENT RISC — regula 1-2%, RR, position sizing, disciplină.",
    ]
    for line in guide:
        safe_write(ws, r, 1, line)
        r += 1

    log.info("  OK GHID INVATARE actualizat")


def main():
    log.info("=" * 70)
    log.info("  PORNIRE ACTUALIZARE AUTOMATA")
    log.info(f"  {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}")
    log.info(
        f"  Active: {len(ACTIVE)} total  "
        f"({len(INDICI)} indici | {len(ACTIUNI)} actiuni | {len(CRYPTO)} crypto | "
        f"{len(VALUTE)} valute | {len(MATERII_PRIME)} materii prime)"
    )
    log.info("=" * 70)

    if not EXCEL_PATH.exists():
        log.error(f"  Fișierul Excel nu există: {EXCEL_PATH}")
        sys.exit(1)

    all_data = {}
    log.info("\n  Preia date pentru toate activele...\n")

    for idx, (name, ticker) in enumerate(ACTIVE.items(), 1):
        d = get_full_data(name, ticker)
        if d:
            all_data[ticker] = d
            log.info(
                f"  [{idx:3}/{len(ACTIVE)}] {name:30} {d['inchidere']:12.4f}  "
                f"({d['var_zi_pct']:+.2f}%)  RSI={d['rsi']:5.1f}  [{d['semnal']}]"
            )
        else:
            log.warning(f"  [{idx:3}/{len(ACTIVE)}] {name:30} -- EROARE / N/A --")

    log.info("\n  Preia Fear & Greed...")
    fear_greed = get_fear_greed()
    log.info(f"  Fear & Greed: {fear_greed['display']}")

    log.info("\n  Preia macro live...")
    macro_live = {}

    for name, ticker in MACRO_TICKERS.items():
        d = get_full_data(name, ticker)
        if d:
            macro_live[name] = d["inchidere"]
            all_data[ticker] = d
            log.info(f"  {name:20} = {d['inchidere']:.4f}")

    for label, ticker in {"EUR/USD": "EURUSD=X", "Oil Brent": "BZ=F", "Oil WTI": "CL=F"}.items():
        d = all_data.get(ticker)
        if d:
            macro_live[label] = d["inchidere"]

    try:
        wb = load_workbook(EXCEL_PATH)

        if "PRETURI VOLUME" in wb.sheetnames:
            update_preturi_volume(wb["PRETURI VOLUME"], all_data)

        if "INDICATORI TEHNICI" in wb.sheetnames:
            update_indicatori_tehnici(wb["INDICATORI TEHNICI"], all_data)

        if "SEMNALE INTRARE" in wb.sheetnames:
            update_semnale(wb["SEMNALE INTRARE"], all_data)

        if "INDICATORI MACRO" in wb.sheetnames:
            update_macro(wb["INDICATORI MACRO"], macro_live, fear_greed)

        if "COMPETITORI SECTOR" in wb.sheetnames:
            update_competitori_sector(wb["COMPETITORI SECTOR"], all_data)

        if "RISCURI OPORTUNITATI" in wb.sheetnames:
            update_riscuri_oportunitati(wb["RISCURI OPORTUNITATI"])

        if "CALENDAR ECONOMIC" in wb.sheetnames:
            update_calendar_economic(wb["CALENDAR ECONOMIC"])

        if "JURNAL TRANZACTII" in wb.sheetnames:
            update_jurnal_stats(wb["JURNAL TRANZACTII"])

        if "ISTORIC TRENDING" in wb.sheetnames:
            update_historic(wb["ISTORIC TRENDING"], all_data, fear_greed)

        if "REZUMAT EXECUTIV" in wb.sheetnames:
            update_rezumat_executiv(wb["REZUMAT EXECUTIV"], all_data, fear_greed)

        activ_sel = get_selected_asset_from_dashboard(wb)

        if "DASHBOARD" in wb.sheetnames:
            update_dashboard(wb["DASHBOARD"], all_data, fear_greed, activ_sel)

        if "FISA ACTIV" in wb.sheetnames:
            update_fisa_activ(wb["FISA ACTIV"], all_data, activ_sel)

        if "GHID INVATARE" in wb.sheetnames:
            update_ghid_invatre(wb["GHID INVATARE"], all_data)

        wb.save(EXCEL_PATH)

        log.info("\n" + "=" * 70)
        log.info(f"  SALVAT LOCAL: {EXCEL_PATH.name}")
        log.info(f"  Procesate: {len(all_data)}/{len(ACTIVE) + len(MACRO_TICKERS)} simboluri cu date")
        log.info(f"  Finalizat: {datetime.now().strftime('%H:%M:%S')}")
        log.info("=" * 70)

    except Exception as e:
        log.error(f"  EROARE FATALA: {e}", exc_info=True)
        sys.exit(1)


if __name__ == "__main__":
    main()