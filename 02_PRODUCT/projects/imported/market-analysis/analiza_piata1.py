# -*- coding: utf-8 -*-
"""
Analiza Piata Profesionala - Script complet local

Rulare zilnic:
    python analiza_piata.py

Dependinte:
    pip install yfinance pandas openpyxl requests

Nota:
- Datele sunt preluate din Yahoo Finance via yfinance.download(period="1y", interval="1d").
- Indicatorii sunt calculati local in pandas.
- Fisierul Excel este generat integral de acest script, pornind de la un template gol.
- Office tinta: Excel 2021 (XLOOKUP disponibil, fara LAMBDA).
"""

from __future__ import annotations

import sys
import math
import time
import glob
import shutil
import logging
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Any

import requests
import pandas as pd
import yfinance as yf

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule


# ============================================================================
# CONFIG
# ============================================================================

# Calea template-ului initial (gol sau aproape gol, scriptul il populeaza)
TEMPLATE_PATH = r"C:\Users\Marius\Desktop\Analiza_Piata_Profesionala.xlsx"

# Parametri yfinance
YF_PERIOD = "1y"
YF_INTERVAL = "1d"

# Cate fisiere cu timestamp pastrezi local
KEEP_LAST_FILES = 30

# Log file
LOG_FILE = "analiza.log"


# ============================================================================
# ACTIVE - LISTE COMPLETE (5 categorii + macro context)
# ============================================================================

ASSETS_INDICI = {
    "S&P 500": "^GSPC",
    "NASDAQ 100": "^NDX",
    "NASDAQ Comp.": "^IXIC",
    "Dow Jones": "^DJI",
    "Russell 2000": "^RUT",
    "DAX Germany": "^GDAXI",
    "FTSE 100": "^FTSE",
    "CAC 40": "^FCHI",
    "Nikkei 225": "^N225",
    "Hang Seng": "^HSI",
    "Shanghai Comp.": "000001.SS",
    "MSCI World ETF": "URTH",
    "MSCI EM ETF": "EEM",
    "BET Romania": "BET.RO",
}

ASSETS_STOCKS_ETF = {
    "Apple": "AAPL",
    "Microsoft": "MSFT",
    "NVIDIA": "NVDA",
    "Alphabet": "GOOGL",
    "Amazon": "AMZN",
    "Meta": "META",
    "Tesla": "TSLA",
    "AMD": "AMD",
    "Intel": "INTC",
    "Broadcom": "AVGO",
    "ASML": "ASML",
    "Taiwan Semi": "TSM",
    "Palantir": "PLTR",
    "Salesforce": "CRM",
    "Oracle": "ORCL",
    "JPMorgan": "JPM",
    "Goldman Sachs": "GS",
    "Berkshire B": "BRK-B",
    "Visa": "V",
    "Mastercard": "MA",
    "ExxonMobil": "XOM",
    "Chevron": "CVX",
    "Shell": "SHEL",
    "Caterpillar": "CAT",
    "Boeing": "BA",
    "SPY": "SPY",
    "QQQ": "QQQ",
    "GLD ETF": "GLD",
    "TLT Bond ETF": "TLT",
    "ARKK": "ARKK",
}

ASSETS_CRYPTO = {
    "Bitcoin": "BTC-USD",
    "Ethereum": "ETH-USD",
    "BNB": "BNB-USD",
    "Solana": "SOL-USD",
    "XRP": "XRP-USD",
    "Cardano": "ADA-USD",
    "Avalanche": "AVAX-USD",
    "Dogecoin": "DOGE-USD",
    "Chainlink": "LINK-USD",
    "Polkadot": "DOT-USD",
    "Litecoin": "LTC-USD",
    "Shiba Inu": "SHIB-USD",
    "Polygon": "MATIC-USD",
    "Uniswap": "UNI-USD",
    "Cosmos": "ATOM-USD",
    "Stellar": "XLM-USD",
    "Monero": "XMR-USD",
    "Tron": "TRX-USD",
    "Filecoin": "FIL-USD",
    "Aave": "AAVE-USD",
    "Arbitrum": "ARB-USD",
    "Optimism": "OP-USD",
    "Render": "RNDR-USD",
    "Sui": "SUI-USD",
    "Near Protocol": "NEAR-USD",
}

ASSETS_FOREX = {
    "EUR/USD": "EURUSD=X",
    "GBP/USD": "GBPUSD=X",
    "USD/JPY": "USDJPY=X",
    "USD/CHF": "USDCHF=X",
    "AUD/USD": "AUDUSD=X",
    "USD/CAD": "USDCAD=X",
    "EUR/RON": "EURRON=X",
    "USD/RON": "USDRON=X",
    "GBP/RON": "GBPRON=X",
    "EUR/GBP": "EURGBP=X",
    "USD/CNY": "USDCNY=X",
    "USD/TRY": "USDTRY=X",
}

ASSETS_COMMODITIES = {
    "Gold": "GC=F",
    "Silver": "SI=F",
    "Platinum": "PL=F",
    "Palladium": "PA=F",
    "Oil WTI": "CL=F",
    "Oil Brent": "BZ=F",
    "Natural Gas": "NG=F",
    "Copper": "HG=F",
    "Corn": "ZC=F",
    "Wheat": "ZW=F",
    "Soybean": "ZS=F",
    "Coffee": "KC=F",
    "Sugar": "SB=F",
    "Cotton": "CT=F",
}

ASSETS_MACRO_CONTEXT = {
    "VIX": "^VIX",
    "Yield 10Y": "^TNX",
    "Yield 2Y": "^IRX",
    "USD Index": "DX-Y.NYB",
}

CATEGORIES = [
    ("INDICI BURSIERI", ASSETS_INDICI),
    ("ACTIUNI & ETF", ASSETS_STOCKS_ETF),
    ("CRYPTO", ASSETS_CRYPTO),
    ("VALUTE FOREX", ASSETS_FOREX),
    ("MATERII PRIME", ASSETS_COMMODITIES),
]

ALL_ASSETS: Dict[str, str] = {}
for cat, d in CATEGORIES:
    ALL_ASSETS.update(d)

ALL_TICKERS = list(ALL_ASSETS.values())


# ============================================================================
# STILURI EXCEL
# ============================================================================

FONT_GLOBAL = Font(name="Arial", size=10)

FILL_HDR = PatternFill("solid", fgColor="1F4E79")
FONT_HDR = Font(name="Arial", size=10, bold=True, color="FFFFFF")
ALIGN_HDR = Alignment(horizontal="center", vertical="center", wrap_text=True)

FILL_ROW_HDR = PatternFill("solid", fgColor="D6DCE4")
FONT_ROW_HDR = Font(name="Arial", size=10, bold=True, color="000000")

FILL_INPUT = PatternFill("solid", fgColor="D9E1F2")
FILL_CALC = PatternFill("solid", fgColor="F2F2F2")
FONT_CALC = Font(name="Arial", size=10, color="595959")

FILL_BUY = PatternFill("solid", fgColor="C6EFCE")
FILL_WAIT = PatternFill("solid", fgColor="FFEB9C")
FILL_SELL = PatternFill("solid", fgColor="FFC7CE")

FILL_TITLE = PatternFill("solid", fgColor="0D2137")
FONT_TITLE = Font(name="Arial", size=14, bold=True, color="FFFFFF")

FILL_CAT = PatternFill("solid", fgColor="1F4E79")
FONT_CAT = Font(name="Arial", size=11, bold=True, color="FFFFFF")

FILL_SELECTOR = PatternFill("solid", fgColor="FFF2CC")
FONT_SELECTOR = Font(name="Arial", size=13, bold=True, color="7D5A00")

FILL_SELECTOR_LBL = PatternFill("solid", fgColor="ED7D31")
FONT_SELECTOR_LBL = Font(name="Arial", size=11, bold=True, color="FFFFFF")

THIN = Side(style="thin", color="9E9E9E")
BORDER_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

TAB_COLORS = {
    "DASHBOARD": "FF1F4E79",
    "REZUMAT EXECUTIV": "FF1F4E79",
    "SEMNALE INTRARE": "FF375623",
    "INDICATORI TEHNICI": "FF595959",
    "INDICATORI MACRO": "FF595959",
    "COMPETITORI SECTOR": "FF595959",
    "PRETURI VOLUME": "FF595959",
    "RISCURI OPORTUNITATI": "FF9C0006",
    "CALENDAR ECONOMIC": "FF595959",
    "JURNAL TRANZACTII": "FF595959",
    "ISTORIC TRENDING": "FF595959",
    "GHID INVATARE": "FF154360",
    "LEGENDA": "FF595959",
    "FISA ACTIV": "FFED7D31",
    "LIST_ACTIVE": "FF595959",
}


# ============================================================================
# UTILE
# ============================================================================

def setup_logging() -> logging.Logger:
    logger = logging.getLogger("analiza_piata")
    logger.setLevel(logging.INFO)
    logger.handlers.clear()

    fmt = logging.Formatter("%(asctime)s  %(levelname)s  %(message)s", datefmt="%H:%M:%S")
    sh = logging.StreamHandler(sys.stdout)
    sh.setFormatter(fmt)
    logger.addHandler(sh)

    fh = logging.FileHandler(LOG_FILE, encoding="utf-8")
    fh.setFormatter(fmt)
    logger.addHandler(fh)

    return logger


def safe_float(x: Any) -> Optional[float]:
    try:
        if x is None:
            return None
        v = float(x)
        if math.isnan(v) or math.isinf(v):
            return None
        return v
    except Exception:
        return None


def fp(x: Any, nd: int = 4) -> str:
    v = safe_float(x)
    if v is None:
        return ""
    return f"{v:.{nd}f}"


def ensure_dir(p: Path) -> None:
    p.mkdir(parents=True, exist_ok=True)


def unmerge_all(ws) -> None:
    merged = list(ws.merged_cells.ranges)
    for r in merged:
        try:
            ws.unmerge_cells(str(r))
        except Exception:
            pass


def apply_cell_style(cell, fill=None, font=None, align=None, border=True):
    if fill is not None:
        cell.fill = fill
    if font is not None:
        cell.font = font
    if align is not None:
        cell.alignment = align
    if border:
        cell.border = BORDER_THIN


def write_row(ws, row: int, values: List[Any], col_start: int = 1):
    for i, v in enumerate(values):
        c = ws.cell(row=row, column=col_start + i)
        c.value = v if v is not None else ""
        c.font = FONT_GLOBAL
        c.border = BORDER_THIN


def set_cols(ws, widths: Dict[int, float]):
    for idx, w in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = w


def add_table_header(ws, row: int, headers: List[str], col_start: int = 1):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=col_start + i)
        c.value = h
        apply_cell_style(c, fill=FILL_HDR, font=FONT_HDR, align=ALIGN_HDR)
    ws.row_dimensions[row].height = 22


def style_category_row(ws, row: int, col_start: int, col_end: int, text: str):
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=col_start)
    cell.value = f"— {text} —"
    apply_cell_style(
        cell,
        fill=FILL_CAT,
        font=FONT_CAT,
        align=Alignment(horizontal="left", vertical="center", indent=1),
    )
    for c in range(col_start + 1, col_end + 1):
        apply_cell_style(ws.cell(row=row, column=c), fill=FILL_CAT, font=FONT_CAT)
    ws.row_dimensions[row].height = 20


def set_tab_colors(wb: Workbook) -> None:
    for sn in wb.sheetnames:
        ws = wb[sn]
        color = TAB_COLORS.get(sn)
        if color:
            ws.sheet_properties.tabColor = color


# ============================================================================
# FETCH DATE
# ============================================================================

def yfinance_bulk_download(tickers: List[str], logger: logging.Logger) -> pd.DataFrame:
    logger.info(f"yfinance.download pentru {len(tickers)} active...")
    df = yf.download(
        tickers=tickers,
        period=YF_PERIOD,
        interval=YF_INTERVAL,
        auto_adjust=False,
        group_by="ticker",
        threads=True,
        progress=False,
    )
    if df is None or df.empty:
        raise ValueError("Nu s-au primit date de la Yahoo Finance")
    return df


def extract_ticker_ohlcv(df_bulk: pd.DataFrame, ticker: str) -> Optional[pd.DataFrame]:
    if df_bulk is None or df_bulk.empty:
        return None
    if isinstance(df_bulk.columns, pd.MultiIndex):
        if ticker not in df_bulk.columns.get_level_values(0):
            return None
        sub = df_bulk[ticker].copy()
    else:
        sub = df_bulk.copy()

    sub = sub.rename(
        columns={
            "Open": "open",
            "High": "high",
            "Low": "low",
            "Close": "close",
            "Adj Close": "adj_close",
            "Volume": "volume",
        }
    )
    if not {"open", "high", "low", "close", "volume"}.issubset(sub.columns):
        return None
    sub = sub.dropna(subset=["close"]).copy()
    if len(sub) < 5:
        return None
    sub.index = pd.to_datetime(sub.index)
    sub.sort_index(inplace=True)
    return sub


def get_fear_greed(logger: logging.Logger) -> Dict[str, Any]:
    url = "https://api.alternative.me/fng/?limit=2"
    try:
        r = requests.get(url, timeout=10)
        r.raise_for_status()
        j = r.json()
        data = j.get("data", [])
        cur = data[0] if data else {}
        prev = data[1] if len(data) > 1 else {}
        return {
            "value": safe_float(cur.get("value")),
            "classification": cur.get("value_classification", ""),
            "timestamp": cur.get("timestamp", ""),
            "prev_value": safe_float(prev.get("value")),
            "prev_timestamp": prev.get("timestamp", ""),
        }
    except Exception as e:
        logger.warning(f"Fear&Greed fetch failed: {e}")
        return {
            "value": None,
            "classification": "",
            "timestamp": "",
            "prev_value": None,
            "prev_timestamp": "",
        }


def get_macro_yfinance(logger: logging.Logger) -> Dict[str, float]:
    out: Dict[str, float] = {}
    try:
        df = yfinance_bulk_download(list(ASSETS_MACRO_CONTEXT.values()), logger)
        for name, tick in ASSETS_MACRO_CONTEXT.items():
            sub = extract_ticker_ohlcv(df, tick)
            if sub is None:
                continue
            out[name] = safe_float(sub["close"].iloc[-1])
    except Exception as e:
        logger.warning(f"Macro yfinance failed: {e}")
    return out


# ============================================================================
# INDICATORI
# ============================================================================

def calc_sma(series: pd.Series, n: int) -> pd.Series:
    return series.rolling(n, min_periods=n).mean()


def calc_ema(series: pd.Series, n: int) -> pd.Series:
    return series.ewm(span=n, adjust=False, min_periods=n).mean()


def calc_rsi(close: pd.Series, n: int = 14) -> pd.Series:
    delta = close.diff()
    gain = delta.where(delta > 0, 0.0)
    loss = (-delta).where(delta < 0, 0.0)
    avg_gain = gain.ewm(alpha=1 / n, adjust=False, min_periods=n).mean()
    avg_loss = loss.ewm(alpha=1 / n, adjust=False, min_periods=n).mean()
    rs = avg_gain / avg_loss.replace(0, pd.NA)
    rsi = 100 - (100 / (1 + rs))
    return rsi.astype("float")


def rsi_status(rsi_val: Optional[float]) -> str:
    v = safe_float(rsi_val)
    if v is None:
        return ""
    if v < 30:
        return "Presiune excesivă vânzare"
    if v < 45:
        return "Presiune moderată vânzare"
    if v < 55:
        return "Echilibru"
    if v <= 70:
        return "Momentum ascendent"
    return "Presiune excesivă cumpărare"


def calc_macd(close: pd.Series, fast: int = 12, slow: int = 26, signal: int = 9) -> Tuple[pd.Series, pd.Series, pd.Series]:
    ema_fast = calc_ema(close, fast)
    ema_slow = calc_ema(close, slow)
    macd = ema_fast - ema_slow
    sig = calc_ema(macd, signal)
    hist = macd - sig
    return macd, sig, hist


def macd_impuls(macd: Optional[float], sig: Optional[float], hist: Optional[float], prev_hist: Optional[float]) -> str:
    m = safe_float(macd)
    s = safe_float(sig)
    h = safe_float(hist)
    ph = safe_float(prev_hist)
    if m is None or s is None or h is None or ph is None:
        return ""
    if m > s and ph < 0 and h >= 0:
        return "Impuls pozitiv nou"
    if m > s and h > 0:
        return "Impuls pozitiv activ"
    if m < s and ph > 0 and h <= 0:
        return "Impuls negativ nou"
    if m < s and h < 0:
        return "Impuls negativ activ"
    return ""


def ma_cross(ma50: Optional[float], ma200: Optional[float]) -> str:
    a = safe_float(ma50)
    b = safe_float(ma200)
    if a is None or b is None:
        return "Neutru"
    if a > b:
        return "Golden Cross"
    if a < b:
        return "Death Cross"
    return "Neutru"


def calc_bbands(close: pd.Series, n: int = 20, k: float = 2.0) -> Tuple[pd.Series, pd.Series, pd.Series]:
    ma = close.rolling(n, min_periods=n).mean()
    sd = close.rolling(n, min_periods=n).std(ddof=0)
    upper = ma + k * sd
    lower = ma - k * sd
    width = (upper - lower) / ma.replace(0, pd.NA)
    return upper, lower, width


def calc_true_range(df: pd.DataFrame) -> pd.Series:
    prev_close = df["close"].shift(1)
    tr1 = df["high"] - df["low"]
    tr2 = (df["high"] - prev_close).abs()
    tr3 = (df["low"] - prev_close).abs()
    tr = pd.concat([tr1, tr2, tr3], axis=1).max(axis=1)
    return tr


def calc_atr(df: pd.DataFrame, n: int = 14) -> pd.Series:
    tr = calc_true_range(df)
    atr = tr.ewm(alpha=1 / n, adjust=False, min_periods=n).mean()
    return atr


def calc_stoch(df: pd.DataFrame, n: int = 14, d: int = 3) -> Tuple[pd.Series, pd.Series]:
    low_n = df["low"].rolling(n, min_periods=n).min()
    high_n = df["high"].rolling(n, min_periods=n).max()
    k = 100 * (df["close"] - low_n) / (high_n - low_n).replace(0, pd.NA)
    dline = k.rolling(d, min_periods=d).mean()
    return k, dline


def calc_rvol(volume: pd.Series, n: int = 20) -> pd.Series:
    vol_ma = volume.rolling(n, min_periods=n).mean()
    return volume / vol_ma.replace(0, pd.NA)


def calc_momentum(close: pd.Series, n: int = 10) -> pd.Series:
    return (close / close.shift(n) - 1.0) * 100.0


def classify_trend(price: Optional[float], ma50v: Optional[float]) -> str:
    p = safe_float(price)
    m = safe_float(ma50v)
    if p is None or m is None:
        return "Sideways"
    if p > m * 1.01:
        return "Bullish"
    if p < m * 0.99:
        return "Bearish"
    return "Sideways"


def score_signal(rsi_v: Optional[float], macd_imp: str, ma_cross_v: str, rvol_v: Optional[float]) -> int:
    score = 0
    r = safe_float(rsi_v)
    rv = safe_float(rvol_v)

    if r is not None:
        if r < 35:
            score += 2
        elif r < 45:
            score += 1
        elif r > 75:
            score -= 2
        elif r > 65:
            score -= 1

    if macd_imp == "Impuls pozitiv nou":
        score += 2
    elif macd_imp == "Impuls pozitiv activ":
        score += 1
    elif macd_imp == "Impuls negativ nou":
        score -= 2
    elif macd_imp == "Impuls negativ activ":
        score -= 1

    if ma_cross_v == "Golden Cross":
        score += 2
    elif ma_cross_v == "Death Cross":
        score -= 2

    if rv is not None:
        if rv > 1.5:
            score += 1
        elif rv < 0.6:
            score -= 1

    return score


def signal_from_score(score: int) -> str:
    if score >= 3:
        return "BUY"
    if score <= -3:
        return "SELL"
    return "WAIT"


def confl_from_score(score: int) -> int:
    return min(abs(int(score)), 5)


def sl_tp(entry: Optional[float], atr: Optional[float], side: str) -> Tuple[Optional[float], Optional[float]]:
    e = safe_float(entry)
    a = safe_float(atr)
    if e is None or a is None:
        return None, None
    if side == "BUY":
        return e - 1.5 * a, e + 3.0 * a
    if side == "SELL":
        return e + 1.5 * a, e - 3.0 * a
    return None, None


def probability(conf: int, rvol_v: Optional[float]) -> int:
    rv = safe_float(rvol_v)
    base = 35 + conf * 10
    if rv is not None and rv > 1.2:
        base += 5
    return int(min(90, base))


def key_levels(df: pd.DataFrame, n: int = 60) -> Tuple[Optional[float], Optional[float]]:
    if df is None or df.empty or len(df) < 10:
        return None, None
    low = safe_float(df["low"].tail(n).min())
    high = safe_float(df["high"].tail(n).max())
    return low, high


def compute_changes(close: pd.Series) -> Tuple[Optional[float], Optional[float], Optional[float]]:
    if close is None or len(close) < 2:
        return None, None, None
    last = close.iloc[-1]
    day = (last / close.iloc[-2] - 1) * 100.0 if len(close) >= 2 else None
    wk = (last / close.iloc[-6] - 1) * 100.0 if len(close) >= 6 else None
    mo = (last / close.iloc[-22] - 1) * 100.0 if len(close) >= 22 else None
    return safe_float(day), safe_float(wk), safe_float(mo)


# ============================================================================
# MODEL DE DATE INTERN
# ============================================================================

@dataclass
class AssetResult:
    date: str
    category: str
    name: str
    ticker: str
    open: Optional[float]
    high: Optional[float]
    low: Optional[float]
    close: Optional[float]
    vol: Optional[float]
    chg_day_pct: Optional[float]
    chg_wk_pct: Optional[float]
    chg_mo_pct: Optional[float]

    ma20: Optional[float]
    ma50: Optional[float]
    ma200: Optional[float]
    rsi14: Optional[float]
    rsi_status: str

    macd: Optional[float]
    macd_signal: Optional[float]
    macd_hist: Optional[float]
    macd_impuls: str

    bb_upper: Optional[float]
    bb_lower: Optional[float]
    bb_width: Optional[float]

    atr14: Optional[float]
    stoch_k: Optional[float]
    stoch_d: Optional[float]

    rvol: Optional[float]
    momentum10: Optional[float]

    trend: str
    support: Optional[float]
    resistance: Optional[float]
    ma_cross: str

    score: int
    signal: str
    condition: str
    confl: int
    entry: Optional[float]
    sl: Optional[float]
    tp: Optional[float]
    prob: Optional[int]
    status: str
    note: str


def analyze_one_asset(cat: str, name: str, ticker: str, df: pd.DataFrame, logger: logging.Logger) -> AssetResult:
    d = df.copy()
    close = d["close"]

    ma20 = calc_sma(close, 20)
    ma50 = calc_sma(close, 50)
    ma200 = calc_sma(close, 200)

    rsi = calc_rsi(close, 14)
    macd, sig, hist = calc_macd(close, 12, 26, 9)

    bb_u, bb_l, bb_w = calc_bbands(close, 20, 2.0)
    atr = calc_atr(d, 14)
    st_k, st_d = calc_stoch(d, 14, 3)
    rvol = calc_rvol(d["volume"], 20)
    mom10 = calc_momentum(close, 10)

    day, wk, mo = compute_changes(close)

    last = d.iloc[-1]
    prev_hist = hist.iloc[-2] if len(hist) >= 2 else None

    last_close = safe_float(last["close"])
    last_open = safe_float(last.get("open"))
    last_high = safe_float(last.get("high"))
    last_low = safe_float(last.get("low"))
    last_vol = safe_float(last.get("volume"))

    last_ma20 = safe_float(ma20.iloc[-1])
    last_ma50 = safe_float(ma50.iloc[-1])
    last_ma200 = safe_float(ma200.iloc[-1])

    last_rsi = safe_float(rsi.iloc[-1])
    rsi_st = rsi_status(last_rsi)

    last_macd = safe_float(macd.iloc[-1])
    last_sig = safe_float(sig.iloc[-1])
    last_hist = safe_float(hist.iloc[-1])
    macd_imp = macd_impuls(last_macd, last_sig, last_hist, safe_float(prev_hist))

    last_bbu = safe_float(bb_u.iloc[-1])
    last_bbl = safe_float(bb_l.iloc[-1])
    last_bbw = safe_float(bb_w.iloc[-1])

    last_atr = safe_float(atr.iloc[-1])
    last_k = safe_float(st_k.iloc[-1])
    last_d = safe_float(st_d.iloc[-1])

    last_rvol = safe_float(rvol.iloc[-1])
    last_mom10 = safe_float(mom10.iloc[-1])

    tr = classify_trend(last_close, last_ma50)
    sup, res = key_levels(d, 60)
    cross = ma_cross(last_ma50, last_ma200)

    sc = score_signal(last_rsi, macd_imp, cross, last_rvol)
    sig_txt = signal_from_score(sc)
    conf = confl_from_score(sc)

    entry = last_close
    slv, tpv = sl_tp(entry, last_atr, sig_txt)
    prob = probability(conf, last_rvol)

    if sig_txt == "BUY":
        cond = "Scor ≥ 3 (confluențe tehnice pozitive)"
    elif sig_txt == "SELL":
        cond = "Scor ≤ -3 (confluențe tehnice negative)"
    else:
        cond = f"Confluențe insuficiente ({conf}/5) sau semnale mixte"

    status = "Activ"
    note = f"Auto-generat {datetime.now().strftime('%H:%M')}"

    return AssetResult(
        date=datetime.now().strftime("%d.%m.%Y"),
        category=cat,
        name=name,
        ticker=ticker,
        open=last_open,
        high=last_high,
        low=last_low,
        close=last_close,
        vol=last_vol,
        chg_day_pct=day,
        chg_wk_pct=wk,
        chg_mo_pct=mo,
        ma20=last_ma20,
        ma50=last_ma50,
        ma200=last_ma200,
        rsi14=last_rsi,
        rsi_status=rsi_st,
        macd=last_macd,
        macd_signal=last_sig,
        macd_hist=last_hist,
        macd_impuls=macd_imp,
        bb_upper=last_bbu,
        bb_lower=last_bbl,
        bb_width=last_bbw,
        atr14=last_atr,
        stoch_k=last_k,
        stoch_d=last_d,
        rvol=last_rvol,
        momentum10=last_mom10,
        trend=tr,
        support=sup,
        resistance=res,
        ma_cross=cross,
        score=sc,
        signal=sig_txt,
        condition=cond,
        confl=conf,
        entry=entry,
        sl=slv,
        tp=tpv,
        prob=prob,
        status=status,
        note=note,
    )


def analyze_all_assets(df_bulk: pd.DataFrame, logger: logging.Logger) -> List[AssetResult]:
    results: List[AssetResult] = []
    total = sum(len(d) for _, d in CATEGORIES)
    idx = 0
    for cat, dct in CATEGORIES:
        for name, ticker in dct.items():
            idx += 1
            sub = extract_ticker_ohlcv(df_bulk, ticker)
            if sub is None:
                logger.warning(f"[{idx}/{total}] Skip {name} ({ticker}) — date insuficiente")
                continue
            try:
                res = analyze_one_asset(cat, name, ticker, sub, logger)
                results.append(res)
                logger.info(
                    f"[{idx}/{total}] {name} {ticker} | close={fp(res.close,4)} | day={fp(res.chg_day_pct,2)}% | RSI={fp(res.rsi14,1)} | {res.signal}"
                )
            except Exception as e:
                logger.warning(f"[{idx}/{total}] Eroare {name} ({ticker}): {e}")
    return results


# ============================================================================
# TEMPLATE EXCEL
# ============================================================================

def create_template(path: Path, logger: logging.Logger) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    sheet_names = [
        "DASHBOARD",
        "REZUMAT EXECUTIV",
        "SEMNALE INTRARE",
        "INDICATORI TEHNICI",
        "INDICATORI MACRO",
        "COMPETITORI SECTOR",
        "PRETURI VOLUME",
        "RISCURI OPORTUNITATI",
        "CALENDAR ECONOMIC",
        "JURNAL TRANZACTII",
        "ISTORIC TRENDING",
        "GHID INVATARE",
        "LEGENDA",
        "FISA ACTIV",
        "LIST_ACTIVE",
    ]
    for sn in sheet_names:
        wb.create_sheet(sn)
    wb["LIST_ACTIVE"].sheet_state = "hidden"
    set_tab_colors(wb)
    wb.save(path)
    logger.info(f"Template creat: {path}")


def ensure_template_exists(logger: logging.Logger) -> Path:
    p = Path(TEMPLATE_PATH)
    if p.exists():
        return p
    ensure_dir(p.parent)
    create_template(p, logger)
    return p


def copy_template_with_timestamp(template_path: Path, logger: logging.Logger) -> Path:
    folder = template_path.parent
    ts = datetime.now().strftime("%Y-%m-%d_%H-%M")
    out_path = folder / f"Analiza_Piata_{ts}.xlsx"
    shutil.copy2(str(template_path), str(out_path))
    logger.info(f"Output workbook: {out_path}")
    return out_path


def cleanup_old_files(folder: Path, logger: logging.Logger) -> None:
    files = sorted(folder.glob("Analiza_Piata_*.xlsx"), key=lambda p: p.stat().st_mtime)
    if len(files) <= KEEP_LAST_FILES:
        return
    for p in files[:-KEEP_LAST_FILES]:
        try:
            p.unlink()
            logger.info(f"Deleted old file: {p.name}")
        except Exception as e:
            logger.warning(f"Cleanup failed {p}: {e}")


# ============================================================================
# POPULARE EXCEL - SHEET-URI
# ============================================================================

def dv_list_active(wb: Workbook) -> None:
    ws = wb["LIST_ACTIVE"]
    ws.sheet_state = "hidden"
    ws.delete_rows(1, ws.max_row)
    for i, name in enumerate(ALL_ASSETS.keys(), start=1):
        ws.cell(row=i, column=1).value = name
    ws.column_dimensions["A"].width = 28

    start = 1
    end = len(ALL_ASSETS)
    rng = f"LIST_ACTIVE!$A${start}:$A${end}"
    dv = DataValidation(type="list", formula1=f"={rng}", allow_blank=False)
    dv.error = "Selecteaza un activ din lista."
    dv.errorTitle = "Input invalid"

    ws_dash = wb["DASHBOARD"]
    ws_dash.add_data_validation(dv)
    dv.add(ws_dash["I2"])


def clear_sheet(ws):
    ws.delete_rows(1, ws.max_row)
    ws.delete_cols(1, ws.max_column)
    unmerge_all(ws)


def upd_preturi_volume(wb: Workbook, results: List[AssetResult]) -> None:
    ws = wb["PRETURI VOLUME"]
    clear_sheet(ws)

    headers = [
        "Data",
        "Activ",
        "Deschidere",
        "Maxim",
        "Minim",
        "Inchidere",
        "Var Zi (%)",
        "Var Sapt (%)",
        "Var Luna (%)",
        "Volum",
        "Medie Vol 20z",
        "RVOL",
        "Sezon",
        "Factor sezonier (%)",
        "Trend",
    ]
    add_table_header(ws, 1, headers)
    ws.freeze_panes = "A2"

    row = 2
    col_end = len(headers)
    for cat, _d in CATEGORIES:
        style_category_row(ws, row, 1, col_end, cat)
        row += 1
        for r in [x for x in results if x.category == cat]:
            vals = [
                r.date,
                r.name,
                r.open,
                r.high,
                r.low,
                r.close,
                r.chg_day_pct,
                r.chg_wk_pct,
                r.chg_mo_pct,
                r.vol,
                "",
                r.rvol,
                "",
                "",
                r.trend,
            ]
            write_row(ws, row, vals)
            ws.row_dimensions[row].height = 18
            row += 1

    set_cols(
        ws,
        {
            1: 12,
            2: 22,
            3: 12,
            4: 12,
            5: 12,
            6: 12,
            7: 12,
            8: 12,
            9: 12,
            10: 16,
            11: 16,
            12: 10,
            13: 12,
            14: 18,
            15: 12,
        },
    )


def upd_indicatori_tehnici(wb: Workbook, results: List[AssetResult]) -> None:
    ws = wb["INDICATORI TEHNICI"]
    clear_sheet(ws)

    headers = [
        "Data",
        "Activ",
        "Pret",
        "MA20",
        "MA50",
        "MA200",
        "RSI(14)",
        "RSI Status",
        "MACD",
        "MACD Signal",
        "MACD Hist",
        "BB Superior",
        "BB Inferior",
        "BB Latime",
        "ATR",
        "Stoch %K",
        "Stoch %D",
        "Volum",
        "RVOL",
        "Trend",
        "Suport cheie",
        "Rezistenta",
        "MA Cross",
    ]
    add_table_header(ws, 1, headers)
    ws.freeze_panes = "A2"

    row = 2
    col_end = len(headers)
    for cat, _d in CATEGORIES:
        style_category_row(ws, row, 1, col_end, cat)
        row += 1
        for r in [x for x in results if x.category == cat]:
            vals = [
                r.date,
                r.name,
                r.close,
                r.ma20,
                r.ma50,
                r.ma200,
                r.rsi14,
                r.rsi_status,
                r.macd,
                r.macd_signal,
                r.macd_hist,
                r.bb_upper,
                r.bb_lower,
                r.bb_width,
                r.atr14,
                r.stoch_k,
                r.stoch_d,
                r.vol,
                r.rvol,
                r.trend,
                r.support,
                r.resistance,
                r.ma_cross,
            ]
            write_row(ws, row, vals)
            ws.row_dimensions[row].height = 18
            row += 1

    set_cols(
        ws,
        {
            1: 12,
            2: 22,
            3: 12,
            4: 12,
            5: 12,
            6: 12,
            7: 10,
            8: 24,
            9: 12,
            10: 12,
            11: 12,
            12: 14,
            13: 14,
            14: 12,
            15: 12,
            16: 10,
            17: 10,
            18: 16,
            19: 10,
            20: 12,
            21: 14,
            22: 14,
            23: 12,
        },
    )


def upd_semnale_intrare(wb: Workbook, results: List[AssetResult]) -> None:
    ws = wb["SEMNALE INTRARE"]
    clear_sheet(ws)

    headers = [
        "Data",
        "Activ",
        "Semnal",
        "Conditie",
        "RSI(14)",
        "Impuls MACD",
        "MA Cross",
        "Volum vs Medie",
        "Momentum 10z",
        "Confluente",
        "Entry",
        "Stop Loss",
        "Take Profit",
        "RR Ratio",
        "Probabilitate",
        "Status",
        "Note",
    ]
    add_table_header(ws, 1, headers)
    ws.freeze_panes = "A2"

    row = 2
    col_end = len(headers)
    for cat, _d in CATEGORIES:
        style_category_row(ws, row, 1, col_end, cat)
        row += 1
        for r in [x for x in results if x.category == cat]:
            rr_formula = f'=IFERROR((M{row}-K{row})/(K{row}-L{row}),"N/A")'
            vals = [
                r.date,
                r.name,
                r.signal,
                r.condition,
                r.rsi14,
                r.macd_impuls,
                r.ma_cross,
                "",
                r.momentum10,
                r.confl,
                r.entry,
                r.sl,
                r.tp,
                rr_formula,
                (r.prob / 100.0 if r.prob is not None else ""),
                r.status,
                r.note,
            ]
            write_row(ws, row, vals)
            ws.cell(row=row, column=15).number_format = "0%"
            ws.row_dimensions[row].height = 18
            row += 1

    set_cols(
        ws,
        {
            1: 12,
            2: 22,
            3: 10,
            4: 36,
            5: 10,
            6: 18,
            7: 14,
            8: 14,
            9: 12,
            10: 12,
            11: 12,
            12: 12,
            13: 12,
            14: 10,
            15: 14,
            16: 10,
            17: 18,
        },
    )


def upd_indicatori_macro(wb: Workbook, fg: Dict[str, Any], macro_yf: Dict[str, float]) -> None:
    ws = wb["INDICATORI MACRO"]
    clear_sheet(ws)

    headers = [
        "Indicator",
        "Valoare curenta",
        "Valoare anterioara",
        "Delta Abs",
        "Delta %",
        "Consens",
        "Dev vs Estimare",
        "Impact piata",
        "Data publicare",
        "Frecventa",
        "Trending",
        "Status",
        "Note",
    ]
    add_table_header(ws, 1, headers)
    ws.freeze_panes = "A2"

    rows = []

    def add(ind, cur, prev, freq="Zilnic", impact="Neutru", trending="→", status="Neutru", note=""):
        curv = safe_float(cur)
        prevv = safe_float(prev)
        da = (curv - prevv) if (curv is not None and prevv is not None) else None
        dp = (da / prevv * 100.0) if (da is not None and prevv not in (None, 0)) else None
        rows.append(
            [
                ind,
                curv,
                prevv,
                da,
                dp,
                "",
                "",
                impact,
                datetime.now().strftime("%d.%m.%Y %H:%M"),
                freq,
                trending,
                status,
                note,
            ]
        )

    add("USD Index (DXY)", macro_yf.get("USD Index"), None, status="Neutru", trending="↑")
    add("Yield 10Y US (%)", macro_yf.get("Yield 10Y"), None, status="Neutru", trending="↑")
    add("Yield 2Y US (%)", macro_yf.get("Yield 2Y"), None, status="Neutru", trending="→")
    add("VIX", macro_yf.get("VIX"), None, status="Neutru", trending="↓")
    add(
        "Fear & Greed (0-100)",
        fg.get("value"),
        fg.get("prev_value"),
        freq="Zilnic",
        status="Neutru",
        trending="↑",
        note=fg.get("classification", ""),
    )

    row = 2
    for r in rows:
        write_row(ws, row, r)
        ws.row_dimensions[row].height = 18
        row += 1

    set_cols(
        ws,
        {
            1: 26,
            2: 16,
            3: 16,
            4: 12,
            5: 12,
            6: 12,
            7: 16,
            8: 14,
            9: 18,
            10: 12,
            11: 10,
            12: 12,
            13: 30,
        },
    )


def upd_dashboard(wb: Workbook) -> None:
    ws = wb["DASHBOARD"]
    clear_sheet(ws)

    ws.merge_cells("A1:P1")
    ws["A1"] = "DASHBOARD — ANALIZA DE PIATA PROFESIONALA"
    apply_cell_style(
        ws["A1"],
        fill=FILL_TITLE,
        font=FONT_TITLE,
        align=Alignment(horizontal="left", vertical="center"),
    )
    ws.row_dimensions[1].height = 30

    ws["A2"] = "Data actualizare:"
    ws["B2"] = datetime.now().strftime("%d.%m.%Y %H:%M")

    ws["H2"] = "Activ selectat:"
    apply_cell_style(
        ws["H2"],
        fill=FILL_SELECTOR_LBL,
        font=FONT_SELECTOR_LBL,
        align=Alignment(horizontal="center", vertical="center"),
    )

    ws["I2"] = list(ALL_ASSETS.keys())[0] if ALL_ASSETS else ""
    apply_cell_style(
        ws["I2"],
        fill=FILL_SELECTOR,
        font=FONT_SELECTOR,
        align=Alignment(horizontal="left", vertical="center"),
    )

    ws.row_dimensions[2].height = 22

    set_cols(
        ws,
        {
            1: 14,
            2: 14,
            3: 12,
            4: 12,
            5: 12,
            6: 12,
            7: 12,
            8: 14,
            9: 22,
            10: 12,
            11: 12,
            12: 12,
            13: 12,
            14: 12,
            15: 12,
            16: 12,
        },
    )


def upd_fisa_activ(wb: Workbook) -> None:
    ws = wb["FISA ACTIV"]
    clear_sheet(ws)

    ws.freeze_panes = "A3"

    ws.merge_cells("A1:H1")
    ws["A1"] = "FISA COMPLETA ACTIV — date integrate"
    apply_cell_style(
        ws["A1"],
        fill=FILL_TITLE,
        font=FONT_TITLE,
        align=Alignment(horizontal="left", vertical="center"),
    )
    ws.row_dimensions[1].height = 28

    ws["A2"] = "Activ analizat:"
    ws["B2"] = "=DASHBOARD!I2"
    apply_cell_style(ws["A2"], fill=FILL_SELECTOR_LBL, font=FONT_SELECTOR_LBL)
    apply_cell_style(ws["B2"], fill=FILL_SELECTOR, font=FONT_SELECTOR)
    ws.row_dimensions[2].height = 22


def upd_legenda(wb: Workbook) -> None:
    ws = wb["LEGENDA"]
    clear_sheet(ws)

    ws.merge_cells("A1:H1")
    ws["A1"] = "LEGENDA — CULORI & ABREVIERI"
    apply_cell_style(
        ws["A1"],
        fill=FILL_TITLE,
        font=FONT_TITLE,
        align=Alignment(horizontal="left", vertical="center"),
    )
    ws.row_dimensions[1].height = 28


def upd_rezumat_executiv(wb: Workbook, results: List[AssetResult], fg: Dict[str, Any]) -> None:
    ws = wb["REZUMAT EXECUTIV"]
    clear_sheet(ws)

    ws.merge_cells("A1:H1")
    ws["A1"] = "REZUMAT EXECUTIV — VIZIUNE DE ANSAMBLU"
    apply_cell_style(
        ws["A1"],
        fill=FILL_TITLE,
        font=FONT_TITLE,
        align=Alignment(horizontal="left", vertical="center"),
    )
    ws.row_dimensions[1].height = 28

    ws["A2"] = "Activ selectat:"
    ws["B2"] = "=DASHBOARD!I2"

    buys = sum(1 for r in results if r.signal == "BUY")
    sells = sum(1 for r in results if r.signal == "SELL")
    waits = sum(1 for r in results if r.signal == "WAIT")
    total = max(1, len(results))

    tendinta = "Neutra"
    if buys / total > 0.5:
        tendinta = "Dominanta BUY"
    elif sells / total > 0.5:
        tendinta = "Dominanta SELL"

    r1 = ["Indicator", "Valoare", "", "Trending", "Observatii", "", "", ""]
    add_table_header(ws, 4, r1)
    ws["A5"] = "Tendinta generala"
    ws["B5"] = tendinta
    ws["D5"] = ""
    ws["E5"] = f"BUY={buys} / SELL={sells} / WAIT={waits}"

    ws["A6"] = "Sentiment (Fear & Greed)"
    ws["B6"] = fg.get("value")
    ws["D6"] = fg.get("classification", "")

    ws["A7"] = "Risc sistemic"
    ws["B7"] = "Moderate"

    set_cols(
        ws,
        {1: 22, 2: 24, 3: 8, 4: 14, 5: 36, 6: 8, 7: 8, 8: 8},
    )


def upd_ghid_invat(wb: Workbook, results: List[AssetResult]) -> None:
    ws = wb["GHID INVATARE"]
    clear_sheet(ws)

    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 28

    ws.merge_cells("A1:C1")
    buys = sum(1 for r in results if r.signal == "BUY")
    sells = sum(1 for r in results if r.signal == "SELL")
    waits = sum(1 for r in results if r.signal == "WAIT")
    ws["A1"] = (
        f"GHID DE INVATARE ZILNIC — ANALIZA DE PIATA PROFESIONALA | "
        f"{datetime.now().strftime('%d.%m.%Y %H:%M')} | BUY={buys} SELL={sells} WAIT={waits}"
    )
    apply_cell_style(
        ws["A1"],
        fill=FILL_TITLE,
        font=FONT_TITLE,
        align=Alignment(horizontal="center", vertical="center", wrap_text=True),
    )
    ws.row_dimensions[1].height = 32


def upd_all_sheets(
    wb: Workbook,
    results: List[AssetResult],
    fg: Dict[str, Any],
    macro_yf: Dict[str, float],
) -> None:
    set_tab_colors(wb)
    dv_list_active(wb)

    upd_preturi_volume(wb, results)
    upd_indicatori_tehnici(wb, results)
    upd_semnale_intrare(wb, results)
    upd_indicatori_macro(wb, fg, macro_yf)
    upd_dashboard(wb)
    upd_fisa_activ(wb)
    upd_legenda(wb)
    upd_rezumat_executiv(wb, results, fg)
    upd_ghid_invat(wb, results)


# ============================================================================
# MAIN
# ============================================================================

def main() -> int:
    logger = setup_logging()
    t0 = time.time()
    try:
        template = ensure_template_exists(logger)
        out_path = copy_template_with_timestamp(template, logger)

        df_bulk = yfinance_bulk_download(ALL_TICKERS, logger)
        fg = get_fear_greed(logger)
        macro_yf = get_macro_yfinance(logger)

        results = analyze_all_assets(df_bulk, logger)

        wb = load_workbook(out_path)
        upd_all_sheets(wb, results, fg, macro_yf)
        wb.save(out_path)

        cleanup_old_files(out_path.parent, logger)

        elapsed = time.time() - t0
        logger.info(
            f"DONE | ok={len(results)}/{len(ALL_TICKERS)} | elapsed={elapsed:.1f}s | file={out_path}"
        )
        return 0
    except Exception as e:
        logger.exception(f"FATAL: {e}")
        return 2


if __name__ == "__main__":
    sys.exit(main())