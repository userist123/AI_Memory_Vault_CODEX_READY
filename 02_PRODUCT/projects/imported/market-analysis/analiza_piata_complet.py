import os
"""
╔══════════════════════════════════════════════════════════════════════════════╗
║       ANALIZĂ PIAȚĂ + GHID DE ÎNVĂȚARE ZILNIC — Script Unificat            ║
║       80+ active: Indici | Acțiuni | Crypto | Valute | Materii prime       ║
╚══════════════════════════════════════════════════════════════════════════════╝

INSTALARE (o singură dată):
    pip install yfinance openpyxl requests pandas

RULARE:
    python analiza_piata_complet.py

CE FACE:
    1. Preia date live pentru 80+ active (prețuri, volume, indicatori)
    2. Calculează RSI, MACD, MA, Bollinger, ATR, Stochastic pentru fiecare
    3. Generează semnale BUY / SELL / WAIT cu SL și TP automat
    4. Actualizează sheet-urile: PREȚURI, INDICATORI TEHNICI, SEMNALE,
       INDICATORI MACRO, DASHBOARD, ISTORIC
    5. Generează sheet-ul GHID INVATARE cu explicații detaliate pentru
       fiecare activ: DE CE s-a mișcat, CE oportunitate există,
       CE pattern grafic este vizibil, CE lecție practică se extrage
    6. Adaugă ghid permanent de citire grafice la finalul sheet-ului

AUTOMATIZARE:
    Windows  → Task Scheduler → ora 08:30 L-V
    Mac/Linux → crontab: 30 8 * * 1-5 python3 /cale/analiza_piata_complet.py
"""

import sys
import logging
from datetime import datetime
from pathlib import Path

import requests
import pandas as pd
import yfinance as yf
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════════════════════
# CONFIGURARE
# ══════════════════════════════════════════════════════════════════════════════

EXCEL_TEMPLATE = Path(r"C:\Users\Marius\Desktop\Analiza_Piata_Profesionala.xlsx")
FRED_API_KEY   = os.environ.get("FRED_API_KEY", "")   # Gratuit la fred.stlouisfed.org — lasă "" dacă nu ai

# ── ACTIVE ────────────────────────────────────────────────────────────────────
INDICI = {
    "S&P 500": "^GSPC",        "NASDAQ 100": "^NDX",       "NASDAQ Comp.": "^IXIC",
    "Dow Jones": "^DJI",       "Russell 2000": "^RUT",     "DAX Germany": "^GDAXI",
    "FTSE 100": "^FTSE",       "CAC 40": "^FCHI",          "Nikkei 225": "^N225",
    "Hang Seng": "^HSI",       "Shanghai Comp.": "000001.SS",
    "MSCI World ETF": "URTH",  "MSCI EM ETF": "EEM",       "BET Romania": "BET.RO",
}
ACTIUNI = {
    "Apple": "AAPL",           "Microsoft": "MSFT",        "NVIDIA": "NVDA",
    "Alphabet": "GOOGL",       "Amazon": "AMZN",           "Meta": "META",
    "Tesla": "TSLA",           "AMD": "AMD",               "Intel": "INTC",
    "Broadcom": "AVGO",        "ASML": "ASML",             "Taiwan Semi": "TSM",
    "Palantir": "PLTR",        "Salesforce": "CRM",        "Oracle": "ORCL",
    "JPMorgan": "JPM",         "Goldman Sachs": "GS",      "Berkshire B": "BRK-B",
    "Visa": "V",               "Mastercard": "MA",         "ExxonMobil": "XOM",
    "Chevron": "CVX",          "Shell": "SHEL",            "Caterpillar": "CAT",
    "Boeing": "BA",            "SPY": "SPY",               "QQQ": "QQQ",
    "GLD ETF": "GLD",          "TLT Bond ETF": "TLT",      "ARKK": "ARKK",
}
CRYPTO = {
    "Bitcoin": "BTC-USD",      "Ethereum": "ETH-USD",      "BNB": "BNB-USD",
    "Solana": "SOL-USD",       "XRP": "XRP-USD",           "Cardano": "ADA-USD",
    "Avalanche": "AVAX-USD",   "Dogecoin": "DOGE-USD",     "Chainlink": "LINK-USD",
    "Polkadot": "DOT-USD",     "Litecoin": "LTC-USD",      "Shiba Inu": "SHIB-USD",
    "Polygon": "MATIC-USD",    "Uniswap": "UNI-USD",       "Cosmos": "ATOM-USD",
    "Stellar": "XLM-USD",      "Monero": "XMR-USD",        "Tron": "TRX-USD",
    "Filecoin": "FIL-USD",     "Aave": "AAVE-USD",         "Arbitrum": "ARB-USD",
    "Optimism": "OP-USD",      "Render": "RNDR-USD",       "Sui": "SUI-USD",
    "Near Protocol": "NEAR-USD",
}
VALUTE = {
    "EUR/USD": "EURUSD=X",     "GBP/USD": "GBPUSD=X",     "USD/JPY": "USDJPY=X",
    "USD/CHF": "USDCHF=X",     "AUD/USD": "AUDUSD=X",     "USD/CAD": "USDCAD=X",
    "EUR/RON": "EURRON=X",     "USD/RON": "USDRON=X",     "GBP/RON": "GBPRON=X",
    "EUR/GBP": "EURGBP=X",     "USD/CNY": "USDCNY=X",     "USD/TRY": "USDTRY=X",
}
MATERII_PRIME = {
    "Gold": "GC=F",            "Silver": "SI=F",           "Platinum": "PL=F",
    "Palladium": "PA=F",       "Oil WTI": "CL=F",          "Oil Brent": "BZ=F",
    "Natural Gas": "NG=F",     "Copper": "HG=F",           "Corn": "ZC=F",
    "Wheat": "ZW=F",           "Soybean": "ZS=F",          "Coffee": "KC=F",
    "Sugar": "SB=F",           "Cotton": "CT=F",
}
MACRO_TICKERS = {
    "VIX": "^VIX",             "Yield 10Y US": "^TNX",     "Yield 2Y US": "^IRX",
    "Yield 30Y US": "^TYX",    "USD Index": "DX-Y.NYB",
}

ACTIVE     = {**INDICI, **ACTIUNI, **CRYPTO, **VALUTE, **MATERII_PRIME}
CATEGORII  = [
    ("INDICI BURSIERI", INDICI),
    ("ACTIUNI & ETF",   ACTIUNI),
    ("CRYPTOCURRENCY",  CRYPTO),
    ("VALUTE FOREX",    VALUTE),
    ("MATERII PRIME",   MATERII_PRIME),
]

# ══════════════════════════════════════════════════════════════════════════════
# LOGGING
# ══════════════════════════════════════════════════════════════════════════════

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(Path(__file__).parent / "analiza.log", encoding="utf-8"),
    ],
)
log = logging.getLogger(__name__)

# ══════════════════════════════════════════════════════════════════════════════
# STILURI EXCEL
# ══════════════════════════════════════════════════════════════════════════════

def fill(h): return PatternFill("solid", start_color=h, fgColor=h)
def fnt(bold=False, color="000000", size=10, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)
def aln(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

NO_FILL = PatternFill(fill_type=None)

# Culori sistem
F_HDR   = fill("1F4E79"); FN_HDR  = fnt(bold=True,  color="FFFFFF", size=10)
F_ROWH  = fill("D6DCE4"); FN_ROWH = fnt(bold=True,  color="000000", size=10)
F_INPUT = fill("D9E1F2")
F_CALC  = fill("FFFFFF")
F_REF   = fill("F2F2F2"); FN_REF  = fnt(color="595959")
F_GRN   = fill("C6EFCE"); FN_GRN  = fnt(color="375623", bold=True)
F_YLW   = fill("FFEB9C"); FN_YLW  = fnt(color="9C6500", bold=True)
F_RED   = fill("FFC7CE"); FN_RED  = fnt(color="9C0006", bold=True)
F_ORG   = fill("FCE4D6"); FN_ORG  = fnt(color="833C00")
F_DASH  = fill("1A3A5C"); FN_DASH = fnt(bold=True,  color="FFFFFF", size=14)
F_CAT   = fill("1F4E79"); FN_CAT  = fnt(bold=True,  color="FFFFFF", size=11)
# Ghid
F_GT    = fill("0A1628"); FN_GT   = fnt(bold=True,  color="FFFFFF", size=14)
F_GC    = fill("154360"); FN_GC   = fnt(bold=True,  color="FFFFFF", size=11)
F_GS    = fill("1A5276"); FN_GS   = fnt(bold=True,  color="FFFFFF", size=10)
F_GTXT  = fill("EBF5FB"); FN_GTXT = fnt(color="1A252F", size=10)
F_ASEC  = fill("D6DCE4"); FN_ASEC = fnt(bold=True,  color="1F4E79", size=10)
F_WHITE = fill("FFFFFF")
F_SEP   = fill("1F4E79")

# ══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def fp(val, dec=4):
    """Format price — sigur, returnează 'N/A' dacă None/NaN."""
    if val is None: return "N/A"
    try:
        f = float(val)
        if f != f: return "N/A"
        return f"{{:,.{dec}f}}".format(f)
    except Exception:
        return "N/A"

def fpc(val, dec=2):
    """Format percent — sigur."""
    if val is None: return "N/A"
    try:
        f = float(val)
        return f"{'+' if f >= 0 else ''}{f:.{dec}f}%"
    except Exception:
        return "N/A"

def safe(val, default=0.0):
    if val is None: return default
    try:
        f = float(val)
        return default if f != f else f
    except Exception:
        return default

def set_h(ws, row, h): ws.row_dimensions[row].height = h

def write_hdr_row(ws, row, headers, col=1):
    for i, h in enumerate(headers):
        c = ws.cell(row, col + i, h)
        c.fill = F_HDR; c.font = FN_HDR
        c.alignment = aln("center"); c.border = None

def clear_rows(ws, start, end=500):
    for row in ws.iter_rows(min_row=start, max_row=end):
        for cell in row:
            cell.value = None; cell.fill = NO_FILL

def cat_header(ws, row, label, ncols):
    ws.cell(row, 1, label).fill = F_HDR
    ws.cell(row, 1).font = FN_HDR
    ws.cell(row, 1).alignment = aln("left")
    for c in range(2, ncols + 1):
        ws.cell(row, c).fill = F_HDR

def color_signal(cell, s):
    if s == "BUY":    cell.fill = F_GRN; cell.font = FN_GRN
    elif s == "SELL": cell.fill = F_RED; cell.font = FN_RED
    else:             cell.fill = F_YLW; cell.font = FN_YLW

def color_pnl(cell, v):
    if v > 0:   cell.fill = F_GRN; cell.font = FN_GRN
    elif v < 0: cell.fill = F_RED; cell.font = FN_RED
    else:       cell.fill = F_YLW; cell.font = FN_YLW

def color_rsi(cell, rsi):
    if rsi < 30:        cell.fill = F_GRN; cell.font = FN_GRN
    elif rsi < 50:      cell.fill = F_YLW; cell.font = FN_YLW
    elif rsi < 70:      cell.fill = fill("E2EFDA"); cell.font = fnt(color="375623")
    else:               cell.fill = F_RED; cell.font = FN_RED

def color_trend(cell, t):
    if t == "Bullish":   cell.fill = F_GRN; cell.font = FN_GRN
    elif t == "Bearish": cell.fill = F_RED; cell.font = FN_RED
    else:                cell.fill = F_YLW; cell.font = FN_YLW

# ══════════════════════════════════════════════════════════════════════════════
# CALCUL INDICATORI TEHNICI
# ══════════════════════════════════════════════════════════════════════════════

def calc_rsi(prices, period=14):
    d = prices.diff()
    g = d.clip(lower=0).rolling(period).mean()
    l = (-d.clip(upper=0)).rolling(period).mean()
    rs = g / l.replace(0, 1e-10)
    v = (100 - 100 / (1 + rs)).iloc[-1]
    return round(float(v), 2) if pd.notna(v) else 50.0

def calc_macd(prices):
    e12 = prices.ewm(span=12, adjust=False).mean()
    e26 = prices.ewm(span=26, adjust=False).mean()
    ml  = e12 - e26
    sl  = ml.ewm(span=9, adjust=False).mean()
    hs  = ml - sl
    m, s, h = float(ml.iloc[-1]), float(sl.iloc[-1]), float(hs.iloc[-1])
    ph = float(hs.iloc[-2]) if len(hs) > 1 else 0
    cross = ("Bullish Cross" if m > s and ph < 0 else
             "Pozitiv"       if m > s else
             "Bearish Cross" if ph > 0 else "Negativ")
    return {"macd": round(m, 6), "signal": round(s, 6),
            "hist": round(h, 6), "cross": cross}

def calc_ma(prices):
    n = len(prices)
    def ma(p): return round(float(prices.tail(p).mean()), 6) if n >= p else None
    m20, m50, m200 = ma(20), ma(50), ma(200)
    cross = ("Golden Cross" if m50 and m200 and m50 > m200 else
             "Death Cross"  if m50 and m200 and m50 < m200 else "Neutru")
    return {"ma20": m20, "ma50": m50, "ma200": m200, "macross": cross}

def calc_boll(prices, p=20):
    if len(prices) < p: return {"bb_sup": None, "bb_inf": None, "bb_w": None}
    m  = prices.rolling(p).mean()
    sd = prices.rolling(p).std()
    s  = float((m + 2 * sd).iloc[-1])
    i  = float((m - 2 * sd).iloc[-1])
    return {"bb_sup": round(s, 6), "bb_inf": round(i, 6), "bb_w": round(s - i, 6)}

def calc_atr(hist, p=14):
    hi, lo, cl = hist["High"], hist["Low"], hist["Close"]
    tr  = pd.concat([(hi - lo), (hi - cl.shift()).abs(),
                     (lo - cl.shift()).abs()], axis=1).max(axis=1)
    v   = tr.rolling(p).mean().iloc[-1]
    return round(float(v), 6) if pd.notna(v) else 0.0

def calc_stoch(hist, p=14):
    if len(hist) < p: return {"k": 50.0, "d": 50.0}
    lo, hi, cl = hist["Low"], hist["High"], hist["Close"]
    l14 = lo.rolling(p).min(); h14 = hi.rolling(p).max()
    k   = (cl - l14) / (h14 - l14).replace(0, 1e-10) * 100
    d   = k.rolling(3).mean()
    return {
        "k": round(float(k.iloc[-1]), 2) if pd.notna(k.iloc[-1]) else 50.0,
        "d": round(float(d.iloc[-1]), 2) if pd.notna(d.iloc[-1]) else 50.0,
    }

def calc_signal(rsi, macd_cross, macross, rvol):
    sc = 0
    if rsi < 35:                        sc += 2
    elif rsi < 45:                      sc += 1
    elif rsi > 75:                      sc -= 2
    elif rsi > 65:                      sc -= 1
    if "Bullish Cross" in macd_cross:   sc += 2
    elif "Pozitiv" in macd_cross:       sc += 1
    elif "Bearish Cross" in macd_cross: sc -= 2
    elif "Negativ" in macd_cross:       sc -= 1
    if macross == "Golden Cross":       sc += 2
    elif macross == "Death Cross":      sc -= 2
    if rvol > 1.5:                      sc += 1
    elif rvol < 0.6:                    sc -= 1
    conf = min(abs(sc), 5)
    if sc >= 3:    return "BUY",  conf
    elif sc <= -3: return "SELL", conf
    else:          return "WAIT", conf

# ══════════════════════════════════════════════════════════════════════════════
# PRELUARE DATE
# ══════════════════════════════════════════════════════════════════════════════

def get_data(name: str, ticker: str) -> dict:
    try:
        t    = yf.Ticker(ticker)
        hist = t.history(period="1y", auto_adjust=True)
        if hist.empty or len(hist) < 5:
            log.warning(f"  ⚠  {name} — date insuficiente")
            return {}

        cl = hist["Close"]
        n  = len(cl)
        lt = hist.iloc[-1]
        pv = hist.iloc[-2] if n > 1 else lt

        now   = float(lt["Close"])
        prev  = float(pv["Close"])
        c5d   = float(hist.iloc[-min(5,  n)]["Close"])
        c20d  = float(hist.iloc[-min(20, n)]["Close"])

        var_zi   = (now - prev) / prev   * 100 if prev  else 0
        var_sapt = (now - c5d)  / c5d   * 100 if c5d   else 0
        var_luna = (now - c20d) / c20d  * 100 if c20d  else 0

        vol     = int(lt.get("Volume", 0))
        avg_vol = int(hist["Volume"].tail(min(20, n)).mean()) if "Volume" in hist else 0
        rvol    = round(vol / avg_vol, 2) if avg_vol > 0 else 1.0

        rsi    = calc_rsi(cl)
        macd   = calc_macd(cl)
        ma     = calc_ma(cl)
        boll   = calc_boll(cl)
        atr    = calc_atr(hist)
        stoch  = calc_stoch(hist)
        mom10  = round(float(cl.pct_change(10).iloc[-1] * 100), 2) if n > 10 else 0.0

        if ma["ma50"] and now > ma["ma50"] * 1.01:   trend = "Bullish"
        elif ma["ma50"] and now < ma["ma50"] * 0.99: trend = "Bearish"
        else:                                          trend = "Sideways"

        rsi_st = ("Supravandut ▲" if rsi < 30 else "Slab" if rsi < 45 else
                  "Neutru" if rsi < 55 else "Puternic" if rsi < 70 else "Supravandut ▼")

        semnal, conf = calc_signal(rsi, macd["cross"], ma["macross"], rvol)

        sl = (round(now - 1.5 * atr, 6) if semnal == "BUY"  else
              round(now + 1.5 * atr, 6) if semnal == "SELL" else
              round(now - 2.0 * atr, 6))
        tp = (round(now + 3.0 * atr, 6) if semnal == "BUY"  else
              round(now - 3.0 * atr, 6) if semnal == "SELL" else
              round(now + 2.0 * atr, 6))
        prob = min(90, 35 + conf * 10 + (5 if rvol > 1.2 else 0))

        return {
            "name": name, "ticker": ticker,
            "data": datetime.now().strftime("%d.%m.%Y"),
            "now": round(now, 6),
            "open": round(float(lt.get("Open", now)), 6),
            "high": round(float(lt.get("High", now)), 6),
            "low":  round(float(lt.get("Low",  now)), 6),
            "var_zi": round(var_zi, 4), "var_sapt": round(var_sapt, 4),
            "var_luna": round(var_luna, 4),
            "vol": vol, "avg_vol": avg_vol, "rvol": rvol,
            "rsi": rsi, "rsi_st": rsi_st,
            "macd": macd["macd"], "macd_sig": macd["signal"],
            "macd_hist": macd["hist"], "macd_cross": macd["cross"],
            "ma20": ma["ma20"], "ma50": ma["ma50"], "ma200": ma["ma200"],
            "macross": ma["macross"],
            "bb_sup": boll["bb_sup"], "bb_inf": boll["bb_inf"], "bb_w": boll["bb_w"],
            "atr": atr, "stoch_k": stoch["k"], "stoch_d": stoch["d"],
            "mom10": mom10, "trend": trend,
            "semnal": semnal, "conf": conf, "sl": sl, "tp": tp, "prob": prob,
            "n": n,
        }
    except Exception as e:
        log.error(f"  ✗  {name} ({ticker}): {e}")
        return {}

def get_fear_greed():
    try:
        r = requests.get("https://api.alternative.me/fng/?limit=1", timeout=10)
        d = r.json()["data"][0]
        v = int(d["value"])
        return {"value": v, "label": d["value_classification"],
                "display": f"{v} — {d['value_classification']}",
                "status": "Pozitiv" if v > 60 else ("Negativ" if v < 40 else "Neutru")}
    except Exception as e:
        log.warning(f"  Fear & Greed: {e}")
        return {"value": None, "display": "N/A", "status": "Neutru"}

def get_fred(sid):
    if not FRED_API_KEY: return None
    try:
        url = (f"https://api.stlouisfed.org/fred/series/observations"
               f"?series_id={sid}&api_key={FRED_API_KEY}"
               f"&file_type=json&sort_order=desc&limit=1")
        obs = requests.get(url, timeout=10).json().get("observations", [])
        if obs and obs[0]["value"] != ".":
            return float(obs[0]["value"])
    except Exception: pass
    return None

# ══════════════════════════════════════════════════════════════════════════════
# ACTUALIZARE SHEET-URI PRINCIPALE
# ══════════════════════════════════════════════════════════════════════════════

def upd_preturi(ws, all_data):
    clear_rows(ws, 3)
    r = 3
    for cat, dic in CATEGORII:
        cat_header(ws, r, f"--- {cat} ---", 15); r += 1
        for name, ticker in dic.items():
            d = all_data.get(ticker, {})
            ws.cell(r, 1, d.get("data", datetime.now().strftime("%d.%m.%Y")))
            ws.cell(r, 2, name).font = fnt(bold=True)
            ws.cell(r, 3, ticker)
            if d:
                for j, (col, key, fmt) in enumerate([
                    (4,"open","#,##0.0000"), (5,"high","#,##0.0000"),
                    (6,"low","#,##0.0000"),  (7,"now","#,##0.0000"),
                    (8,"var_zi","0.00%"),    (9,"var_sapt","0.00%"),
                    (10,"var_luna","0.00%"), (11,"vol","#,##0"),
                    (12,"avg_vol","#,##0"),  (13,"rvol","0.00x"),
                ]):
                    v = d[key]
                    c = ws.cell(r, col, v / 100 if "var" in key else v)
                    c.number_format = fmt
                ws.cell(r, 14, d["semnal"]); color_signal(ws.cell(r, 14), d["semnal"])
                ws.cell(r, 15, d["trend"]);  color_trend(ws.cell(r, 15),  d["trend"])
                for col in [8, 9, 10]: color_pnl(ws.cell(r, col), d[["var_zi","var_sapt","var_luna"][col-8]])
            else:
                ws.cell(r, 7, "N/A").fill = F_REF
            r += 1
    log.info(f"  ✓ Preturi Volume — {r-3} randuri")

def upd_tehnic(ws, all_data):
    clear_rows(ws, 3)
    r = 3
    for cat, dic in CATEGORII:
        cat_header(ws, r, f"--- {cat} ---", 23); r += 1
        for name, ticker in dic.items():
            d = all_data.get(ticker, {})
            ws.cell(r, 1, d.get("data", datetime.now().strftime("%d.%m.%Y")))
            ws.cell(r, 2, name).font = fnt(bold=True)
            ws.cell(r, 3, ticker)
            if d:
                rows_data = [
                    (4,d["now"],"#,##0.0000"), (5,d["ma20"],"#,##0.0000"),
                    (6,d["ma50"],"#,##0.0000"),(7,d["ma200"],"#,##0.0000"),
                    (8,d["rsi"],"0.00"),        (9,d["rsi_st"],None),
                    (10,d["macd"],"0.000000"), (11,d["macd_sig"],"0.000000"),
                    (12,d["macd_hist"],"0.000000"),(13,d["macd_cross"],None),
                    (14,d["bb_sup"],"#,##0.0000"),(15,d["bb_inf"],"#,##0.0000"),
                    (16,d["bb_w"],"#,##0.0000"),(17,d["atr"],"0.0000"),
                    (18,d["stoch_k"],"0.00"),   (19,d["stoch_d"],"0.00"),
                    (20,d["vol"],"#,##0"),       (21,d["rvol"],"0.00x"),
                    (22,d["trend"],None),         (23,d["macross"],None),
                ]
                for col, val, fmt in rows_data:
                    if val is None: continue
                    c = ws.cell(r, col, val)
                    if fmt: c.number_format = fmt
                color_rsi(ws.cell(r, 8),   d["rsi"])
                color_trend(ws.cell(r, 22), d["trend"])
                mc = ws.cell(r, 23)
                if d["macross"] == "Golden Cross": mc.fill = F_GRN; mc.font = FN_GRN
                elif d["macross"] == "Death Cross": mc.fill = F_RED; mc.font = FN_RED
                else:                               mc.fill = F_YLW; mc.font = FN_YLW
            r += 1
    log.info(f"  ✓ Indicatori Tehnici — {r-3} randuri")

def upd_semnale(ws, all_data):
    clear_rows(ws, 3)
    r = 3
    for cat, dic in CATEGORII:
        cat_header(ws, r, f"--- {cat} ---", 17); r += 1
        for name, ticker in dic.items():
            d = all_data.get(ticker, {})
            if not d: continue
            now   = d["now"]
            cond  = f"RSI={d['rsi']:.0f} | {d['macd_cross']} | {d['macross']} | RVOL={d['rvol']:.1f}x"
            vol_t = "Crescut" if d["rvol"] > 1.3 else ("Scazut" if d["rvol"] < 0.8 else "Normal")
            ws.cell(r, 1, d["data"]); ws.cell(r, 2, name).font = fnt(bold=True)
            ws.cell(r, 3, d["semnal"]);  color_signal(ws.cell(r, 3), d["semnal"])
            ws.cell(r, 4, cond)
            ws.cell(r, 5, d["rsi"]).number_format   = "0.00"; color_rsi(ws.cell(r, 5), d["rsi"])
            ws.cell(r, 6, d["macd_cross"])
            ws.cell(r, 7, d["macross"])
            ws.cell(r, 8, vol_t)
            ws.cell(r, 9, d["mom10"] / 100).number_format = "0.00%"
            ws.cell(r, 10, d["conf"])
            ws.cell(r, 11, now).number_format   = "#,##0.0000"
            ws.cell(r, 12, d["sl"]).number_format= "#,##0.0000"
            ws.cell(r, 13, d["tp"]).number_format= "#,##0.0000"
            ws.cell(r, 14).value = f"=IFERROR((M{r}-K{r})/(K{r}-L{r}),\"N/A\")"
            ws.cell(r, 14).number_format = "0.00x"
            p = ws.cell(r, 15, d["prob"]); p.number_format = "0%"
            if d["prob"] >= 65:   p.fill = F_GRN; p.font = FN_GRN
            elif d["prob"] >= 50: p.fill = F_YLW; p.font = FN_YLW
            else:                 p.fill = F_RED; p.font = FN_RED
            ws.cell(r, 16, "Activ")
            ws.cell(r, 17, f"Auto {datetime.now().strftime('%H:%M')}")
            r += 1
    log.info(f"  ✓ Semnale Intrare — {r-3} randuri")

def upd_macro(ws, macro_live, fg):
    today = datetime.now().strftime("%d.%m.%Y %H:%M")
    mp = {
        "VIX":         macro_live.get("VIX",         {}).get("now"),
        "Yield 10Y":   macro_live.get("Yield 10Y US", {}).get("now"),
        "Yield 2Y":    macro_live.get("Yield 2Y US",  {}).get("now"),
        "USD Index":   macro_live.get("USD Index",    {}).get("now"),
        "Fear & Greed":fg.get("value"),
    }
    if FRED_API_KEY:
        mp["Rata dobanzii"] = get_fred("FEDFUNDS")
        mp["CPI"]           = get_fred("CPIAUCSL")
        mp["Somaj"]         = get_fred("UNRATE")
    upd = 0
    for r in range(3, 50):
        ind = ws.cell(r, 1).value
        if not ind: break
        for k, v in mp.items():
            if k.lower() in str(ind).lower() and v is not None:
                prev = ws.cell(r, 2).value
                try:
                    if prev and float(str(prev).replace(",", ".")) != float(v):
                        ws.cell(r, 3).value = prev
                except Exception: pass
                ws.cell(r, 2).value = round(float(v), 4)
                ws.cell(r, 9).value = today
                upd += 1; break
    log.info(f"  ✓ Indicatori Macro — {upd} valori")

def upd_dashboard(ws, all_data, fg):
    ws["D2"] = datetime.now().strftime("%d.%m.%Y %H:%M")
    buy = sum(1 for d in all_data.values() if d and d.get("semnal") == "BUY")
    sel = sum(1 for d in all_data.values() if d and d.get("semnal") == "SELL")
    tot = len([d for d in all_data.values() if d])
    trend = ("Bullish ▲" if tot and buy/tot > 0.55 else
             "Bearish ▼" if tot and sel/tot > 0.55 else "Mixt / Neutru →")
    log.info(f"  ✓ Dashboard — BUY:{buy} SELL:{sel} WAIT:{tot-buy-sel} | {trend}")

def upd_istoric(ws, all_data, fg):
    luna = datetime.now().strftime("%b %Y")
    for row in ws.iter_rows(min_row=3, max_col=1):
        if row[0].value == luna:
            log.info(f"  ✓ Istoric — {luna} exista"); return
    last = 2
    for row in ws.iter_rows(min_row=3, max_col=1):
        if row[0].value: last = row[0].row
    r = last + 1
    rsi_vals = [d["rsi"] for d in all_data.values() if d and d.get("rsi")]
    avg_rsi  = round(sum(rsi_vals) / len(rsi_vals), 1) if rsi_vals else 50
    buy  = sum(1 for d in all_data.values() if d and d.get("semnal") == "BUY")
    sell = sum(1 for d in all_data.values() if d and d.get("semnal") == "SELL")
    sm   = "BUY" if buy > sell else ("SELL" if sell > buy else "WAIT")
    ws.cell(r, 1, luna); ws.cell(r, 2, avg_rsi); ws.cell(r, 8, sm)
    sp500 = all_data.get("^GSPC", {}).get("now")
    if sp500: ws.cell(r, 4, sp500)
    if fg.get("value"): ws.cell(r, 11, fg["value"])
    color_signal(ws.cell(r, 8), sm)
    log.info(f"  ✓ Istoric — {luna} RSI={avg_rsi} Semnal={sm}")

# ══════════════════════════════════════════════════════════════════════════════
# MOTOR EXPLICAȚII — Ghid de Învățare
# ══════════════════════════════════════════════════════════════════════════════

def txt_miscare(d: dict) -> str:
    name = d["name"]; pret = d["now"]; var_zi = d["var_zi"]
    var_sapt = d["var_sapt"]; var_luna = d["var_luna"]
    rsi = d["rsi"]; mc = d["macd_cross"]; mh = d["macd_hist"]
    macross = d["macross"]; rvol = d["rvol"]
    ma50 = d["ma50"]; ma200 = d["ma200"]
    bb_sup = d["bb_sup"]; bb_inf = d["bb_inf"]; bb_w = d["bb_w"]
    stoch_k = d["stoch_k"]; mom10 = d["mom10"]
    semnal = d["semnal"]; n = d["n"]

    L = []
    # 1. Mișcarea zilei
    dir_ = "crescut" if var_zi > 0 else "scăzut"
    int_ = ("semnificativ" if abs(var_zi) > 3 else "moderat" if abs(var_zi) > 1 else "ușor")
    L.append(
        f"{name} a {dir_} {int_} cu {abs(var_zi):.2f}% astăzi (preț: {fp(pret)}).\n"
        f"Săptămână: {fpc(var_sapt)} | Lună: {fpc(var_luna)}."
    )
    # 2. Volum
    if rvol > 1.5:
        L.append(
            f"VOLUM EXCEPȚIONAL: {rvol:.1f}x media 20 zile.\n"
            f"Mișcarea este CONFIRMATĂ de participare instituțională. "
            f"Fondurile și algoritmii sunt activi pe acest activ azi."
        )
    elif rvol < 0.7:
        L.append(
            f"VOLUM SCĂZUT: {rvol:.1f}x media — ATENȚIE!\n"
            f"Mișcările pe volum mic sunt adesea false (fake moves). "
            f"Probabilitate crescută ca mișcarea să se inverseze."
        )
    else:
        L.append(f"Volum în linie cu media ({rvol:.1f}x) — participare normală.")
    # 3. RSI
    if rsi < 30:
        L.append(
            f"RSI = {rsi:.1f} — SUPRAVÂNDUT (sub 30).\n"
            f"Activul a scăzut prea rapid. Presiune acumulată de cumpărare. "
            f"Nu înseamnă revenire imediată, dar probabilitatea unui bounce crește."
        )
    elif rsi < 50:
        L.append(f"RSI = {rsi:.1f} — Zonă slabă. Bears domină, presiunea scade treptat.")
    elif rsi < 70:
        L.append(
            f"RSI = {rsi:.1f} — Momentum POZITIV. Bulls în control. "
            f"Zona 55-70 este cea mai profitabilă pentru trend-following."
        )
    else:
        L.append(
            f"RSI = {rsi:.1f} — SUPRAEVALUAT (peste 70).\n"
            f"Activul a crescut prea rapid. Probabilitate crescută de corecție."
        )
    # 4. MACD
    if "Bullish Cross" in mc:
        L.append(
            f"MACD: BULLISH CROSSOVER confirmat (histogram: {mh:+.4f}).\n"
            f"Linia MACD a depășit linia Signal — momentum în favoarea cumpărătorilor."
        )
    elif "Bearish Cross" in mc:
        L.append(
            f"MACD: BEARISH CROSSOVER confirmat (histogram: {mh:+.4f}).\n"
            f"Momentum bullish s-a epuizat — vânzătorii preiau controlul."
        )
    else:
        L.append(f"MACD {'pozitiv' if 'Pozitiv' in mc else 'negativ'} (histogram: {mh:+.4f}).")
    # 5. MA
    ma50_s  = fp(ma50, 2)
    ma200_s = fp(ma200, 2)
    if macross == "Golden Cross" and ma50 and ma200:
        L.append(
            f"GOLDEN CROSS ACTIV: MA50 ({ma50_s}) > MA200 ({ma200_s}).\n"
            f"Cel mai puternic semnal bullish pe termen lung. "
            f"Trendul major este ascendent — scăderile sunt oportunități de cumpărare."
        )
    elif macross == "Death Cross" and ma50 and ma200:
        L.append(
            f"DEATH CROSS ACTIV: MA50 ({ma50_s}) < MA200 ({ma200_s}).\n"
            f"Semnal bearish major. Creșterile pot fi oportunități de vânzare."
        )
    elif ma50 and not ma200:
        L.append(
            f"MA50 disponibilă ({ma50_s}), MA200 indisponibilă ({n} zile date).\n"
            f"Urmărește relația prețului față de MA50 ca referință principală."
        )
    else:
        L.append(f"MA50 ({ma50_s}) și MA200 ({ma200_s}) — piață în tranziție.")
    # 6. Bollinger
    if bb_sup and bb_inf and bb_w:
        p = float(pret)
        if p >= float(bb_sup) * 0.99:
            L.append(f"Preț la BANDA SUPERIOARĂ Bollinger ({fp(bb_sup)}) — risc corecție pe termen scurt.")
        elif p <= float(bb_inf) * 1.01:
            L.append(f"Preț la BANDA INFERIOARĂ Bollinger ({fp(bb_inf)}) — zonă potențial suport.")
        elif float(bb_w) / p < 0.03:
            L.append(
                f"BOLLINGER SQUEEZE (lățime: {fp(bb_w)}) — volatilitate minimă.\n"
                f"Explozie de volatilitate iminentă. Pregătește ordine pending în ambele direcții."
            )
    # 7. Stochastic
    sk = d["stoch_k"]; sd_v = d["stoch_d"]
    if sk < 20:
        L.append(f"Stochastic %K = {sk:.1f} — OVERSOLD. Semnal potențial de cumpărare pe termen scurt.")
    elif sk > 80:
        L.append(f"Stochastic %K = {sk:.1f} — OVERBOUGHT. Semnal potențial de vânzare pe termen scurt.")
    # 8. Momentum
    if abs(mom10) > 5:
        L.append(f"Momentum 10 zile: {fpc(mom10)} — {'puternic pozitiv' if mom10 > 0 else 'puternic negativ'}.")
    # 9. Concluzie
    if semnal == "BUY":
        L.append(
            "CONCLUZIE: Configurație FAVORABILĂ CUMPĂRĂRII. "
            "Respectă SL strict în caz de invalidare."
        )
    elif semnal == "SELL":
        L.append(
            "CONCLUZIE: Configurație de PRESIUNE VÂNZARE. "
            "Prudență cu pozițiile long existente."
        )
    else:
        L.append("CONCLUZIE: Semnale MIXTE — AȘTEPTARE recomandată. Nu forța o tranzacție.")
    return "\n\n".join(L)


def txt_oportunitate(d: dict) -> str:
    semnal = d["semnal"]; pret = d["now"]; sl = d["sl"]; tp = d["tp"]
    atr = d["atr"]; rvol = d["rvol"]; conf = d["conf"]; prob = d["prob"]
    rr = abs(float(tp) - float(pret)) / abs(float(sl) - float(pret)) if abs(float(sl) - float(pret)) > 0 else 0
    risc_pct = abs(float(pret) - float(sl)) / float(pret) * 100 if float(pret) > 0 else 0
    rr_e = ("EXCELENT ★★★" if rr >= 3 else "BUN ★★" if rr >= 2 else
            "ACCEPTABIL ★" if rr >= 1.5 else "SLAB — reconsideră")
    L = []
    if semnal in ("BUY", "SELL"):
        tip = "CUMPĂRARE" if semnal == "BUY" else "SHORT/VÂNZARE"
        L.append(
            f"OPORTUNITATE DE {tip} — {conf} confluențe | Probabilitate: {prob}%\n\n"
            f"  Entry               :  {fp(pret, 4)}\n"
            f"  Stop Loss (SL)      :  {fp(sl,   4)}   ← risc {risc_pct:.1f}%\n"
            f"  Take Profit (TP)    :  {fp(tp,   4)}   ← recompensă {abs(float(tp)-float(pret))/float(pret)*100:.1f}%\n"
            f"  Risk/Reward Ratio   :  {rr:.2f}x   {rr_e}\n"
            f"  ATR (volatilitate)  :  {fp(atr, 4)}"
        )
        if rr >= 2:
            L.append(
                f"RR de {rr:.1f}x — matematic profitabil chiar cu win rate de "
                f"{'34%' if rr >= 2 else '40%'}. Nu muta TP-ul prematur din frică."
            )
        else:
            L.append(
                f"RR de {rr:.1f}x este sub pragul ideal de 2x. "
                f"Evaluează dacă poți ajusta SL/TP pentru un raport mai bun."
            )
        if semnal == "SELL":
            L.append(
                "ATENȚIE SHORT: Pierderea teoretică este nelimitată (prețul poate urca oricât). "
                "Necesită cont de marjă și experiență. Alternativa: ETF inversat sau opțiuni PUT."
            )
    else:
        L.append(
            f"NU EXISTĂ OPORTUNITATE CLARĂ — Confluențe: {conf}/5 (sub pragul minim 3).\n\n"
            f"Ce să urmărești:\n"
            f"  • RSI sub 40 (BUY) sau peste 70 (SELL)\n"
            f"  • MACD Crossover în direcția dorită\n"
            f"  • Volum peste 1.2x medie\n"
            f"  • Golden Cross confirmat pentru semnale BUY"
        )
    if rvol > 1.5 and semnal != "WAIT":
        L.append(
            f"CONFIRMARE VOLUM: {rvol:.1f}x medie — semnalul este de calitate superioară. "
            f"Mișcările cu volum mare sunt mult mai puțin susceptibile de a fi false."
        )
    return "\n\n".join(L)


def txt_pattern(d: dict) -> str:
    pret = d["now"]; rsi = d["rsi"]; mc = d["macd_cross"]; mh = d["macd_hist"]
    macross = d["macross"]; sk = d["stoch_k"]; sd_v = d["stoch_d"]
    bb_sup = d["bb_sup"]; bb_inf = d["bb_inf"]; bb_w = d["bb_w"]
    ma50 = d["ma50"]; ma200 = d["ma200"]; var_zi = d["var_zi"]; rvol = d["rvol"]

    ma50_s  = fp(ma50, 2);   ma200_s = fp(ma200, 2)
    bbs_s   = fp(bb_sup, 4); bbi_s   = fp(bb_inf, 4); bbw_s = fp(bb_w, 4)
    P = []

    if macross == "Golden Cross" and ma50 and ma200:
        dp = (float(ma50) - float(ma200)) / float(ma200) * 100
        P.append(
            f"PATTERN ACTIV: GOLDEN CROSS\n"
            f"MA50 ({ma50_s}) > MA200 ({ma200_s}) cu {dp:.2f}%.\n\n"
            f"Cel mai urmărit semnal bullish pe termen lung. Instituțiile și "
            f"fondurile de indici cumpără automat la Golden Cross. Istoric, "
            f"S&P 500 a generat +15% medie în 12 luni după Golden Cross.\n\n"
            f"PE GRAFIC: Linia MA50 taie MA200 de jos în sus. "
            f"Confirmă cu candle bullish și volum crescut la momentul crossover-ului."
        )
    elif macross == "Death Cross" and ma50 and ma200:
        dp = (float(ma200) - float(ma50)) / float(ma200) * 100
        P.append(
            f"PATTERN ACTIV: DEATH CROSS\n"
            f"MA50 ({ma50_s}) < MA200 ({ma200_s}) cu {dp:.2f}%.\n\n"
            f"Semnal bearish major. Algoritmii de trailing stop se activează automat.\n\n"
            f"CAPCANA: Death Cross apare DUPĂ ce prețul a coborât deja 15-25%. "
            f"Nu vinde în panică la semnal — caută un retest al MA50 (acum rezistență) "
            f"pentru un entry short mai bun cu risc redus."
        )

    if bb_w and bb_sup and bb_inf and float(bb_w) / float(pret) < 0.04:
        target_buy  = fp(float(bb_sup) * 1.005, 4)
        target_sell = fp(float(bb_inf) * 0.995, 4)
        P.append(
            f"PATTERN ACTIV: BOLLINGER BAND SQUEEZE\n"
            f"Lățime benzi: {bbw_s} (sub 4% din preț) — compresie extremă.\n\n"
            f"Explozie de volatilitate IMINENTĂ — direcția necunoscută, amplitudinea va fi mare.\n\n"
            f"STRATEGIE PRACTICĂ:\n"
            f"  BUY STOP  deasupra {bbs_s} (target: {target_buy})\n"
            f"  SELL STOP sub {bbi_s} (target: {target_sell})\n"
            f"Ordinul activat = direcția breakout. Anulează imediat celălalt."
        )

    if "Bullish Cross" in mc:
        P.append(
            f"PATTERN ACTIV: MACD BULLISH CROSSOVER\n"
            f"Histogramă: {mh:+.6f} (trecere în teritoriu pozitiv).\n\n"
            f"Barele histogramei au trecut din ROȘU în VERDE. "
            f"Cel mai puternic: crossover LA LINIA ZERO.\n\n"
            f"Confirmă cu RVOL curent: {d['rvol']:.1f}x "
            f"{'✓ Confirmat' if d['rvol'] > 1.2 else '⚠ Slab'}."
        )
    elif "Bearish Cross" in mc:
        P.append(
            f"PATTERN ACTIV: MACD BEARISH CROSSOVER\n"
            f"Histogramă: {mh:+.6f} (trecere în teritoriu negativ).\n\n"
            f"Barele histogramei din VERDE în ROȘU. "
            f"Semnal de prudență pentru deținătorii de poziții long."
        )

    if rsi < 32 and var_zi < -2 and rvol > 1.3:
        P.append(
            f"PATTERN POSIBIL: SELLING CLIMAX\n"
            f"Scădere {var_zi:.1f}% + RSI {rsi:.1f} + Volum {rvol:.1f}x.\n\n"
            f"Toți cei care voiau să vândă au vândut. Potențial punct de inversare.\n"
            f"PE GRAFIC: Caută Hammer sau Pin Bar (fitil lung jos, corp sus). "
            f"Confirmă cu volum scăzut a doua zi."
        )
    elif rsi > 75 and var_zi > 2 and rvol > 1.3:
        P.append(
            f"PATTERN POSIBIL: BUYING CLIMAX\n"
            f"Creștere {var_zi:.1f}% + RSI {rsi:.1f} + Volum {rvol:.1f}x.\n\n"
            f"Euforia la maxim — smart money DISTRIBUIE pozițiile acum. "
            f"PE GRAFIC: Shooting Star sau Bearish Engulfing = semnal de inversare."
        )

    if sk < 25 and sk > sd_v:
        P.append(f"STOCHASTIC OVERSOLD CROSSOVER: %K ({sk:.1f}) > %D din zona sub 25. Semnal BUY pe termen scurt.")
    elif sk > 75 and sk < sd_v:
        P.append(f"STOCHASTIC OVERBOUGHT CROSSOVER: %K ({sk:.1f}) < %D din zona peste 75. Semnal SELL pe termen scurt.")

    if ma50:
        dist = (float(pret) - float(ma50)) / float(ma50) * 100
        if abs(dist) < 0.8:
            P.append(
                f"PATTERN: TEST MA50\n"
                f"Prețul ({fp(pret)}) la {dist:+.2f}% față de MA50 ({ma50_s}).\n\n"
                f"MA50 = cel mai urmărit nivel dinamic de suport/rezistență.\n"
                f"BOUNCE = continuă trendul | BREAKDOWN = semnal bearish | WHIPSAW = piață slabă.\n"
                f"Fii atent la candle-ul de ÎNCHIDERE față de MA50!"
            )

    if not P:
        P.append(
            "Nu există un pattern tehnic dominant clar astăzi.\n"
            "Activ în consolidare sau așteptare. Monitorizează zilnic."
        )
    return "\n\n".join(P)


def txt_lectie(d: dict) -> str:
    semnal = d["semnal"]; rsi = d["rsi"]; macross = d["macross"]
    rvol = d["rvol"]; conf = d["conf"]; pret = d["now"]; sl = d["sl"]; tp = d["tp"]
    rr   = abs(float(tp) - float(pret)) / abs(float(sl) - float(pret)) if abs(float(sl) - float(pret)) > 0 else 0
    L = []

    if semnal == "WAIT" and conf < 2:
        L.append(
            "LECȚIE — PUTEREA 'NU TRANZACȚIONEZ':\n\n"
            "Marii profesioniști stau în cash 60-70% din timp și atacă DOAR când setup-ul este perfect. "
            "Jack Schwager (Market Wizards, 1989) a intervievat cei mai buni traderi din lume — "
            "toți aveau în comun: SELECTIVITATE EXTREMĂ.\n\n"
            "Aplicare: stabilește-ți un checklist cu 5 condiții. "
            "Intri DOAR dacă toate 5 sunt bifate. Astăzi, acest activ nu le bifează. "
            "Răbdarea este o strategie validă și profitabilă."
        )
    if 0 < rr < 1.5:
        L.append(
            f"LECȚIE — MATEMATICA RISK/REWARD:\n\n"
            f"RR curent: {rr:.1f}x — sub pragul profesionist de 2x.\n\n"
            f"De ce contează matematic:\n"
            f"  RR=1.0x + win rate 50% → breakeven (zero profit)\n"
            f"  RR=1.5x + win rate 40% → PROFITABIL\n"
            f"  RR=2.0x + win rate 34% → PROFITABIL\n"
            f"  RR=3.0x + win rate 26% → PROFITABIL\n\n"
            f"Concluzie: cu RR bun, câștigând mai puțin de jumătate din tranzacții "
            f"rămâi profitabil. Aceasta este matematica longevității în trading."
        )
    if rvol < 0.7:
        L.append(
            "LECȚIE — VOLUMUL NU MINTE:\n\n"
            "O mișcare pe volum sub 70% din medie este suspectă — posibilă manipulare "
            "sau thin market (piață subțire).\n\n"
            "Regula de aur: 'Nu cumpăra breakout-uri pe volum mic.'\n"
            "Dacă prețul urcă pe volum scăzut, vânzătorii nu sunt convinși că e scump — "
            "cumpărătorii ridică prețul singuri. Mișcarea nu are substanță."
        )
    if rsi < 30:
        L.append(
            "LECȚIE — 'FALLING KNIFE': CUM NU TE RĂNEȘTI:\n\n"
            "RSI sub 30 atrage gândul 'e ieftin, cumpăr'. Aceasta este capcana clasică.\n\n"
            "RSI sub 30 = condiție NECESARĂ, nu SUFICIENTĂ. Ai nevoie de confirmare:\n"
            "  1. Candle bullish solid (Hammer, Pin Bar, Engulfing verde)\n"
            "  2. MACD bullish crossover\n"
            "  3. Nivel clar de suport ținut\n"
            "  4. Volum scăzut pe continuarea scăderii (vânzătorii se epuizează)\n\n"
            "Fără confirmări: 'ieftin poate deveni și mai ieftin'."
        )
    elif rsi > 70:
        L.append(
            "LECȚIE — NU SHORT-UI ÎNTR-UN BULL MARKET:\n\n"
            "NVIDIA în 2023-2024: RSI peste 70 timp de 8 luni consecutive, +400% în acea perioadă. "
            "Activele pot rămâne 'supraevaluate' luni întregi în bull market-uri puternice.\n\n"
            "Short pe RSI ridicat funcționează în piețe laterale sau bearish. "
            "În uptrend puternic = rețetă pentru pierderi mari.\n\n"
            "Dacă ești long: folosește trailing stop pentru a lăsa profitul să curgă."
        )
    if macross == "Golden Cross":
        L.append(
            "LECȚIE — GOLDEN CROSS: ANTICIPARE VS. REACȚIE:\n\n"
            "Golden Cross este puternic dar lagging — apare DUPĂ ce prețul a urcat 10-30%.\n\n"
            "Profesioniștii ANTICIPEAZĂ Golden Cross-ul:\n"
            "  1. MA200 se aplatizează (oprește scăderea)\n"
            "  2. MA50 inversează și începe să urce\n"
            "  3. RSI revine din sub 30 → zona 40-50\n"
            "  4. MACD formează bullish crossover la zero line\n"
            "  5. Volumul crește pe zilele UP\n\n"
            "4 din 5 bifate ÎNAINTE de crossover = momentul optim, nu după."
        )
    elif macross == "Death Cross":
        L.append(
            "LECȚIE — DEATH CROSS: GREȘEALA CLASICĂ:\n\n"
            "Death Cross apare adesea la MIJLOCUL corecției, nu la început. "
            "Vânzarea în panică la Death Cross = vinzi la jumătatea scăderii.\n\n"
            "Strategia corectă: reduce expunerea la PRIMELE semnale "
            "(MA50 se aplatizează, MACD bearish cross, volum crescut pe zile negative). "
            "Nu aști Death Cross-ul complet format."
        )
    # Lecție psihologie — mereu
    L.append(
        "LECȚIE PSIHOLOGIE — LOSS AVERSION:\n\n"
        "Studiile lui Daniel Kahneman (Nobel 2002) arată că durerea unei pierderi de 100 lei "
        "este de 2x mai intensă decât bucuria unui câștig de 100 lei.\n\n"
        "Cum afectează trading-ul:\n"
        "  → Ții pozițiile perdante prea mult ('o să revină')\n"
        "  → Vinzi câștigătoarele prea devreme (frică să nu pierzi profitul)\n"
        "  → Faci revenge trading după pierdere (emoție, nu analiză)\n\n"
        "Soluția: PLAN SCRIS înainte de intrare — Entry, SL, TP, sizing. "
        "Odată în poziție, lasă planul să lucreze, nu emoțiile."
    )
    return "\n\n" + ("─" * 45 + "\n\n").join(L[:2])


# ══════════════════════════════════════════════════════════════════════════════
# GHID PERMANENT DE CITIRE GRAFICE
# ══════════════════════════════════════════════════════════════════════════════

GHID = [
    {
        "titlu": "CANDLESTICK (LUMÂNĂRI JAPONEZE)",
        "sub": [
            ("Ce este o lumânare",
             "Reprezintă o perioadă de timp cu 4 prețuri: Open, High, Low, Close (OHLC).\n"
             "VERDE (bullish): Închidere > Deschidere → prețul a URCAT.\n"
             "ROȘU (bearish):  Închidere < Deschidere → prețul a COBORÂT.\n"
             "Mărimea corpului = forța mișcării. Corp mare = presiune unidirecțională puternică."),
            ("Umbrele (fitilele)",
             "Umbră LUNGĂ sus: prețul a urcat la maxim dar Bears l-au respins.\n"
             "Umbră LUNGĂ jos: prețul a coborât la minim dar Bulls l-au apărat.\n"
             "Umbră scurtă/absentă: direcție clară, presiune uniformă."),
            ("Pattern-uri esențiale",
             "HAMMER: corp mic sus, umbră lungă jos → reversal BULLISH\n"
             "SHOOTING STAR: corp mic jos, umbră lungă sus → reversal BEARISH\n"
             "DOJI: corp inexistent → indeczie pură, potențial inversare\n"
             "BULLISH ENGULFING: candle verde mare înghite candle roșu → BUY puternic\n"
             "BEARISH ENGULFING: candle roșu mare înghite candle verde → SELL puternic\n"
             "MARUBOZU: corp mare fără umbre → trend extrem de puternic"),
            ("Cum aplici",
             "1. Caută pattern-uri LA NIVELURI CHEIE (suport, rezistență, MA50, MA200)\n"
             "2. Confirmă cu VOLUMUL: Hammer pe volum mare = semnal puternic\n"
             "3. Așteptă CONFIRMAREA: candle-ul următor să confirme direcția\n"
             "4. Pattern-urile pe timeframe ZILNIC sau SĂPTĂMÂNAL sunt cele mai fiabile"),
        ]
    },
    {
        "titlu": "RSI — Relative Strength Index (0-100)",
        "sub": [
            ("Ce măsoară",
             "Viteza și amplitudinea mișcărilor din ultimele 14 perioade.\n"
             "Răspunde la: 'A urcat/coborât prețul prea repede față de norma sa?'"),
            ("Zonele RSI",
             "0-30   → SUPRAVÂNDUT: prețul a scăzut prea mult, prea repede\n"
             "30-45  → Zonă slabă: Bears domină, presiunea scade\n"
             "45-55  → Neutru: echilibru, piața caută direcție\n"
             "55-70  → Puternic: Bulls domină, zona cea mai profitabilă pentru trend-following\n"
             "70-100 → SUPRAEVALUAT: prețul a crescut prea mult, prea repede"),
            ("RSI Divergence — cel mai puternic semnal",
             "BULLISH DIVERGENCE:\n"
             "  Prețul face un nou MINIM, dar RSI face un minim MAI MARE\n"
             "  → Presiunea de vânzare slăbește deși prețul coboară\n"
             "  → Semnal puternic de inversare bullish iminent\n\n"
             "BEARISH DIVERGENCE:\n"
             "  Prețul face un nou MAXIM, dar RSI face un maxim MAI MIC\n"
             "  → Momentumul bullish slăbește → inversare sau corecție majoră"),
            ("Greșeli frecvente",
             "✗ A vinde DOAR pe RSI > 70 (poate rămâne acolo luni în bull market)\n"
             "✗ A cumpăra DOAR pe RSI < 30 (poate scădea la 10 în bear market sever)\n"
             "✗ Folosirea RSI izolat fără confirmare de la alt indicator\n"
             "✗ RSI pe 5 minute = zgomot. RSI zilnic sau săptămânal = semnal."),
        ]
    },
    {
        "titlu": "MACD — Moving Average Convergence/Divergence",
        "sub": [
            ("Componentele",
             "Linia MACD  = EMA(12) - EMA(26)  [linia rapidă]\n"
             "Linia Signal = EMA(9) a MACD       [linia lentă]\n"
             "Histogramă  = MACD - Signal        [barele verzi/roșii]\n\n"
             "Răspunde la: 'Care este direcția și forța momentumului?'"),
            ("Semnalele principale",
             "BULLISH CROSSOVER: MACD > Signal → histograma din roșu în verde → BUY\n"
             "BEARISH CROSSOVER: MACD < Signal → histograma din verde în roșu → SELL\n"
             "ZERO LINE CROSS: MACD din negativ în pozitiv = confirmare trend bullish major\n"
             "HISTOGRAMĂ CREȘTE = momentum se accelerează\n"
             "HISTOGRAMĂ SCADE = momentum slăbește (prețul poate urca încă)"),
            ("Cel mai puternic semnal combinat",
             "MACD Bullish Crossover LA ZERO LINE\n"
             "+ Golden Cross MA50/MA200\n"
             "+ Volum crescut (RVOL > 1.5x)\n"
             "+ RSI în zona 40-55\n"
             "= Unul dintre cele mai fiabile setup-uri din analiza tehnică.\n"
             "Apare de câteva ori pe an pe activele majore — merită așteptat."),
            ("Limitări",
             "LAGGING indicator: confirmă, nu prezice.\n"
             "Cel mai eficient pe timeframe ZILNIC și SĂPTĂMÂNAL.\n"
             "Pe timeframe sub 1 oră generează mult zgomot.\n"
             "Funcționează cel mai bine în trenduri clare, mai puțin în range."),
        ]
    },
    {
        "titlu": "MEDII MOBILE — MA20, MA50, MA200",
        "sub": [
            ("Ce reprezintă fiecare",
             "MA20  (~1 lună)   = tendința pe termen scurt  → day traderi, scalpers\n"
             "MA50  (~2.5 luni) = tendința pe termen mediu  → swing traderi, fonduri\n"
             "MA200 (~10 luni)  = tendința pe termen lung   → investitori, fonduri pensii"),
            ("Regulile esențiale",
             "1. Preț DEASUPRA MA200 = Bull Market\n"
             "2. Preț DEDESUBT MA200 = Bear Market\n"
             "3. GOLDEN CROSS (MA50 > MA200) = semnal bullish major pe termen lung\n"
             "4. DEATH CROSS  (MA50 < MA200) = semnal bearish major pe termen lung\n"
             "5. MA50 = SUPORT DINAMIC în uptrend și REZISTENȚĂ DINAMICĂ în downtrend\n"
             "6. Profecia care se autoîmplinește: cu cât mai mulți urmăresc un nivel, "
             "cu atât devine mai puternic."),
            ("Utilizare practică",
             "UPTREND: Cumpără la pullback-uri la MA50 cu candle bullish de confirmare\n"
             "  SL: sub MA50 sau sub ultimul minim semnificativ\n\n"
             "DOWNTREND: Short la retestul MA50 (bear rally) cu candle bearish\n"
             "  SL: deasupra MA50\n\n"
             "TRANZIȚIE (MA50 lângă MA200): Evită tranzacțiile — risc maxim de whipsaw."),
            ("EMA vs SMA",
             "SMA: medie simplă, toate zilele au greutate egală → mai stabilă, mai lentă\n"
             "EMA: zilele recente au greutate mai mare → reacționează mai rapid\n\n"
             "Recomandare: EMA20, EMA50 pentru semnale operative | SMA200 pentru tendința majoră"),
        ]
    },
    {
        "titlu": "BOLLINGER BANDS",
        "sub": [
            ("Structura",
             "Banda Superioară = MA20 + (2 × deviație standard)\n"
             "Banda Mijlocie   = MA20\n"
             "Banda Inferioară = MA20 - (2 × deviație standard)\n\n"
             "95% din timp prețul se află ÎNTRE benzile externe.\n"
             "Benzile se LĂRGESC la volatilitate ridicată și se ÎNGUSTEAZĂ când scade."),
            ("Strategia Bounce",
             "Funcționează cel mai bine în piețe LATERALE (range trading):\n"
             "  BUY: la banda INFERIOARĂ → țintă MA20 sau banda superioară\n"
             "  SELL: la banda SUPERIOARĂ → țintă MA20 sau banda inferioară\n\n"
             "NU aplica în trenduri puternice! În uptrend, prețul 'plimbă' "
             "banda superioară săptămâni întregi."),
            ("Strategia Squeeze Breakout",
             "Squeeze = benzi foarte înguste (sub 4% din preț) = compresie extremă.\n\n"
             "Execuție:\n"
             "1. Identifică squeeze-ul vizual\n"
             "2. BUY STOP puțin deasupra benzii superioare\n"
             "3. SELL STOP puțin sub banda inferioară\n"
             "4. Ordinul activat = direcția breakout → anulează celălalt imediat\n\n"
             "Avantaj: intri DUPĂ confirmare, nu ghicești direcția."),
            ("Lățimea benzilor ca filtru",
             "Benzi LARGI   = volatilitate ridicată, mișcări mari, risc crescut\n"
             "Benzi ÎNGUSTE = volatilitate scăzută, piața 'se odihnește' → explozie iminentă\n"
             "Preț deasupra benzii superioare constant = trend bullish extrem de puternic"),
        ]
    },
    {
        "titlu": "VOLUM & RVOL",
        "sub": [
            ("De ce contează",
             "PREȚUL spune unde s-a dus piața.\n"
             "VOLUMUL spune CÂT DE CONVINSĂ era piața de acea mișcare.\n\n"
             "Axioma nr. 1: 'Volumul precedă prețul.'\n"
             "Mișcare fără volum = neconvingătoare, probabilitate inversare.\n"
             "Mișcare cu volum = participare reală, probabilitate continuare."),
            ("Scala RVOL",
             "RVOL > 2.0x   → Excepțional: eveniment major, instituții active\n"
             "RVOL 1.5-2.0x → Ridicat: mișcare de calitate superioară\n"
             "RVOL 1.0-1.5x → Normal-ridicat: mișcare validă\n"
             "RVOL 0.7-1.0x → Normal: nimic special\n"
             "RVOL < 0.7x   → Scăzut: mișcare suspectă, probabilitate fake-out"),
            ("Regulile de aur",
             "UPTREND SĂNĂTOS: zile UP volum mare + zile DOWN volum mic\n"
             "DOWNTREND SĂNĂTOS: zile DOWN volum mare + zile UP volum mic\n\n"
             "SEMN DE SLĂBIRE: prețul face noi maxime DAR volumul scade\n"
             "= Distribuție: smart money vinde în timp ce retail cumpără\n\n"
             "BREAKOUT VOLUM MIC = CAPCANĂ\n"
             "BREAKOUT VOLUM MARE = REAL, urmărește continuation"),
            ("Acumulare vs Distribuție",
             "ACUMULARE (smart money cumpără discret):\n"
             "  Preț lateral + zile UP volum mare + zile DOWN volum mic\n\n"
             "DISTRIBUȚIE (smart money vinde discret):\n"
             "  Preț lateral + zile DOWN volum mare + zile UP volum mic\n\n"
             "Metodologia Wyckoff (1930) descrie aceste faze în detaliu — "
             "valabilă și astăzi pe orice piață."),
        ]
    },
    {
        "titlu": "SUPORT & REZISTENȚĂ",
        "sub": [
            ("Definiție și logică",
             "SUPORT: nivel unde cumpărătorii opresc și inversează scăderea.\n"
             "REZISTENȚĂ: nivel unde vânzătorii opresc și inversează creșterea.\n\n"
             "De ce există:\n"
             "  → Memorie colectivă: traders care au cumpărat la un nivel cumpără din nou\n"
             "  → Algoritmi programați la niveluri istorice cheie\n"
             "  → Ordine limită plasate în masă la niveluri rotunde (psihologice)"),
            ("Cum identifici nivelurile",
             "1. Maxime și minime anterioare semnificative (min. 2-3 atingeri)\n"
             "2. Zone de consolidare unde prețul a petrecut timp\n"
             "3. Niveluri psihologice rotunde: 1000, 50.000, 100, 1.00\n"
             "4. Medii mobile (MA50, MA200) — suport/rezistență dinamice\n"
             "5. Retracement Fibonacci: 38.2%, 50%, 61.8% din mișcarea anterioară\n\n"
             "Regula: cu cât mai multe atingeri, cu atât mai puternic nivelul."),
            ("Role Reversal — regula de aur",
             "SUPORTUL SPART → devine REZISTENȚĂ\n"
             "REZISTENȚA SPARTĂ → devine SUPORT\n\n"
             "Logica: traders care au cumpărat la suportul spart și sunt în pierdere "
             "VOR VINDE când prețul revine la breakeven → generează presiune la acel nivel.\n\n"
             "Role reversal = punct de intrare cu RISC REDUS și POTENȚIAL RIDICAT."),
            ("Stop Loss optim",
             "NU plasa SL exact la nivelul de suport — piața testează cu câteva procente.\n\n"
             "REGULA: SL = suport MINUS 1-3% (sau 1-2x ATR).\n"
             "Exemplu: suport la 100 → SL la 97-98 (breathing room).\n\n"
             "SL prea strâns = ești scos înainte ca prețul să revină în direcția ta. "
             "Frustrant și costisitor."),
        ]
    },
    {
        "titlu": "MANAGEMENTUL RISCULUI",
        "sub": [
            ("Regula 1-2% per tranzacție",
             "NU risca mai mult de 1-2% din capitalul total pe o singură tranzacție.\n\n"
             "FORMULA position sizing:\n"
             "  Mărime = Risc acceptat ÷ (Entry - Stop Loss)\n\n"
             "Exemplu: Capital 10.000 EUR, risc 1% (100 EUR), Entry 50, SL 48\n"
             "  Mărime = 100 ÷ (50-48) = 50 acțiuni\n\n"
             "De ce contează: 10 pierderi consecutive la 2% = -18% (nu -20%).\n"
             "La 10% per tranzacție: 10 pierderi = -65% (greu de recuperat)."),
            ("RR — matematica supraviețuirii",
             "RR = (TP - Entry) ÷ (Entry - SL)\n\n"
             "Win rate minim necesar pentru profitabilitate:\n"
             "  RR=1.0x → >50%\n"
             "  RR=1.5x → >40%\n"
             "  RR=2.0x → >34%\n"
             "  RR=3.0x → >26%\n\n"
             "Concluzie: cu RR bun, câștigând mai puțin de jumătate din tranzacții "
             "rămâi profitabil. Aceasta este matematica longevității."),
            ("Diversificare și corelație",
             "Max 10-15% din portofoliu pe un singur activ.\n"
             "Max 25-30% pe o singură clasă de active.\n\n"
             "ATENȚIE LA CORELAȚIE:\n"
             "BTC și ETH: corelație 85-90% → nu sunt cu adevărat diversificate\n"
             "NVDA, AMD, INTC: corelație 70-80% — același risc sector\n"
             "Gold și S&P500: NEGATIV corelate în criză → gold = hedge real\n\n"
             "Diversificare adevărată = active care NU se mișcă în același timp."),
            ("Cele 10 reguli de aur",
             "1.  Planifică tranzacția, tranzacționează planul.\n"
             "2.  Niciodată RR sub 1.5x — ideal 2x sau mai mult.\n"
             "3.  Niciodată risc peste 2% din capital per tranzacție.\n"
             "4.  Confirmă breakout-urile cu VOLUMUL.\n"
             "5.  Nu face averaging down neselectiv.\n"
             "6.  Lasă câștigătoarele să curgă, taie pierdătoarele rapid.\n"
             "7.  Nu tranzacționa din plictiseală sau FOMO.\n"
             "8.  Nu revenge trade după pierdere — ia o pauză.\n"
             "9.  Ține jurnal de tranzacții — dacă nu îl măsori, nu îl îmbunătățești.\n"
             "10. Piața există și mâine — capitalul protejat = oportunități viitoare."),
        ]
    },
]

# ══════════════════════════════════════════════════════════════════════════════
# CONSTRUIRE SHEET GHID INVATARE
# ══════════════════════════════════════════════════════════════════════════════

def write_bloc(ws, d, r):
    semnal = d["semnal"]
    f_head = (fill("1E6B3C") if semnal == "BUY" else
              fill("8B0000") if semnal == "SELL" else fill("7D5A00"))
    set_h(ws, r, 24)
    txt = (f"  {d['name']}   |   Preț: {fp(d['now'], 4)}"
           f"   |   Zi: {fpc(d['var_zi'])}"
           f"   |   RSI: {d['rsi']:.1f}"
           f"   |   Confluențe: {d['conf']}/5"
           f"   |   SEMNAL: {semnal}  ")
    c = ws.cell(r, 1, txt)
    c.fill = f_head; c.font = FN_ACTIV
    c.alignment = aln("left", "center")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    r += 1

    secs = [
        ("  DE CE S-A MIȘCAT ASTĂZI",  txt_miscare(d)),
        ("  OPORTUNITATE DE TRADING",   txt_oportunitate(d)),
        ("  PATTERN GRAFIC DETECTAT",   txt_pattern(d)),
        ("  LECȚIA ZILEI",              txt_lectie(d)),
    ]
    for titlu, continut in secs:
        set_h(ws, r, 18)
        h = ws.cell(r, 1, titlu)
        h.fill = F_ASEC; h.font = FN_ASEC
        h.alignment = aln("left", "center")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        r += 1
        nl = max(continut.count("\n") + 1, 3)
        set_h(ws, r, max(nl * 13 + 16, 50))
        c = ws.cell(r, 1, continut)
        c.fill = F_WHITE; c.font = fnt(size=10)
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        r += 1

    set_h(ws, r, 6)
    for col in range(1, 4): ws.cell(r, col).fill = F_SEP
    r += 2
    return r


def write_ghid(ws, r):
    set_h(ws, r, 55)
    c = ws.cell(r, 1,
                "GHID COMPLET DE CITIRE GRAFICE & INDICATORI TEHNICI\n"
                "Referință permanentă — citește zilnic pentru a-ți forma ochiul de trader")
    c.fill = F_GT; c.font = FN_GT
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    r += 2

    for cap in GHID:
        set_h(ws, r, 28)
        c = ws.cell(r, 1, f"  {cap['titlu']}")
        c.fill = F_GC; c.font = FN_GC; c.alignment = aln("left", "center")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        r += 1

        for sub, txt in cap["sub"]:
            set_h(ws, r, 20)
            h = ws.cell(r, 1, f"    ▸  {sub}")
            h.fill = F_GS; h.font = FN_GS; h.alignment = aln("left", "center")
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
            r += 1
            nl = max(txt.count("\n") + 1, 3)
            set_h(ws, r, max(nl * 13 + 16, 60))
            c = ws.cell(r, 1, txt)
            c.fill = F_GTXT; c.font = FN_GTXT
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
            r += 1

        set_h(ws, r, 6)
        for col in range(1, 4): ws.cell(r, col).fill = fill("2E75B6")
        r += 2
    return r


def build_ghid_sheet(wb, all_data):
    SHEET = "GHID INVATARE"
    if SHEET in wb.sheetnames: del wb[SHEET]
    ws = wb.create_sheet(SHEET)
    ws.sheet_properties.tabColor = "154360"
    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 28
    ws.freeze_panes = "A2"

    buy  = sum(1 for d in all_data.values() if d and d.get("semnal") == "BUY")
    sell = sum(1 for d in all_data.values() if d and d.get("semnal") == "SELL")
    wait = sum(1 for d in all_data.values() if d and d.get("semnal") == "WAIT")
    ok   = len([d for d in all_data.values() if d])

    r = 1
    set_h(ws, r, 55)
    c = ws.cell(r, 1,
                f"GHID DE INVATARE ZILNIC — ANALIZA DE PIATA PROFESIONALA\n"
                f"Generat: {datetime.now().strftime('%d.%m.%Y  %H:%M')}   "
                f"|   {ok} active   "
                f"|   BUY: {buy}   SELL: {sell}   WAIT: {wait}")
    c.fill = F_DASH; c.font = FN_DASH
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    r += 2

    for cat, dic in CATEGORII:
        activi = [(nm, tk) for nm, tk in dic.items() if tk in all_data and all_data[tk]]
        if not activi: continue
        set_h(ws, r, 28)
        c = ws.cell(r, 1, f"  ══════  {cat}  ══════")
        c.fill = F_CAT; c.font = FN_CAT; c.alignment = aln("center", "center")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        r += 2
        for nm, tk in activi:
            log.info(f"  Ghid bloc: {nm:25s} [{all_data[tk]['semnal']}]")
            r = write_bloc(ws, all_data[tk], r)
        r += 1

    r += 2
    r = write_ghid(ws, r)
    log.info(f"  ✓ GHID INVATARE — sheet complet")

# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

# ══════════════════════════════════════════════════════════════════════════════
# FIȘĂ ACTIV — selector dinamic + sheet complet
# ══════════════════════════════════════════════════════════════════════════════

REF_ACTIV = "DASHBOARD!I2"   # Celula cu activul selectat de utilizator

def get_all_asset_names(wb):
    """Extrage toate numele de active din INDICATORI TEHNICI."""
    ws = wb["INDICATORI TEHNICI"]
    assets = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        name = row[1]
        if name and "---" not in str(name) and name != "Activ" and name not in assets:
            assets.append(str(name))
    return assets


def update_dashboard_selector(wb, assets):
    """Actualizează TOATE celulele din Dashboard cu formule dinamice XLOOKUP."""
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import CellIsRule

    ws  = wb["DASHBOARD"]
    REF = REF_ACTIV   # = "DASHBOARD!I2"

    # ── Sheet ascuns cu lista activelor ──────────────────────────────────
    if "LIST_ACTIVE" not in wb.sheetnames:
        ws_list = wb.create_sheet("LIST_ACTIVE")
        ws_list.sheet_state = "hidden"
    else:
        ws_list = wb["LIST_ACTIVE"]
    for i, a in enumerate(assets, 1):
        ws_list.cell(i, 1, a)

    # ── Helper: dezactivează un merge dacă există ─────────────────────────
    def unmerge_safe(range_str):
        try:
            ws.merged_cells.remove(range_str)
        except Exception:
            pass

    def unmerge_row(row, col_min, col_max):
        to_rm = [str(m) for m in ws.merged_cells.ranges
                 if m.min_row == row and m.min_col >= col_min and m.max_col <= col_max]
        for mr in to_rm:
            try: ws.merged_cells.remove(mr)
            except Exception: pass

    def set_cell(row, col, val, f_fill=None, f_font=None, h_align="center",
                 fmt=None, merge_to=None):
        c = ws.cell(row, col, val)
        if f_fill: c.fill  = f_fill
        if f_font: c.font  = f_font
        c.alignment = aln(h_align, "center")
        if fmt: c.number_format = fmt
        if merge_to:
            ws.merge_cells(start_row=row, start_column=col,
                           end_row=row, end_column=merge_to)
        return c

    # ════════════════════════════════════════════════════════════════════
    # ROW 2 — Selector activ (H2 eticheta, I2:M2 dropdown)
    # ════════════════════════════════════════════════════════════════════
    unmerge_row(2, 8, 13)
    ws.row_dimensions[2].height = 28

    set_cell(2, 8, "Activ selectat:",
             fill("ED7D31"), fnt(bold=True, color="FFFFFF", size=11), "right")

    sel = ws["I2"]
    if not sel.value or str(sel.value).startswith("Data"):
        sel.value = assets[0]
    sel.fill = fill("FFF2CC"); sel.font = fnt(bold=True, color="7D5A00", size=13)
    sel.alignment = aln("center", "center")
    ws.merge_cells("I2:M2")

    dv = DataValidation(
        type="list",
        formula1=f"LIST_ACTIVE!$A$1:$A${len(assets)}",
        allow_blank=False, showDropDown=False,
        showErrorMessage=True, errorTitle="Activ invalid",
        error="Selecteaza un activ din lista."
    )
    dv.sqref = "I2:M2"
    ws.add_data_validation(dv)

    # ════════════════════════════════════════════════════════════════════
    # RÂNDURILE 4-6 — KPI CARDS (Trend / Volatilitate / Volum / Moment)
    # Fiecare card ocupă 4 coloane: A-D, E-H, I-L, M-P
    # R4 = label card (hardcodat, nu se schimbă)
    # R5 = VALOAREA principală  → DINAMICĂ
    # R6 = subtitlu/detaliu     → DINAMIC
    # ════════════════════════════════════════════════════════════════════
    ws.row_dimensions[4].height = 18
    ws.row_dimensions[5].height = 30
    ws.row_dimensions[6].height = 18

    # Card 1 — TREND GENERAL (A5:D5 și A6:D6)
    unmerge_row(5, 1, 4); unmerge_row(6, 1, 4)
    # Valoare: Trend din INDICATORI TEHNICI col T
    set_cell(5, 1,
             f"=IFERROR(XLOOKUP({REF},'INDICATORI TEHNICI'!B:B,'INDICATORI TEHNICI'!T:T,\"N/A\",0,-1),\"N/A\")",
             fill("C6EFCE"), fnt(bold=True, color="375623", size=14), "center",
             merge_to=4)
    # Subtitlu: MA Cross (Golden/Death)
    set_cell(6, 1,
             f"=IFERROR(XLOOKUP({REF},'INDICATORI TEHNICI'!B:B,'INDICATORI TEHNICI'!W:W,\"\",0,-1),\"\")",
             fill("C6EFCE"), fnt(color="375623", size=9, italic=True), "center",
             merge_to=4)

    # Card 2 — VOLATILITATE (E5:H5 și E6:H6)
    unmerge_row(5, 5, 8); unmerge_row(6, 5, 8)
    # Valoare: ATR formatat
    set_cell(5, 5,
             f"=IFERROR(\"ATR: \"&TEXT(XLOOKUP({REF},'INDICATORI TEHNICI'!B:B,'INDICATORI TEHNICI'!O:O,\"\",0,-1),\"0.0000\"),\"N/A\")",
             fill("FFEB9C"), fnt(bold=True, color="9C6500", size=12), "center",
             merge_to=8)
    # Subtitlu: BB Lățime
    set_cell(6, 5,
             f"=IFERROR(\"BB: \"&TEXT(XLOOKUP({REF},'INDICATORI TEHNICI'!B:B,'INDICATORI TEHNICI'!N:N,\"\",0,-1),\"0.0000\"),\"\")",
             fill("FFEB9C"), fnt(color="9C6500", size=9, italic=True), "center",
             merge_to=8)

    # Card 3 — VOLUM RELATIV (I5:L5 și I6:L6)
    unmerge_row(5, 9, 12); unmerge_row(6, 9, 12)
    # Valoare: RVOL
    set_cell(5, 9,
             f"=IFERROR(TEXT(XLOOKUP({REF},'INDICATORI TEHNICI'!B:B,'INDICATORI TEHNICI'!S:S,\"\",0,-1),\"0.00\")&\"x medie\",\"N/A\")",
             fill("C6EFCE"), fnt(bold=True, color="375623", size=12), "center",
             merge_to=12)
    # Subtitlu: Volum absolut
    set_cell(6, 9,
             f"=IFERROR(\"Vol: \"&TEXT(XLOOKUP({REF},'INDICATORI TEHNICI'!B:B,'INDICATORI TEHNICI'!R:R,\"\",0,-1),\"#,##0\"),\"\")",
             fill("C6EFCE"), fnt(color="375623", size=9, italic=True), "center",
             merge_to=12)

    # Card 4 — MOMENT INTRARE (M5:P5 și M6:P6)
    unmerge_row(5, 13, 16); unmerge_row(6, 13, 16)
    # Valoare: Semnal BUY/SELL/WAIT
    set_cell(5, 13,
             f"=IFERROR(XLOOKUP({REF},'SEMNALE INTRARE'!B:B,'SEMNALE INTRARE'!C:C,\"N/A\",0,-1),\"N/A\")",
             fill("C6EFCE"), fnt(bold=True, color="375623", size=14), "center",
             merge_to=16)
    # Subtitlu: Confluențe
    set_cell(6, 13,
             f"=IFERROR(XLOOKUP({REF},'SEMNALE INTRARE'!B:B,'SEMNALE INTRARE'!J:J,0,0,-1)&\"/5 confluente\",\"\")",
             fill("C6EFCE"), fnt(color="375623", size=9, italic=True), "center",
             merge_to=16)

    # Conditional formatting pe KPI cards R5 (Trend și Moment)
    for rng in ["A5:D5", "M5:P5"]:
        ws.conditional_formatting.add(rng, CellIsRule("equal", ['"Bullish"'],  fill=fill("C6EFCE"), font=fnt(bold=True,color="375623",size=14)))
        ws.conditional_formatting.add(rng, CellIsRule("equal", ['"Bearish"'],  fill=fill("FFC7CE"), font=fnt(bold=True,color="9C0006",size=14)))
        ws.conditional_formatting.add(rng, CellIsRule("equal", ['"Sideways"'], fill=fill("FFEB9C"), font=fnt(bold=True,color="9C6500",size=14)))
        ws.conditional_formatting.add(rng, CellIsRule("equal", ['"BUY"'],      fill=fill("C6EFCE"), font=fnt(bold=True,color="375623",size=14)))
        ws.conditional_formatting.add(rng, CellIsRule("equal", ['"SELL"'],     fill=fill("FFC7CE"), font=fnt(bold=True,color="9C0006",size=14)))
        ws.conditional_formatting.add(rng, CellIsRule("equal", ['"WAIT"'],     fill=fill("FFEB9C"), font=fnt(bold=True,color="9C6500",size=14)))

    # ════════════════════════════════════════════════════════════════════
    # RÂNDURILE 9-17 — SIGNAL BOX (complet dinamic)
    # ════════════════════════════════════════════════════════════════════
    sig_map = {
        9:  ("SEMNAL ACTIV",
             f"=IFERROR(XLOOKUP({REF},'SEMNALE INTRARE'!B:B,'SEMNALE INTRARE'!C:C,\"N/A\",0,-1),\"N/A\")"),
        10: ("Activ analizat",
             f"={REF}"),
        11: ("Entry Price",
             f"=IFERROR(XLOOKUP({REF},'SEMNALE INTRARE'!B:B,'SEMNALE INTRARE'!K:K,\"\",0,-1),\"\")"),
        12: ("Stop Loss (SL)",
             f"=IFERROR(XLOOKUP({REF},'SEMNALE INTRARE'!B:B,'SEMNALE INTRARE'!L:L,\"\",0,-1),\"\")"),
        13: ("Take Profit (TP)",
             f"=IFERROR(XLOOKUP({REF},'SEMNALE INTRARE'!B:B,'SEMNALE INTRARE'!M:M,\"\",0,-1),\"\")"),
        14: ("Risk/Reward Ratio",
             "=IFERROR((G13-G11)/(G11-G12),\"N/A\")"),
        15: ("Confluențe aliniate",
             f"=IFERROR(XLOOKUP({REF},'SEMNALE INTRARE'!B:B,'SEMNALE INTRARE'!J:J,\"\",0,-1)&\"/5\",\"\")"),
        16: ("Probabilitate (%)",
             f"=IFERROR(XLOOKUP({REF},'SEMNALE INTRARE'!B:B,'SEMNALE INTRARE'!O:O,\"\",0,-1),\"\")"),
        17: ("Condiție declanșare",
             f"=IFERROR(XLOOKUP({REF},'SEMNALE INTRARE'!B:B,'SEMNALE INTRARE'!D:D,\"\",0,-1),\"\")"),
    }
    fmt_map = {11:"#,##0.0000", 12:"#,##0.0000", 13:"#,##0.0000",
               14:"0.00\"x\"", 16:"0%"}
    fill_val = {9:fill("C6EFCE"), 12:fill("FFC7CE"), 13:fill("C6EFCE")}
    font_val = {9:fnt(bold=True,color="375623",size=12),
                12:fnt(bold=True,color="9C0006",size=11),
                13:fnt(bold=True,color="375623",size=11)}

    for row, (label, formula) in sig_map.items():
        ws.row_dimensions[row].height = 24
        unmerge_row(row, 1, 6); unmerge_row(row, 7, 16)
        # Label
        lc = ws.cell(row, 1, label)
        lc.fill = fill("D6DCE4"); lc.font = fnt(bold=True, size=10)
        lc.alignment = aln("right", "center")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        # Value
        vc = ws.cell(row, 7, formula)
        vc.fill = fill_val.get(row, fill("F2F2F2"))
        vc.font = font_val.get(row, fnt(size=11))
        vc.alignment = aln("left", "center")
        ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=16)
        if row in fmt_map: vc.number_format = fmt_map[row]

    # CF pe semnal R9
    ws.conditional_formatting.add("G9:P9", CellIsRule("equal", ['"BUY"'],  fill=fill("C6EFCE"), font=fnt(bold=True,color="375623",size=12)))
    ws.conditional_formatting.add("G9:P9", CellIsRule("equal", ['"SELL"'], fill=fill("FFC7CE"), font=fnt(bold=True,color="9C0006",size=12)))
    ws.conditional_formatting.add("G9:P9", CellIsRule("equal", ['"WAIT"'], fill=fill("FFEB9C"), font=fnt(bold=True,color="9C6500",size=12)))

    # ════════════════════════════════════════════════════════════════════
    # RÂNDURILE 21-26 — REZUMAT INDICATORI (complet dinamic)
    # Structura: col A=indicator, B:D=valoare, E:H=status, I:P=observatie
    # ════════════════════════════════════════════════════════════════════
    ind_rows = {
        21: ("RSI(14)",
             f"=IFERROR(TEXT(XLOOKUP({REF},'INDICATORI TEHNICI'!B:B,'INDICATORI TEHNICI'!G:G,\"\",0,-1),\"0.0\"),\"\")",
             f"=IFERROR(XLOOKUP({REF},'INDICATORI TEHNICI'!B:B,'INDICATORI TEHNICI'!H:H,\"\",0,-1),\"\")",
             f"=IFERROR(IF(XLOOKUP({REF},'INDICATORI TEHNICI'!B:B,'INDICATORI TEHNICI'!G:G,0,0,-1)<30,\"Supravandut — potential BUY\",IF(XLOOKUP({REF},'INDICATORI TEHNICI'!B:B,'INDICATORI TEHNICI'!G:G,0,0,-1)>70,\"Supraevaluat — potential SELL\",\"Zona neutra 30-70\")),""\")"),
        22: ("MACD Cross",
             f"=IFERROR(XLOOKUP({REF},'INDICATORI TEHNICI'!B:B,'INDICATORI TEHNICI'!W:W,\"\",0,-1),\"\")",
             f"=IFERROR(IF(ISNUMBER(SEARCH(\"Bullish\",XLOOKUP({REF},'INDICATORI TEHNICI'!B:B,'INDICATORI TEHNICI'!W:W,\"\",0,-1))),\"BUY\",IF(ISNUMBER(SEARCH(\"Bearish\",XLOOKUP({REF},'INDICATORI TEHNICI'!B:B,'INDICATORI TEHNICI'!W:W,\"\",0,-1))),\"SELL\",\"WAIT\")),\"\")",
             f"=IFERROR(\"Hist: \"&TEXT(XLOOKUP({REF},'INDICATORI TEHNICI'!B:B,'INDICATORI TEHNICI'!K:K,\"\",0,-1),\"0.000000\"),\"\")"),
        23: ("MA50 vs MA200",
             f"=IFERROR(XLOOKUP({REF},'INDICATORI TEHNICI'!B:B,'INDICATORI TEHNICI'!V:V,\"\",0,-1),\"\")",
             f"=IFERROR(IF(XLOOKUP({REF},'INDICATORI TEHNICI'!B:B,'INDICATORI TEHNICI'!V:V,\"\",0,-1)=\"Golden Cross\",\"BUY\",IF(XLOOKUP({REF},'INDICATORI TEHNICI'!B:B,'INDICATORI TEHNICI'!V:V,\"\",0,-1)=\"Death Cross\",\"SELL\",\"WAIT\")),\"\")",
             f"=IFERROR(\"MA50: \"&TEXT(XLOOKUP({REF},'INDICATORI TEHNICI'!B:B,'INDICATORI TEHNICI'!E:E,\"\",0,-1),\"#,##0.00\")&\" | MA200: \"&TEXT(XLOOKUP({REF},'INDICATORI TEHNICI'!B:B,'INDICATORI TEHNICI'!F:F,\"\",0,-1),\"#,##0.00\"),\"\")"),
        24: ("VIX (global)",
             "=IFERROR(XLOOKUP(\"VIX\",'INDICATORI MACRO'!A:A,'INDICATORI MACRO'!B:B,\"\"),\"\")",
             "=IFERROR(XLOOKUP(\"VIX\",'INDICATORI MACRO'!A:A,'INDICATORI MACRO'!L:L,\"\"),\"\")",
             "=IFERROR(IF(XLOOKUP(\"VIX\",'INDICATORI MACRO'!A:A,'INDICATORI MACRO'!B:B,0)<15,\"Volatilitate scazuta\",IF(XLOOKUP(\"VIX\",'INDICATORI MACRO'!A:A,'INDICATORI MACRO'!B:B,0)<25,\"Volatilitate moderata\",\"Volatilitate ridicata\")),\"\")"),
        25: ("Fear & Greed",
             "=IFERROR(XLOOKUP(\"Fear & Greed (0-100)\",'INDICATORI MACRO'!A:A,'INDICATORI MACRO'!B:B,\"\"),\"\")",
             "=IFERROR(IF(XLOOKUP(\"Fear & Greed (0-100)\",'INDICATORI MACRO'!A:A,'INDICATORI MACRO'!B:B,0)>60,\"BUY\",IF(XLOOKUP(\"Fear & Greed (0-100)\",'INDICATORI MACRO'!A:A,'INDICATORI MACRO'!B:B,0)<40,\"SELL\",\"WAIT\")),\"\")",
             "=IFERROR(IF(XLOOKUP(\"Fear & Greed (0-100)\",'INDICATORI MACRO'!A:A,'INDICATORI MACRO'!B:B,0)>60,\"Greed — sentiment favorabil\",IF(XLOOKUP(\"Fear & Greed (0-100)\",'INDICATORI MACRO'!A:A,'INDICATORI MACRO'!B:B,0)<40,\"Fear — sentiment negativ\",\"Neutru\")),\"\")"),
        26: ("RVOL activ",
             f"=IFERROR(TEXT(XLOOKUP({REF},'INDICATORI TEHNICI'!B:B,'INDICATORI TEHNICI'!S:S,\"\",0,-1),\"0.00\")&\"x\",\"\")",
             f"=IFERROR(IF(XLOOKUP({REF},'INDICATORI TEHNICI'!B:B,'INDICATORI TEHNICI'!S:S,0,0,-1)>1.3,\"Pozitiv\",IF(XLOOKUP({REF},'INDICATORI TEHNICI'!B:B,'INDICATORI TEHNICI'!S:S,0,0,-1)<0.7,\"Negativ\",\"Neutru\")),\"\")",
             f"=IFERROR(\"Vol: \"&TEXT(XLOOKUP({REF},'INDICATORI TEHNICI'!B:B,'INDICATORI TEHNICI'!R:R,\"\",0,-1),\"#,##0\")&\" | Medie: \"&TEXT(XLOOKUP({REF},'PRETURI VOLUME'!B:B,'PRETURI VOLUME'!L:L,0,0,-1),\"#,##0\"),\"\")"),
    }

    for row, (label, val, status, obs) in ind_rows.items():
        ws.row_dimensions[row].height = 20
        # Dezactivează merge-urile existente
        unmerge_row(row, 2, 4); unmerge_row(row, 5, 8); unmerge_row(row, 9, 16)

        # Col A — indicator label
        ws.cell(row, 1, label).fill  = fill("D6DCE4")
        ws.cell(row, 1).font         = fnt(bold=True, size=10)
        ws.cell(row, 1).alignment    = aln("left")

        # Col B:D — valoare
        cv = ws.cell(row, 2, val)
        cv.fill = fill("F2F2F2"); cv.font = fnt(size=10)
        cv.alignment = aln("center")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)

        # Col E:H — status (cu CF)
        cs = ws.cell(row, 5, status)
        cs.fill = fill("F2F2F2"); cs.font = fnt(size=10, bold=True)
        cs.alignment = aln("center")
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=8)

        # Col I:P — observatie
        co = ws.cell(row, 9, obs)
        co.fill = fill("FFFFFF"); co.font = fnt(size=9, italic=True, color="595959")
        co.alignment = aln("left")
        ws.merge_cells(start_row=row, start_column=9, end_row=row, end_column=16)

    # CF pe status col (E21:H26)
    for rng in ["E21:H26"]:
        ws.conditional_formatting.add(rng, CellIsRule("equal", ['"BUY"'],     fill=fill("C6EFCE"), font=fnt(bold=True,color="375623")))
        ws.conditional_formatting.add(rng, CellIsRule("equal", ['"SELL"'],    fill=fill("FFC7CE"), font=fnt(bold=True,color="9C0006")))
        ws.conditional_formatting.add(rng, CellIsRule("equal", ['"WAIT"'],    fill=fill("FFEB9C"), font=fnt(bold=True,color="9C6500")))
        ws.conditional_formatting.add(rng, CellIsRule("equal", ['"Pozitiv"'], fill=fill("C6EFCE"), font=fnt(bold=True,color="375623")))
        ws.conditional_formatting.add(rng, CellIsRule("equal", ['"Negativ"'], fill=fill("FFC7CE"), font=fnt(bold=True,color="9C0006")))
        ws.conditional_formatting.add(rng, CellIsRule("equal", ['"Neutru"'],  fill=fill("FFEB9C"), font=fnt(bold=True,color="9C6500")))

    log.info("  ✓ DASHBOARD — TOATE celulele actualizate dinamic (KPI + Signal Box + Rezumat)")


def build_fisa_activ(wb, assets):
    """Creează sheet-ul FISA ACTIV cu toate datele legate de activul selectat."""
    from openpyxl.utils import get_column_letter as gcl

    SHEET = "FISA ACTIV"
    if SHEET in wb.sheetnames:
        del wb[SHEET]
    idx = wb.sheetnames.index("DASHBOARD") + 1
    ws  = wb.create_sheet(SHEET, idx)
    ws.sheet_properties.tabColor = "ED7D31"

    for col, w in {"A":22,"B":16,"C":14,"D":14,"E":14,"F":14,"G":14,"H":14}.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A3"

    def mc(r1,c1,r2,c2,val,f=None,fn=None,ha="left"):
        cell = ws.cell(r1,c1,val)
        if f:  cell.fill = f
        if fn: cell.font = fn
        cell.alignment = aln(ha,"center")
        if r2>r1 or c2>c1:
            ws.merge_cells(start_row=r1,start_column=c1,end_row=r2,end_column=c2)

    def sec(row, label, bg="1F4E79"):
        ws.row_dimensions[row].height = 20
        mc(row,1,row,8,f"  {label}",fill(bg),fnt(bold=True,color="FFFFFF",size=10),"left")

    def tabel_hdr(row, hdrs):
        ws.row_dimensions[row].height = 18
        for ci, h in enumerate(hdrs, 1):
            c = ws.cell(row, ci, h)
            c.fill = fill("1F4E79"); c.font = fnt(bold=True,color="FFFFFF",size=10)
            c.alignment = aln("center")

    def ref_row(row, src_sheet, src_cols, col_start=1, merge_to=None):
        """Afișează un rând de date din alt sheet."""
        ws.row_dimensions[row].height = 18
        for ci, sc in enumerate(src_cols, col_start):
            frm = f"=IFERROR('{src_sheet}'!{gcl(sc)}{sc if isinstance(sc,int) else sc},\"\")"
            c = ws.cell(row, ci, frm)
            c.fill = fill("FFFFFF"); c.font = fnt(size=10)
            c.alignment = aln("left")

    def kv_row(row, label, formula, fmt=None, span=8):
        ws.row_dimensions[row].height = 20
        ck = ws.cell(row,1,label)
        ck.fill=fill("D6DCE4"); ck.font=fnt(bold=True,size=10)
        ck.alignment=aln("left"); ck.border=_tb()
        cv = ws.cell(row,2,formula)
        cv.fill=fill("F2F2F2"); cv.font=fnt(size=10)
        cv.alignment=aln("left"); cv.border=_tb()
        if span>1: ws.merge_cells(start_row=row,start_column=2,end_row=row,end_column=span)
        if fmt: cv.number_format=fmt

    def kv2(row, lbl1, frm1, lbl2, frm2, fmt1=None, fmt2=None):
        ws.row_dimensions[row].height = 20
        ws.cell(row,1,lbl1).fill=fill("D6DCE4"); ws.cell(row,1).font=fnt(bold=True,size=10)
        ws.cell(row,1).alignment=aln("left"); ws.cell(row,1).border=_tb()
        c1=ws.cell(row,2,frm1); c1.fill=fill("F2F2F2"); c1.font=fnt(size=10)
        c1.alignment=aln("left"); c1.border=_tb()
        ws.merge_cells(start_row=row,start_column=2,end_row=row,end_column=4)
        if fmt1: c1.number_format=fmt1
        ws.cell(row,5,lbl2).fill=fill("D6DCE4"); ws.cell(row,5).font=fnt(bold=True,size=10)
        ws.cell(row,5).alignment=aln("left"); ws.cell(row,5).border=_tb()
        c2=ws.cell(row,6,frm2); c2.fill=fill("F2F2F2"); c2.font=fnt(size=10)
        c2.alignment=aln("left"); c2.border=_tb()
        ws.merge_cells(start_row=row,start_column=6,end_row=row,end_column=8)
        if fmt2: c2.number_format=fmt2

    def _tb():
        s = Side(style="thin", color="BFBFBF")
        return Border(left=s,right=s,top=s,bottom=s)

    REF = REF_ACTIV
    r = 1

    # ── TITLU ─────────────────────────────────────────────────────────────
    ws.row_dimensions[r].height = 45
    mc(r,1,r,8,"FIȘĂ COMPLETĂ ACTIV — toate datele pentru activul selectat",
       fill("0D2137"),fnt(bold=True,color="FFFFFF",size=14),"center"); r+=1

    ws.row_dimensions[r].height = 30
    mc(r,1,r,3,"Activ analizat:",fill("ED7D31"),fnt(bold=True,color="FFFFFF",size=11),"right")
    mc(r,4,r,8,f"={REF}",fill("FFF2CC"),fnt(bold=True,color="7D5A00",size=13),"center"); r+=1

    ws.row_dimensions[r].height = 20
    mc(r,1,r,8,
       "Schimba activul in DASHBOARD → celula portocalie (dreapta sus) → toate sectiunile se actualizeaza automat",
       fill("FFF9E6"),fnt(italic=True,color="9C6500",size=9),"center"); r+=2

    # ─────────────────────────────────────────────────────────────────────
    # 1. SEMNAL DE INTRARE
    # ─────────────────────────────────────────────────────────────────────
    sec(r, "1 — SEMNAL DE INTRARE"); r+=1
    kv_row(r,"Semnal (BUY/SELL/WAIT)",
           f"=IFERROR(XLOOKUP({REF},'SEMNALE INTRARE'!B:B,'SEMNALE INTRARE'!C:C,\"N/A\",0,-1),\"N/A\")"); r+=1
    kv_row(r,"Condiție declanșare",
           f"=IFERROR(XLOOKUP({REF},'SEMNALE INTRARE'!B:B,'SEMNALE INTRARE'!D:D,\"\",0,-1),\"\")"); r+=1
    kv2(r,"Entry Price",
        f"=IFERROR(XLOOKUP({REF},'SEMNALE INTRARE'!B:B,'SEMNALE INTRARE'!K:K,\"\",0,-1),\"\")","#,##0.0000",
        "Confluențe",
        f"=IFERROR(XLOOKUP({REF},'SEMNALE INTRARE'!B:B,'SEMNALE INTRARE'!J:J,0,0,-1)&\"/5\",\"\")"); r+=1
    kv2(r,"Stop Loss (SL)",
        f"=IFERROR(XLOOKUP({REF},'SEMNALE INTRARE'!B:B,'SEMNALE INTRARE'!L:L,\"\",0,-1),\"\")","#,##0.0000",
        "Take Profit (TP)",
        f"=IFERROR(XLOOKUP({REF},'SEMNALE INTRARE'!B:B,'SEMNALE INTRARE'!M:M,\"\",0,-1),\"\")","#,##0.0000"); r+=1
    kv2(r,"Risk/Reward Ratio",
        f"=IFERROR((XLOOKUP({REF},'SEMNALE INTRARE'!B:B,'SEMNALE INTRARE'!M:M,0,0,-1)"
        f"-XLOOKUP({REF},'SEMNALE INTRARE'!B:B,'SEMNALE INTRARE'!K:K,0,0,-1))"
        f"/(XLOOKUP({REF},'SEMNALE INTRARE'!B:B,'SEMNALE INTRARE'!K:K,0,0,-1)"
        f"-XLOOKUP({REF},'SEMNALE INTRARE'!B:B,'SEMNALE INTRARE'!L:L,0,0,-1)),\"N/A\")","0.00\"x\"",
        "Probabilitate (%)",
        f"=IFERROR(XLOOKUP({REF},'SEMNALE INTRARE'!B:B,'SEMNALE INTRARE'!O:O,\"\",0,-1),\"\")","0%"); r+=1
    kv2(r,"Status semnal",
        f"=IFERROR(XLOOKUP({REF},'SEMNALE INTRARE'!B:B,'SEMNALE INTRARE'!P:P,\"\",0,-1),\"\")","",
        "Ultima actualizare",
        f"=IFERROR(XLOOKUP({REF},'SEMNALE INTRARE'!B:B,'SEMNALE INTRARE'!A:A,\"\",0,-1),\"\")"); r+=2

    # ─────────────────────────────────────────────────────────────────────
    # 2. INDICATORI TEHNICI
    # ─────────────────────────────────────────────────────────────────────
    sec(r, "2 — INDICATORI TEHNICI"); r+=1
    teh_pairs = [
        ("Preț curent",  f"=IFERROR(XLOOKUP({REF},'INDICATORI TEHNICI'!B:B,'INDICATORI TEHNICI'!C:C,\"\",0,-1),\"\")","#,##0.0000",
         "Trend",        f"=IFERROR(XLOOKUP({REF},'INDICATORI TEHNICI'!B:B,'INDICATORI TEHNICI'!T:T,\"\",0,-1),\"\")",""),
        ("MA20",         f"=IFERROR(XLOOKUP({REF},'INDICATORI TEHNICI'!B:B,'INDICATORI TEHNICI'!D:D,\"\",0,-1),\"\")","#,##0.0000",
         "MA Cross",     f"=IFERROR(XLOOKUP({REF},'INDICATORI TEHNICI'!B:B,'INDICATORI TEHNICI'!W:W,\"\",0,-1),\"\")",""),
        ("MA50",         f"=IFERROR(XLOOKUP({REF},'INDICATORI TEHNICI'!B:B,'INDICATORI TEHNICI'!E:E,\"\",0,-1),\"\")","#,##0.0000",
         "MACD Cross",   f"=IFERROR(XLOOKUP({REF},'INDICATORI TEHNICI'!B:B,'INDICATORI TEHNICI'!V:V,\"\",0,-1),\"\")",""),
        ("MA200",        f"=IFERROR(XLOOKUP({REF},'INDICATORI TEHNICI'!B:B,'INDICATORI TEHNICI'!F:F,\"\",0,-1),\"\")","#,##0.0000",
         "MACD",         f"=IFERROR(XLOOKUP({REF},'INDICATORI TEHNICI'!B:B,'INDICATORI TEHNICI'!I:I,\"\",0,-1),\"\")","0.000000"),
        ("RSI(14)",      f"=IFERROR(XLOOKUP({REF},'INDICATORI TEHNICI'!B:B,'INDICATORI TEHNICI'!G:G,\"\",0,-1),\"\")","0.00",
         "MACD Signal",  f"=IFERROR(XLOOKUP({REF},'INDICATORI TEHNICI'!B:B,'INDICATORI TEHNICI'!J:J,\"\",0,-1),\"\")","0.000000"),
        ("RSI Status",   f"=IFERROR(XLOOKUP({REF},'INDICATORI TEHNICI'!B:B,'INDICATORI TEHNICI'!H:H,\"\",0,-1),\"\")","",
         "MACD Hist",    f"=IFERROR(XLOOKUP({REF},'INDICATORI TEHNICI'!B:B,'INDICATORI TEHNICI'!K:K,\"\",0,-1),\"\")","0.000000"),
        ("BB Superior",  f"=IFERROR(XLOOKUP({REF},'INDICATORI TEHNICI'!B:B,'INDICATORI TEHNICI'!L:L,\"\",0,-1),\"\")","#,##0.0000",
         "BB Inferior",  f"=IFERROR(XLOOKUP({REF},'INDICATORI TEHNICI'!B:B,'INDICATORI TEHNICI'!M:M,\"\",0,-1),\"\")","#,##0.0000"),
        ("BB Lățime",    f"=IFERROR(XLOOKUP({REF},'INDICATORI TEHNICI'!B:B,'INDICATORI TEHNICI'!N:N,\"\",0,-1),\"\")","#,##0.0000",
         "ATR",          f"=IFERROR(XLOOKUP({REF},'INDICATORI TEHNICI'!B:B,'INDICATORI TEHNICI'!O:O,\"\",0,-1),\"\")","0.0000"),
        ("Stoch %K",     f"=IFERROR(XLOOKUP({REF},'INDICATORI TEHNICI'!B:B,'INDICATORI TEHNICI'!P:P,\"\",0,-1),\"\")","0.00",
         "Stoch %D",     f"=IFERROR(XLOOKUP({REF},'INDICATORI TEHNICI'!B:B,'INDICATORI TEHNICI'!Q:Q,\"\",0,-1),\"\")","0.00"),
        ("Volum",        f"=IFERROR(XLOOKUP({REF},'INDICATORI TEHNICI'!B:B,'INDICATORI TEHNICI'!R:R,\"\",0,-1),\"\")","#,##0",
         "RVOL",         f"=IFERROR(XLOOKUP({REF},'INDICATORI TEHNICI'!B:B,'INDICATORI TEHNICI'!S:S,\"\",0,-1),\"\")","0.00\"x\""),
        ("Suport cheie", f"=IFERROR(XLOOKUP({REF},'INDICATORI TEHNICI'!B:B,'INDICATORI TEHNICI'!U:U,\"\",0,-1),\"\")","#,##0.0000",
         "Rezistență",   f"=IFERROR(XLOOKUP({REF},'INDICATORI TEHNICI'!B:B,'INDICATORI TEHNICI'!V:V,\"\",0,-1),\"\")","#,##0.0000"),
    ]
    for t in teh_pairs:
        kv2(r, t[0],t[1],t[2], t[3],t[4],t[5]); r+=1
    r+=1

    # ─────────────────────────────────────────────────────────────────────
    # 3. PREȚURI & VOLUME
    # ─────────────────────────────────────────────────────────────────────
    sec(r, "3 — PREȚURI & VOLUME"); r+=1
    pv_pairs = [
        ("Deschidere",    f"=IFERROR(XLOOKUP({REF},'PRETURI VOLUME'!B:B,'PRETURI VOLUME'!D:D,\"\",0,-1),\"\")","#,##0.0000",
         "Închidere",     f"=IFERROR(XLOOKUP({REF},'PRETURI VOLUME'!B:B,'PRETURI VOLUME'!G:G,\"\",0,-1),\"\")","#,##0.0000"),
        ("Maxim",         f"=IFERROR(XLOOKUP({REF},'PRETURI VOLUME'!B:B,'PRETURI VOLUME'!E:E,\"\",0,-1),\"\")","#,##0.0000",
         "Minim",         f"=IFERROR(XLOOKUP({REF},'PRETURI VOLUME'!B:B,'PRETURI VOLUME'!F:F,\"\",0,-1),\"\")","#,##0.0000"),
        ("Var. Zi (%)",   f"=IFERROR(XLOOKUP({REF},'PRETURI VOLUME'!B:B,'PRETURI VOLUME'!H:H,\"\",0,-1),\"\")","0.00%",
         "Var. Săpt (%)", f"=IFERROR(XLOOKUP({REF},'PRETURI VOLUME'!B:B,'PRETURI VOLUME'!I:I,\"\",0,-1),\"\")","0.00%"),
        ("Var. Lună (%)", f"=IFERROR(XLOOKUP({REF},'PRETURI VOLUME'!B:B,'PRETURI VOLUME'!J:J,\"\",0,-1),\"\")","0.00%",
         "Semnal",        f"=IFERROR(XLOOKUP({REF},'PRETURI VOLUME'!B:B,'PRETURI VOLUME'!N:N,\"\",0,-1),\"\")",""),
        ("Volum",         f"=IFERROR(XLOOKUP({REF},'PRETURI VOLUME'!B:B,'PRETURI VOLUME'!K:K,\"\",0,-1),\"\")","#,##0",
         "RVOL",          f"=IFERROR(XLOOKUP({REF},'PRETURI VOLUME'!B:B,'PRETURI VOLUME'!M:M,\"\",0,-1),\"\")","0.00\"x\""),
        ("Medie Vol.20z", f"=IFERROR(XLOOKUP({REF},'PRETURI VOLUME'!B:B,'PRETURI VOLUME'!L:L,\"\",0,-1),\"\")","#,##0",
         "Trend",         f"=IFERROR(XLOOKUP({REF},'PRETURI VOLUME'!B:B,'PRETURI VOLUME'!O:O,\"\",0,-1),\"\")",""),
    ]
    for t in pv_pairs:
        kv2(r,t[0],t[1],t[2],t[3],t[4],t[5]); r+=1
    r+=1

    # ─────────────────────────────────────────────────────────────────────
    # 4. INDICATORI MACRO
    # ─────────────────────────────────────────────────────────────────────
    sec(r, "4 — INDICATORI MACRO (context piață)"); r+=1
    tabel_hdr(r, ["Indicator","Val. curentă","Val. anterioară","Δ%","Impact","Trending","Status","Data"]); r+=1
    for src_r in range(3, 20):
        ws.row_dimensions[r].height = 18
        for ci in range(1, 9):
            frm = f"=IFERROR('INDICATORI MACRO'!{gcl(ci)}{src_r},\"\")"
            c = ws.cell(r, ci, frm)
            c.fill = fill("FFFFFF"); c.font = fnt(size=10); c.alignment = aln("left")
        r+=1
    r+=1

    # ─────────────────────────────────────────────────────────────────────
    # 5. COMPETITORI & SECTOR
    # ─────────────────────────────────────────────────────────────────────
    sec(r, "5 — COMPETITORI & SECTOR"); r+=1
    tabel_hdr(r, ["Entitate","Sector","Cotă %","YoY","Preț","Trending","Revenue","Status"]); r+=1
    for src_r in range(3, 10):
        ws.row_dimensions[r].height = 18
        for ci, sc in enumerate([1,2,3,4,5,6,7,13], 1):
            frm = f"=IFERROR('COMPETITORI SECTOR'!{gcl(sc)}{src_r},\"\")"
            c = ws.cell(r, ci, frm)
            c.fill = fill("FFFFFF"); c.font = fnt(size=10); c.alignment = aln("left")
        r+=1
    r+=1

    # ─────────────────────────────────────────────────────────────────────
    # 6. RISCURI & OPORTUNITĂȚI
    # ─────────────────────────────────────────────────────────────────────
    sec(r, "6 — RISCURI & OPORTUNITĂȚI"); r+=1
    tabel_hdr(r, ["ID","Tip","Categorie","Descriere","Impact","Prob %","Scor","Orizont"]); r+=1
    for src_r in range(3, 15):
        ws.row_dimensions[r].height = 18
        for ci in range(1, 9):
            frm = f"=IFERROR('RISCURI OPORTUNITATI'!{gcl(ci)}{src_r},\"\")"
            c = ws.cell(r, ci, frm)
            c.fill = fill("FFFFFF"); c.font = fnt(size=10); c.alignment = aln("left")
        r+=1
    r+=1

    # ─────────────────────────────────────────────────────────────────────
    # 7. CALENDAR ECONOMIC
    # ─────────────────────────────────────────────────────────────────────
    sec(r, "7 — CALENDAR ECONOMIC"); r+=1
    ws.row_dimensions[r].height = 20
    mc(r,1,r,8,
       "Evenimentele cu impact mare sunt evidențiate în sheet-ul CALENDAR ECONOMIC. "
       "Filtrează coloana 'Activ afectat' după activul tău selectat.",
       fill("FCE4D6"), fnt(italic=True,color="833C00",size=9),"left"); r+=1
    tabel_hdr(r, ["Data & Ora","Eveniment","Țară","Impact","Anterior","Estimare","Actual","Activ"]); r+=1
    for src_r in range(3, 12):
        ws.row_dimensions[r].height = 18
        for ci, sc in enumerate([1,2,3,4,5,6,7,10], 1):
            frm = f"=IFERROR('CALENDAR ECONOMIC'!{gcl(sc)}{src_r},\"\")"
            c = ws.cell(r, ci, frm)
            c.fill = fill("FFFFFF"); c.font = fnt(size=10); c.alignment = aln("left")
        r+=1
    r+=1

    # ─────────────────────────────────────────────────────────────────────
    # 8. JURNAL TRANZACȚII — statistici filtrate pe activ
    # ─────────────────────────────────────────────────────────────────────
    sec(r, "8 — JURNAL TRANZACȚII (statistici activ selectat)"); r+=1

    # Statistici
    ws.row_dimensions[r].height = 22
    for ci, (lbl, frm, fmt) in enumerate([
        ("Total tranzacții",
         f"=IFERROR(COUNTIF('JURNAL TRANZACTII'!D:D,{REF}),0)", "0"),
        ("Win Rate",
         f"=IFERROR(COUNTIFS('JURNAL TRANZACTII'!D:D,{REF},'JURNAL TRANZACTII'!N:N,\">0\")"
         f"/COUNTIF('JURNAL TRANZACTII'!D:D,{REF}),0)","0.0%"),
        ("P&L Total ($)",
         f"=IFERROR(SUMIF('JURNAL TRANZACTII'!D:D,{REF},'JURNAL TRANZACTII'!N:N),0)","#,##0.00"),
        ("P&L Mediu/Trade",
         f"=IFERROR(AVERAGEIF('JURNAL TRANZACTII'!D:D,{REF},'JURNAL TRANZACTII'!N:N),0)","#,##0.00"),
    ], 0):
        col = ci * 2 + 1
        ws.cell(r, col, lbl).fill = fill("D6DCE4")
        ws.cell(r, col).font = fnt(bold=True, size=9)
        ws.cell(r, col).alignment = aln("left")
        cv = ws.cell(r, col+1, frm)
        cv.fill = fill("F2F2F2"); cv.font = fnt(size=10)
        cv.alignment = aln("center")
        if fmt: cv.number_format = fmt
    r+=1

    tabel_hdr(r, ["ID","Data","Activ","L/S","Entry","SL","TP","P&L $"]); r+=1
    for src_r in range(3, 15):
        ws.row_dimensions[r].height = 18
        for ci, sc in enumerate([1,2,4,5,7,8,9,14], 1):
            frm = f"=IFERROR('JURNAL TRANZACTII'!{gcl(sc)}{src_r},\"\")"
            c = ws.cell(r, ci, frm)
            c.fill = fill("FFFFFF"); c.font = fnt(size=10); c.alignment = aln("left")
        r+=1
    ws.row_dimensions[r].height = 20
    mc(r,1,r,8,
       "Sfat: Filtrează coloana D (Activ) în sheet-ul JURNAL TRANZACTII după activul selectat pentru istoricul complet.",
       fill("DEEAF1"), fnt(italic=True,color="1F4E79",size=9),"left"); r+=2

    # ─────────────────────────────────────────────────────────────────────
    # 9. REZUMAT EXECUTIV
    # ─────────────────────────────────────────────────────────────────────
    sec(r, "9 — REZUMAT EXECUTIV"); r+=1
    tabel_hdr(r, ["Indicator","Valoare","","Trending","Observații","","",""]); r+=1
    for src_r in range(7, 15):
        ws.row_dimensions[r].height = 20
        frm1 = f"=IFERROR('REZUMAT EXECUTIV'!A{src_r},\"\")"
        frm2 = f"=IFERROR('REZUMAT EXECUTIV'!B{src_r},\"\")"
        frm4 = f"=IFERROR('REZUMAT EXECUTIV'!D{src_r},\"\")"
        frm5 = f"=IFERROR('REZUMAT EXECUTIV'!E{src_r},\"\")"
        for ci, frm in [(1,frm1),(2,frm2),(4,frm4),(5,frm5)]:
            c = ws.cell(r, ci, frm)
            c.fill = fill("FFFFFF"); c.font = fnt(size=10); c.alignment = aln("left")
        ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=3)
        ws.merge_cells(start_row=r,start_column=5,end_row=r,end_column=8)
        r+=1
    r+=1

    # ─────────────────────────────────────────────────────────────────────
    # 10. ISTORIC & TRENDING
    # ─────────────────────────────────────────────────────────────────────
    sec(r, "10 — ISTORIC & TRENDING (context 24 luni)"); r+=1
    tabel_hdr(r, ["Luna/An","RSI Medie","Preț S&P500","PIB YoY","CPI","VIX","Semnal Luna","Fear&Greed"]); r+=1
    for src_r in range(3, 27):
        ws.row_dimensions[r].height = 18
        for ci, sc in enumerate([1,2,4,5,6,7,8,11], 1):
            frm = f"=IFERROR('ISTORIC TRENDING'!{gcl(sc)}{src_r},\"\")"
            c = ws.cell(r, ci, frm)
            c.fill = fill("FFFFFF"); c.font = fnt(size=10); c.alignment = aln("left")
        r+=1
    r+=1

    # ── Footer ────────────────────────────────────────────────────────────
    ws.row_dimensions[r].height = 25
    mc(r,1,r,8,
       f"Fișă generată automat — actualizare zilnică prin script | "
       f"Selectează activul în DASHBOARD pentru a vedea datele corecte",
       fill("E2EFDA"),fnt(italic=True,color="375623",size=9),"center")

    log.info(f"  ✓ FISA ACTIV — creată cu {r} rânduri pentru {len(assets)} active disponibile")


# ══════════════════════════════════════════════════════════════════════════════
# NOTIFICĂRI — WhatsApp + Google Drive
# ══════════════════════════════════════════════════════════════════════════════

def send_whatsapp(all_data, fear_greed, elapsed):
    """Trimite rezumatul pe WhatsApp via CallMeBot."""
    if not WA_PHONE or not WA_APIKEY:
        return
    ok   = sum(1 for d in all_data.values() if d)
    buy  = sum(1 for d in all_data.values() if d and d.get("semnal") == "BUY")
    sell = sum(1 for d in all_data.values() if d and d.get("semnal") == "SELL")
    wait = ok - buy - sell

    top_buy = sorted(
        [d for d in all_data.values() if d and d.get("semnal") == "BUY"],
        key=lambda x: (x.get("conf",0), x.get("var_zi",0)), reverse=True
    )[:3]
    top_sell = sorted(
        [d for d in all_data.values() if d and d.get("semnal") == "SELL"],
        key=lambda x: (x.get("conf",0), abs(x.get("var_zi",0))), reverse=True
    )[:3]

    lines = [
        f"ANALIZA PIATA — {datetime.now().strftime('%d.%m.%Y %H:%M')}",
        f"Active: {ok} | BUY:{buy} | SELL:{sell} | WAIT:{wait}",
        f"Fear&Greed: {fear_greed.get('display','N/A')}",
        "",
    ]
    if top_buy:
        lines.append("TOP BUY:")
        for d in top_buy:
            lines.append(f"  {d['name']} {fpc(d['var_zi'])} RSI={d['rsi']:.0f} {d['macross']} RVOL={d['rvol']:.1f}x")
    if top_sell:
        lines.append("\nTOP SELL:")
        for d in top_sell:
            lines.append(f"  {d['name']} {fpc(d['var_zi'])} RSI={d['rsi']:.0f} {d['macross']}")
    lines.append(f"\nDurata: {elapsed//60}m {elapsed%60}s | Excel actualizat")

    try:
        r = requests.get("https://api.callmebot.com/whatsapp.php",
                         params={"phone":WA_PHONE,"text":"\n".join(lines),"apikey":WA_APIKEY},
                         timeout=15)
        log.info(f"  WhatsApp: {'trimis ✓' if r.status_code==200 else f'eroare {r.status_code}'}")
    except Exception as e:
        log.error(f"  WhatsApp failed: {e}")


def upload_gdrive(excel_path):
    """Uploadează Excel pe Google Drive via rclone."""
    import subprocess
    if not Path(RCLONE_EXE).exists():
        log.warning("  rclone negasit — skip upload Drive")
        return
    try:
        res = subprocess.run(
            [RCLONE_EXE, "copy", str(excel_path), GDRIVE_DEST, "--update", "--verbose"],
            capture_output=True, text=True, timeout=60
        )
        log.info(f"  Google Drive: {'upload ✓' if res.returncode==0 else f'eroare: {res.stderr[:100]}'}")
    except Exception as e:
        log.error(f"  Google Drive failed: {e}")


def main():
    start = datetime.now()

    # Fișier cu timestamp — template-ul original rămâne intact
    import shutil
    stamp      = start.strftime("%Y-%m-%d_%H-%M")
    EXCEL_PATH = EXCEL_TEMPLATE.parent / f"Analiza_Piata_{stamp}.xlsx"

    log.info("=" * 65)
    log.info("  ANALIZA PIATA + FISA ACTIV + GHID INVATARE — Script Unificat")
    log.info(f"  {start.strftime('%d.%m.%Y  %H:%M:%S')}")
    log.info(f"  Active: {len(ACTIVE)} total "
             f"({len(INDICI)} indici | {len(ACTIUNI)} actiuni | "
             f"{len(CRYPTO)} crypto | {len(VALUTE)} valute | "
             f"{len(MATERII_PRIME)} materii prime)")
    log.info(f"  Fisier output: {EXCEL_PATH.name}")
    log.info("=" * 65)

    if not EXCEL_TEMPLATE.exists():
        log.error(f"Template Excel negasit: {EXCEL_TEMPLATE}")
        sys.exit(1)

    # Copiaza template-ul cu noul nume
    shutil.copy2(str(EXCEL_TEMPLATE), str(EXCEL_PATH))
    log.info(f"  Template copiat: {EXCEL_PATH.name}")

    # ── 1. Preia toate datele ─────────────────────────────────────
    log.info(f"\n  [1/3] Preia date pentru {len(ACTIVE)} active...\n")
    all_data = {}
    total    = len(ACTIVE)
    for i, (name, ticker) in enumerate(ACTIVE.items(), 1):
        d = get_data(name, ticker)
        all_data[ticker] = d
        if d:
            log.info(f"  [{i:3d}/{total}] {name:25s} "
                     f"{fp(d['now'], 4):>14}  "
                     f"{fpc(d['var_zi']):>8}  "
                     f"RSI={d['rsi']:5.1f}  [{d['semnal']}]  "
                     f"({d['n']}z)")
        else:
            log.warning(f"  [{i:3d}/{total}] {name:25s}  -- N/A --")

    log.info("\n  Preia Fear & Greed Index...")
    fg = get_fear_greed()
    log.info(f"  Fear & Greed: {fg.get('display', 'N/A')}")

    log.info("\n  Preia date macro...")
    macro_live = {}
    for name, ticker in MACRO_TICKERS.items():
        d = get_data(name, ticker)
        if d:
            macro_live[name] = d
            log.info(f"  {name:25s} = {fp(d['now'], 4)}")

    ok    = sum(1 for d in all_data.values() if d)
    fails = total - ok
    log.info(f"\n  Procesate: {ok}/{total}" + (f"  ({fails} erori)" if fails else "  ✓"))

    # ── 2. Actualizează sheet-urile de analiză ────────────────────
    log.info("\n  [2/4] Actualizare sheet-uri Excel...\n")
    wb = load_workbook(str(EXCEL_PATH))

    tasks_analiza = [
        ("PRETURI VOLUME",     upd_preturi,  [all_data]),
        ("INDICATORI TEHNICI", upd_tehnic,   [all_data]),
        ("SEMNALE INTRARE",    upd_semnale,  [all_data]),
        ("INDICATORI MACRO",   upd_macro,    [macro_live, fg]),
        ("DASHBOARD",          upd_dashboard,[all_data, fg]),
        ("ISTORIC TRENDING",   upd_istoric,  [all_data, fg]),
    ]
    for sname, fn, args in tasks_analiza:
        if sname not in wb.sheetnames:
            log.warning(f"  Sheet '{sname}' lipsa — skip")
            continue
        try:
            fn(wb[sname], *args)
        except Exception as e:
            log.error(f"  EROARE {sname}: {e}")

    # ── 3. Dashboard selector + Fișă Activ ───────────────────────
    log.info("\n  [3/4] Dashboard selector + Fisa Activ...\n")
    try:
        assets = get_all_asset_names(wb)
        log.info(f"  Active disponibile in selector: {len(assets)}")
        update_dashboard_selector(wb, assets)
        build_fisa_activ(wb, assets)
    except Exception as e:
        log.error(f"  EROARE Fisa Activ: {e}")

    # ── 4. Ghid Învățare ──────────────────────────────────────────
    log.info("\n  [4/4] Generare Ghid Invatare...\n")
    try:
        build_ghid_sheet(wb, all_data)
    except Exception as e:
        log.error(f"  EROARE Ghid Invatare: {e}")

    # ── Salvează ──────────────────────────────────────────────────
    wb.save(str(EXCEL_PATH))
    elapsed = (datetime.now() - start).seconds

    # Curăță fișierele mai vechi de 30 pe Desktop
    fisiere = sorted(
        EXCEL_TEMPLATE.parent.glob("Analiza_Piata_*.xlsx"),
        key=lambda f: f.stat().st_mtime
    )
    for f in fisiere[:-30]:
        try:
            f.unlink()
            log.info(f"  Sters (>30 zile): {f.name}")
        except Exception:
            pass

    # ── Notificări ────────────────────────────────────────────────
    upload_gdrive(EXCEL_PATH)
    send_whatsapp(all_data, fg, elapsed)

    log.info("\n" + "=" * 65)
    log.info(f"  SALVAT CU SUCCES: {EXCEL_PATH.name}")
    log.info(f"  Active OK: {ok}/{total}"
             + (f"  |  {fails} erori" if fails else ""))
    log.info(f"  Sheet-uri: Preturi | Tehnic | Semnale | Macro | Dashboard")
    log.info(f"            | Fisa Activ | Ghid Invatare | Istoric")
    log.info(f"  Durata totala: {elapsed // 60}m {elapsed % 60}s")
    log.info(f"  Finalizat: {datetime.now().strftime('%H:%M:%S')}")
    log.info("=" * 65)


if __name__ == "__main__":
    main()