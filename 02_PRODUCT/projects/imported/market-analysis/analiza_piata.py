#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ANALIZĂ DE PIAȚĂ PROFESIONALĂ — SISTEM COMPLET
Script Python pentru analiza zilnică a 80+ active
Versiune: 3.0 | Python 3.10+ | openpyxl + yfinance + pandas
"""

import os, sys, glob, json, math, time, shutil, logging, subprocess
from copy import copy
from datetime import datetime, timedelta
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed

import numpy as np
import pandas as pd
import yfinance as yf
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.worksheet.datavalidation import DataValidation

TEMPLATE_PATH = r"C:\Users\Marius\Desktop\Analiza_Piata_Profesionala.xlsx"
OUTPUT_DIR = os.path.dirname(TEMPLATE_PATH) or "."
MAX_FILES = 30
YFINANCE_PERIOD = "1y"
WA_PHONE = ""
WA_APIKEY = ""
RCLONE_EXE = r"C:\rclone\rclone.exe"
GDRIVE_DEST = "gdrive:Analiza Piata"
FRED_API_KEY = ""

NOW = datetime.now()
TODAY_STR = NOW.strftime("%Y-%m-%d")
TIMESTAMP_STR = NOW.strftime("%Y-%m-%d_%H-%M")
EXCEL_PATH = os.path.join(OUTPUT_DIR, f"Analiza_Piata_{TIMESTAMP_STR}.xlsx")

log = logging.getLogger("AnalPiata")
log.setLevel(logging.DEBUG)
_fmt = logging.Formatter("%(asctime)s  %(levelname)-7s  %(message)s", datefmt="%H:%M:%S")
_sh = logging.StreamHandler(sys.stdout)
_sh.setLevel(logging.INFO)
_sh.setFormatter(_fmt)
log.addHandler(_sh)
_fh = logging.FileHandler("analiza.log", encoding="utf-8")
_fh.setLevel(logging.DEBUG)
_fh.setFormatter(_fmt)
log.addHandler(_fh)

INDICI = {"S&P 500":"^GSPC","NASDAQ 100":"^NDX","NASDAQ Comp.":"^IXIC","Dow Jones":"^DJI",
    "Russell 2000":"^RUT","DAX Germany":"^GDAXI","FTSE 100":"^FTSE","CAC 40":"^FCHI",
    "Nikkei 225":"^N225","Hang Seng":"^HSI","Shanghai Comp.":"000001.SS",
    "MSCI World ETF":"URTH","MSCI EM ETF":"EEM","BET Romania":"BET.RO"}

ACTIUNI_ETF = {"Apple":"AAPL","Microsoft":"MSFT","NVIDIA":"NVDA","Alphabet":"GOOGL",
    "Amazon":"AMZN","Meta":"META","Tesla":"TSLA","AMD":"AMD","Intel":"INTC",
    "Broadcom":"AVGO","ASML":"ASML","Taiwan Semi":"TSM","Palantir":"PLTR",
    "Salesforce":"CRM","Oracle":"ORCL","JPMorgan":"JPM","Goldman Sachs":"GS",
    "Berkshire B":"BRK-B","Visa":"V","Mastercard":"MA","ExxonMobil":"XOM",
    "Chevron":"CVX","Shell":"SHEL","Caterpillar":"CAT","Boeing":"BA",
    "SPY ETF":"SPY","QQQ ETF":"QQQ","GLD ETF":"GLD","TLT Bond ETF":"TLT","ARKK ETF":"ARKK"}

CRYPTO = {"Bitcoin":"BTC-USD","Ethereum":"ETH-USD","BNB":"BNB-USD","Solana":"SOL-USD",
    "XRP":"XRP-USD","Cardano":"ADA-USD","Avalanche":"AVAX-USD","Dogecoin":"DOGE-USD",
    "Chainlink":"LINK-USD","Polkadot":"DOT-USD","Litecoin":"LTC-USD","Shiba Inu":"SHIB-USD",
    "Polygon":"MATIC-USD","Uniswap":"UNI-USD","Cosmos":"ATOM-USD","Stellar":"XLM-USD",
    "Monero":"XMR-USD","Tron":"TRX-USD","Filecoin":"FIL-USD","Aave":"AAVE-USD",
    "Arbitrum":"ARB-USD","Optimism":"OP-USD","Render":"RNDR-USD","Sui":"SUI-USD","Near":"NEAR-USD"}

FOREX = {"EUR/USD":"EURUSD=X","GBP/USD":"GBPUSD=X","USD/JPY":"USDJPY=X",
    "USD/CHF":"USDCHF=X","AUD/USD":"AUDUSD=X","USD/CAD":"USDCAD=X",
    "EUR/RON":"EURRON=X","USD/RON":"USDRON=X","GBP/RON":"GBPRON=X",
    "EUR/GBP":"EURGBP=X","USD/CNY":"USDCNY=X","USD/TRY":"USDTRY=X"}

MATERII_PRIME = {"Gold":"GC=F","Silver":"SI=F","Platinum":"PL=F","Palladium":"PA=F",
    "Oil WTI":"CL=F","Oil Brent":"BZ=F","Natural Gas":"NG=F","Copper":"HG=F",
    "Corn":"ZC=F","Wheat":"ZW=F","Soybean":"ZS=F","Coffee":"KC=F","Sugar":"SB=F","Cotton":"CT=F"}

MACRO_TICKERS = {"VIX":"^VIX","Yield 10Y":"^TNX","Yield 2Y":"^IRX","USD Index":"DX-Y.NYB"}

ALL_CATEGORIES = [("INDICI BURSIERI",INDICI),("ACȚIUNI & ETF",ACTIUNI_ETF),
    ("CRYPTO",CRYPTO),("VALUTE FOREX",FOREX),("MATERII PRIME",MATERII_PRIME)]

ALL_ASSETS = {}
for _cn, _cd in ALL_CATEGORIES:
    ALL_ASSETS.update(_cd)
CATEGORY_FOR_ASSET = {}
for _cn, _cd in ALL_CATEGORIES:
    for _an in _cd:
        CATEGORY_FOR_ASSET[_an] = _cn

# ═══════════════════════════════════════════════════════════════════════════
# CULORI ȘI STILURI
# ═══════════════════════════════════════════════════════════════════════════
C_HEADER_BG="1F4E79"; C_HEADER_FG="FFFFFF"; C_SUBHEADER_BG="D6DCE4"; C_SUBHEADER_FG="000000"
C_INPUT_BG="D9E1F2"; C_CALC_BG="F2F2F2"; C_CALC_FG="595959"
C_BUY_BG="C6EFCE"; C_BUY_FG="375623"; C_WAIT_BG="FFEB9C"; C_WAIT_FG="9C6500"
C_SELL_BG="FFC7CE"; C_SELL_FG="9C0006"; C_ALERT_BG="FCE4D6"; C_ALERT_FG="833C00"
C_TITLE_BG="0D2137"; C_TITLE_FG="FFFFFF"; C_SEP_BG="1F4E79"; C_SEP_FG="FFFFFF"
C_SELECTOR_BG="FFF2CC"; C_SELECTOR_FG="7D5A00"
C_LABEL_SEL_BG="ED7D31"; C_LABEL_SEL_FG="FFFFFF"

FILL_HEADER=PatternFill("solid",fgColor=C_HEADER_BG)
FILL_SUBHEADER=PatternFill("solid",fgColor=C_SUBHEADER_BG)
FILL_INPUT=PatternFill("solid",fgColor=C_INPUT_BG)
FILL_CALC=PatternFill("solid",fgColor=C_CALC_BG)
FILL_BUY=PatternFill("solid",fgColor=C_BUY_BG)
FILL_WAIT=PatternFill("solid",fgColor=C_WAIT_BG)
FILL_SELL=PatternFill("solid",fgColor=C_SELL_BG)
FILL_ALERT=PatternFill("solid",fgColor=C_ALERT_BG)
FILL_TITLE=PatternFill("solid",fgColor=C_TITLE_BG)
FILL_SEP=PatternFill("solid",fgColor=C_SEP_BG)
FILL_SELECTOR=PatternFill("solid",fgColor=C_SELECTOR_BG)
FILL_LABEL_SEL=PatternFill("solid",fgColor=C_LABEL_SEL_BG)

FONT_HEADER=Font(name="Arial",bold=True,color=C_HEADER_FG,size=10)
FONT_SUBHEADER=Font(name="Arial",bold=True,color=C_SUBHEADER_FG,size=10)
FONT_TITLE=Font(name="Arial",bold=True,color=C_TITLE_FG,size=14)
FONT_SEP=Font(name="Arial",bold=True,color=C_SEP_FG,size=11)
FONT_NORMAL=Font(name="Arial",size=10)
FONT_CALC=Font(name="Arial",size=10,color=C_CALC_FG)
FONT_BUY=Font(name="Arial",bold=True,color=C_BUY_FG,size=10)
FONT_WAIT=Font(name="Arial",bold=True,color=C_WAIT_FG,size=10)
FONT_SELL=Font(name="Arial",bold=True,color=C_SELL_FG,size=10)
FONT_ALERT=Font(name="Arial",color=C_ALERT_FG,size=10)
FONT_SELECTOR=Font(name="Arial",bold=True,color=C_SELECTOR_FG,size=13)
FONT_LABEL_SEL=Font(name="Arial",bold=True,color=C_LABEL_SEL_FG,size=11)

ALIGN_CENTER=Alignment(horizontal="center",vertical="center",wrap_text=True)
ALIGN_LEFT=Alignment(horizontal="left",vertical="center",wrap_text=True)
ALIGN_RIGHT=Alignment(horizontal="right",vertical="center",wrap_text=True)
THIN_BORDER=Border(left=Side(style="thin",color="B4C6E7"),right=Side(style="thin",color="B4C6E7"),
    top=Side(style="thin",color="B4C6E7"),bottom=Side(style="thin",color="B4C6E7"))

TAB_COLORS={"DASHBOARD":"1F4E79","REZUMAT EXECUTIV":"1F4E79","SEMNALE INTRARE":"375623",
    "INDICATORI TEHNICI":"595959","INDICATORI MACRO":"595959","COMPETITORI SECTOR":"595959",
    "PRETURI VOLUME":"595959","RISCURI OPORTUNITATI":"9C0006","CALENDAR ECONOMIC":"595959",
    "JURNAL TRANZACTII":"595959","ISTORIC TRENDING":"595959","GHID INVATARE":"154360",
    "LEGENDA":"595959","FISA ACTIV":"ED7D31"}

COMP_MAP = {
    "INDICI BURSIERI":[("S&P 500","^GSPC"),("NASDAQ 100","^NDX"),("Dow Jones","^DJI"),
        ("DAX Germany","^GDAXI"),("FTSE 100","^FTSE"),("Nikkei 225","^N225")],
    "ACȚIUNI & ETF":[("Apple","AAPL"),("Microsoft","MSFT"),("NVIDIA","NVDA"),
        ("Alphabet","GOOGL"),("Amazon","AMZN"),("Meta","META")],
    "CRYPTO":[("Bitcoin","BTC-USD"),("Ethereum","ETH-USD"),("BNB","BNB-USD"),
        ("Solana","SOL-USD"),("XRP","XRP-USD"),("Cardano","ADA-USD")],
    "VALUTE FOREX":[("EUR/USD","EURUSD=X"),("GBP/USD","GBPUSD=X"),("USD/JPY","USDJPY=X"),
        ("USD/CHF","USDCHF=X"),("AUD/USD","AUDUSD=X"),("USD/CAD","USDCAD=X")],
    "MATERII PRIME":[("Gold","GC=F"),("Silver","SI=F"),("Oil WTI","CL=F"),
        ("Oil Brent","BZ=F"),("Natural Gas","NG=F"),("Copper","HG=F")],
}

RISKS_MAP = {
    "INDICI BURSIERI":[
        ("R01","Risc","Recesiune SUA — contracție economică afectează indicii",5,35,"6-12 luni","Monitorizare GDP, yield curve"),
        ("R02","Risc","Inflație persistentă — presiune pe dobânzi Fed",4,45,"3-6 luni","Urmărire CPI, PCE"),
        ("R03","Risc","Breakdown MA200 — confirmare bear market tehnic",4,30,"1-3 luni","Verificare MA200 pe S&P, NASDAQ"),
        ("R04","Risc","Tensiuni geopolitice — impact pe sentiment",3,40,"Imprevizibil","Monitorizare VIX, Gold, USD"),
        ("O01","Oportunitate","Rally sezonier Q4 — sezonalitate pozitivă",3,55,"1-3 luni","Acumulare treptată pe dip-uri"),
        ("O02","Oportunitate","Pivot Fed — relaxare monetară stimulează piața",4,40,"6-12 luni","Monitorizare dot plot"),
    ],
    "ACȚIUNI & ETF":[
        ("R01","Risc","Reglementări antitrust Big Tech",4,40,"6-12 luni","Monitorizare DOJ/FTC"),
        ("R02","Risc","Comprimarea marjelor — inflație costuri",3,50,"3-6 luni","Analiză quarterly earnings"),
        ("R03","Risc","Miss earnings season",4,35,"1-3 luni","Calendar earnings, whisper numbers"),
        ("R04","Risc","RSI extins overbought — risc corecție 5-10%",3,45,"1-4 săptămâni","Trailing stop loss"),
        ("O01","Oportunitate","AI capex boom — investiții masive AI",5,65,"12-24 luni","Focus NVDA, AVGO, MSFT"),
        ("O02","Oportunitate","Buybacks record — repurchase agresiv",3,60,"6-12 luni","Monitorizare anunțuri"),
    ],
    "CRYPTO":[
        ("R01","Risc","Reglementări SEC — acțiuni vs exchange-uri",5,50,"3-12 luni","Monitorizare decizii SEC"),
        ("R02","Risc","Volatilitate extremă — drawdown 30-50%",5,70,"Permanent","Position sizing redus 2-5%"),
        ("R03","Risc","Hack exchange sau DeFi",4,25,"Imprevizibil","Self-custody, diversificare"),
        ("R04","Risc","Sentiment risk-off global",4,45,"1-3 luni","Monitorizare VIX, corelație BTC-NASDAQ"),
        ("O01","Oportunitate","Bitcoin Halving cycle — reducere supply",5,60,"12-18 luni","DCA pe BTC/ETH"),
        ("O02","Oportunitate","Instituționalizare — ETF-uri spot",4,55,"6-12 luni","Monitorizare flows ETF"),
    ],
    "VALUTE FOREX":[
        ("R01","Risc","Divergență politici monetare Fed vs ECB vs BOJ",4,60,"3-6 luni","Monitorizare rate diferențiale"),
        ("R02","Risc","Intervenție bancă centrală BOJ/SNB",3,35,"Imprevizibil","Alertă niveluri cheie"),
        ("R03","Risc","Flight to safety USD — criză globală",4,40,"1-6 luni","Monitorizare DXY, Gold, VIX"),
        ("R04","Risc","Volatilitate RON — presiune curs leu",3,45,"3-6 luni","Hedging poziții RON"),
        ("O01","Oportunitate","Carry trade — diferențiale dobândă",3,50,"6-12 luni","Perechi cu spread rate mare"),
        ("O02","Oportunitate","Mean reversion — perechi la echilibru",3,55,"1-3 luni","BB + RSI pe perechi majore"),
    ],
    "MATERII PRIME":[
        ("R01","Risc","Încetinire economică — scădere cerere",4,45,"6-12 luni","Monitorizare PMI global"),
        ("R02","Risc","Perturbări supply chain",3,35,"Imprevizibil","Diversificare surse"),
        ("R03","Risc","USD puternic — presiune pe prețuri USD",4,50,"3-6 luni","Corelație inversă DXY-Gold"),
        ("R04","Risc","Tranziție energetică — impact fossil fuels",3,55,"12-36 luni","Rebalansare spre metale rare"),
        ("O01","Oportunitate","Gold safe haven — beneficiază în instabilitate",4,60,"6-12 luni","Alocare 5-10% GLD/fizic"),
        ("O02","Oportunitate","Supply squeeze petrol — OPEC+ reduceri",4,45,"3-6 luni","Monitorizare EIA, OPEC+"),
    ],
}

CALENDAR_MAP = {
    "INDICI BURSIERI":[("FOMC Rate Decision","SUA","Ridicat","Toți indicii"),("Non-Farm Payrolls","SUA","Ridicat","S&P 500, NASDAQ"),
        ("CPI Inflation","SUA","Ridicat","Toți indicii"),("GDP Growth Rate","SUA","Mediu","S&P 500, Dow Jones"),
        ("PMI Manufacturing","SUA","Mediu","Industriali, Russell"),("Earnings Season","SUA","Mediu","S&P 500, NASDAQ")],
    "ACȚIUNI & ETF":[("Earnings Report Big Tech","SUA","Ridicat","AAPL, MSFT, NVDA, GOOGL"),("FOMC Rate Decision","SUA","Ridicat","Toate acțiunile"),
        ("CPI Inflation","SUA","Ridicat","Growth stocks, ARKK"),("Non-Farm Payrolls","SUA","Mediu","Financials, JPM, GS"),
        ("PCE Price Index","SUA","Mediu","Consumer sector"),("Retail Sales MoM","SUA","Mediu","AMZN, consumer")],
    "CRYPTO":[("Bitcoin Halving Impact","Global","Ridicat","BTC, ETH, alts"),("FOMC Rate Decision","SUA","Ridicat","BTC, ETH corelat NASDAQ"),
        ("SEC Crypto Ruling","SUA","Ridicat","XRP, SOL, exchange tokens"),("CPI Inflation","SUA","Mediu","BTC ca hedge inflație"),
        ("ETH Network Upgrade","Global","Mediu","ETH, L2 tokens"),("Macro Risk-Off Event","Global","Mediu","Tot piața crypto")],
    "VALUTE FOREX":[("FOMC Rate Decision","SUA","Ridicat","Toate perechile USD"),("ECB Rate Decision","UE","Ridicat","EUR/USD, EUR/RON"),
        ("BOE Rate Decision","UK","Ridicat","GBP/USD, GBP/RON"),("BOJ Policy Statement","Japonia","Mediu","USD/JPY"),
        ("CPI Inflation SUA","SUA","Mediu","USD index"),("Non-Farm Payrolls","SUA","Mediu","EUR/USD, GBP/USD")],
    "MATERII PRIME":[("OPEC+ Meeting","Global","Ridicat","Oil WTI, Oil Brent"),("EIA Crude Inventories","SUA","Ridicat","CL=F, BZ=F"),
        ("FOMC Rate Decision","SUA","Ridicat","Gold, Silver, DXY"),("China PMI Manufacturing","China","Mediu","Copper, Soy"),
        ("USD Index Movement","SUA","Mediu","Toate materiile prime"),("Geopolitical Risk Event","Global","Mediu","Gold, Oil, Gas")],
}

# ═══════════════════════════════════════════════════════════════════════════
# HELPER FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════════
def fp(val, decimals=4):
    if val is None: return ""
    try:
        v = float(val)
        if math.isnan(v) or math.isinf(v): return ""
        return round(v, decimals)
    except (ValueError, TypeError): return ""

def safe_str(val, default=""): return str(val) if val is not None else default

def safe_write(ws, row, col, value, font=None, fill=None, alignment=None, number_format=None, border=None):
    cell = ws.cell(row=row, column=col)
    if value is not None: cell.value = value
    if font: cell.font = font
    if fill: cell.fill = fill
    cell.alignment = alignment if alignment else ALIGN_CENTER
    if number_format: cell.number_format = number_format
    cell.border = border if border else THIN_BORDER
    return cell

def write_header_row(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        safe_write(ws, row, start_col + i, h, font=FONT_HEADER, fill=FILL_HEADER)

def write_separator(ws, row, text, max_col=16):
    for c in range(1, max_col + 1):
        safe_write(ws, row, c, "", font=FONT_SEP, fill=FILL_SEP)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    safe_write(ws, row, 1, f"— {text} —", font=FONT_SEP, fill=FILL_SEP)
    ws.row_dimensions[row].height = 22

def apply_signal_fill(ws, row, col, value):
    if value is None: return
    v = str(value).upper()
    if v in ("BUY",) or "BULLISH" in v or "POZITIV" in v or "GOLDEN" in v:
        safe_write(ws, row, col, value, font=FONT_BUY, fill=FILL_BUY)
    elif v in ("SELL",) or "BEARISH" in v or "NEGATIV" in v or "DEATH" in v:
        safe_write(ws, row, col, value, font=FONT_SELL, fill=FILL_SELL)
    elif v in ("WAIT",) or "SIDEWAYS" in v or "NEUTRU" in v or "ECHILIBRU" in v:
        safe_write(ws, row, col, value, font=FONT_WAIT, fill=FILL_WAIT)
    else:
        safe_write(ws, row, col, value)

def add_cf_signal(ws, cl, mn, mx):
    rng = f"{cl}{mn}:{cl}{mx}"
    for w, bg, fg in [("BUY",C_BUY_BG,C_BUY_FG),("SELL",C_SELL_BG,C_SELL_FG),("WAIT",C_WAIT_BG,C_WAIT_FG)]:
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=[f'"{w}"'],
            fill=PatternFill("solid",fgColor=bg), font=Font(bold=True,color=fg)))

def add_cf_trend(ws, cl, mn, mx):
    rng = f"{cl}{mn}:{cl}{mx}"
    for w, bg, fg in [("Bullish",C_BUY_BG,C_BUY_FG),("Bearish",C_SELL_BG,C_SELL_FG),("Sideways",C_WAIT_BG,C_WAIT_FG)]:
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=[f'"{w}"'],
            fill=PatternFill("solid",fgColor=bg), font=Font(bold=True,color=fg)))

def add_cf_pozitiv_negativ(ws, cl, mn, mx):
    rng = f"{cl}{mn}:{cl}{mx}"
    for w, bg, fg in [("Pozitiv",C_BUY_BG,C_BUY_FG),("Negativ",C_SELL_BG,C_SELL_FG),("Neutru",C_WAIT_BG,C_WAIT_FG)]:
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=[f'"{w}"'],
            fill=PatternFill("solid",fgColor=bg), font=Font(bold=True,color=fg)))

def add_cf_rsi(ws, cl, mn, mx):
    rng = f"{cl}{mn}:{cl}{mx}"
    ws.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["30"],
        fill=PatternFill("solid",fgColor=C_BUY_BG)))
    ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThan", formula=["70"],
        fill=PatternFill("solid",fgColor=C_SELL_BG)))
    ws.conditional_formatting.add(rng, CellIsRule(operator="between", formula=["30","50"],
        fill=PatternFill("solid",fgColor=C_WAIT_BG)))

def add_cf_rvol_cs(ws, cl, mn, mx):
    rng = f"{cl}{mn}:{cl}{mx}"
    ws.conditional_formatting.add(rng, ColorScaleRule(start_type="num",start_value=0.5,start_color="FFC7CE",
        mid_type="num",mid_value=1.0,mid_color="FFEB9C",end_type="num",end_value=2.0,end_color="C6EFCE"))

def add_cf_score_cs(ws, cl, mn, mx):
    rng = f"{cl}{mn}:{cl}{mx}"
    ws.conditional_formatting.add(rng, ColorScaleRule(start_type="num",start_value=0,start_color="C6EFCE",
        mid_type="num",mid_value=10,mid_color="FFEB9C",end_type="num",end_value=20,end_color="FFC7CE"))

def add_cf_pnl(ws, cl, mn, mx):
    rng = f"{cl}{mn}:{cl}{mx}"
    ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThan", formula=["0"],
        fill=PatternFill("solid",fgColor=C_BUY_BG), font=Font(color=C_BUY_FG)))
    ws.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["0"],
        fill=PatternFill("solid",fgColor=C_SELL_BG), font=Font(color=C_SELL_FG)))

def set_col_widths(ws, wd):
    for c, w in wd.items(): ws.column_dimensions[c].width = w

def unmerge_all(ws):
    for m in list(ws.merged_cells.ranges): ws.unmerge_cells(str(m))

# ═══════════════════════════════════════════════════════════════════════════
# DATA FETCHING
# ═══════════════════════════════════════════════════════════════════════════
def get_yf_data(ticker, period=YFINANCE_PERIOD):
    try:
        t = yf.Ticker(ticker)
        df = t.history(period=period)
        if df is None or len(df) < 5:
            log.warning(f"Date insuficiente {ticker}")
            return None
        return df
    except Exception as e:
        log.warning(f"Eroare yfinance {ticker}: {e}")
        return None

def get_all_data_parallel(assets_dict, max_workers=10):
    results = {}
    with ThreadPoolExecutor(max_workers=max_workers) as ex:
        futures = {ex.submit(get_yf_data, tk): nm for nm, tk in assets_dict.items()}
        for f in as_completed(futures):
            nm = futures[f]
            try:
                df = f.result()
                if df is not None: results[nm] = df
            except Exception as e: log.warning(f"Eroare paralel {nm}: {e}")
    return results

def get_fear_greed():
    try:
        r = requests.get("https://api.alternative.me/fng/?limit=1", timeout=10)
        d = r.json()
        return int(d["data"][0]["value"]), d["data"][0]["value_classification"]
    except Exception as e:
        log.warning(f"Fear & Greed API error: {e}")
        return None, None

def get_fred_data(series_id):
    if not FRED_API_KEY: return None, None
    try:
        r = requests.get(f"https://api.stlouisfed.org/fred/series/observations",
            params={"series_id":series_id,"api_key":FRED_API_KEY,"file_type":"json","sort_order":"desc","limit":2}, timeout=10)
        obs = r.json().get("observations",[])
        if obs: return float(obs[0]["value"]), float(obs[1]["value"]) if len(obs)>1 else None
    except: pass
    return None, None

# ═══════════════════════════════════════════════════════════════════════════
# CALCUL INDICATORI TEHNICI
# ═══════════════════════════════════════════════════════════════════════════
def calc_rsi(series, period=14):
    delta = series.diff()
    gain = delta.where(delta > 0, 0.0)
    loss = (-delta).where(delta < 0, 0.0)
    avg_gain = gain.rolling(window=period, min_periods=period).mean()
    avg_loss = loss.rolling(window=period, min_periods=period).mean()
    for i in range(period, len(series)):
        avg_gain.iloc[i] = (avg_gain.iloc[i-1]*(period-1)+gain.iloc[i])/period
        avg_loss.iloc[i] = (avg_loss.iloc[i-1]*(period-1)+loss.iloc[i])/period
    rs = avg_gain / avg_loss
    return 100 - (100 / (1 + rs))

def calc_rsi_status(rsi):
    if rsi is None or math.isnan(rsi): return "N/A"
    if rsi < 30: return "Presiune excesivă vânzare"
    elif rsi < 45: return "Presiune moderată vânzare"
    elif rsi <= 55: return "Echilibru"
    elif rsi <= 70: return "Momentum ascendent"
    else: return "Presiune excesivă cumpărare"

def calc_macd(series, fast=12, slow=26, signal=9):
    ef = series.ewm(span=fast,adjust=False).mean()
    es = series.ewm(span=slow,adjust=False).mean()
    ml = ef - es
    sl = ml.ewm(span=signal,adjust=False).mean()
    return ml, sl, ml - sl

def calc_macd_status(macd, sig, hist, prev_hist):
    if any(v is None or (isinstance(v,float) and math.isnan(v)) for v in [macd,sig,hist]): return "N/A"
    if macd > sig:
        if prev_hist is not None and not math.isnan(prev_hist) and prev_hist <= 0 and hist > 0:
            return "Impuls pozitiv nou"
        return "Impuls pozitiv activ"
    else:
        if prev_hist is not None and not math.isnan(prev_hist) and prev_hist >= 0 and hist < 0:
            return "Impuls negativ nou"
        return "Impuls negativ activ"

def calc_bollinger(series, period=20, sd=2):
    sma = series.rolling(window=period).mean()
    std = series.rolling(window=period).std()
    return sma + std*sd, sma - std*sd, (sma + std*sd) - (sma - std*sd)

def calc_atr(df, period=14):
    h, l, c = df["High"], df["Low"], df["Close"].shift(1)
    tr = pd.concat([h-l, (h-c).abs(), (l-c).abs()], axis=1).max(axis=1)
    return tr.rolling(window=period).mean()

def calc_stochastic(df, k_period=14, d_period=3):
    lm = df["Low"].rolling(window=k_period).min()
    hm = df["High"].rolling(window=k_period).max()
    k = 100 * (df["Close"] - lm) / (hm - lm)
    return k, k.rolling(window=d_period).mean()

def calc_ma_cross(ma50, ma200):
    if ma50 is None or ma200 is None: return "Neutru"
    try:
        if math.isnan(ma50) or math.isnan(ma200): return "Neutru"
    except: return "Neutru"
    return "Golden Cross" if ma50 > ma200 else ("Death Cross" if ma50 < ma200 else "Neutru")

def calc_trend(price, ma50):
    if price is None or ma50 is None: return "Sideways"
    try:
        if math.isnan(price) or math.isnan(ma50): return "Sideways"
    except: return "Sideways"
    if price > ma50*1.01: return "Bullish"
    elif price < ma50*0.99: return "Bearish"
    return "Sideways"

def calc_rvol(vol, period=20):
    av = vol.rolling(window=period).mean()
    if av.iloc[-1] and av.iloc[-1] > 0: return vol.iloc[-1] / av.iloc[-1]
    return 1.0

def calc_momentum(series, period=10):
    if len(series) < period+1: return None
    return ((series.iloc[-1] - series.iloc[-period-1]) / series.iloc[-period-1]) * 100

def calc_support_resistance(df, lookback=20):
    r = df.tail(lookback)
    return r["Low"].min(), r["High"].max()

def calc_signal_score(rsi, macd_st, ma_cross, rvol):
    s = 0
    if rsi is not None and not math.isnan(rsi):
        if rsi < 35: s += 2
        elif rsi < 45: s += 1
        elif rsi > 75: s -= 2
        elif rsi > 65: s -= 1
    if macd_st == "Impuls pozitiv nou": s += 2
    elif macd_st == "Impuls pozitiv activ": s += 1
    elif macd_st == "Impuls negativ nou": s -= 2
    elif macd_st == "Impuls negativ activ": s -= 1
    if ma_cross == "Golden Cross": s += 2
    elif ma_cross == "Death Cross": s -= 2
    if rvol is not None and not math.isnan(rvol):
        if rvol > 1.5: s += 1
        elif rvol < 0.6: s -= 1
    return s

def calc_signal(score):
    if score >= 3: return "BUY"
    elif score <= -3: return "SELL"
    return "WAIT"

def calc_confluences(score): return min(abs(score), 5)

def calc_sl_tp(entry, atr, signal):
    if entry is None or atr is None: return None, None
    try:
        if math.isnan(entry) or math.isnan(atr): return None, None
    except: return None, None
    if signal == "BUY": return entry - 1.5*atr, entry + 3.0*atr
    elif signal == "SELL": return entry + 1.5*atr, entry - 3.0*atr
    return entry - 1.5*atr, entry + 3.0*atr

def calc_probability(conf, rvol):
    p = 35 + conf * 10
    if rvol is not None and not math.isnan(rvol) and rvol > 1.2: p += 5
    return min(90, p) / 100.0

def calc_condition(rsi, macd_st, ma_cross, trend, rvol):
    parts = []
    if rsi is not None and not math.isnan(rsi): parts.append(f"RSI={rsi:.0f}")
    if macd_st and macd_st != "N/A": parts.append(macd_st)
    if ma_cross and ma_cross != "Neutru": parts.append(ma_cross)
    if trend: parts.append(f"Trend={trend}")
    if rvol is not None and not math.isnan(rvol): parts.append(f"RVOL={rvol:.2f}x")
    return " | ".join(parts) if parts else "Date insuficiente"

# ═══════════════════════════════════════════════════════════════════════════
# PROCESS ASSET
# ═══════════════════════════════════════════════════════════════════════════
def process_asset(name, df):
    r = {"name": name, "error": False}
    try:
        close = df["Close"]
        price = close.iloc[-1]
        r["price"] = fp(price)
        r["open"] = fp(df["Open"].iloc[-1])
        r["high"] = fp(df["High"].iloc[-1])
        r["low"] = fp(df["Low"].iloc[-1])
        r["close"] = fp(price)
        vol = df["Volume"].iloc[-1] if "Volume" in df.columns else 0
        r["volume"] = int(vol) if vol and not math.isnan(vol) else 0
        ma20 = close.rolling(20).mean().iloc[-1] if len(close) >= 20 else None
        ma50 = close.rolling(50).mean().iloc[-1] if len(close) >= 50 else None
        ma200 = close.rolling(200).mean().iloc[-1] if len(close) >= 200 else None
        r["ma20"], r["ma50"], r["ma200"] = fp(ma20), fp(ma50), fp(ma200)
        rsi_s = calc_rsi(close)
        rsi_val = rsi_s.iloc[-1] if len(rsi_s) > 0 else None
        r["rsi"] = fp(rsi_val, 1)
        r["rsi_status"] = calc_rsi_status(rsi_val) if rsi_val and not math.isnan(rsi_val) else "N/A"
        ml, sl, hist = calc_macd(close)
        mv, sv, hv = (ml.iloc[-1] if len(ml)>0 else None), (sl.iloc[-1] if len(sl)>0 else None), (hist.iloc[-1] if len(hist)>0 else None)
        ph = hist.iloc[-2] if len(hist) > 1 else None
        r["macd"], r["macd_signal"], r["macd_hist"] = fp(mv), fp(sv), fp(hv)
        r["macd_status"] = calc_macd_status(mv, sv, hv, ph)
        bb_u, bb_l, bb_w = calc_bollinger(close)
        r["bb_upper"], r["bb_lower"], r["bb_width"] = fp(bb_u.iloc[-1]), fp(bb_l.iloc[-1]), fp(bb_w.iloc[-1])
        atr_s = calc_atr(df)
        atr_val = atr_s.iloc[-1] if len(atr_s) > 0 else None
        r["atr"] = fp(atr_val)
        k, d = calc_stochastic(df)
        r["stoch_k"] = fp(k.iloc[-1], 1) if len(k) > 0 else ""
        r["stoch_d"] = fp(d.iloc[-1], 1) if len(d) > 0 else ""
        rvol = calc_rvol(df["Volume"]) if "Volume" in df.columns and df["Volume"].sum() > 0 else 1.0
        r["rvol"] = fp(rvol, 2)
        r["avg_vol_20"] = int(df["Volume"].rolling(20).mean().iloc[-1]) if "Volume" in df.columns and len(df) >= 20 else 0
        r["ma_cross"] = calc_ma_cross(ma50, ma200)
        r["trend"] = calc_trend(price, ma50)
        r["momentum"] = fp(calc_momentum(close), 2)
        sup, res = calc_support_resistance(df)
        r["support"], r["resistance"] = fp(sup), fp(res)
        vz = ((close.iloc[-1]-close.iloc[-2])/close.iloc[-2]) if len(close)>1 else 0
        vs = ((close.iloc[-1]-close.iloc[-5])/close.iloc[-5]) if len(close)>5 else 0
        vl = ((close.iloc[-1]-close.iloc[-21])/close.iloc[-21]) if len(close)>21 else 0
        r["var_zi"], r["var_sapt"], r["var_luna"] = fp(vz*100,2), fp(vs*100,2), fp(vl*100,2)
        score = calc_signal_score(rsi_val, r["macd_status"], r["ma_cross"], rvol)
        r["score"] = score
        signal = calc_signal(score)
        r["signal"] = signal
        r["confluences"] = calc_confluences(score)
        sl_v, tp_v = calc_sl_tp(price, atr_val if atr_val and not math.isnan(atr_val) else None, signal)
        r["entry"], r["sl"], r["tp"] = fp(price), fp(sl_v), fp(tp_v)
        r["probability"] = calc_probability(r["confluences"], rvol if isinstance(rvol, float) else None)
        r["condition"] = calc_condition(rsi_val, r["macd_status"], r["ma_cross"], r["trend"], rvol)
        seasons = {1:"Iarnă",2:"Iarnă",3:"Primăvară",4:"Primăvară",5:"Primăvară",6:"Vară",
                   7:"Vară",8:"Vară",9:"Toamnă",10:"Toamnă",11:"Toamnă",12:"Iarnă"}
        r["season"] = seasons.get(NOW.month, "N/A")
        r["season_factor"], r["status"] = "", "Activ"
    except Exception as e:
        log.warning(f"Eroare procesare {name}: {e}")
        r["error"] = True
        r["signal"] = "WAIT"
        r["condition"] = f"Eroare: {str(e)[:50]}"
        for k in ["price","open","high","low","close","volume","ma20","ma50","ma200","rsi","rsi_status",
                   "macd","macd_signal","macd_hist","macd_status","bb_upper","bb_lower","bb_width",
                   "atr","stoch_k","stoch_d","rvol","avg_vol_20","ma_cross","trend","momentum",
                   "support","resistance","var_zi","var_sapt","var_luna","score","confluences",
                   "entry","sl","tp","probability","season","season_factor","status"]:
            r.setdefault(k, "")
    return r

# ═══════════════════════════════════════════════════════════════════════════
# TEMPLATE CREATION
# ═══════════════════════════════════════════════════════════════════════════
def create_template(path):
    log.info("Creare template gol...")
    wb = Workbook()
    sns = ["DASHBOARD","REZUMAT EXECUTIV","SEMNALE INTRARE","INDICATORI TEHNICI",
        "INDICATORI MACRO","COMPETITORI SECTOR","PRETURI VOLUME","RISCURI OPORTUNITATI",
        "CALENDAR ECONOMIC","JURNAL TRANZACTII","ISTORIC TRENDING","GHID INVATARE","LEGENDA","FISA ACTIV","LIST_ACTIVE"]
    wb.active.title = sns[0]
    for sn in sns[1:]: wb.create_sheet(sn)
    for sn in sns:
        if sn in TAB_COLORS: wb[sn].sheet_properties.tabColor = TAB_COLORS[sn]
    ws_list = wb["LIST_ACTIVE"]
    r = 1
    for cn, cd in ALL_CATEGORIES:
        for an in cd:
            ws_list.cell(row=r, column=1, value=an)
            r += 1
    d = os.path.dirname(path)
    if d and not os.path.exists(d): os.makedirs(d, exist_ok=True)
    wb.save(path)
    log.info(f"Template creat: {path}")
    return path

# ═══════════════════════════════════════════════════════════════════════════
# SHEET UPDATERS — PRETURI VOLUME
# ═══════════════════════════════════════════════════════════════════════════
def upd_preturi_volume(wb, all_data):
    ws = wb["PRETURI VOLUME"]; unmerge_all(ws)
    hdrs = ["Data","Produs/Activ","Deschidere","Maxim","Minim","Închidere",
        "Var. Zi (%)","Var. Săpt (%)","Var. Lună (%)","Volum","Medie Vol 20z","RVOL","Sezon","Factor sezonier (%)","Trend"]
    safe_write(ws,1,1,"💰 PREȚURI, VOLUME & SEZONALITATE",font=FONT_TITLE,fill=FILL_TITLE,alignment=ALIGN_LEFT)
    for c in range(2,16): safe_write(ws,1,c,"",fill=FILL_TITLE)
    ws.merge_cells("A1:O1"); ws.row_dimensions[1].height = 30
    write_header_row(ws,2,hdrs); ws.row_dimensions[2].height = 20
    row = 3
    for cn, cd in ALL_CATEGORIES:
        write_separator(ws,row,cn,15); row += 1
        for an in cd:
            d = all_data.get(an,{})
            vals = [TODAY_STR,an,d.get("open",""),d.get("high",""),d.get("low",""),d.get("close",""),
                d.get("var_zi",""),d.get("var_sapt",""),d.get("var_luna",""),d.get("volume",""),
                d.get("avg_vol_20",""),d.get("rvol",""),d.get("season",""),d.get("season_factor",""),d.get("trend","")]
            for c, v in enumerate(vals, 1):
                if c == 15: apply_signal_fill(ws,row,c,v)
                else: safe_write(ws,row,c,v,font=FONT_NORMAL)
            ws.row_dimensions[row].height = 18; row += 1
    add_cf_trend(ws,"O",3,row-1); add_cf_rvol_cs(ws,"L",3,row-1)
    add_cf_pnl(ws,"G",3,row-1); add_cf_pnl(ws,"H",3,row-1); add_cf_pnl(ws,"I",3,row-1)
    set_col_widths(ws,{"A":12,"B":18,"C":14,"D":14,"E":14,"F":14,"G":12,"H":12,"I":12,"J":14,"K":14,"L":10,"M":10,"N":14,"O":12})
    ws.freeze_panes = "A3"
    log.info(f"PRETURI VOLUME: {row-3} rânduri")

# ═══════════════════════════════════════════════════════════════════════════
# INDICATORI TEHNICI
# ═══════════════════════════════════════════════════════════════════════════
def upd_indicatori_tehnici(wb, all_data):
    ws = wb["INDICATORI TEHNICI"]; unmerge_all(ws)
    hdrs = ["Data","Activ","Preț","MA20","MA50","MA200","RSI(14)","RSI Status",
        "MACD","MACD Signal","MACD Hist","BB Superior","BB Inferior","BB Lățime","ATR",
        "Stoch %K","Stoch %D","Volum","RVOL","Trend","Suport cheie","Rezistență","MA Cross"]
    safe_write(ws,1,1,"📉 INDICATORI TEHNICI",font=FONT_TITLE,fill=FILL_TITLE,alignment=ALIGN_LEFT)
    for c in range(2,24): safe_write(ws,1,c,"",fill=FILL_TITLE)
    ws.merge_cells("A1:W1"); ws.row_dimensions[1].height = 30
    write_header_row(ws,2,hdrs); ws.row_dimensions[2].height = 20
    row = 3
    for cn, cd in ALL_CATEGORIES:
        write_separator(ws,row,cn,23); row += 1
        for an in cd:
            d = all_data.get(an,{})
            vals = [TODAY_STR,an,d.get("price",""),d.get("ma20",""),d.get("ma50",""),d.get("ma200",""),
                d.get("rsi",""),d.get("rsi_status",""),d.get("macd",""),d.get("macd_signal",""),
                d.get("macd_hist",""),d.get("bb_upper",""),d.get("bb_lower",""),d.get("bb_width",""),
                d.get("atr",""),d.get("stoch_k",""),d.get("stoch_d",""),d.get("volume",""),
                d.get("rvol",""),d.get("trend",""),d.get("support",""),d.get("resistance",""),d.get("ma_cross","")]
            for c, v in enumerate(vals, 1):
                if c in (8,20,23): apply_signal_fill(ws,row,c,v)
                else: safe_write(ws,row,c,v,font=FONT_NORMAL)
            ws.row_dimensions[row].height = 18; row += 1
    add_cf_rsi(ws,"G",3,row-1); add_cf_trend(ws,"T",3,row-1); add_cf_rvol_cs(ws,"S",3,row-1)
    set_col_widths(ws,{"A":12,"B":18,"C":14,"D":14,"E":14,"F":14,"G":10,"H":26,"I":12,"J":12,
        "K":12,"L":14,"M":14,"N":12,"O":12,"P":10,"Q":10,"R":14,"S":10,"T":12,"U":14,"V":14,"W":14})
    ws.freeze_panes = "A3"
    log.info(f"INDICATORI TEHNICI: {row-3} rânduri")

# ═══════════════════════════════════════════════════════════════════════════
# SEMNALE INTRARE
# ═══════════════════════════════════════════════════════════════════════════
def upd_semnale_intrare(wb, all_data):
    ws = wb["SEMNALE INTRARE"]; unmerge_all(ws)
    hdrs = ["Data","Activ","Semnal","Condiție declanșare","RSI(14)","Impuls MACD","MA Cross",
        "Volum vs Medie","Momentum 10z","Confluențe","Entry","Stop Loss","Take Profit",
        "RR Ratio","Probabilitate","Status","Note"]
    safe_write(ws,1,1,"📈 SEMNALE DE INTRARE PE PIAȚĂ",font=FONT_TITLE,fill=FILL_TITLE,alignment=ALIGN_LEFT)
    for c in range(2,18): safe_write(ws,1,c,"",fill=FILL_TITLE)
    ws.merge_cells("A1:Q1"); ws.row_dimensions[1].height = 30
    write_header_row(ws,2,hdrs); ws.row_dimensions[2].height = 20
    row = 3
    for cn, cd in ALL_CATEGORIES:
        write_separator(ws,row,cn,17); row += 1
        for an in cd:
            d = all_data.get(an,{})
            rv = d.get("rvol","")
            vvm = f"{rv}x" if rv != "" else ""
            vals = [TODAY_STR,an,d.get("signal","WAIT"),d.get("condition",""),d.get("rsi",""),
                d.get("macd_status",""),d.get("ma_cross",""),vvm,d.get("momentum",""),
                d.get("confluences",""),d.get("entry",""),d.get("sl",""),d.get("tp","")]
            for c, v in enumerate(vals, 1):
                if c == 3: apply_signal_fill(ws,row,c,v)
                else: safe_write(ws,row,c,v,font=FONT_NORMAL)
            safe_write(ws,row,14,f'=IFERROR((M{row}-K{row})/(K{row}-L{row}),"N/A")',font=FONT_NORMAL,number_format='0.00"x"')
            safe_write(ws,row,15,d.get("probability",""),font=FONT_NORMAL,number_format='0%')
            safe_write(ws,row,16,d.get("status","Activ"),font=FONT_NORMAL)
            safe_write(ws,row,17,"",font=FONT_NORMAL)
            ws.row_dimensions[row].height = 18; row += 1
    add_cf_signal(ws,"C",3,row-1); add_cf_signal(ws,"P",3,row-1)
    set_col_widths(ws,{"A":12,"B":18,"C":10,"D":40,"E":10,"F":22,"G":14,"H":14,"I":12,
        "J":12,"K":14,"L":14,"M":14,"N":10,"O":12,"P":10,"Q":20})
    ws.freeze_panes = "A3"
    log.info(f"SEMNALE INTRARE: {row-3} rânduri")

# ═══════════════════════════════════════════════════════════════════════════
# INDICATORI MACRO
# ═══════════════════════════════════════════════════════════════════════════
def upd_indicatori_macro(wb, all_data, macro_data, fg_val, fg_class):
    ws = wb["INDICATORI MACRO"]; unmerge_all(ws)
    hdrs = ["Indicator","Valoare curentă","Valoare anterioară","Δ Absolut","Δ %","Consens",
        "Dev. vs Estimare","Impact piață","Data publicare","Frecvență","Trending","Status","Note"]
    safe_write(ws,1,1,"🌍 INDICATORI MACROECONOMICI",font=FONT_TITLE,fill=FILL_TITLE,alignment=ALIGN_LEFT)
    for c in range(2,14): safe_write(ws,1,c,"",fill=FILL_TITLE)
    ws.merge_cells("A1:M1"); ws.row_dimensions[1].height = 30
    write_header_row(ws,2,hdrs); ws.row_dimensions[2].height = 20
    vix_d = macro_data.get("VIX",{})
    y10_d = macro_data.get("Yield 10Y",{})
    dxy_d = macro_data.get("USD Index",{})
    eurusd_d = all_data.get("EUR/USD",{})
    brent_d = all_data.get("Oil Brent",{})
    wti_d = all_data.get("Oil WTI",{})
    fr_c, fr_p = get_fred_data("FEDFUNDS") if FRED_API_KEY else (None, None)
    cpi_c, cpi_p = get_fred_data("CPIAUCSL") if FRED_API_KEY else (None, None)
    un_c, un_p = get_fred_data("UNRATE") if FRED_API_KEY else (None, None)
    macro_rows = [
        ("PIB YoY (%)",2.8,2.4,2.6,"Ridicat","Trimestrial","Bullish dacă > estimare"),
        ("CPI YoY (%)",fp(cpi_c,1) if cpi_c else 3.2,3.4 if not cpi_p else fp(cpi_p,1),3.3,"Ridicat","Lunar","Inflație sub estimare = bullish"),
        ("Core CPI (%)",3.8,3.9,3.8,"Ridicat","Lunar","Exclud food & energy"),
        ("Rata dobânzii (%)",fp(fr_c,2) if fr_c else 5.25,fp(fr_p,2) if fr_p else 5.25,5.25,"Ridicat","6 săptămâni","Fed Funds Rate"),
        ("Rata șomajului (%)",fp(un_c,1) if un_c else 3.9,fp(un_p,1) if un_p else 3.8,3.9,"Ridicat","Lunar","Sub 4% = solidă"),
        ("PMI Manufacturing",49.2,49.5,49.8,"Mediu","Lunar","Sub 50 = contracție"),
        ("PMI Services",52.3,52.0,52.5,"Mediu","Lunar","Peste 50 = expansiune"),
        ("Retail Sales MoM (%)",0.3,0.5,0.4,"Mediu","Lunar","Consum personal"),
        ("Balanță comercială ($B)",-68.5,-65.2,-67.0,"Scăzut","Lunar","Deficit comercial SUA"),
        ("USD Index (DXY)",fp(dxy_d.get("price","")) if dxy_d else 104.5,"","","Ridicat","Continuu","Forța dolarului"),
        ("EUR/USD",fp(eurusd_d.get("price","")) if eurusd_d else 1.085,"","","Ridicat","Continuu","Pereche principală"),
        ("Petrol Brent ($)",fp(brent_d.get("price","")) if brent_d else 82.5,"","","Mediu","Continuu","Referință internațional"),
        ("Petrol WTI ($)",fp(wti_d.get("price","")) if wti_d else 78.3,"","","Mediu","Continuu","Referință SUA"),
        ("Yield 10Y US (%)",fp(y10_d.get("price",""),2) if y10_d else 4.25,"","","Ridicat","Continuu","Randament 10 ani"),
        ("VIX",fp(vix_d.get("price",""),1) if vix_d else 18.5,"","","Ridicat","Continuu","Indicele fricii"),
        ("Fear & Greed (0-100)",fg_val if fg_val else "","","","Mediu","Zilnic",fg_class if fg_class else ""),
    ]
    row = 3
    for mr in macro_rows:
        ind, cur, prev, cons, imp, freq, note = mr
        safe_write(ws,row,1,ind,font=FONT_NORMAL,alignment=ALIGN_LEFT)
        safe_write(ws,row,2,cur if cur != "" else "",font=FONT_NORMAL)
        safe_write(ws,row,3,prev if prev != "" else "",font=FONT_NORMAL)
        safe_write(ws,row,4,f'=IFERROR(B{row}-C{row},"")',font=FONT_CALC)
        safe_write(ws,row,5,f'=IFERROR((B{row}-C{row})/ABS(C{row}),"")',font=FONT_CALC,number_format='0.0%')
        safe_write(ws,row,6,cons,font=FONT_NORMAL)
        safe_write(ws,row,7,f'=IFERROR(B{row}-F{row},"")',font=FONT_CALC)
        safe_write(ws,row,8,imp,font=FONT_NORMAL)
        safe_write(ws,row,9,TODAY_STR,font=FONT_NORMAL)
        safe_write(ws,row,10,freq,font=FONT_NORMAL)
        if cur != "" and prev != "":
            try:
                trending = "↑ Creștere" if float(cur)>float(prev) else ("↓ Scădere" if float(cur)<float(prev) else "→ Stabil")
            except: trending = "→ Stabil"
        else: trending = ""
        apply_signal_fill(ws,row,11,trending)
        if ind == "VIX":
            try: status = "Pozitiv" if float(cur)<20 else ("Negativ" if float(cur)>30 else "Neutru")
            except: status = "Neutru"
        elif ind == "Fear & Greed (0-100)":
            try: status = "Pozitiv" if int(cur)>60 else ("Negativ" if int(cur)<30 else "Neutru")
            except: status = "Neutru"
        else: status = "Neutru"
        apply_signal_fill(ws,row,12,status)
        safe_write(ws,row,13,note,font=FONT_NORMAL,alignment=ALIGN_LEFT)
        ws.row_dimensions[row].height = 18; row += 1
    set_col_widths(ws,{"A":22,"B":16,"C":16,"D":12,"E":10,"F":12,"G":14,"H":12,"I":14,"J":12,"K":14,"L":10,"M":30})
    ws.freeze_panes = "A3"
    log.info("INDICATORI MACRO: actualizat")

# ═══════════════════════════════════════════════════════════════════════════
# COMPETITORI SECTOR
# ═══════════════════════════════════════════════════════════════════════════
def upd_competitori_sector(wb, all_data, macro_data):
    ws = wb["COMPETITORI SECTOR"]; unmerge_all(ws)
    hdrs = ["Entitate","Sector","Cotă piață (%)","Schimbare YoY","Preț mediu","Trending",
        "Revenue ($B)","Marjă (%)","Puncte forte","Puncte slabe","Scor (1-10)","Risc (1-5)","Status","Data","Note"]
    safe_write(ws,1,1,"🏭 ANALIZĂ COMPETITORI & SECTOR",font=FONT_TITLE,fill=FILL_TITLE,alignment=ALIGN_LEFT)
    for c in range(2,16): safe_write(ws,1,c,"",fill=FILL_TITLE)
    ws.merge_cells("A1:O1"); ws.row_dimensions[1].height = 30
    write_header_row(ws,2,hdrs); ws.row_dimensions[2].height = 20
    sm = {"INDICI BURSIERI":"Index","ACȚIUNI & ETF":"Technology","CRYPTO":"Crypto","VALUTE FOREX":"Forex","MATERII PRIME":"Commodities"}
    row = 3
    for cn, cl in COMP_MAP.items():
        write_separator(ws,row,cn,15); row += 1
        for cname, ctk in cl:
            d = all_data.get(cname,{})
            trend = d.get("trend","Sideways")
            trending = "↑ Creștere" if trend=="Bullish" else ("↓ Scădere" if trend=="Bearish" else "→ Stabil")
            vals = [f"{cname} ({ctk})",sm.get(cn,""),"","",d.get("price",""),trending,"","","","","","","Activ",TODAY_STR,""]
            for c, v in enumerate(vals, 1):
                if c == 6: apply_signal_fill(ws,row,c,v)
                elif c in (3,4,7,8,11,12): safe_write(ws,row,c,v,font=FONT_NORMAL,fill=FILL_INPUT)
                else: safe_write(ws,row,c,v,font=FONT_NORMAL)
            ws.row_dimensions[row].height = 18; row += 1
    add_cf_score_cs(ws,"K",3,row-1)
    set_col_widths(ws,{"A":22,"B":14,"C":14,"D":12,"E":14,"F":14,"G":14,"H":10,"I":20,"J":20,"K":10,"L":10,"M":10,"N":12,"O":20})
    ws.freeze_panes = "A3"
    log.info("COMPETITORI SECTOR: actualizat")

# ═══════════════════════════════════════════════════════════════════════════
# RISCURI & OPORTUNITATI
# ═══════════════════════════════════════════════════════════════════════════
def upd_riscuri_oportunitati(wb):
    ws = wb["RISCURI OPORTUNITATI"]; unmerge_all(ws)
    hdrs = ["ID","Tip","Categorie","Descriere","Impact (1-5)","Probabilitate (%)","Scor prioritate",
        "Orizont","Acțiuni recomandate","Owner","Status","Data","Note"]
    safe_write(ws,1,1,"⚠️ RISCURI & OPORTUNITĂȚI",font=FONT_TITLE,fill=FILL_TITLE,alignment=ALIGN_LEFT)
    for c in range(2,14): safe_write(ws,1,c,"",fill=FILL_TITLE)
    ws.merge_cells("A1:M1"); ws.row_dimensions[1].height = 30
    write_header_row(ws,2,hdrs); ws.row_dimensions[2].height = 20
    row = 3
    for cn, rl in RISKS_MAP.items():
        write_separator(ws,row,cn,13); row += 1
        for ri in rl:
            rid, tip, desc, imp, prob, oriz, act = ri
            scor = round(imp * prob / 100, 1)
            vals = [rid,tip,cn,desc,imp,prob,scor,oriz,act,"","Activ",TODAY_STR,""]
            for c, v in enumerate(vals, 1):
                if c == 2:
                    if v == "Risc": safe_write(ws,row,c,v,font=FONT_SELL,fill=FILL_SELL)
                    else: safe_write(ws,row,c,v,font=FONT_BUY,fill=FILL_BUY)
                elif c in (10,13): safe_write(ws,row,c,v,font=FONT_NORMAL,fill=FILL_INPUT)
                else: safe_write(ws,row,c,v,font=FONT_NORMAL)
            ws.row_dimensions[row].height = 18; row += 1
    add_cf_score_cs(ws,"G",3,row-1)
    set_col_widths(ws,{"A":6,"B":14,"C":18,"D":50,"E":12,"F":14,"G":14,"H":14,"I":40,"J":14,"K":10,"L":12,"M":20})
    ws.freeze_panes = "A3"
    log.info("RISCURI OPORTUNITATI: actualizat")

# ═══════════════════════════════════════════════════════════════════════════
# CALENDAR ECONOMIC
# ═══════════════════════════════════════════════════════════════════════════
def upd_calendar_economic(wb):
    ws = wb["CALENDAR ECONOMIC"]; unmerge_all(ws)
    hdrs = ["Data & Ora","Eveniment","Țară","Impact","Anterior","Estimare","Actual","Deviere","Impact real","Activ afectat","Note"]
    safe_write(ws,1,1,"📅 CALENDAR ECONOMIC — EVENIMENTE DE PIAȚĂ",font=FONT_TITLE,fill=FILL_TITLE,alignment=ALIGN_LEFT)
    for c in range(2,12): safe_write(ws,1,c,"",fill=FILL_TITLE)
    ws.merge_cells("A1:K1"); ws.row_dimensions[1].height = 30
    write_header_row(ws,2,hdrs); ws.row_dimensions[2].height = 20
    row = 3; eo = 1
    for ci, (cn, el) in enumerate(CALENDAR_MAP.items()):
        write_separator(ws,row,cn,11); row += 1
        for i, (ev, co, imp, note) in enumerate(el):
            ed = NOW + timedelta(days=eo); eo += 3 + i*2
            ds = ed.strftime("%d.%m.%Y %H:%M")
            id_str = "⚡⚡⚡ Ridicat" if imp=="Ridicat" else "⚡⚡ Mediu"
            vals = [ds,ev,co,id_str,"","","","","",note,""]
            for c, v in enumerate(vals, 1):
                if c in (5,6,7): safe_write(ws,row,c,v,font=FONT_NORMAL,fill=FILL_INPUT)
                elif c == 4:
                    if "Ridicat" in str(v): safe_write(ws,row,c,v,font=FONT_SELL,fill=FILL_ALERT)
                    else: safe_write(ws,row,c,v,font=FONT_WAIT,fill=FILL_WAIT)
                else: safe_write(ws,row,c,v,font=FONT_NORMAL)
            ws.row_dimensions[row].height = 18; row += 1
        eo = (ci + 1) * 5
    set_col_widths(ws,{"A":18,"B":28,"C":10,"D":16,"E":12,"F":12,"G":12,"H":10,"I":14,"J":35,"K":20})
    ws.freeze_panes = "A3"
    log.info("CALENDAR ECONOMIC: actualizat")

# ═══════════════════════════════════════════════════════════════════════════
# JURNAL TRANZACTII
# ═══════════════════════════════════════════════════════════════════════════
def upd_jurnal_tranzactii(wb):
    ws = wb["JURNAL TRANZACTII"]; unmerge_all(ws)
    hdrs = ["ID","Data","Ora","Activ","L/S","Setup","Entry","SL","TP","Mărime poz","Risc $",
        "Exit","Data ieșire","P&L $","P&L %","RR Realizat","Cal exec (1-10)","Emoție","Plan?","Lecție","Link/SS"]
    safe_write(ws,1,1,"📜 JURNAL TRANZACȚII — HEDGE FUND STYLE",font=FONT_TITLE,fill=FILL_TITLE,alignment=ALIGN_LEFT)
    for c in range(2,22): safe_write(ws,1,c,"",fill=FILL_TITLE)
    ws.merge_cells("A1:U1"); ws.row_dimensions[1].height = 30
    write_header_row(ws,2,hdrs); ws.row_dimensions[2].height = 20
    for r in range(3,23):
        for c in range(1,22):
            cell = ws.cell(row=r,column=c)
            if cell.value is None:
                cell.fill = FILL_INPUT; cell.border = THIN_BORDER; cell.font = FONT_NORMAL; cell.alignment = ALIGN_CENTER
        ws.row_dimensions[r].height = 18
    sr = 25
    safe_write(ws,sr,1,"📊 STATISTICI JURNAL",font=FONT_SEP,fill=FILL_SEP)
    for c in range(2,8): safe_write(ws,sr,c,"",fill=FILL_SEP)
    ws.merge_cells(f"A{sr}:G{sr}"); ws.row_dimensions[sr].height = 22
    stats = [("Nr. tranzacții",f'=COUNTA(A3:A22)'),("Win Rate",f'=IFERROR(COUNTIF(N3:N22,">0")/COUNTA(N3:N22),"N/A")'),
        ("P&L Total",f'=IFERROR(SUM(N3:N22),"N/A")'),("P&L Mediu",f'=IFERROR(AVERAGE(N3:N22),"N/A")'),
        ("Best Trade",f'=IFERROR(MAX(N3:N22),"N/A")'),("Worst Trade",f'=IFERROR(MIN(N3:N22),"N/A")'),
        ("RR Mediu",f'=IFERROR(AVERAGE(P3:P22),"N/A")'),("Calitate Exec Medie",f'=IFERROR(AVERAGE(Q3:Q22),"N/A")')]
    for i, (lb, fm) in enumerate(stats):
        r = sr+1+i
        safe_write(ws,r,1,lb,font=FONT_SUBHEADER,fill=FILL_SUBHEADER,alignment=ALIGN_LEFT)
        ws.merge_cells(f"A{r}:C{r}")
        safe_write(ws,r,4,fm,font=FONT_CALC,fill=FILL_CALC)
        ws.merge_cells(f"D{r}:G{r}"); ws.row_dimensions[r].height = 18
    add_cf_pnl(ws,"N",3,22); add_cf_signal(ws,"E",3,22)
    set_col_widths(ws,{"A":6,"B":12,"C":8,"D":14,"E":8,"F":14,"G":12,"H":12,"I":12,"J":12,"K":10,
        "L":12,"M":12,"N":10,"O":10,"P":10,"Q":12,"R":12,"S":8,"T":25,"U":14})
    ws.freeze_panes = "A3"
    log.info("JURNAL TRANZACTII: actualizat")

# ═══════════════════════════════════════════════════════════════════════════
# ISTORIC TRENDING
# ═══════════════════════════════════════════════════════════════════════════
def upd_istoric_trending(wb, all_data, macro_data, fg_val):
    ws = wb["ISTORIC TRENDING"]; unmerge_all(ws)
    hdrs = ["Luna/An","RSI Medie","Preț activ selectat","Preț S&P500","PIB YoY","CPI","VIX",
        "Semnal Lună","Vol Mediu (B)","Trend dominant","Fear&Greed","Rate Dobândă","Yield 10Y",
        "Gold","Oil WTI","EUR/USD","Note"]
    safe_write(ws,1,1,"🕰 ISTORIC & TRENDING — SNAPSHOT LUNAR 24 LUNI",font=FONT_TITLE,fill=FILL_TITLE,alignment=ALIGN_LEFT)
    for c in range(2,18): safe_write(ws,1,c,"",fill=FILL_TITLE)
    ws.merge_cells("A1:Q1"); ws.row_dimensions[1].height = 30
    write_header_row(ws,2,hdrs); ws.row_dimensions[2].height = 20
    rsi_vals = [float(d["rsi"]) for d in all_data.values() if d.get("rsi") not in (None,"") and d.get("rsi") is not None]
    rsi_avg = round(np.mean(rsi_vals),1) if rsi_vals else ""
    bc = sum(1 for d in all_data.values() if d.get("signal")=="BUY")
    sc = sum(1 for d in all_data.values() if d.get("signal")=="SELL")
    tot = len(all_data)
    sig_l = "BUY" if bc>sc and bc>tot*0.3 else ("SELL" if sc>bc and sc>tot*0.3 else "WAIT")
    bul = sum(1 for d in all_data.values() if d.get("trend")=="Bullish")
    bear = sum(1 for d in all_data.values() if d.get("trend")=="Bearish")
    td = "Bullish" if bul>bear else ("Bearish" if bear>bul else "Sideways")
    sp = all_data.get("S&P 500",{}); gd = all_data.get("Gold",{}); wt = all_data.get("Oil WTI",{})
    eu = all_data.get("EUR/USD",{}); vx = macro_data.get("VIX",{})
    cm = NOW.strftime("%b %Y")
    em = set()
    for r in range(3, ws.max_row+1):
        v = ws.cell(row=r,column=1).value
        if v: em.add(str(v))
    if cm not in em:
        row = max(ws.max_row+1, 3)
        vals = [cm,rsi_avg,'=IFERROR(XLOOKUP(DASHBOARD!I2,\'INDICATORI TEHNICI\'!B:B,\'INDICATORI TEHNICI\'!C:C),"")',
            sp.get("price",""),2.8,3.2,vx.get("price",""),sig_l,"",td,fg_val if fg_val else "",
            5.25,"",gd.get("price",""),wt.get("price",""),eu.get("price",""),"Automat"]
        for c, v in enumerate(vals, 1):
            if c in (8,10): apply_signal_fill(ws,row,c,v)
            else: safe_write(ws,row,c,v,font=FONT_NORMAL)
        ws.row_dimensions[row].height = 18
    add_cf_signal(ws,"H",3,ws.max_row); add_cf_trend(ws,"J",3,ws.max_row)
    set_col_widths(ws,{"A":12,"B":10,"C":16,"D":14,"E":10,"F":8,"G":8,"H":12,"I":12,"J":14,"K":12,"L":12,"M":10,"N":12,"O":12,"P":12,"Q":20})
    ws.freeze_panes = "A3"
    log.info("ISTORIC TRENDING: actualizat")

# ═══════════════════════════════════════════════════════════════════════════
# REZUMAT EXECUTIV
# ═══════════════════════════════════════════════════════════════════════════
def upd_rezumat_executiv(wb, all_data, macro_data, fg_val, fg_class):
    ws = wb["REZUMAT EXECUTIV"]; unmerge_all(ws)
    safe_write(ws,1,1,"📋 REZUMAT EXECUTIV — MARKET OVERVIEW",font=FONT_TITLE,fill=FILL_TITLE,alignment=ALIGN_LEFT)
    for c in range(2,9): safe_write(ws,1,c,"",fill=FILL_TITLE)
    ws.merge_cells("A1:H1"); ws.row_dimensions[1].height = 30
    safe_write(ws,2,1,"Activ selectat:",font=FONT_SUBHEADER,fill=FILL_SUBHEADER)
    ws.merge_cells("A2:C2")
    safe_write(ws,2,4,"=DASHBOARD!I2",font=FONT_SELECTOR,fill=FILL_SELECTOR)
    ws.merge_cells("D2:H2"); ws.row_dimensions[2].height = 22
    safe_write(ws,3,1,"Data ultimei actualizări:",font=FONT_NORMAL); ws.merge_cells("A3:C3")
    safe_write(ws,3,4,NOW.strftime("%d.%m.%Y %H:%M"),font=FONT_NORMAL); ws.merge_cells("D3:H3")
    write_header_row(ws,5,["Indicator","Valoare","","Trending","Observații","","",""])
    ws.merge_cells("B5:C5"); ws.merge_cells("E5:H5"); ws.row_dimensions[5].height = 20
    bc = sum(1 for d in all_data.values() if d.get("signal")=="BUY")
    sc = sum(1 for d in all_data.values() if d.get("signal")=="SELL")
    wc = sum(1 for d in all_data.values() if d.get("signal")=="WAIT")
    tot = len(all_data)
    bul = sum(1 for d in all_data.values() if d.get("trend")=="Bullish")
    bear = sum(1 for d in all_data.values() if d.get("trend")=="Bearish")
    if tot>0:
        if bc>sc: tendinta, t_trend = f"Bullish ({bc}/{tot} BUY)", "↑ Pozitiv"
        elif sc>bc: tendinta, t_trend = f"Bearish ({sc}/{tot} SELL)", "↓ Negativ"
        else: tendinta, t_trend = f"Neutru ({wc}/{tot} WAIT)", "→ Neutru"
    else: tendinta, t_trend = "N/A", "→ Neutru"
    vix_d = macro_data.get("VIX",{})
    vix_p = vix_d.get("price","")
    try:
        vv = float(vix_p)
        vol_str = f"VIX = {vv:.1f}"
        vol_t = "↓ Scăzut" if vv<15 else ("↑ Ridicat" if vv>25 else "→ Moderat")
    except: vol_str, vol_t = "N/A", "→ Neutru"
    fg_d = f"{fg_val} — {fg_class}" if fg_val else "N/A"
    fg_t = "↑ Greed" if fg_val and fg_val>60 else ("↓ Fear" if fg_val and fg_val<30 else "→ Neutru")
    srows = [
        ("Tendință generală",tendinta,t_trend,f"BUY:{bc} SELL:{sc} WAIT:{wc}"),
        ("Volatilitate",vol_str,vol_t,"VIX sub 15=calm, 15-25=normal, >25=volatil"),
        ("Volum piață",f"Bullish:{bul} Bearish:{bear}","↑ Activ" if bul>bear else "↓ Slăbit","Raport trend-uri active"),
        ("Sentiment (F&G)",fg_d,fg_t,"0-25=Frică extremă, 75-100=Lăcomie extremă"),
        ("Risc sistemic","Moderat","→ Monitorizare","VIX, spread-uri, corelații"),
    ]
    row = 6
    for lb, val, tr, obs in srows:
        safe_write(ws,row,1,lb,font=FONT_SUBHEADER,fill=FILL_SUBHEADER,alignment=ALIGN_LEFT)
        safe_write(ws,row,2,val,font=FONT_NORMAL); ws.merge_cells(f"B{row}:C{row}")
        apply_signal_fill(ws,row,4,tr)
        safe_write(ws,row,5,obs,font=FONT_NORMAL,alignment=ALIGN_LEFT); ws.merge_cells(f"E{row}:H{row}")
        ws.row_dimensions[row].height = 20; row += 1
    row += 1
    safe_write(ws,row,1,"📌 CONCLUZII GENERATE AUTOMAT",font=FONT_SEP,fill=FILL_SEP)
    for c in range(2,9): safe_write(ws,row,c,"",fill=FILL_SEP)
    ws.merge_cells(f"A{row}:H{row}"); ws.row_dimensions[row].height = 22; row += 1
    concs = []
    if bc>sc: concs.append(f"1. Piața prezintă bias BULLISH cu {bc} semnale BUY din {tot} active.")
    elif sc>bc: concs.append(f"1. Piața prezintă bias BEARISH cu {sc} semnale SELL din {tot} active.")
    else: concs.append(f"1. Piața este în ECHILIBRU cu {wc} active WAIT.")
    try:
        vv = float(vix_p)
        if vv>25: concs.append(f"2. Volatilitate RIDICATĂ (VIX={vv:.1f}) — reducere expunere recomandată.")
        elif vv<15: concs.append(f"2. Volatilitate SCĂZUTĂ (VIX={vv:.1f}) — trend-following favorabil.")
        else: concs.append(f"2. Volatilitate NORMALĂ (VIX={vv:.1f}) — trading activ permis.")
    except: concs.append("2. Volatilitate: date insuficiente.")
    if fg_val:
        if fg_val>75: concs.append(f"3. Sentiment LĂCOMIE EXTREMĂ (F&G={fg_val}) — atenție reversări.")
        elif fg_val<25: concs.append(f"3. Sentiment FRICĂ EXTREMĂ (F&G={fg_val}) — potențial acumulare.")
        else: concs.append(f"3. Sentiment NEUTRU (F&G={fg_val}) — piața așteaptă catalizatori.")
    else: concs.append("3. Sentiment: date indisponibile.")
    concs.append(f"4. Distribuție: Bullish={bul}, Bearish={bear}, Sideways={tot-bul-bear}.")
    concs.append("5. Recomandare: Focus pe active cu confluențe ≥3 și RVOL > 1.2x.")
    for cn in concs:
        safe_write(ws,row,1,cn,font=FONT_NORMAL,alignment=ALIGN_LEFT)
        ws.merge_cells(f"A{row}:H{row}"); ws.row_dimensions[row].height = 20; row += 1
    set_col_widths(ws,{"A":22,"B":18,"C":10,"D":14,"E":16,"F":14,"G":14,"H":14})
    ws.freeze_panes = "A3"
    log.info("REZUMAT EXECUTIV: actualizat")

# ═══════════════════════════════════════════════════════════════════════════
# DASHBOARD
# ═══════════════════════════════════════════════════════════════════════════
def upd_dashboard(wb, all_data, macro_data, fg_val, fg_class):
    ws = wb["DASHBOARD"]; unmerge_all(ws)
    safe_write(ws,1,1,"📊 DASHBOARD — ANALIZĂ DE PIAȚĂ",font=FONT_TITLE,fill=FILL_TITLE,alignment=ALIGN_LEFT)
    for c in range(2,17): safe_write(ws,1,c,"",fill=FILL_TITLE)
    ws.merge_cells("A1:P1"); ws.row_dimensions[1].height = 35
    safe_write(ws,2,1,"Data actualizare:",font=FONT_SUBHEADER,fill=FILL_SUBHEADER); ws.merge_cells("A2:C2")
    safe_write(ws,2,4,NOW.strftime("%d.%m.%Y %H:%M"),font=FONT_NORMAL); ws.merge_cells("D2:G2")
    safe_write(ws,2,8,"Activ selectat:",font=FONT_LABEL_SEL,fill=FILL_LABEL_SEL)
    fa = list(ALL_ASSETS.keys())[0] if ALL_ASSETS else "S&P 500"
    safe_write(ws,2,9,fa,font=FONT_SELECTOR,fill=FILL_SELECTOR); ws.merge_cells("I2:M2")
    dv = DataValidation(type="list",formula1="=LIST_ACTIVE!$A$1:$A$95",allow_blank=True,showDropDown=False)
    ws.add_data_validation(dv); dv.add(ws["I2"])
    ws.row_dimensions[2].height = 28; ws.row_dimensions[3].height = 6
    cr = [(1,4),(5,8),(9,12),(13,16)]
    ct = ["TREND GENERAL","VOLATILITATE","VOLUM RELATIV","MOMENT INTRARE"]
    f5s = [
        '=IFERROR(XLOOKUP($I$2,\'INDICATORI TEHNICI\'!$B:$B,\'INDICATORI TEHNICI\'!$T:$T,"",-1,-1),"")',
        '=IFERROR("ATR: "&TEXT(XLOOKUP($I$2,\'INDICATORI TEHNICI\'!$B:$B,\'INDICATORI TEHNICI\'!$O:$O,"",-1,-1),"0.0000"),"")',
        '=IFERROR(TEXT(XLOOKUP($I$2,\'INDICATORI TEHNICI\'!$B:$B,\'INDICATORI TEHNICI\'!$S:$S,"",-1,-1),"0.00")&"x medie","")',
        '=IFERROR(XLOOKUP($I$2,\'SEMNALE INTRARE\'!$B:$B,\'SEMNALE INTRARE\'!$C:$C,"",-1,-1),"")',
    ]
    f6s = [
        '=IFERROR(XLOOKUP($I$2,\'INDICATORI TEHNICI\'!$B:$B,\'INDICATORI TEHNICI\'!$W:$W,"",-1,-1),"")',
        '=IFERROR("BB: "&TEXT(XLOOKUP($I$2,\'INDICATORI TEHNICI\'!$B:$B,\'INDICATORI TEHNICI\'!$N:$N,"",-1,-1),"0.0000"),"")',
        '=IFERROR("Vol: "&TEXT(XLOOKUP($I$2,\'INDICATORI TEHNICI\'!$B:$B,\'INDICATORI TEHNICI\'!$R:$R,"",-1,-1),"#,##0"),"")',
        '=IFERROR(XLOOKUP($I$2,\'SEMNALE INTRARE\'!$B:$B,\'SEMNALE INTRARE\'!$J:$J,"",-1,-1)&"/5 conf","")',
    ]
    for idx, (cs, ce) in enumerate(cr):
        cl, el = get_column_letter(cs), get_column_letter(ce)
        safe_write(ws,4,cs,ct[idx],font=FONT_HEADER,fill=FILL_HEADER)
        for cc in range(cs+1,ce+1): safe_write(ws,4,cc,"",fill=FILL_HEADER)
        ws.merge_cells(f"{cl}4:{el}4")
        safe_write(ws,5,cs,f5s[idx],font=Font(name="Arial",bold=True,size=12))
        for cc in range(cs+1,ce+1): safe_write(ws,5,cc,"")
        ws.merge_cells(f"{cl}5:{el}5")
        safe_write(ws,6,cs,f6s[idx],font=FONT_NORMAL)
        for cc in range(cs+1,ce+1): safe_write(ws,6,cc,"")
        ws.merge_cells(f"{cl}6:{el}6")
        rng = f"{cl}5:{el}5"
        for w, bg, fg in [("Bullish",C_BUY_BG,C_BUY_FG),("Bearish",C_SELL_BG,C_SELL_FG),("Sideways",C_WAIT_BG,C_WAIT_FG),
                          ("BUY",C_BUY_BG,C_BUY_FG),("SELL",C_SELL_BG,C_SELL_FG),("WAIT",C_WAIT_BG,C_WAIT_FG)]:
            ws.conditional_formatting.add(rng, CellIsRule(operator="equal",formula=[f'"{w}"'],
                fill=PatternFill("solid",fgColor=bg),font=Font(bold=True,color=fg)))
    ws.row_dimensions[4].height = 22; ws.row_dimensions[5].height = 28; ws.row_dimensions[6].height = 22; ws.row_dimensions[7].height = 6
    safe_write(ws,8,1,"SEMNAL PRINCIPAL DE TRADING",font=FONT_HEADER,fill=FILL_HEADER)
    for c in range(2,17): safe_write(ws,8,c,"",fill=FILL_HEADER)
    ws.merge_cells("A8:P8"); ws.row_dimensions[8].height = 25
    sig_rows = [
        (9,"SEMNAL ACTIV",'=IFERROR(XLOOKUP($I$2,\'SEMNALE INTRARE\'!$B:$B,\'SEMNALE INTRARE\'!$C:$C,"",-1,-1),"")'),
        (10,"Activ analizat",'=$I$2'),
        (11,"Entry Price",'=IFERROR(XLOOKUP($I$2,\'SEMNALE INTRARE\'!$B:$B,\'SEMNALE INTRARE\'!$K:$K,"",-1,-1),"")'),
        (12,"Stop Loss (SL)",'=IFERROR(XLOOKUP($I$2,\'SEMNALE INTRARE\'!$B:$B,\'SEMNALE INTRARE\'!$L:$L,"",-1,-1),"")'),
        (13,"Take Profit (TP)",'=IFERROR(XLOOKUP($I$2,\'SEMNALE INTRARE\'!$B:$B,\'SEMNALE INTRARE\'!$M:$M,"",-1,-1),"")'),
        (14,"Risk/Reward Ratio",'=IFERROR((G13-G11)/(G11-G12),"N/A")'),
        (15,"Confluențe",'=IFERROR(XLOOKUP($I$2,\'SEMNALE INTRARE\'!$B:$B,\'SEMNALE INTRARE\'!$J:$J,"",-1,-1)&"/5","")'),
        (16,"Probabilitate",'=IFERROR(XLOOKUP($I$2,\'SEMNALE INTRARE\'!$B:$B,\'SEMNALE INTRARE\'!$O:$O,"",-1,-1),"")'),
        (17,"Condiție",'=IFERROR(XLOOKUP($I$2,\'SEMNALE INTRARE\'!$B:$B,\'SEMNALE INTRARE\'!$D:$D,"",-1,-1),"")'),
    ]
    for r, lb, fm in sig_rows:
        safe_write(ws,r,1,lb,font=FONT_SUBHEADER,fill=FILL_SUBHEADER,alignment=ALIGN_LEFT)
        for c in range(2,7): safe_write(ws,r,c,"",fill=FILL_SUBHEADER)
        ws.merge_cells(f"A{r}:F{r}")
        sl_f = FILL_SELL if r==12 else (FILL_BUY if r==13 else None)
        nf = '#,##0.0000' if r in (11,12,13) else ('0.00"x"' if r==14 else ('0%' if r==16 else None))
        safe_write(ws,r,7,fm,font=Font(name="Arial",bold=True,size=11),fill=sl_f,number_format=nf)
        for c in range(8,17): safe_write(ws,r,c,"",fill=sl_f)
        ws.merge_cells(f"G{r}:P{r}"); ws.row_dimensions[r].height = 22
    for w, bg, fg in [("BUY",C_BUY_BG,C_BUY_FG),("SELL",C_SELL_BG,C_SELL_FG),("WAIT",C_WAIT_BG,C_WAIT_FG)]:
        ws.conditional_formatting.add("G9:P9", CellIsRule(operator="equal",formula=[f'"{w}"'],
            fill=PatternFill("solid",fgColor=bg),font=Font(bold=True,color=fg,size=14)))
    ws.row_dimensions[18].height = 6
    safe_write(ws,19,1,"REZUMAT INDICATORI — ACTIVUL SELECTAT",font=FONT_HEADER,fill=FILL_HEADER)
    for c in range(2,17): safe_write(ws,19,c,"",fill=FILL_HEADER)
    ws.merge_cells("A19:P19"); ws.row_dimensions[19].height = 22
    for c, h in [(1,"Indicator"),(3,"Valoare"),(5,"Status"),(9,"Observație")]:
        safe_write(ws,20,c,h,font=FONT_HEADER,fill=FILL_SUBHEADER)
    ws.merge_cells("A20:B20"); ws.merge_cells("C20:D20"); ws.merge_cells("E20:H20"); ws.merge_cells("I20:P20")
    ws.row_dimensions[20].height = 20
    ind_rows = [
        (21,"RSI(14)",'=IFERROR(TEXT(XLOOKUP($I$2,\'INDICATORI TEHNICI\'!$B:$B,\'INDICATORI TEHNICI\'!$G:$G,"",-1,-1),"0.0"),"")',
         '=IFERROR(XLOOKUP($I$2,\'INDICATORI TEHNICI\'!$B:$B,\'INDICATORI TEHNICI\'!$H:$H,"",-1,-1),"")',"Condiție RSI curentă"),
        (22,"Impuls MACD",'=IFERROR(XLOOKUP($I$2,\'INDICATORI TEHNICI\'!$B:$B,\'INDICATORI TEHNICI\'!$I:$I,"",-1,-1),"")',
         '=IFERROR(IF(ISNUMBER(SEARCH("pozitiv",XLOOKUP($I$2,\'INDICATORI TEHNICI\'!$B:$B,\'INDICATORI TEHNICI\'!$H:$H,"",-1,-1))),"BUY","SELL"),"")',"Histogramă MACD"),
        (23,"MA50 vs MA200",'=IFERROR(XLOOKUP($I$2,\'INDICATORI TEHNICI\'!$B:$B,\'INDICATORI TEHNICI\'!$W:$W,"",-1,-1),"")',
         '=IFERROR(IF(XLOOKUP($I$2,\'INDICATORI TEHNICI\'!$B:$B,\'INDICATORI TEHNICI\'!$W:$W,"",-1,-1)="Golden Cross","BUY","SELL"),"")',"Crossover MA"),
        (24,"Trend activ",'=IFERROR(XLOOKUP($I$2,\'INDICATORI TEHNICI\'!$B:$B,\'INDICATORI TEHNICI\'!$T:$T,"",-1,-1),"")',
         '=IFERROR(XLOOKUP($I$2,\'INDICATORI TEHNICI\'!$B:$B,\'INDICATORI TEHNICI\'!$T:$T,"",-1,-1),"")',"RVOL + ATR context"),
        (25,"VIX (global)",'=IFERROR(XLOOKUP("VIX",\'INDICATORI MACRO\'!$A:$A,\'INDICATORI MACRO\'!$B:$B,""),"")',
         '=IFERROR(IF(XLOOKUP("VIX",\'INDICATORI MACRO\'!$A:$A,\'INDICATORI MACRO\'!$B:$B,"")>25,"Negativ",IF(XLOOKUP("VIX",\'INDICATORI MACRO\'!$A:$A,\'INDICATORI MACRO\'!$B:$B,"")<15,"Pozitiv","Neutru")),"")',"Volatilitate implicită"),
        (26,"Fear & Greed",'=IFERROR(XLOOKUP("Fear & Greed (0-100)",\'INDICATORI MACRO\'!$A:$A,\'INDICATORI MACRO\'!$B:$B,""),"")',
         '=IFERROR(IF(XLOOKUP("Fear & Greed (0-100)",\'INDICATORI MACRO\'!$A:$A,\'INDICATORI MACRO\'!$B:$B,"")>60,"BUY",IF(XLOOKUP("Fear & Greed (0-100)",\'INDICATORI MACRO\'!$A:$A,\'INDICATORI MACRO\'!$B:$B,"")<30,"SELL","WAIT")),"")',"Sentiment global"),
        (27,"RVOL activ",'=IFERROR(TEXT(XLOOKUP($I$2,\'INDICATORI TEHNICI\'!$B:$B,\'INDICATORI TEHNICI\'!$S:$S,"",-1,-1),"0.00")&"x","")',
         '=IFERROR(IF(XLOOKUP($I$2,\'INDICATORI TEHNICI\'!$B:$B,\'INDICATORI TEHNICI\'!$S:$S,"",-1,-1)>1.2,"Pozitiv",IF(XLOOKUP($I$2,\'INDICATORI TEHNICI\'!$B:$B,\'INDICATORI TEHNICI\'!$S:$S,"",-1,-1)<0.7,"Negativ","Neutru")),"")',"Volum relativ"),
    ]
    for r, lb, vf, sf, obs in ind_rows:
        safe_write(ws,r,1,lb,font=FONT_SUBHEADER,fill=FILL_SUBHEADER,alignment=ALIGN_LEFT)
        safe_write(ws,r,2,"",fill=FILL_SUBHEADER); ws.merge_cells(f"A{r}:B{r}")
        safe_write(ws,r,3,vf,font=FONT_NORMAL); safe_write(ws,r,4,""); ws.merge_cells(f"C{r}:D{r}")
        safe_write(ws,r,5,sf,font=FONT_NORMAL)
        for cc in range(6,9): safe_write(ws,r,cc,"")
        ws.merge_cells(f"E{r}:H{r}")
        safe_write(ws,r,9,obs,font=FONT_NORMAL,alignment=ALIGN_LEFT)
        for cc in range(10,17): safe_write(ws,r,cc,"")
        ws.merge_cells(f"I{r}:P{r}"); ws.row_dimensions[r].height = 20
    for r in range(21,28):
        rng = f"E{r}:H{r}"
        for w, bg, fg in [("BUY",C_BUY_BG,C_BUY_FG),("SELL",C_SELL_BG,C_SELL_FG),("WAIT",C_WAIT_BG,C_WAIT_FG),
                          ("Pozitiv",C_BUY_BG,C_BUY_FG),("Negativ",C_SELL_BG,C_SELL_FG),("Neutru",C_WAIT_BG,C_WAIT_FG),
                          ("Bullish",C_BUY_BG,C_BUY_FG),("Bearish",C_SELL_BG,C_SELL_FG),("Sideways",C_WAIT_BG,C_WAIT_FG)]:
            ws.conditional_formatting.add(rng, CellIsRule(operator="equal",formula=[f'"{w}"'],
                fill=PatternFill("solid",fgColor=bg),font=Font(bold=True,color=fg)))
    set_col_widths(ws,{"A":14,"B":10,"C":12,"D":12,"E":12,"F":12,"G":14,"H":12,"I":14,"J":12,"K":12,"L":12,"M":12,"N":10,"O":10,"P":10})
    ws.freeze_panes = "A3"
    log.info("DASHBOARD: actualizat")


# ═══════════════════════════════════════════════════════════════════════════════
# FISA ACTIV — FULLY DYNAMIC WITH XLOOKUP
# ═══════════════════════════════════════════════════════════════════════════════

def upd_fisa_activ(wb, all_data, macro_data):
    """Update FISA ACTIV sheet — fully dynamic with XLOOKUP."""
    ws = wb["FISA ACTIV"]
    unmerge_all(ws)

    safe_write(ws, 1, 1, "FIȘĂ COMPLETĂ ACTIV — toate datele pentru activul selectat",
               font=FONT_TITLE, fill=FILL_TITLE, alignment=ALIGN_LEFT)
    for c in range(2, 9):
        safe_write(ws, 1, c, "", font=FONT_TITLE, fill=FILL_TITLE)
    ws.merge_cells("A1:H1")
    ws.row_dimensions[1].height = 30

    safe_write(ws, 2, 1, "Activ analizat:", font=FONT_LABEL_SEL, fill=FILL_LABEL_SEL)
    ws.merge_cells("A2:C2")
    safe_write(ws, 2, 4, "=DASHBOARD!I2", font=FONT_SELECTOR, fill=FILL_SELECTOR)
    ws.merge_cells("D2:H2")
    ws.row_dimensions[2].height = 24

    safe_write(ws, 3, 1, "Schimbă activul în DASHBOARD → celula portocalie → toate secțiunile se actualizează automat",
               font=Font(name="Arial", italic=True, size=9, color="7D5A00"),
               fill=PatternFill("solid", fgColor="FFF2CC"), alignment=ALIGN_LEFT)
    ws.merge_cells("A3:H3")
    ws.row_dimensions[3].height = 18

    row = 5

    def section_header(r, title):
        safe_write(ws, r, 1, title, font=FONT_SEP, fill=FILL_SEP)
        for c in range(2, 9):
            safe_write(ws, r, c, "", font=FONT_SEP, fill=FILL_SEP)
        ws.merge_cells(f"A{r}:H{r}")
        ws.row_dimensions[r].height = 22
        return r + 1

    def xlookup_si(col):
        return f'=IFERROR(XLOOKUP($D$2,\'SEMNALE INTRARE\'!$B:$B,\'SEMNALE INTRARE\'!${col}:${col},"",-1,-1),"")'

    def xlookup_it(col):
        return f'=IFERROR(XLOOKUP($D$2,\'INDICATORI TEHNICI\'!$B:$B,\'INDICATORI TEHNICI\'!${col}:${col},"",-1,-1),"")'

    def xlookup_pv(col):
        return f'=IFERROR(XLOOKUP($D$2,\'PRETURI VOLUME\'!$B:$B,\'PRETURI VOLUME\'!${col}:${col},"",-1,-1),"")'

    def pair_row(r, label1, formula1, label2, formula2, nf1=None, nf2=None):
        safe_write(ws, r, 1, label1, font=FONT_SUBHEADER, fill=FILL_SUBHEADER, alignment=ALIGN_LEFT)
        ws.merge_cells(f"A{r}:B{r}")
        safe_write(ws, r, 3, formula1, font=FONT_NORMAL, number_format=nf1)
        ws.merge_cells(f"C{r}:D{r}")
        safe_write(ws, r, 5, label2, font=FONT_SUBHEADER, fill=FILL_SUBHEADER, alignment=ALIGN_LEFT)
        ws.merge_cells(f"E{r}:F{r}")
        safe_write(ws, r, 7, formula2, font=FONT_NORMAL, number_format=nf2)
        ws.merge_cells(f"G{r}:H{r}")
        ws.row_dimensions[r].height = 18
        return r + 1

    # 1. SEMNAL DE INTRARE
    row = section_header(row, "1. SEMNAL DE INTRARE")
    signal_fields = [
        ("Semnal", xlookup_si("C"), "Condiție", xlookup_si("D")),
        ("Entry", xlookup_si("K"), "Stop Loss", xlookup_si("L")),
        ("Take Profit", xlookup_si("M"),
         "RR Ratio", f'=IFERROR((XLOOKUP($D$2,\'SEMNALE INTRARE\'!$B:$B,\'SEMNALE INTRARE\'!$M:$M,"",-1,-1)-XLOOKUP($D$2,\'SEMNALE INTRARE\'!$B:$B,\'SEMNALE INTRARE\'!$K:$K,"",-1,-1))/(XLOOKUP($D$2,\'SEMNALE INTRARE\'!$B:$B,\'SEMNALE INTRARE\'!$K:$K,"",-1,-1)-XLOOKUP($D$2,\'SEMNALE INTRARE\'!$B:$B,\'SEMNALE INTRARE\'!$L:$L,"",-1,-1)),"N/A")'),
        ("Probabilitate", xlookup_si("O"), "Status", xlookup_si("P")),
        ("Confluențe", xlookup_si("J"), "Ultima actualizare", TODAY_STR),
    ]
    for l1, f1, l2, f2 in signal_fields:
        row = pair_row(row, l1, f1, l2, f2)

    row += 1

    # 2. INDICATORI TEHNICI
    row = section_header(row, "2. INDICATORI TEHNICI")
    tech_pairs = [
        ("Preț", xlookup_it("C"), "Trend", xlookup_it("T")),
        ("MA20", xlookup_it("D"), "MA Cross", xlookup_it("W")),
        ("MA50", xlookup_it("E"), "Impuls MACD", xlookup_it("I")),
        ("MA200", xlookup_it("F"), "MACD Hist", xlookup_it("K")),
        ("RSI(14)", xlookup_it("G"), "RSI Status", xlookup_it("H")),
        ("BB Superior", xlookup_it("L"), "BB Inferior", xlookup_it("M")),
        ("BB Lățime", xlookup_it("N"), "ATR", xlookup_it("O")),
        ("Stoch %K", xlookup_it("P"), "Stoch %D", xlookup_it("Q")),
        ("Volum", xlookup_it("R"), "RVOL", xlookup_it("S")),
        ("Suport cheie", xlookup_it("U"), "Rezistență", xlookup_it("V")),
    ]
    for l1, f1, l2, f2 in tech_pairs:
        row = pair_row(row, l1, f1, l2, f2)

    row += 1

    # 3. PREȚURI & VOLUME
    row = section_header(row, "3. PREȚURI & VOLUME")
    pv_pairs = [
        ("Deschidere", xlookup_pv("C"), "Închidere", xlookup_pv("F")),
        ("Maxim", xlookup_pv("D"), "Minim", xlookup_pv("E")),
        ("Var Zi (%)", xlookup_pv("G"), "Var Săpt (%)", xlookup_pv("H")),
        ("Var Lună (%)", xlookup_pv("I"), "Semnal", xlookup_si("C")),
        ("Volum", xlookup_pv("J"), "RVOL", xlookup_pv("L")),
        ("Medie Vol 20z", xlookup_pv("K"), "Trend", xlookup_pv("O")),
    ]
    for l1, f1, l2, f2 in pv_pairs:
        row = pair_row(row, l1, f1, l2, f2)

    row += 1

    # 4. INDICATORI MACRO
    row = section_header(row, "4. INDICATORI MACRO")
    macro_indicators = ["VIX", "Yield 10Y US (%)", "USD Index (DXY)", "EUR/USD",
                        "Fear & Greed (0-100)", "CPI YoY (%)", "Rata dobânzii (%)",
                        "PIB YoY (%)"]
    for mi in macro_indicators:
        f_val = f'=IFERROR(XLOOKUP("{mi}",\'INDICATORI MACRO\'!$A:$A,\'INDICATORI MACRO\'!$B:$B,""),"")'
        f_status = f'=IFERROR(XLOOKUP("{mi}",\'INDICATORI MACRO\'!$A:$A,\'INDICATORI MACRO\'!$L:$L,""),"")'
        safe_write(ws, row, 1, mi, font=FONT_NORMAL, alignment=ALIGN_LEFT)
        ws.merge_cells(f"A{row}:D{row}")
        safe_write(ws, row, 5, f_val, font=FONT_NORMAL)
        ws.merge_cells(f"E{row}:F{row}")
        safe_write(ws, row, 7, f_status, font=FONT_NORMAL)
        ws.merge_cells(f"G{row}:H{row}")
        ws.row_dimensions[row].height = 18
        row += 1

    row += 1

    # 5-7: Reference sections
    for sec_num, sec_title, sec_note in [
        (5, "COMPETITORI & SECTOR", "Consultați sheet-ul COMPETITORI SECTOR pentru detalii complete"),
        (6, "RISCURI & OPORTUNITĂȚI", "Consultați sheet-ul RISCURI OPORTUNITATI pentru detalii complete"),
        (7, "CALENDAR ECONOMIC", "Consultați sheet-ul CALENDAR ECONOMIC pentru detalii complete"),
    ]:
        row = section_header(row, f"{sec_num}. {sec_title}")
        safe_write(ws, row, 1, sec_note, font=Font(name="Arial", italic=True, size=10, color="595959"),
                   alignment=ALIGN_LEFT)
        ws.merge_cells(f"A{row}:H{row}")
        ws.row_dimensions[row].height = 18
        row += 2

    # 8. JURNAL TRANZACTII — statistici filtrate
    row = section_header(row, "8. JURNAL TRANZACȚII — STATISTICI ACTIVUL SELECTAT")
    jurnal_stats = [
        ("Nr. tranzacții activ", f'=IFERROR(COUNTIF(\'JURNAL TRANZACTII\'!D:D,$D$2),"0")'),
        ("Win Rate activ", f'=IFERROR(COUNTIFS(\'JURNAL TRANZACTII\'!D:D,$D$2,\'JURNAL TRANZACTII\'!N:N,">0")/COUNTIF(\'JURNAL TRANZACTII\'!D:D,$D$2),"N/A")'),
        ("P&L Total activ", f'=IFERROR(SUMIF(\'JURNAL TRANZACTII\'!D:D,$D$2,\'JURNAL TRANZACTII\'!N:N),"N/A")'),
        ("P&L Mediu activ", f'=IFERROR(AVERAGEIF(\'JURNAL TRANZACTII\'!D:D,$D$2,\'JURNAL TRANZACTII\'!N:N),"N/A")'),
    ]
    for label, formula in jurnal_stats:
        safe_write(ws, row, 1, label, font=FONT_SUBHEADER, fill=FILL_SUBHEADER, alignment=ALIGN_LEFT)
        ws.merge_cells(f"A{row}:D{row}")
        safe_write(ws, row, 5, formula, font=FONT_NORMAL)
        ws.merge_cells(f"E{row}:H{row}")
        ws.row_dimensions[row].height = 18
        row += 1

    row += 1

    # 9-10: Reference sections
    for sec_num, sec_title, sec_note in [
        (9, "REZUMAT EXECUTIV", "Consultați sheet-ul REZUMAT EXECUTIV pentru analiza completă de piață"),
        (10, "ISTORIC & TRENDING", "Consultați sheet-ul ISTORIC TRENDING pentru snapshot-urile lunare"),
    ]:
        row = section_header(row, f"{sec_num}. {sec_title}")
        safe_write(ws, row, 1, sec_note, font=Font(name="Arial", italic=True, size=10, color="595959"),
                   alignment=ALIGN_LEFT)
        ws.merge_cells(f"A{row}:H{row}")
        row += 2

    safe_write(ws, row, 1, "⚙️ Toate datele se actualizează automat la schimbarea activului din DASHBOARD → celula portocalie I2",
               font=Font(name="Arial", italic=True, size=9, color="595959"), alignment=ALIGN_LEFT)
    ws.merge_cells(f"A{row}:H{row}")

    set_col_widths(ws, {"A": 16, "B": 12, "C": 16, "D": 16, "E": 16, "F": 12, "G": 16, "H": 16})
    ws.freeze_panes = "A3"
    log.info("FISA ACTIV: actualizat cu formule XLOOKUP")


# ═══════════════════════════════════════════════════════════════════════════════
# GHID INVATARE — TEXT GENERATORS
# ═══════════════════════════════════════════════════════════════════════════════

def text_why_moved(d):
    """Generate 'DE CE S-A MIȘCAT ASTĂZI' text."""
    lines = []
    price = d.get("price", "")
    var_zi = d.get("var_zi", "")
    var_sapt = d.get("var_sapt", "")
    var_luna = d.get("var_luna", "")
    rvol = d.get("rvol", "")
    rsi = d.get("rsi", "")
    macd_status = d.get("macd_status", "")
    ma_cross = d.get("ma_cross", "")
    bb_width = d.get("bb_width", "")
    stoch_k = d.get("stoch_k", "")
    trend = d.get("trend", "")

    if var_zi != "" and var_zi is not None:
        try:
            vz = float(var_zi)
            if abs(vz) > 3:
                intensity = "semnificativ"
            elif abs(vz) > 1:
                intensity = "moderat"
            else:
                intensity = "ușor"
            direction = "crescut" if vz > 0 else "scăzut"
            lines.append(f"• Prețul a {direction} {intensity} cu {abs(vz):.2f}% astăzi.")
        except (ValueError, TypeError):
            pass

    if var_sapt != "" and var_sapt is not None:
        try:
            lines.append(f"  Variație săptămânală: {float(var_sapt):+.2f}%")
        except (ValueError, TypeError):
            pass

    if var_luna != "" and var_luna is not None:
        try:
            lines.append(f"  Variație lunară: {float(var_luna):+.2f}%")
        except (ValueError, TypeError):
            pass

    if rvol != "" and rvol is not None:
        try:
            rv = float(rvol)
            if rv > 1.5:
                lines.append(f"• Volum EXCEPȚIONAL ({rv:.2f}x medie) — interes instituțional ridicat.")
            elif rv > 1.0:
                lines.append(f"• Volum peste medie ({rv:.2f}x) — participare activă pe piață.")
            elif rv < 0.7:
                lines.append(f"• Volum SCĂZUT ({rv:.2f}x medie) — lipsa de interes, atenție la breakout-uri false.")
            else:
                lines.append(f"• Volum normal ({rv:.2f}x medie).")
        except (ValueError, TypeError):
            pass

    if rsi != "" and rsi is not None:
        try:
            rsi_v = float(rsi)
            rsi_s = d.get("rsi_status", "")
            lines.append(f"• RSI(14) = {rsi_v:.1f} → {rsi_s}")
        except (ValueError, TypeError):
            pass

    if macd_status and macd_status != "N/A":
        lines.append(f"• MACD: {macd_status}")

    if ma_cross and ma_cross != "Neutru":
        ma50 = d.get("ma50", "")
        ma200 = d.get("ma200", "")
        if ma50 != "" and ma200 != "":
            try:
                dist = abs(float(ma50) - float(ma200)) / float(ma200) * 100
                lines.append(f"• {ma_cross} — distanță MA50 vs MA200: {dist:.1f}%")
            except (ValueError, TypeError, ZeroDivisionError):
                lines.append(f"• {ma_cross}")
        else:
            lines.append(f"• {ma_cross}")

    if bb_width != "" and bb_width is not None and price != "" and price is not None:
        try:
            bw = float(bb_width)
            p = float(price)
            if p > 0:
                bb_pct = bw / p * 100
                if bb_pct < 4:
                    lines.append(f"• Bollinger Squeeze detectat (lățime {bb_pct:.1f}% din preț) — breakout iminent.")
                else:
                    lines.append(f"• Bollinger Bands: lățime {bb_pct:.1f}% din preț.")
        except (ValueError, TypeError, ZeroDivisionError):
            pass

    if stoch_k != "" and stoch_k is not None:
        try:
            sk = float(stoch_k)
            lines.append(f"• Stochastic %K = {sk:.1f}")
        except (ValueError, TypeError):
            pass

    if trend:
        lines.append(f"• Concluzie: Trend {trend}.")

    return "\n".join(lines) if lines else "Date insuficiente pentru analiză."


def text_opportunity(d):
    """Generate 'OPORTUNITATE DE TRADING' text."""
    lines = []
    signal = d.get("signal", "WAIT")
    entry = d.get("entry", "")
    sl = d.get("sl", "")
    tp = d.get("tp", "")
    prob = d.get("probability", "")
    rvol = d.get("rvol", "")

    if signal in ("BUY", "SELL"):
        lines.append(f"SEMNAL: {signal}")
        if entry != "":
            lines.append(f"  Entry: {entry}")
        if sl != "":
            lines.append(f"  Stop Loss: {sl}")
            if entry != "" and signal == "BUY":
                try:
                    risk_pct = abs(float(entry) - float(sl)) / float(entry) * 100
                    lines.append(f"  Risc: {risk_pct:.2f}%")
                except (ValueError, TypeError, ZeroDivisionError):
                    pass
        if tp != "":
            lines.append(f"  Take Profit: {tp}")
            if entry != "" and signal == "BUY":
                try:
                    reward_pct = abs(float(tp) - float(entry)) / float(entry) * 100
                    lines.append(f"  Câștig potențial: {reward_pct:.2f}%")
                except (ValueError, TypeError, ZeroDivisionError):
                    pass
        if entry != "" and sl != "" and tp != "":
            try:
                rr = abs(float(tp) - float(entry)) / abs(float(entry) - float(sl))
                rating = "Excelent" if rr >= 3 else ("Bun" if rr >= 2 else ("Acceptabil" if rr >= 1.5 else "Slab"))
                lines.append(f"  RR = {rr:.2f}x — Rating: {rating}")
            except (ValueError, TypeError, ZeroDivisionError):
                pass
        if prob != "":
            try:
                lines.append(f"  Probabilitate estimată: {float(prob) * 100:.0f}%")
            except (ValueError, TypeError):
                pass
        if rvol != "":
            try:
                rv = float(rvol)
                if rv > 1.5:
                    lines.append("  ✓ Confirmare volum: RVOL > 1.5x — semnal puternic!")
            except (ValueError, TypeError):
                pass
        if signal == "SELL":
            lines.append("  ⚠️ WARNING: Pozițiile SHORT au risc de pierdere nelimitată teoretic.")
    else:
        lines.append("SEMNAL: WAIT — Nu există semnal valid de intrare.")
        lines.append("Ce să urmărești pentru un semnal valid:")
        lines.append("  • RSI sub 35 sau peste 75")
        lines.append("  • MACD Impuls pozitiv/negativ NOU")
        lines.append("  • Golden/Death Cross pe MA50 vs MA200")
        lines.append("  • RVOL > 1.5x pentru confirmare volum")
        lines.append("  • Confluențe ≥ 3 pentru probabilitate ridicată")

    return "\n".join(lines)


def text_pattern(d):
    """Generate 'PATTERN GRAFIC DETECTAT' text."""
    lines = []
    patterns_found = False

    ma_cross = d.get("ma_cross", "")
    ma50 = d.get("ma50", "")
    ma200 = d.get("ma200", "")
    if ma_cross in ("Golden Cross", "Death Cross"):
        try:
            dist = abs(float(ma50) - float(ma200)) / float(ma200) * 100
            lines.append(f"✦ {ma_cross} detectat — distanță MA50 vs MA200: {dist:.1f}%")
            patterns_found = True
        except (ValueError, TypeError, ZeroDivisionError):
            lines.append(f"✦ {ma_cross} detectat")
            patterns_found = True

    bb_width = d.get("bb_width", "")
    price = d.get("price", "")
    if bb_width != "" and price != "":
        try:
            bw_pct = float(bb_width) / float(price) * 100
            if bw_pct < 4:
                lines.append(f"✦ Bollinger Squeeze (lățime {bw_pct:.1f}%) — strategie pending orders ambele direcții.")
                patterns_found = True
        except (ValueError, TypeError, ZeroDivisionError):
            pass

    macd_status = d.get("macd_status", "")
    if "nou" in str(macd_status).lower():
        hist = d.get("macd_hist", "")
        lines.append(f"✦ {macd_status} — histogramă MACD: {hist}")
        patterns_found = True

    rsi = d.get("rsi", "")
    var_zi = d.get("var_zi", "")
    rvol = d.get("rvol", "")
    try:
        rsi_v = float(rsi) if rsi != "" else None
        vz_v = float(var_zi) if var_zi != "" else None
        rv_v = float(rvol) if rvol != "" else None
        if rsi_v and vz_v and rv_v:
            if rsi_v < 32 and vz_v < -2 and rv_v > 1.3:
                lines.append("✦ CLIMAX DE VÂNZARE detectat (RSI<32 + var.zi<-2% + RVOL>1.3)")
                patterns_found = True
            elif rsi_v > 75 and vz_v > 2 and rv_v > 1.3:
                lines.append("✦ CLIMAX DE CUMPĂRARE detectat (RSI>75 + var.zi>2% + RVOL>1.3)")
                patterns_found = True
    except (ValueError, TypeError):
        pass

    stoch_k = d.get("stoch_k", "")
    stoch_d = d.get("stoch_d", "")
    try:
        sk = float(stoch_k) if stoch_k != "" else None
        sd = float(stoch_d) if stoch_d != "" else None
        if sk and sd:
            if sk < 20 and sk > sd:
                lines.append("✦ Stochastic crossover bullish din zona oversold (<20)")
                patterns_found = True
            elif sk > 80 and sk < sd:
                lines.append("✦ Stochastic crossover bearish din zona overbought (>80)")
                patterns_found = True
    except (ValueError, TypeError):
        pass

    ma50_v = d.get("ma50", "")
    if price != "" and ma50_v != "":
        try:
            dist_ma50 = abs(float(price) - float(ma50_v)) / float(ma50_v) * 100
            if dist_ma50 < 0.8:
                lines.append(f"✦ Test MA50 — prețul la {dist_ma50:.1f}% de MA50")
                patterns_found = True
        except (ValueError, TypeError, ZeroDivisionError):
            pass

    if not patterns_found:
        lines.append("Niciun pattern major detectat — consolidare, monitorizează zilnic.")

    return "\n".join(lines)


def text_lesson(d):
    """Generate 'LECȚIA ZILEI' adaptive text."""
    lines = []
    signal = d.get("signal", "WAIT")
    conf = d.get("confluences", 0)
    rsi = d.get("rsi", "")
    rvol = d.get("rvol", "")
    ma_cross = d.get("ma_cross", "")

    if signal == "WAIT" and (isinstance(conf, int) and conf < 2):
        lines.append("📚 LECȚIE: Răbdarea ca strategie de trading")
        lines.append("Cei mai buni traderi petrec 80% din timp AȘTEPTÂND.")
        lines.append("Un setup cu confluențe < 2 nu merită riscul.")
        lines.append("Regula: Nu forța trade-uri. Piața va fi acolo și mâine.")

    try:
        entry = d.get("entry", "")
        sl = d.get("sl", "")
        tp = d.get("tp", "")
        if entry != "" and sl != "" and tp != "":
            rr = abs(float(tp) - float(entry)) / abs(float(entry) - float(sl))
            if rr < 1.5:
                lines.append("📚 LECȚIE: De ce RR < 1.5x distruge contul")
                lines.append("  Win Rate 50% + RR 1.0x = Break-even (comisioane = pierdere)")
                lines.append("  Win Rate 50% + RR 2.0x = Profitabil pe termen lung")
                lines.append("  Win Rate 40% + RR 3.0x = Foarte profitabil")
                lines.append("  Concluzie: RR minim acceptabil = 1.5x, ideal ≥ 2.0x")
    except (ValueError, TypeError, ZeroDivisionError):
        pass

    try:
        rv = float(rvol) if rvol != "" else None
        if rv and rv < 0.7:
            lines.append("📚 LECȚIE: Volumul ca confirmare")
            lines.append("Un breakout fără volum este un breakout FALS în >60% din cazuri.")
            lines.append("RVOL < 0.7x = lipsă de convingere pe piață.")
            lines.append("Așteaptă RVOL > 1.2x pentru confirmare validă.")
    except (ValueError, TypeError):
        pass

    try:
        rsi_v = float(rsi) if rsi != "" else None
        if rsi_v and rsi_v < 30:
            lines.append("📚 LECȚIE: Presiune excesivă vânzare ≠ Cumpărare imediată")
            lines.append("RSI < 30 indică presiune excesivă, NU semnal automat de cumpărare.")
            lines.append("Într-un bear market, RSI poate rămâne sub 30 săptămâni întregi.")
            lines.append("Caută DIVERGENȚĂ BULLISH: preț face low nou, RSI face low mai sus.")
        elif rsi_v and rsi_v > 70:
            lines.append("📚 LECȚIE: RSI > 70 în bull market")
            lines.append("În trend-uri puternice, RSI poate rămâne peste 70 luni de zile.")
            lines.append("Nu vinde doar pentru că RSI este 'overbought'.")
            lines.append("Caută: scăderea SUB 70 = potențial semnal de vânzare.")
    except (ValueError, TypeError):
        pass

    if ma_cross == "Golden Cross":
        lines.append("📚 LECȚIE: Golden Cross — anticipare vs reacție")
        lines.append("Golden Cross este un indicator TARDIV — confirmă trend-ul, nu îl prezice.")
        lines.append("Când apare, mișcarea a avansat deja 15-30%.")
        lines.append("Utilizare optimă: CONFIRMAREA unui trend existent, nu ca singur semnal.")
    elif ma_cross == "Death Cross":
        lines.append("📚 LECȚIE: Death Cross — eroarea clasică")
        lines.append("Death Cross apare DE OBICEI la MIJLOCUL corecției.")
        lines.append("Vânzarea la Death Cross = vânzare la niveluri deja scăzute.")
        lines.append("Alternativă: folosește-l ca FILTRU (nu tranzacționa long cât există Death Cross).")

    lines.append("")
    lines.append("🧠 PSIHOLOGIE: Loss Aversion (Daniel Kahneman)")
    lines.append("Durerea unei pierderi este de 2.5x mai puternică decât plăcerea unui câștig egal.")
    lines.append("De aceea traderii: (1) închid profiturile prea devreme, (2) țin pierderile prea mult.")
    lines.append("Soluția: SL și TP predefinite ÎNAINTE de intrare. Zero emoție la execuție.")

    return "\n".join(lines)


# ═══════════════════════════════════════════════════════════════════════════════
# GHID INVATARE — SHEET UPDATER
# ═══════════════════════════════════════════════════════════════════════════════

def upd_ghid_invatare(wb, all_data):
    """Update GHID INVATARE sheet."""
    ws = wb["GHID INVATARE"]
    unmerge_all(ws)

    set_col_widths(ws, {"A": 60, "B": 50, "C": 28})

    row = 1
    safe_write(ws, row, 1, "GHID DE ÎNVĂȚARE ZILNIC — ANALIZĂ DE PIAȚĂ PROFESIONALĂ",
               font=Font(name="Arial", bold=True, color="FFFFFF", size=14),
               fill=PatternFill("solid", fgColor="0D2137"), alignment=ALIGN_CENTER)
    safe_write(ws, row, 2, "", fill=PatternFill("solid", fgColor="0D2137"))
    safe_write(ws, row, 3, "", fill=PatternFill("solid", fgColor="0D2137"))
    ws.merge_cells("A1:C1")
    ws.row_dimensions[row].height = 35
    row += 1

    buy_count = sum(1 for d in all_data.values() if d.get("signal") == "BUY")
    sell_count = sum(1 for d in all_data.values() if d.get("signal") == "SELL")
    wait_count = sum(1 for d in all_data.values() if d.get("signal") == "WAIT")
    stats_text = f"Data: {NOW.strftime('%d.%m.%Y %H:%M')} | Active: {len(all_data)} | BUY: {buy_count} | SELL: {sell_count} | WAIT: {wait_count}"
    safe_write(ws, row, 1, stats_text,
               font=Font(name="Arial", size=10, color="AED6F1"),
               fill=PatternFill("solid", fgColor="154360"), alignment=ALIGN_CENTER)
    safe_write(ws, row, 2, "", fill=PatternFill("solid", fgColor="154360"))
    safe_write(ws, row, 3, "", fill=PatternFill("solid", fgColor="154360"))
    ws.merge_cells(f"A{row}:C{row}")
    ws.row_dimensions[row].height = 22
    row += 1

    # Per-asset analysis sections
    for cat_name, cat_dict in ALL_CATEGORIES:
        safe_write(ws, row, 1, f"  ══════  {cat_name}  ══════",
                   font=Font(name="Arial", bold=True, color="FFFFFF", size=12),
                   fill=PatternFill("solid", fgColor="1F4E79"), alignment=ALIGN_CENTER)
        safe_write(ws, row, 2, "", fill=PatternFill("solid", fgColor="1F4E79"))
        safe_write(ws, row, 3, "", fill=PatternFill("solid", fgColor="1F4E79"))
        ws.merge_cells(f"A{row}:C{row}")
        ws.row_dimensions[row].height = 28
        row += 1

        for aname in cat_dict:
            d = all_data.get(aname, {})
            signal = d.get("signal", "WAIT")
            if signal == "BUY":
                banner_fill = PatternFill("solid", fgColor=C_BUY_BG)
                banner_font = Font(name="Arial", bold=True, color=C_BUY_FG, size=11)
            elif signal == "SELL":
                banner_fill = PatternFill("solid", fgColor=C_SELL_BG)
                banner_font = Font(name="Arial", bold=True, color=C_SELL_FG, size=11)
            else:
                banner_fill = PatternFill("solid", fgColor=C_WAIT_BG)
                banner_font = Font(name="Arial", bold=True, color=C_WAIT_FG, size=11)

            conf = d.get("confluences", 0)
            price_str = d.get("price", "N/A")
            var_zi_str = d.get("var_zi", "N/A")
            rsi_str = d.get("rsi", "N/A")
            banner_text = f"{aname} | Preț: {price_str} | Var.zi: {var_zi_str}% | RSI: {rsi_str} | Conf: {conf}/5 | {signal}"

            safe_write(ws, row, 1, banner_text, font=banner_font, fill=banner_fill, alignment=ALIGN_LEFT)
            safe_write(ws, row, 2, "", fill=banner_fill)
            safe_write(ws, row, 3, "", fill=banner_fill)
            ws.merge_cells(f"A{row}:C{row}")
            ws.row_dimensions[row].height = 24
            row += 1

            sections = [
                ("A. DE CE S-A MIȘCAT ASTĂZI", text_why_moved(d)),
                ("B. OPORTUNITATE DE TRADING", text_opportunity(d)),
                ("C. PATTERN GRAFIC DETECTAT", text_pattern(d)),
                ("D. LECȚIA ZILEI", text_lesson(d)),
            ]

            for sec_title, sec_text in sections:
                safe_write(ws, row, 1, sec_title,
                           font=Font(name="Arial", bold=True, color="1F4E79", size=10),
                           fill=FILL_SUBHEADER, alignment=ALIGN_LEFT)
                safe_write(ws, row, 2, "", fill=FILL_SUBHEADER)
                safe_write(ws, row, 3, "", fill=FILL_SUBHEADER)
                ws.merge_cells(f"A{row}:C{row}")
                ws.row_dimensions[row].height = 20
                row += 1

                for line in sec_text.split("\n"):
                    if line.strip():
                        safe_write(ws, row, 1, line,
                                   font=FONT_GHID_BODY,
                                   fill=PatternFill("solid", fgColor="FFFFFF"),
                                   alignment=ALIGN_LEFT)
                        safe_write(ws, row, 2, "",
                                   fill=PatternFill("solid", fgColor="FFFFFF"))
                        safe_write(ws, row, 3, "",
                                   fill=PatternFill("solid", fgColor="FFFFFF"))
                        ws.merge_cells(f"A{row}:C{row}")
                        ws.row_dimensions[row].height = 16
                        row += 1

            # Thin separator between assets
            safe_write(ws, row, 1, "", fill=PatternFill("solid", fgColor="1F4E79"))
            safe_write(ws, row, 2, "", fill=PatternFill("solid", fgColor="1F4E79"))
            safe_write(ws, row, 3, "", fill=PatternFill("solid", fgColor="1F4E79"))
            ws.merge_cells(f"A{row}:C{row}")
            ws.row_dimensions[row].height = 4
            row += 1

        row += 1

    # ═══════════════════════════════════════════════════════════════════
    # GHID PERMANENT — 8 CAPITOLE
    # ═══════════════════════════════════════════════════════════════════
    row += 1
    safe_write(ws, row, 1, "GHID COMPLET DE CITIRE GRAFICE & INDICATORI TEHNICI",
               font=Font(name="Arial", bold=True, color="FFFFFF", size=14),
               fill=PatternFill("solid", fgColor="0A1628"), alignment=ALIGN_CENTER)
    safe_write(ws, row, 2, "", fill=PatternFill("solid", fgColor="0A1628"))
    safe_write(ws, row, 3, "", fill=PatternFill("solid", fgColor="0A1628"))
    ws.merge_cells(f"A{row}:C{row}")
    ws.row_dimensions[row].height = 35
    row += 1

    chapters = [
        ("1. CANDLESTICK — Limbajul Graficelor", [
            "Ce este un candlestick: Fiecare lumânare arată 4 prețuri — Open, High, Low, Close.",
            "Corpul = diferența între Open și Close. Umbrele (wicks) = extremele zilei.",
            "Lumânare verde/albă = Close > Open (bullish). Lumânare roșie/neagră = Close < Open (bearish).",
            "",
            "Pattern-uri esențiale:",
            "• Hammer — corp mic + umbra inferioară lungă (2x corp). Apare în downtrend = potențial reversal bullish.",
            "• Shooting Star — corp mic + umbra superioară lungă. Apare în uptrend = potențial reversal bearish.",
            "• Doji — Open ≈ Close (corp foarte mic). Semnifică indecizija pieței.",
            "• Bullish Engulfing — lumânare verde 'înghite' complet lumânarea roșie anterioară.",
            "• Bearish Engulfing — lumânare roșie 'înghite' complet lumânarea verde anterioară.",
            "",
            "Cum aplici: NICIODATĂ nu acționa doar pe un singur candlestick!",
            "Confirmă cu: volum, RSI, nivel S/R. Un hammer la suport + RVOL>1.2x = semnal puternic.",
        ]),
        ("2. RSI — Relative Strength Index", [
            "Ce măsoară: Viteza și magnitudinea variațiilor de preț pe ultimele 14 perioade.",
            "Formula: RSI = 100 - (100 / (1 + RS)), unde RS = Medie Gain / Medie Loss.",
            "",
            "Zone profesionale:",
            "  RSI < 30: Presiune excesivă vânzare — potențial oversold.",
            "  RSI 30-45: Presiune moderată vânzare — piața în teritoriu slab.",
            "  RSI 45-55: Echilibru — nicio direcție dominantă.",
            "  RSI 55-70: Momentum ascendent — forță cumpărători.",
            "  RSI > 70: Presiune excesivă cumpărare — potențial overbought.",
            "",
            "Divergență bullish: Prețul face low nou, RSI face low MAI SUS → forța vânzătorilor slăbește.",
            "Divergență bearish: Prețul face high nou, RSI face high MAI JOS → forța cumpărătorilor slăbește.",
            "",
            "Greșeli frecvente:",
            "• RSI < 30 ≠ 'cumpără imediat'. Într-un bear market, poate rămâne sub 30 săptămâni.",
            "• RSI > 70 ≠ 'vinde imediat'. Trend-urile puternice mențin RSI peste 70 luni întregi.",
            "• Cel mai important semnal: revenirea PRIN 30 sau scăderea PRIN 70.",
        ]),
        ("3. MACD — Moving Average Convergence Divergence", [
            "Componente: Linia MACD (EMA12 - EMA26), Linia Signal (EMA9 a MACD), Histograma (MACD - Signal).",
            "",
            "Semnale principale:",
            "• Impuls pozitiv NOU: MACD traversează Signal de jos în sus + histograma trece în pozitiv.",
            "  Cel mai puternic semnal MACD bullish — moment optim de intrare.",
            "• Impuls pozitiv ACTIV: MACD > Signal, histograma deja pozitivă. Trend bullish în desfășurare.",
            "• Impuls negativ NOU: MACD traversează Signal de sus în jos + histograma trece în negativ.",
            "  Cel mai puternic semnal MACD bearish — moment optim de ieșire sau short.",
            "• Impuls negativ ACTIV: MACD < Signal, histograma deja negativă. Trend bearish în desfășurare.",
            "",
            "Cel mai puternic semnal combinat: MACD Impuls pozitiv nou + RSI revine prin 30 + RVOL > 1.5x",
            "",
            "Limitări: MACD este un indicator TARDIV (lagging) — confirmă, nu prezice.",
            "Nu funcționează bine în piețe laterale (sideways) — generează semnale false.",
        ]),
        ("4. MEDII MOBILE — MA20, MA50, MA200", [
            "MA20: Media ultimelor 20 zile — trend pe termen SCURT.",
            "MA50: Media ultimelor 50 zile — trend pe termen MEDIU. Cel mai important nivel dinamic de S/R.",
            "MA200: Media ultimelor 200 zile — trend pe termen LUNG. 'Linia care separă bull de bear market'.",
            "",
            "Regulile esențiale:",
            "• Preț > MA50 > MA200 = Bull market confirmat.",
            "• Preț < MA50 < MA200 = Bear market confirmat.",
            "• Golden Cross (MA50 traversează MA200 în sus) = semnal bullish pe termen lung.",
            "• Death Cross (MA50 traversează MA200 în jos) = semnal bearish pe termen lung.",
            "",
            "Utilizare practică: Pullback la MA50 în uptrend = oportunitate de cumpărare.",
            "Dacă prețul 'bounce' de pe MA50 cu volum crescut = confirmare trend bullish continuă.",
        ]),
        ("5. BOLLINGER BANDS", [
            "Structura: Banda medie (SMA20) + Banda superioară (SMA20 + 2σ) + Banda inferioară (SMA20 - 2σ).",
            "Statistic, 95% din prețuri se află între benzi.",
            "",
            "Strategia Bounce:",
            "• Prețul atinge banda inferioară = potențial BUY (dacă trend este bullish).",
            "• Prețul atinge banda superioară = potențial SELL (dacă trend este bearish).",
            "• IMPORTANT: Nu funcționează în trend-uri puternice — prețul 'walks the band'.",
            "",
            "Strategia Squeeze Breakout:",
            "• Când benzile se îngustează (lățime < 4% din preț) = volatilitate la minim.",
            "• Breakout INEVITABIL urmează — direcția se determină din alt indicator (RSI, MACD).",
            "• Plasează pending orders: BUY la banda superioară, SELL la banda inferioară.",
        ]),
        ("6. VOLUM & RVOL — Relative Volume", [
            "De ce contează volumul: Prețul arată CE se întâmplă, volumul arată CU CÂTĂ CONVINGERE.",
            "",
            "Scala RVOL (volum zilnic / medie 20 zile):",
            "• RVOL < 0.5x: Volum foarte scăzut — evită tranzacționarea, spread-uri mari.",
            "• RVOL 0.5-0.8x: Sub medie — semnale nesigure, breakout-uri false frecvente.",
            "• RVOL 0.8-1.2x: Normal — condiții standard de trading.",
            "• RVOL 1.2-1.5x: Peste medie — interes crescut, semnale mai fiabile.",
            "• RVOL > 1.5x: Excepțional — potențial mișcare instituțională, semnale puternice.",
            "• RVOL > 3.0x: Extreme — probabil știre majoră, atenție la gap-uri.",
            "",
            "Regulile de aur:",
            "• Breakout + RVOL > 1.5x = breakout VALID (probabilitate >70%).",
            "• Breakout + RVOL < 0.7x = breakout FALS (probabilitate >60%).",
            "• Trend up + volum crescut = sănătos. Trend up + volum scăzut = slăbiciune.",
        ]),
        ("7. SUPORT & REZISTENȚĂ", [
            "Definiție și logică:",
            "• Suport = nivel unde cererea depășește oferta → prețul tinde să revină în sus.",
            "• Rezistență = nivel unde oferta depășește cererea → prețul tinde să revină în jos.",
            "",
            "Cum identifici niveluri puternice:",
            "• Prețul a 'bounce' de pe nivel de multiple ori (minim 2-3 atingeri).",
            "• Volumul crescut la nivel = confirmare importanță.",
            "• Numere rotunde (1000, 5000, 100.00) = niveluri psihologice.",
            "• MA50 și MA200 = suport/rezistență dinamic.",
            "",
            "Role Reversal: Când suportul este spart, devine rezistență (și invers).",
            "Aceasta este una dintre cele mai importante reguli în analiza tehnică.",
            "",
            "Stop Loss optim: NU exact la nivelul de suport/rezistență!",
            "• Pentru LONG: SL sub suport cu 1-3% (lasă spațiu pentru 'false break').",
            "• Pentru SHORT: SL peste rezistență cu 1-3%.",
            "• Alternativă: SL bazat pe ATR (1.5x ATR sub/peste nivel).",
        ]),
        ("8. MANAGEMENT RISC — Fundația Profitabilității", [
            "Regula 1-2%: Nu risca mai mult de 1-2% din cont pe o singură tranzacție.",
            "Cu un cont de $10,000: risc maxim per trade = $100-200.",
            "",
            "Formula Position Sizing:",
            "  Mărime poziție = Risc $ / (Entry - SL)",
            "  Exemplu: Risc $200, Entry $50, SL $48 → Poziție = $200 / $2 = 100 acțiuni.",
            "",
            "Tabel RR vs Win Rate (profit pe termen lung):",
            "  RR 1.0x: Ai nevoie de Win Rate > 50% (doar break-even cu comisioane).",
            "  RR 1.5x: Ai nevoie de Win Rate > 40%.",
            "  RR 2.0x: Ai nevoie de Win Rate > 33%. ← SWEET SPOT recomandat.",
            "  RR 3.0x: Ai nevoie de Win Rate > 25%. ← Foarte profitabil chiar cu WR scăzut.",
            "",
            "Cele 10 reguli de aur:",
            "1. Protejează capitalul — fără capital, fără oportunități.",
            "2. SL predefinit ÎNAINTE de fiecare trade.",
            "3. RR minim 1.5x, ideal 2.0x+.",
            "4. Nu muta SL în direcția pierderii.",
            "5. Trailing stop în trend-uri puternice.",
            "6. Nu tranzacționa din emoție — plan, execuție, jurnal.",
            "7. Maximum 1-2% risc per trade.",
            "8. Diversifică între clase de active.",
            "9. Marchează profit parțial la TP1 (50%), lasă restul cu trailing stop.",
            "10. Jurnalul de tranzacții este cel mai important instrument de învățare.",
        ]),
    ]

    for chapter_title, chapter_lines in chapters:
        safe_write(ws, row, 1, chapter_title,
                   font=FONT_GHID_HDR, fill=FILL_GHID_HDR, alignment=ALIGN_LEFT)
        safe_write(ws, row, 2, "", fill=FILL_GHID_HDR)
        safe_write(ws, row, 3, "", fill=FILL_GHID_HDR)
        ws.merge_cells(f"A{row}:C{row}")
        ws.row_dimensions[row].height = 26
        row += 1

        for line in chapter_lines:
            if line == "":
                ws.row_dimensions[row].height = 6
                row += 1
                continue

            if line.startswith("•") or line.startswith("  "):
                bg = "FFFFFF"
            else:
                bg = C_GHID_TEXT_BG

            safe_write(ws, row, 1, line,
                       font=FONT_GHID_BODY,
                       fill=PatternFill("solid", fgColor=bg),
                       alignment=ALIGN_LEFT)
            safe_write(ws, row, 2, "", fill=PatternFill("solid", fgColor=bg))
            safe_write(ws, row, 3, "", fill=PatternFill("solid", fgColor=bg))
            ws.merge_cells(f"A{row}:C{row}")
            ws.row_dimensions[row].height = 16
            row += 1

        row += 1

    ws.freeze_panes = "A3"
    log.info(f"GHID INVATARE: {row} rânduri scrise")


# ═══════════════════════════════════════════════════════════════════════════════
# LEGENDA SHEET
# ═══════════════════════════════════════════════════════════════════════════════

def upd_legenda(wb):
    """Update LEGENDA sheet."""
    ws = wb["LEGENDA"]
    unmerge_all(ws)

    safe_write(ws, 1, 1, "📋 LEGENDĂ — SISTEM DE CULORI ȘI ABREVIERI",
               font=FONT_TITLE, fill=FILL_TITLE, alignment=ALIGN_LEFT)
    for c in range(2, 8):
        safe_write(ws, 1, c, "", font=FONT_TITLE, fill=FILL_TITLE)
    ws.merge_cells("A1:G1")
    ws.row_dimensions[1].height = 30

    safe_write(ws, 3, 1, "Tip celulă", font=FONT_HEADER, fill=FILL_HEADER)
    safe_write(ws, 3, 2, "Fundal", font=FONT_HEADER, fill=FILL_HEADER)
    safe_write(ws, 3, 3, "Text", font=FONT_HEADER, fill=FILL_HEADER)
    safe_write(ws, 3, 4, "Semnificație", font=FONT_HEADER, fill=FILL_HEADER)
    ws.merge_cells("D3:G3")
    ws.row_dimensions[3].height = 20

    color_rows = [
        ("Celule editabile (input manual)", FILL_INPUT, FONT_NORMAL, "Date introduse manual de utilizator"),
        ("Calculat / formulă", FILL_CALC, FONT_CALC, "Valori calculate automat, nu editați"),
        ("BUY / Pozitiv / Bullish", FILL_BUY, FONT_BUY, "Semnal de cumpărare, trend ascendent"),
        ("WAIT / Neutru / Sideways", FILL_WAIT, FONT_WAIT, "Fără semnal, așteptare"),
        ("SELL / Negativ / Bearish", FILL_SELL, FONT_SELL, "Semnal de vânzare, trend descendent"),
        ("Alertă / Warning", FILL_ALERT, FONT_ALERT, "Necesită atenție specială"),
        ("Header coloane", FILL_HEADER, FONT_HEADER, "Titluri de coloană"),
        ("Header rânduri", FILL_SUBHEADER, FONT_SUBHEADER, "Etichete și subtitluri"),
        ("Titlu principal", FILL_TITLE, FONT_TITLE, "Titlul secțiunii/sheet-ului"),
        ("Separator categorie", FILL_SEP, FONT_SEP, "Delimitare între categorii"),
        ("Selector activ", FILL_SELECTOR, FONT_SELECTOR, "Celula dropdown pentru selecție"),
        ("Eticheta selectorului", FILL_LABEL_SEL, FONT_LABEL_SEL, "Label-ul selectorului"),
    ]

    row = 4
    for label, fill, font, desc in color_rows:
        safe_write(ws, row, 1, label, font=FONT_NORMAL, alignment=ALIGN_LEFT)
        safe_write(ws, row, 2, "█████", font=font, fill=fill, alignment=ALIGN_CENTER)
        safe_write(ws, row, 3, "Abc", font=font, fill=fill, alignment=ALIGN_CENTER)
        safe_write(ws, row, 4, desc, font=FONT_NORMAL, alignment=ALIGN_LEFT)
        ws.merge_cells(f"D{row}:G{row}")
        ws.row_dimensions[row].height = 20
        row += 1

    row += 2
    safe_write(ws, row, 1, "ABREVIERI UTILIZATE",
               font=FONT_HEADER, fill=FILL_HEADER)
    for c in range(2, 8):
        safe_write(ws, row, c, "", font=FONT_HEADER, fill=FILL_HEADER)
    ws.merge_cells(f"A{row}:G{row}")
    ws.row_dimensions[row].height = 22
    row += 1

    safe_write(ws, row, 1, "Abreviere", font=FONT_HEADER, fill=FILL_SUBHEADER)
    safe_write(ws, row, 2, "Semnificație", font=FONT_HEADER, fill=FILL_SUBHEADER)
    ws.merge_cells(f"B{row}:G{row}")
    ws.row_dimensions[row].height = 20
    row += 1

    abbreviations = [
        ("RSI", "Relative Strength Index — indicator momentum (0-100)"),
        ("MACD", "Moving Average Convergence Divergence — indicator trend/momentum"),
        ("MA20/50/200", "Moving Average — media mobilă pe 20/50/200 zile"),
        ("BB", "Bollinger Bands — benzi de volatilitate ±2σ"),
        ("ATR", "Average True Range — volatilitate medie pe 14 zile"),
        ("RVOL", "Relative Volume — volum curent / medie 20 zile"),
        ("SL", "Stop Loss — nivel de protecție pierdere"),
        ("TP", "Take Profit — nivel de încasare profit"),
        ("RR", "Risk/Reward Ratio — raportul câștig/risc"),
        ("S/R", "Support/Resistance — suport/rezistență"),
        ("EMA", "Exponential Moving Average — media mobilă exponențială"),
        ("SMA", "Simple Moving Average — media mobilă simplă"),
        ("VIX", "Volatility Index — indicele fricii pieței"),
        ("F&G", "Fear & Greed Index — indicele sentiment 0-100"),
        ("DXY", "US Dollar Index — forța dolarului american"),
        ("CPI", "Consumer Price Index — indicele prețurilor de consum"),
        ("PIB/GDP", "Produs Intern Brut — creștere economică"),
        ("PMI", "Purchasing Managers Index — index manageri achiziții"),
        ("FOMC", "Federal Open Market Committee — comitetul Fed"),
        ("YoY", "Year over Year — an peste an"),
        ("MoM", "Month over Month — lună peste lună"),
        ("P&L", "Profit & Loss — profit și pierdere"),
        ("DCA", "Dollar Cost Averaging — cumpărare medie periodică"),
        ("L/S", "Long/Short — direcția tranzacției"),
    ]

    for abbr, desc in abbreviations:
        safe_write(ws, row, 1, abbr, font=FONT_SUBHEADER, fill=FILL_SUBHEADER, alignment=ALIGN_LEFT)
        safe_write(ws, row, 2, desc, font=FONT_NORMAL, alignment=ALIGN_LEFT)
        ws.merge_cells(f"B{row}:G{row}")
        ws.row_dimensions[row].height = 18
        row += 1

    row += 2
    safe_write(ws, row, 1, "FORMULE UTILIZATE",
               font=FONT_HEADER, fill=FILL_HEADER)
    for c in range(2, 8):
        safe_write(ws, row, c, "", font=FONT_HEADER, fill=FILL_HEADER)
    ws.merge_cells(f"A{row}:G{row}")
    ws.row_dimensions[row].height = 22
    row += 1

    formulas = [
        ("Signal Score", "RSI(±2) + MACD(±2) + MA Cross(±2) + RVOL(±1) → BUY≥3, SELL≤-3, WAIT"),
        ("Stop Loss", "BUY: Entry - 1.5×ATR | SELL: Entry + 1.5×ATR"),
        ("Take Profit", "BUY: Entry + 3.0×ATR | SELL: Entry - 3.0×ATR"),
        ("RR Ratio", "(TP - Entry) / (Entry - SL)"),
        ("Probabilitate", "35% + Confluențe×10% + (RVOL>1.2 ? +5%) | max 90%"),
        ("Trend", "Preț > MA50×1.01 = Bullish | Preț < MA50×0.99 = Bearish | else Sideways"),
        ("Confluențe", "min(abs(Signal Score), 5) — nr indicatori care confirmă"),
    ]

    for fname, fdesc in formulas:
        safe_write(ws, row, 1, fname, font=FONT_SUBHEADER, fill=FILL_SUBHEADER, alignment=ALIGN_LEFT)
        safe_write(ws, row, 2, fdesc, font=FONT_NORMAL, alignment=ALIGN_LEFT)
        ws.merge_cells(f"B{row}:G{row}")
        ws.row_dimensions[row].height = 20
        row += 1

    set_col_widths(ws, {"A": 28, "B": 14, "C": 10, "D": 16, "E": 16, "F": 14, "G": 14})
    ws.freeze_panes = "A3"
    log.info("LEGENDA: actualizat")


# ═══════════════════════════════════════════════════════════════════════════════
# NOTIFICĂRI — WhatsApp & Google Drive (opțional)
# ═══════════════════════════════════════════════════════════════════════════════

def send_whatsapp(message):
    """Send WhatsApp notification via CallMeBot API (optional)."""
    if not WA_PHONE or not WA_APIKEY:
        log.debug("WhatsApp: nu este configurat (WA_PHONE/WA_APIKEY lipsă)")
        return False
    try:
        import urllib.parse
        encoded = urllib.parse.quote_plus(message)
        url = f"https://api.callmebot.com/whatsapp.php?phone={WA_PHONE}&text={encoded}&apikey={WA_APIKEY}"
        r = requests.get(url, timeout=30)
        if r.status_code == 200:
            log.info("WhatsApp: notificare trimisă cu succes")
            return True
        else:
            log.warning(f"WhatsApp: eroare HTTP {r.status_code}")
            return False
    except Exception as e:
        log.warning(f"WhatsApp: eroare trimitere — {e}")
        return False


def upload_gdrive(filepath):
    """Upload file to Google Drive via rclone (optional)."""
    if not os.path.isfile(RCLONE_EXE):
        log.debug(f"rclone: nu s-a găsit la {RCLONE_EXE}")
        return False
    try:
        cmd = [RCLONE_EXE, "copy", filepath, GDRIVE_DEST, "--progress"]
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=120)
        if result.returncode == 0:
            log.info(f"Google Drive: fișier încărcat → {GDRIVE_DEST}")
            return True
        else:
            log.warning(f"rclone eroare: {result.stderr[:200]}")
            return False
    except Exception as e:
        log.warning(f"Google Drive: eroare upload — {e}")
        return False


def cleanup_old_files(directory, pattern="Analiza_Piata_*.xlsx", max_files=MAX_FILES):
    """Keep only the last max_files analysis files."""
    try:
        files = sorted(glob.glob(os.path.join(directory, pattern)))
        if len(files) > max_files:
            to_delete = files[:-max_files]
            for f in to_delete:
                os.remove(f)
                log.info(f"Cleanup: șters {os.path.basename(f)}")
            log.info(f"Cleanup: păstrate ultimele {max_files} fișiere")
    except Exception as e:
        log.warning(f"Cleanup eroare: {e}")


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN — ORCHESTRARE COMPLETĂ
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    """Main orchestration function."""
    log.info("=" * 75)
    log.info("  ANALIZĂ DE PIAȚĂ PROFESIONALĂ — START")
    log.info(f"  Data: {NOW.strftime('%d.%m.%Y %H:%M')}")
    log.info("=" * 75)

    start_time = time.time()

    # ── 1. Pregătire template ──────────────────────────────────────────
    if os.path.isfile(TEMPLATE_PATH):
        log.info(f"Template găsit: {TEMPLATE_PATH}")
        shutil.copy2(TEMPLATE_PATH, EXCEL_PATH)
        log.info(f"Copiat → {EXCEL_PATH}")
    else:
        log.info("Template nu există, se creează unul nou...")
        create_template(EXCEL_PATH)

    # ── 2. Fetch date în paralel ──────────────────────────────────────
    log.info("Descărcare date yfinance (80+ active)...")
    raw_data = get_all_data_parallel(ALL_ASSETS, max_workers=12)
    log.info(f"Date primite: {len(raw_data)}/{len(ALL_ASSETS)} active")

    log.info("Descărcare date macro...")
    macro_raw = get_all_data_parallel(MACRO_TICKERS, max_workers=4)
    log.info(f"Macro date: {len(macro_raw)}/{len(MACRO_TICKERS)}")

    log.info("Descărcare Fear & Greed Index...")
    fg_val, fg_class = get_fear_greed()
    log.info(f"Fear & Greed: {fg_val} ({fg_class})")

    # ── 3. Procesare indicatori ───────────────────────────────────────
    log.info("Calcul indicatori tehnici pentru toate activele...")
    all_data = {}
    for name, df in raw_data.items():
        result = process_asset(name, df)
        all_data[name] = result

    macro_data = {}
    for name, df in macro_raw.items():
        result = process_asset(name, df)
        macro_data[name] = result

    buy_count = sum(1 for d in all_data.values() if d.get("signal") == "BUY")
    sell_count = sum(1 for d in all_data.values() if d.get("signal") == "SELL")
    wait_count = sum(1 for d in all_data.values() if d.get("signal") == "WAIT")
    log.info(f"Semnale: BUY={buy_count} | SELL={sell_count} | WAIT={wait_count}")

    # ── 4. Actualizare Excel ──────────────────────────────────────────
    log.info(f"Deschidere workbook: {EXCEL_PATH}")
    wb = load_workbook(EXCEL_PATH)

    log.info("Actualizare sheet-uri...")

    log.info("  → PRETURI VOLUME")
    upd_preturi_volume(wb, all_data)

    log.info("  → INDICATORI TEHNICI")
    upd_indicatori_tehnici(wb, all_data)

    log.info("  → SEMNALE INTRARE")
    upd_semnale_intrare(wb, all_data)

    log.info("  → INDICATORI MACRO")
    upd_indicatori_macro(wb, all_data, macro_data, fg_val, fg_class)

    log.info("  → COMPETITORI SECTOR")
    upd_competitori_sector(wb, all_data, macro_data)

    log.info("  → RISCURI OPORTUNITATI")
    upd_riscuri_oportunitati(wb)

    log.info("  → CALENDAR ECONOMIC")
    upd_calendar_economic(wb)

    log.info("  → JURNAL TRANZACTII")
    upd_jurnal_tranzactii(wb)

    log.info("  → ISTORIC TRENDING")
    upd_istoric_trending(wb, all_data, macro_data, fg_val)

    log.info("  → REZUMAT EXECUTIV")
    upd_rezumat_executiv(wb, all_data, macro_data, fg_val, fg_class)

    log.info("  → DASHBOARD")
    upd_dashboard(wb, all_data, macro_data, fg_val, fg_class)

    log.info("  → FISA ACTIV")
    upd_fisa_activ(wb, all_data, macro_data)

    log.info("  → GHID INVATARE")
    upd_ghid_invatare(wb, all_data)

    log.info("  → LEGENDA")
    upd_legenda(wb)

    # ── 5. Setare ordine sheet-uri & vizibilitate ─────────────────────
    desired_order = [
        "DASHBOARD", "REZUMAT EXECUTIV", "SEMNALE INTRARE", "INDICATORI TEHNICI",
        "INDICATORI MACRO", "COMPETITORI SECTOR", "PRETURI VOLUME",
        "RISCURI OPORTUNITATI", "CALENDAR ECONOMIC", "JURNAL TRANZACTII",
        "ISTORIC TRENDING", "GHID INVATARE", "LEGENDA", "FISA ACTIV", "LIST_ACTIVE",
    ]
    existing_sheets = wb.sheetnames
    new_order = []
    for sn in desired_order:
        if sn in existing_sheets:
            new_order.append(existing_sheets.index(sn))
    if len(new_order) == len(existing_sheets):
        wb.move_sheet(wb["DASHBOARD"], offset=0)  # Ensure DASHBOARD is first
    # Hide LIST_ACTIVE
    if "LIST_ACTIVE" in wb.sheetnames:
        wb["LIST_ACTIVE"].sheet_state = "hidden"

    # Set DASHBOARD as active
    wb.active = wb.sheetnames.index("DASHBOARD") if "DASHBOARD" in wb.sheetnames else 0

    # ── 6. Salvare ────────────────────────────────────────────────────
    log.info(f"Salvare workbook: {EXCEL_PATH}")
    wb.save(EXCEL_PATH)
    log.info("Workbook salvat cu succes!")

    # ── 7. Recalculare formule (opțional — necesită scripts/recalc.py) ─
    recalc_script = os.path.join(os.path.dirname(os.path.abspath(__file__)), "scripts", "recalc.py")
    if os.path.isfile(recalc_script):
        try:
            log.info("Recalculare formule cu recalc.py...")
            subprocess.run([sys.executable, recalc_script, EXCEL_PATH], timeout=60, check=True)
            log.info("Formule recalculate cu succes")
        except Exception as e:
            log.warning(f"Recalculare formule eșuată: {e}")

    # ── 8. Notificări (opțional) ──────────────────────────────────────
    if WA_PHONE and WA_APIKEY:
        summary = (
            f"📊 Analiză Piață — {NOW.strftime('%d.%m.%Y %H:%M')}\n"
            f"Active: {len(all_data)} | BUY: {buy_count} | SELL: {sell_count} | WAIT: {wait_count}\n"
            f"Fear & Greed: {fg_val} ({fg_class})\n"
            f"Fișier: {os.path.basename(EXCEL_PATH)}"
        )
        send_whatsapp(summary)

    if os.path.isfile(RCLONE_EXE):
        upload_gdrive(EXCEL_PATH)

    # ── 9. Cleanup fișiere vechi ──────────────────────────────────────
    cleanup_old_files(OUTPUT_DIR)

    # ── 10. Raport final ──────────────────────────────────────────────
    elapsed = time.time() - start_time
    log.info("=" * 75)
    log.info(f"  ANALIZĂ COMPLETĂ în {elapsed:.1f} secunde")
    log.info(f"  Fișier: {EXCEL_PATH}")
    log.info(f"  Active procesate: {len(all_data)}/{len(ALL_ASSETS)}")
    log.info(f"  Semnale: BUY={buy_count} | SELL={sell_count} | WAIT={wait_count}")
    log.info(f"  Fear & Greed: {fg_val} ({fg_class})")
    log.info(f"  Sheet-uri actualizate: 14")
    log.info("=" * 75)

    return EXCEL_PATH


if __name__ == "__main__":
    try:
        result_path = main()
        print(f"\n✅ Analiză completă! Fișier: {result_path}")
    except KeyboardInterrupt:
        log.info("Întrerupt de utilizator (Ctrl+C)")
        sys.exit(0)
    except Exception as e:
        log.error(f"Eroare critică: {e}", exc_info=True)
        sys.exit(1)
